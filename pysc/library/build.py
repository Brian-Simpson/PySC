"""The HTH control library: canonical registry of every audited control.

Identity model (per the program owner): a control is identified by WHAT IS
AUDITED — the canonical evaluated-item key derived from the check's substance
(policy field, registry key+item, PowerShell audit target...), NOT its
description text or CIS numbering. WHAT IS EXPECTED is recorded on the entry;
the same control checked by different mechanics (native PASSWORD_POLICY,
net accounts, secedit) is ONE library entry. Consequences:

- same key, same expectation, same file        -> exact DUPLICATE
- same key, different expectations             -> EXPECTATION VARIANCE (finding)
- key absent from the library                  -> NEW / UNKNOWN control

The library persists as control_library.json (tracked in git) and exports to
an Excel workbook for review.
"""

import json
import re
import time
from collections import defaultdict
from pathlib import Path

from pysc.gap.extract import _NIST_REF_RE, extract_control_number
from pysc.gap.harvest import extract_fields
from pysc.library._keys_core import (
    derive_control_keyword,
    derive_evaluated_item_key,
    derive_expected_value,
)
from pysc.parser import extract_variables
from pysc.platforms import PlatformMatcher

LIBRARY_NAME = "control_library.json"
MASTER_WORKBOOK_NAME = "Control_Library.xlsx"

_STRIP_INACTIVE_RE = re.compile(r"(?ms)^\s*#\s*<custom_item>.*?^\s*#\s*</custom_item>\s*$")
_BLOCK_RE = re.compile(r"(?ms)^(?!\s*#)\s*<(?:custom_item|item)>(.*?)^\s*</(?:custom_item|item)>[ \t]*$")
_TIMESTAMP_SUFFIX_RE = re.compile(r"_(\d{8,12})$")


def iter_active_checks(path):
    """Yield field dicts for every active check block in an audit file.

    Tenable @VARIABLE@ placeholders in value_data are resolved to their
    declared defaults so expectations compare meaningfully across raw vendor
    files and resolved baselines.
    """
    text = Path(path).read_text(encoding="utf-8", errors="ignore")
    try:
        variables = extract_variables(text.splitlines())
    except Exception:
        variables = {}
    text = _STRIP_INACTIVE_RE.sub("", text)
    for match in _BLOCK_RE.finditer(text):
        fields = extract_fields(match.group(1))
        if not fields:
            continue
        if variables and "value_data" in fields and "@" in fields["value_data"]:
            for name, default in variables.items():
                fields["value_data"] = fields["value_data"].replace(f"@{name}@", default)
        yield fields


def latest_normalized(normalized_dir):
    """Latest timestamped generation of each normalized audit in a folder."""
    normalized_dir = Path(normalized_dir)
    if not normalized_dir.is_dir():
        return []
    best = {}
    for path in normalized_dir.glob("*.audit"):
        match = _TIMESTAMP_SUFFIX_RE.search(path.stem)
        if not match:
            continue
        stem = path.stem[: match.start()]
        stamp = match.group(1)
        if stem not in best or stamp > best[stem][0]:
            best[stem] = (stamp, path)
    return [path for _stamp, path in sorted(best.values())]


def _occurrence(source, fields):
    description = fields.get("description", "").strip().strip('"')
    return {
        "file": str(source),
        "rule_id": extract_control_number(description) if description else "",
        "type": fields.get("type", "").strip(),
        "description": description,
        "expected": str(derive_expected_value(fields)),
        "value_type": fields.get("value_type", "").strip(),
    }


def build_library(sources, matcher=None):
    """Aggregate controls across audit files -> {key: entry}."""
    entries = {}
    for source in sources:
        source = Path(source)
        platform = matcher.match(source.name) if matcher else None
        for fields in iter_active_checks(source):
            key = derive_evaluated_item_key(fields)
            if not key:
                continue
            entry = entries.setdefault(
                key,
                {
                    "key": key,
                    "readable": derive_control_keyword(fields),
                    "types": set(),
                    "platforms": set(),
                    "expectations": defaultdict(int),
                    "nist_refs": set(),
                    "occurrences": [],
                },
            )
            occurrence = _occurrence(source, fields)
            entry["types"].add(occurrence["type"] or "UNKNOWN")
            if platform:
                entry["platforms"].add(platform)
            entry["expectations"][occurrence["expected"]] += 1
            for ref in _NIST_REF_RE.finditer(fields.get("reference", "")):
                entry["nist_refs"].add(ref.group(1).upper())
            entry["occurrences"].append(occurrence)
    return entries


def duplicates_in_file(entries):
    """[(key, file, expected, count)] where one file repeats key+expectation."""
    rows = []
    for key, entry in entries.items():
        per_file = defaultdict(int)
        for occ in entry["occurrences"]:
            per_file[(occ["file"], occ["expected"])] += 1
        for (file, expected), count in sorted(per_file.items()):
            if count > 1:
                rows.append((key, file, expected, count))
    return rows


def expectation_variances(entries):
    """[(key, {expected: count})] where a control carries >1 distinct expectation."""
    return [
        (key, dict(entry["expectations"]))
        for key, entry in sorted(entries.items())
        if len(entry["expectations"]) > 1
    ]


def save_library(entries, path):
    serializable = {}
    for key, entry in sorted(entries.items()):
        serializable[key] = {
            "readable": entry["readable"],
            "types": sorted(entry["types"]),
            "platforms": sorted(entry["platforms"]),
            "expectations": dict(sorted(entry["expectations"].items())),
            "nist_refs": sorted(entry["nist_refs"]),
            "occurrences": entry["occurrences"],
        }
    payload = {
        "generated": time.strftime("%Y-%m-%d %H:%M:%S"),
        "control_count": len(serializable),
        "controls": serializable,
    }
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=1)
    return path


def load_library(path):
    with open(path, "r", encoding="utf-8") as fh:
        payload = json.load(fh)
    return payload.get("controls", {})


def check_audit_file(library_controls, audit_path):
    """Classify each active check of an audit against the library.

    Statuses: NEW (key unknown), KNOWN (key + expectation match),
    EXPECTATION_DIFFERS (known key, unseen expectation),
    DUPLICATE_IN_FILE (key repeated within this audit).
    """
    seen_in_file = defaultdict(int)
    rows = []
    for fields in iter_active_checks(audit_path):
        key = derive_evaluated_item_key(fields)
        if not key:
            continue
        occurrence = _occurrence(audit_path, fields)
        seen_in_file[key] += 1
        if seen_in_file[key] > 1:
            status = "DUPLICATE_IN_FILE"
        elif key not in library_controls:
            status = "NEW"
        elif occurrence["expected"] in library_controls[key].get("expectations", {}):
            status = "KNOWN"
        else:
            status = "EXPECTATION_DIFFERS"
        rows.append(
            {
                "status": status,
                "key": key,
                "rule_id": occurrence["rule_id"],
                "description": occurrence["description"],
                "expected": occurrence["expected"],
                "library_expectations": sorted(
                    library_controls.get(key, {}).get("expectations", {})
                ),
            }
        )
    return rows


def export_workbook(entries, output_path, policy_rows=None):
    from openpyxl import Workbook

    from pysc.report.excel_util import FILL_PARTIAL, write_sheet

    wb = Workbook()
    ws = wb.active
    ws.title = "Controls"
    control_rows = []
    for key, entry in sorted(entries.items()):
        example = entry["occurrences"][0]
        control_rows.append(
            [
                key,
                entry["readable"],
                " ".join(sorted(entry["types"])),
                " ".join(sorted(entry["platforms"])),
                len(entry["occurrences"]),
                len(entry["expectations"]),
                " | ".join(sorted(entry["expectations"])),
                " ".join(sorted(entry["nist_refs"])),
                example["description"],
                "; ".join(sorted({Path(o["file"]).name for o in entry["occurrences"]})),
            ]
        )
    write_sheet(
        ws,
        [
            "Control Key (what is audited)", "Readable Target", "Check Types",
            "Platforms", "Occurrences", "Distinct Expectations",
            "Expectations (what we expect)", "NIST 800-53r5", "Example Description",
            "Files",
        ],
        control_rows,
    )
    for row_idx, (key, entry) in enumerate(sorted(entries.items()), start=2):
        if len(entry["expectations"]) > 1:
            ws.cell(row=row_idx, column=6).fill = FILL_PARTIAL

    ws2 = wb.create_sheet("Expectation_Variance")
    if policy_rows is None:
        variance_rows = []
        for key, expectations in expectation_variances(entries):
            for expected, count in sorted(expectations.items()):
                files = sorted(
                    {
                        Path(o["file"]).name
                        for o in entries[key]["occurrences"]
                        if o["expected"] == expected
                    }
                )
                variance_rows.append([key, "", "", expected, count, "; ".join(files)])
    else:
        variance_rows = []
        for row in policy_rows:
            for expected, count in sorted(row["all_values"].items()):
                files = sorted(
                    {
                        Path(o["file"]).name
                        for o in entries[row["key"]]["occurrences"]
                        if o["expected"] == expected
                    }
                )
                variance_rows.append(
                    [
                        row["key"], row["status"], row["approved"], expected,
                        count, "; ".join(files),
                    ]
                )
    write_sheet(
        ws2,
        ["Control Key", "Policy Status", "Approved Value", "Expected Value", "Occurrences", "Files"],
        variance_rows,
    )

    ws3 = wb.create_sheet("Duplicates_In_File")
    write_sheet(
        ws3,
        ["Control Key", "File", "Expected Value", "Count"],
        [[k, Path(f).name, e, c] for k, f, e, c in duplicates_in_file(entries)],
    )

    ws4 = wb.create_sheet("All_Occurrences")
    write_sheet(
        ws4,
        ["Control Key", "File", "Rule ID", "Type", "Expected", "Description"],
        [
            [key, Path(o["file"]).name, o["rule_id"], o["type"], o["expected"], o["description"]]
            for key, entry in sorted(entries.items())
            for o in entry["occurrences"]
        ],
    )

    wb.save(output_path)
    return output_path


def default_sources(cfg, include_normalized=False):
    sources = []
    for root_key in ("production_inputs", "vendor_inputs"):
        root = cfg.path(root_key)
        if root and root.is_dir():
            sources.extend(sorted(root.glob("*.audit")))
            if include_normalized:
                sources.extend(latest_normalized(root / "Normalized"))
    return sources


def run_build(cfg, include_normalized=None, progress=print):
    if include_normalized is None:
        include_normalized = bool(
            cfg.data.get("library", {}).get("include_normalized", True)
        )
    sources = default_sources(cfg, include_normalized)
    matcher = PlatformMatcher.from_config(cfg)
    progress(f"Scanning {len(sources)} audit file(s)...")
    entries = build_library(sources, matcher)

    library_path = cfg.root / LIBRARY_NAME
    save_library(entries, library_path)
    progress(f"Library: {library_path} ({len(entries)} controls)")

    from pysc.library.policy import REGISTER_NAME, classify_variances, load_register

    register = load_register(cfg.root / REGISTER_NAME)
    policy_rows = classify_variances(entries, register, cfg.path("production_inputs"))

    # One master workbook, overwritten each build. If it is open in Excel the
    # write fails with PermissionError; fall back to a timestamped copy.
    workbook = cfg.root / MASTER_WORKBOOK_NAME
    try:
        export_workbook(entries, workbook, policy_rows=policy_rows)
        progress(f"Workbook: {workbook}")
    except PermissionError:
        out_dir = cfg.path("report_output")
        out_dir.mkdir(parents=True, exist_ok=True)
        fallback = out_dir / f"Control_Library_{time.strftime('%y%m%d%H%M')}.xlsx"
        export_workbook(entries, fallback, policy_rows=policy_rows)
        progress(
            f"WARNING: {workbook} is locked (open in Excel?) - wrote {fallback}; "
            "close the master file and rebuild to refresh it"
        )

    dupes = duplicates_in_file(entries)
    status_counts = {}
    for row in policy_rows:
        status_counts[row["status"]] = status_counts.get(row["status"], 0) + 1
    total_occurrences = sum(len(e["occurrences"]) for e in entries.values())
    progress(
        f"Occurrences: {total_occurrences} | Unique controls: {len(entries)} | "
        f"In-file duplicates: {len(dupes)}"
    )
    if policy_rows:
        progress(
            "Variances: "
            + " | ".join(f"{k}: {v}" for k, v in sorted(status_counts.items()))
        )
    return entries
