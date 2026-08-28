"""Shared Excel helpers for pysc reports.

sanitize_for_excel ports the CSV/formula-injection guard from the legacy
enterprise report (cells starting with =, +, -, @ are prefixed with ' so Excel
renders them inert).
"""

from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FILL_HEADER = PatternFill("solid", start_color="1F4E78")
FONT_HEADER = Font(bold=True, color="FFFFFF")
FILL_GOOD = PatternFill("solid", start_color="A9D08E")
FILL_PARTIAL = PatternFill("solid", start_color="DDEBF7")
FILL_BAD = PatternFill("solid", start_color="FCE4D6")
FILL_NEUTRAL = PatternFill("solid", start_color="E2EFDA")

_FORMULA_PREFIXES = ("=", "+", "-", "@")


def sanitize_for_excel(value):
    if value is None:
        return ""
    text = ILLEGAL_CHARACTERS_RE.sub("", str(value))
    if text.startswith(_FORMULA_PREFIXES):
        return "'" + text
    return text


def write_sheet(ws, headers, rows, autosize=True, sanitize=True):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(vertical="top")
    for row in rows:
        if sanitize:
            row = [v if isinstance(v, (int, float)) else sanitize_for_excel(v) for v in row]
        ws.append(row)
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions
    if autosize:
        autofit_columns(ws, len(headers))


def autofit_columns(ws, n_columns, max_width=80):
    for idx in range(1, n_columns + 1):
        letter = get_column_letter(idx)
        width = max(
            (len(str(c.value)) for c in ws[letter] if c.value is not None),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_width)
