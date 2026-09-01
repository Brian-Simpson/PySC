"""Unified_Compliance_Matrix workbook: enterprise-wide gap posture.

Sheets: Executive_Summary (per-platform scorecard + coverage bar chart),
Platform_Family_Coverage (platform x NIST family), NIST_Matrix (control x
platform status), Priority_Gaps (risk-weighted), Trend (from history DB).

Priority weighting follows the legacy production gap analysis: families
AC, IA, AU, SC, SI weigh x3; a fully missing control outranks a recoverable
one (x2 vs x1).
"""

from collections import defaultdict

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

from pysc.nist.oscal import OscalCatalog
from pysc.report.excel_util import (
    FILL_BAD,
    FILL_GOOD,
    FILL_PARTIAL,
    write_sheet,
)


def _pct(part, whole):
    return round((part / whole) * 100, 2) if whole else 0.0


def build_matrix(result, output_file, history=None, cis_variances=None, attack_mappings=None):
    """Write the workbook for an EnterpriseGapResult; returns output_file."""
    wb = Workbook()

    # --- Executive_Summary ----------------------------------------------------
    ws = wb.active
    ws.title = "Executive_Summary"
    headers = [
        "Platform", "Baseline Audit", "Active Checks", "Inactive Checks",
        "Base Controls Covered", "Catalog Base Controls", "Coverage %",
        "Recoverable (un-comment)", "Requires New Checks", "Candidate Audits",
    ]
    rows = []
    for code, analysis in sorted(result.analyses.items()):
        total = len(analysis.target_baseline)
        rows.append(
            [
                code,
                analysis.baseline.short_name,
                analysis.baseline.checks_parsed,
                len(analysis.baseline.inactive_checks),
                analysis.baseline_coverage_count,
                total,
                _pct(analysis.baseline_coverage_count, total),
                len(analysis.inactive_coverage_opportunities),
                len(analysis.additional_controls_not_present),
                len(analysis.candidates),
            ]
        )
    for code, candidates in sorted(result.missing_baseline.items()):
        rows.append(
            [code, "NO BASELINE DECLARED", 0, 0, 0, 0, 0.0, 0, 0, len(candidates)]
        )
    write_sheet(ws, headers, rows)
    for row_idx in range(2, 2 + len(rows)):
        if ws.cell(row=row_idx, column=2).value == "NO BASELINE DECLARED":
            ws.cell(row=row_idx, column=2).fill = FILL_BAD

    if result.analyses:
        chart = BarChart()
        chart.title = "Base-control coverage % by platform"
        chart.y_axis.title = "Coverage %"
        chart.height = 8
        chart.width = 24
        n = len(result.analyses)
        data = Reference(ws, min_col=7, min_row=1, max_row=1 + n)
        cats = Reference(ws, min_col=1, min_row=2, max_row=1 + n)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.legend = None
        ws.add_chart(chart, f"A{len(rows) + 4}")

    # --- Platform_Family_Coverage ----------------------------------------------
    ws = wb.create_sheet("Platform_Family_Coverage")
    fam_rows = []
    for code, analysis in sorted(result.analyses.items()):
        by_family = defaultdict(lambda: {"total": 0, "covered": 0, "recoverable": 0})
        for control_id in analysis.target_baseline:
            family, _ = OscalCatalog.family_of(control_id)
            bucket = by_family[family]
            bucket["total"] += 1
            if control_id in analysis.baseline_covered_set:
                bucket["covered"] += 1
            elif control_id in analysis.inactive_coverage_opportunities:
                bucket["recoverable"] += 1
        for family, bucket in sorted(by_family.items()):
            gap = bucket["total"] - bucket["covered"] - bucket["recoverable"]
            fam_rows.append(
                [
                    code, family, bucket["total"], bucket["covered"],
                    bucket["recoverable"], gap, _pct(bucket["covered"], bucket["total"]),
                ]
            )
    write_sheet(
        ws,
        ["Platform", "Family", "Total", "Covered", "Recoverable", "Gap", "Coverage %"],
        fam_rows,
    )

    # --- NIST_Matrix ------------------------------------------------------------
    ws = wb.create_sheet("NIST_Matrix")
    platforms = sorted(result.analyses.keys())
    all_controls = {}
    for analysis in result.analyses.values():
        all_controls.update(analysis.target_baseline)
    headers = ["Control ID", "Control Title", "Family"] + platforms
    matrix_rows = []
    for control_id in sorted(all_controls):
        family, _ = OscalCatalog.family_of(control_id)
        row = [control_id, all_controls[control_id], family]
        for code in platforms:
            analysis = result.analyses[code]
            if control_id in analysis.baseline_covered_set:
                row.append("COVERED")
            elif control_id in analysis.inactive_coverage_opportunities:
                row.append("RECOVERABLE")
            else:
                row.append("MISSING")
        matrix_rows.append(row)
    write_sheet(ws, headers, matrix_rows, autosize=False)
    from pysc.report.excel_util import autofit_columns

    autofit_columns(ws, 3)
    status_fills = {"COVERED": FILL_GOOD, "RECOVERABLE": FILL_PARTIAL, "MISSING": FILL_BAD}
    for row_idx in range(2, 2 + len(matrix_rows)):
        for col_idx in range(4, 4 + len(platforms)):
            cell = ws.cell(row=row_idx, column=col_idx)
            fill = status_fills.get(cell.value)
            if fill:
                cell.fill = fill

    # --- Priority_Gaps ----------------------------------------------------------
    from pysc.report.priority import priority_gap_rows

    ws = wb.create_sheet("Priority_Gaps")
    prio_rows = [
        [
            r["score"], r["platform"], r["control_id"], r["title"],
            r["family"], r["family_name"], r["action"],
        ]
        for r in priority_gap_rows(result)
    ]
    write_sheet(
        ws,
        ["Priority", "Platform", "Control ID", "Control Title", "Family", "Family Name", "Remediation Path"],
        prio_rows,
    )

    # --- CIS_Variances (enterprise deviation register) --------------------------
    ws = wb.create_sheet("CIS_Variances")
    write_sheet(
        ws,
        [
            "Control Key", "Platforms",
            "Enterprise Expected Value", "CIS Expected Value",
            "Enterprise Raw Expression", "CIS Raw Expression(s)",
            "CIS Source Files", "NIST 800-53r5", "Rationale", "Example Description",
        ],
        [
            [
                r["key"], r["platforms"],
                r["hth_display"], r["cis_display"],
                r["hth_value"], r["cis_values"],
                r["cis_sources"], r["nist_refs"], r["rationale"], r["description"],
            ]
            for r in (cis_variances or [])
        ],
    )

    # --- Attack_Vectors (ATT&CK exposure from open gaps) -------------------------
    ws = wb.create_sheet("Attack_Vectors")
    vector_rows = []
    if attack_mappings:
        from pysc.nist.attack import attack_vectors_for_gaps

        for v in attack_vectors_for_gaps(result, attack_mappings):
            vector_rows.append(
                [
                    v["technique_id"], v["technique_name"], v["sub_technique_count"],
                    " ".join(v["controls"]), " ".join(v["platforms"]),
                ]
            )
    write_sheet(
        ws,
        ["Technique", "Attack Vector", "Sub-techniques", "Weakened Mitigations (gap controls)", "Platforms"],
        vector_rows,
    )

    # --- Trend ------------------------------------------------------------------
    ws = wb.create_sheet("Trend")
    trend_rows = []
    if history is not None:
        for run_id, ts, platform, covered, recoverable, total in history.platform_trend():
            trend_rows.append(
                [run_id, ts, platform, covered, recoverable, total, _pct(covered, total)]
            )
    write_sheet(
        ws,
        ["Run", "Timestamp", "Platform", "Covered", "Recoverable", "Total", "Coverage %"],
        trend_rows,
    )

    wb.save(output_file)
    return output_file
