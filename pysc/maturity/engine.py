"""Pass-rate-driven baseline maturity (file-based; no Tenable API).

Workflow: export compliance scan results from the Tenable console to Excel
(one row per check with its fleet pass rate), then:

    pysc maturity --audit <baseline.audit> --pass-rates <export.xlsx>
        [--threshold 90] [--apply]

Checks whose fleet pass rate is below the threshold get comment-out proposals
(a review workbook); --apply writes a new audit with those checks commented
out. Commented checks then surface as "recoverable coverage" in the next gap
run — the maturity loop the broken Comment_Below_%.py twins intended.

Pass-rate workbook expectations (ported from Comment_Below_%.py): a header row
containing a description column and a pass column (named like 'Description' /
'Pass'); pass values may be fractions (0.85) or percentages (85).
"""

import re
import time
from pathlib import Path

from openpyxl import load_workbook

from pysc.gap.harvest import extract_fields

_ACTIVE_BLOCK_RE = re.compile(r"(?ms)^(?!\s*#)\s*<custom_item>.*?^\s*</custom_item>[ \t]*$")


class MaturityError(RuntimeError):
    pass


def _normalize_rate(value):
    """Pass rate as a fraction: accepts 0.85, 85, '85%', '0.85'."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip().rstrip("%")
        if not value:
            return None
        try:
            value = float(value)
        except ValueError:
            return None
    value = float(value)
    if value > 1.0:
        value = value / 100.0
    return value


def load_pass_rates(xlsx_path):
    """{description: pass_rate fraction} from a Tenable results export."""
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise MaturityError(f"Empty workbook: {xlsx_path}")

    desc_idx = pass_idx = None
    for idx, name in enumerate(header):
        label = str(name or "").strip().lower()
        if desc_idx is None and "description" in label:
            desc_idx = idx
        if pass_idx is None and label.startswith("pass"):
            pass_idx = idx
    if desc_idx is None or pass_idx is None:
        raise MaturityError(
            f"Could not find Description/Pass columns in {xlsx_path} "
            f"(header: {list(header)})"
        )

    rates = {}
    for row in rows:
        if row is None or desc_idx >= len(row):
            continue
        description = str(row[desc_idx] or "").strip()
        rate = _normalize_rate(row[pass_idx] if pass_idx < len(row) else None)
        if description and rate is not None:
            rates[description] = rate
    wb.close()
    return rates


def propose(audit_path, pass_rates, threshold=0.90):
    """Comment-out proposals for active checks under the threshold.

    Returns (proposals, unmatched_export_rows):
    - proposals: [{description, pass_rate, block_span}] for active checks whose
      description exactly matches an export row below the threshold
    - unmatched: export descriptions under threshold with no active check
    """
    text = Path(audit_path).read_text(encoding="utf-8", errors="ignore")
    low = {d: r for d, r in pass_rates.items() if r < threshold}

    proposals = []
    matched = set()
    for match in _ACTIVE_BLOCK_RE.finditer(text):
        fields = extract_fields(match.group(0))
        description = fields.get("description", "").strip().strip('"')
        if description in low:
            matched.add(description)
            proposals.append(
                {
                    "description": description,
                    "pass_rate": low[description],
                    "span": match.span(),
                }
            )
    unmatched = sorted(set(low) - matched)
    return proposals, unmatched


def apply_proposals(audit_path, proposals, output_path=None):
    """Write a copy of the audit with proposed blocks commented out."""
    audit_path = Path(audit_path)
    text = audit_path.read_text(encoding="utf-8", errors="ignore")

    if output_path is None:
        stamp = time.strftime("%y%m%d%H%M")
        output_path = audit_path.with_name(
            f"{audit_path.stem}_matured_{stamp}{audit_path.suffix}"
        )

    # Comment from the end so earlier spans stay valid.
    for proposal in sorted(proposals, key=lambda p: p["span"][0], reverse=True):
        start, end = proposal["span"]
        block = text[start:end]
        commented = "\n".join(f"#{line}" for line in block.split("\n"))
        text = text[:start] + commented + text[end:]

    Path(output_path).write_text(text, encoding="utf-8")
    return Path(output_path)


def write_proposal_workbook(proposals, unmatched, threshold, output_path):
    from openpyxl import Workbook

    from pysc.report.excel_util import write_sheet

    wb = Workbook()
    ws = wb.active
    ws.title = "Comment_Out_Proposals"
    write_sheet(
        ws,
        ["Description", "Fleet Pass Rate", "Threshold", "Action"],
        [
            [p["description"], round(p["pass_rate"], 4), threshold, "Comment out"]
            for p in sorted(proposals, key=lambda p: p["pass_rate"])
        ],
    )
    ws2 = wb.create_sheet("Unmatched_Export_Rows")
    write_sheet(
        ws2,
        ["Description (below threshold, no matching active check)"],
        [[d] for d in unmatched],
    )
    wb.save(output_path)
    return output_path
