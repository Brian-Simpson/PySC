"""Normalization core, extracted VERBATIM from the legacy engine.

Generated once by pysc/_devtools/extract_core.py from
pysc/_legacy/all_audits.py; now hand-maintained. Definitions keep
their original text and order. Catalog/merge/gap/threat-intel and CLI
remain in the legacy module until their own extraction phases.
"""

import os
import re
import sys
import subprocess
import time
import shutil
import importlib.util
import urllib.request
from collections import OrderedDict, defaultdict
from datetime import datetime
import json
import csv

try:
    import openpyxl
    from openpyxl import Workbook
except Exception:
    openpyxl = None
    Workbook = None

_PARSER_START_BLOCK_RE = re.compile(r"^\s*<\s*(custom_item|report)\b([^>]*)>\s*$", re.IGNORECASE)


_PARSER_END_CUSTOM_RE = re.compile(r"^\s*</\s*custom_item\s*>\s*$", re.IGNORECASE)


_PARSER_END_REPORT_RE = re.compile(r"^\s*</\s*report\s*>\s*$", re.IGNORECASE)


_PARSER_FIELD_RE = re.compile(r"^\s*([A-Za-z0-9_]+)\s*:\s*(.*)$")


_PARSER_REPORT_TYPE_RE = re.compile(r"\btype\s*:\s*\"?([A-Za-z-]+)\"?", re.IGNORECASE)


_PARSER_VAR_LINE_RE = re.compile(r"^\s*([A-Za-z_][A-Za-z0-9_]*)\s*[:=]\s*(.*?)\s*$")


def _parser_lines(text_or_lines):
    if isinstance(text_or_lines, str):
        return text_or_lines.splitlines()
    return [str(line).rstrip("\n") for line in text_or_lines]


def _parser_unquote(value):
    value = value.strip()
    if len(value) >= 2 and value[0] == value[-1] and value[0] in ('"', "'"):
        return value[1:-1]
    return value


def extract_variables(text_or_lines):
    """Extract top-level and commented XML-style audit variable definitions."""
    lines = _parser_lines(text_or_lines)
    variables = {}

    for line in lines:
        if _PARSER_START_BLOCK_RE.match(line):
            break
        match = _PARSER_VAR_LINE_RE.match(line)
        if not match:
            continue
        key, value = match.group(1), match.group(2)
        if key.lower() not in {'type', 'description', 'info', 'reference'} and not (value.startswith('<') and value.endswith('>')):
            variables[key] = _parser_unquote(value)

    in_variable = False
    variable_name = ''
    variable_default = ''
    for line in lines:
        cleaned = re.sub(r"^\s*#\s*", "", line).strip()
        if cleaned.lower() == '<variable>':
            in_variable, variable_name, variable_default = True, '', ''
        elif in_variable and cleaned.lower() == '</variable>':
            if variable_name and variable_default:
                variables[variable_name] = _parser_unquote(variable_default)
            in_variable = False
        elif in_variable:
            name_match = re.match(r"^<name>(.*?)</name>$", cleaned, flags=re.IGNORECASE)
            default_match = re.match(r"^<default>(.*?)</default>$", cleaned, flags=re.IGNORECASE)
            if name_match:
                variable_name = name_match.group(1).strip()
            elif default_match:
                variable_default = default_match.group(1).strip()
    return variables


def parse_document(text_or_lines):
    """Parse an audit file into text, custom-item, and report nodes."""
    document = []
    in_block = False
    block_type = ''
    block_fields = OrderedDict()
    last_key = ''

    for line in _parser_lines(text_or_lines):
        if not in_block:
            start_match = _PARSER_START_BLOCK_RE.match(line)
            if not start_match:
                document.append({'type': 'text', 'text': line})
                continue
            tag, attributes = start_match.group(1).lower(), start_match.group(2) or ''
            if tag == 'report':
                type_match = _PARSER_REPORT_TYPE_RE.search(attributes)
                report_type = type_match.group(1).strip().lower() if type_match else 'warning'
                block_type = f"report-{'passed' if report_type == 'passed' else 'warning'}"
            else:
                block_type = 'custom_item'
            block_fields, last_key, in_block = OrderedDict(), '', True
            continue

        is_end = (block_type.startswith('report') and _PARSER_END_REPORT_RE.match(line)) or (block_type == 'custom_item' and _PARSER_END_CUSTOM_RE.match(line))
        if is_end:
            document.append({'type': block_type, 'fields': block_fields})
            in_block, block_type, block_fields, last_key = False, '', OrderedDict(), ''
            continue

        field_match = _PARSER_FIELD_RE.match(line)
        if field_match:
            last_key = field_match.group(1).strip()
            block_fields[last_key] = field_match.group(2).strip()
        elif last_key and line.strip():
            block_fields[last_key] = (block_fields[last_key] + ' ' + line.strip()).strip()

    if in_block and block_type:
        document.append({'type': block_type, 'fields': block_fields})
    return document


PARSING_RESULTS_FILENAME = 'Parsing Results.xlsx'


PARSING_RESULTS_BY_FOLDER = OrderedDict()


VALIDATION_RESULTS = OrderedDict()


RUN_TIMESTAMP = time.strftime('%y%m%d%H', time.localtime())


_TS_SUFFIX_RE = re.compile(r'_\d{8}$')


NON_COMPLIANT_OUTPUT = '__NON_COMPLIANT__'


REAL_KEYS = {
    "type", "description", "info", "reference", "see_also", "solution",
    "api_request_type", "request", "xsl_stmt", "not_expect", "show_output",
    "powershell_args", "key_item",
    "value_type", "value_data", "reg_key", "reg_item", "reg_option",
    "audit_policy_subcategory", "right_type", "reg_include_hku_users",
    "check_type", "account_type", "password_policy", "lockout_policy",
    "regex", "expect", "severity",
    "wmi_key",
    "wmi_namespace",
    "wmi_request",
    "wmi_attribute",
    "f5_command",
    "item",
    "json_transform",
    "match_all",
    "sql_expect",
    "sql_request",
    "sql_types",
    "cmd",
    "is_substring",
    "mask",
    "min_occurrences",
    "rpm",
    "string_required",
    "timeout",
    "policy_arn",
    "interface_name",
    "shared_key",
    "where",
    "tmsh",
    "dont_echo_cmd",
    "file_required",
    "powershell_option",
    "operator",
    "only_show_cmd_output",
    "system",
    "reg_ignore_hku_users",
    "reg_type",
}


IGNORED_KEYS = {
    "Impact",
    "Note",
    "4944",
    "4945",
    "4946",
    "4947",
    "4948",
    "4949",
    "4950",
    "4951",
    "4952",
    "4953",
    "4954",
    "4956",
    "4957",
    "4958",
    "5063",
    "5064",
    "5065",
    "5066",
    "5067",
    "5068",
    "5069",
    "5070",
    "6145",
    "Caution",
    "Disabled",
    "Enabled",
    "Important",
    "Warning",
    "Example",
    "MinimumPasswordLength",
    "PasswordReusePrevention",
    "aws_action",
    "NOTE",
    "content",
    "context",
    "Rationale",
    "Steps",
    "WARNING",
    "https",
    "1",
    "2",
    "Source",
    "or",
    "levels",
    "service",
    "status",
    "Notes",
    "Satisfies",
    "days",
    "name",
    "optionally",
    "file",
    "group",
    "owner",
    "required",
    "ignore",
    "IMPORTANT",
    "Run",
    "Default",
    "Mitigated",
    "Name",
    "None",
    "Vulnerable",
    "Run",
}


SEE_ALSO_REPLACEMENT = "See HTH Policies and Standards"


def resolve_variables(text, variables):
    for k, v in variables.items():
        text = text.replace(f"@{k}@", v)
    return text


# Per-platform info sentence limits (detection codes, e.g. 'F5' -> 3).
# Populated from pysc.toml platform profiles via pysc.normalize.apply_platform_overrides;
# empty by default so behavior matches the legacy engine byte-for-byte.
INFO_SENTENCES_BY_PLATFORM = {}


def _info_sentence_limit(platform_hint):
    try:
        return max(1, int(INFO_SENTENCES_BY_PLATFORM.get(platform_hint, 1)))
    except (TypeError, ValueError):
        return 1


def normalize_info(raw, max_sentences=1):
    if not raw:
        return None
    s = raw.strip()
    s = re.sub(r'^[\'"]+', '', s)
    s = re.sub(r'[\'"]+$', '', s)
    s = re.sub(r'\s+', ' ', s)
    if s.startswith('This audit is written to dynamically identify if all paths are present in any order.'):
        s = s.rstrip('.') + '.'
        return f'"{s}"'
    if max_sentences > 1:
        sentence = '.'.join(s.split('.')[:max_sentences]).strip()
    else:
        sentence = s.split('.')[0].strip()
    if not sentence:
        return None
    return f"\"{sentence}.\""


def normalize_reference(raw):
    if not raw:
        return None
    flat = _strip_outer_quotes_once(raw)
    flat = re.sub(r"\s+", " ", flat)
    parts = [p.strip() for p in flat.split(",")]
    controls = []
    for p in parts:
        p = p.strip().strip('"').strip("'")
        m = re.match(r"^(?:NIST\s+)?800-53r5\|(.+)$", p, flags=re.IGNORECASE)
        if m:
            controls.append(m.group(1).strip().strip('"').strip("'"))
    if not controls:
        return None
    seen = set()
    unique = []
    for c in controls:
        if c not in seen:
            seen.add(c)
            unique.append(c)
    return "\"NIST 800-53r5|" + " ".join(unique) + "\""


def normalize_reference_or_passthrough(raw):
    """Prefer normalized NIST references, but retain non-NIST references instead of dropping them."""
    ref = normalize_reference(raw)
    if ref:
        return ref

    cleaned = _strip_outer_quotes_once(raw)
    if cleaned is None:
        cleaned = ''
    return f'"{cleaned}"'


def normalize_description(raw):
    if not raw:
        return None
    s = raw.strip()
    s = s.strip('"')
    s = re.sub(r"^\d+(\.\d+)+\s*", "", s)
    s = re.sub(r"'([^']+)'", r"\1", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = _escape_unescaped_double_quotes(s)
    return f'"{s}"'


def normalize_solution(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None

    # Normalize to a double-quoted value and guarantee a trailing quote.
    if s.startswith("'") and s.endswith("'") and len(s) >= 2:
        s = s[1:-1]
    elif s.startswith("'"):
        s = s[1:]
    elif s.startswith('"') and s.endswith('"') and len(s) >= 2:
        s = s[1:-1]
    elif s.startswith('"'):
        s = s[1:]

    s = s.rstrip("'\"").rstrip()
    s = re.sub(r'\s+', ' ', s).strip()
    s = _escape_unescaped_double_quotes(s)
    if s.endswith('\\'):
        # Prevent escaping the final quote of the rendered audit value.
        s += '\\'
    s = f'"{s}"'
    return s


def _first_int(text):
    if text is None:
        return None
    m = re.search(r'\d+', str(text))
    if not m:
        return None
    try:
        return int(m.group(0))
    except ValueError:
        return None


def _derive_expected_scalar(fields):
    raw = _strip_outer_quotes_once(fields.get('value_data', ''))
    if not raw:
        raw = ''
    else:
        # Range style values like [24..MAX] -> 24.
        m_range = re.search(r'\[(\d+)\s*\.\.', raw)
        if m_range:
            return m_range.group(1)

        # Alternate values like "1" || "2" -> first value.
        m_first = re.search(r'\d+', raw)
        if m_first:
            return m_first.group(0)

    # Fall back to description numbers and avoid control-id prefixes like 1.2.3.
    desc = str(fields.get('description', ''))
    nums = re.findall(r'\d+', desc)
    if nums:
        return nums[-1]

    return raw


def _build_ge_integer_regex(min_value):
    digits = str(int(min_value))
    parts = []

    def rec(prefix, index, tight):
        if index == len(digits):
            parts.append(prefix)
            return

        current = int(digits[index])
        remaining = len(digits) - index - 1

        if tight:
            rec(prefix + str(current), index + 1, True)
            for candidate in range(current + 1, 10):
                if remaining:
                    parts.append(prefix + str(candidate) + rf'\d{{{remaining}}}')
                else:
                    parts.append(prefix + str(candidate))
        else:
            if len(digits) - index > 0:
                parts.append(prefix + rf'\d{{{len(digits) - index}}}')

    rec('', 0, True)

    if len(digits) == 1:
        longer = r'[1-9]\d+'
    else:
        longer = rf'[1-9]\d{{{len(digits)},}}'
    parts.append(longer)

    unique = []
    seen = set()
    for part in parts:
        if part not in seen:
            seen.add(part)
            unique.append(part)
    return '^(?:' + '|'.join(unique) + ')$'


def _build_bracket_range_expected(raw_expected):
    raw = _strip_outer_quotes_once(raw_expected).strip()
    match = re.fullmatch(r'\[(MIN|\d+)\s*\.\.\.?\s*(MAX|\d+)\]', raw, flags=re.IGNORECASE)
    if not match:
        return None

    lower = match.group(1).upper()
    upper = match.group(2).upper()

    lower_num = None if lower == 'MIN' else int(lower)
    upper_num = None if upper == 'MAX' else int(upper)

    if lower_num is not None and upper_num is not None:
        if upper_num < lower_num:
            return raw
        values = [str(i) for i in range(lower_num, upper_num + 1)]
        return '^(?:' + '|'.join(values) + ')$'

    if lower_num is None and upper_num is not None:
        values = [str(i) for i in range(0, upper_num + 1)]
        return '^(?:' + '|'.join(values) + ')$'

    if lower_num is not None and upper_num is None:
        return _build_ge_integer_regex(lower_num)

    return r'^\d+$'


def _build_policy_powershell_args(policy_type, policy_name, expected_value):
    ptype = _norm_upper_token(policy_type)
    pname = _norm_upper_token(policy_name)
    expected = str(expected_value or '').strip()

    if ptype == 'PASSWORD_POLICY' and pname == 'ENFORCE_PASSWORD_HISTORY':
        return (
            "$HisPwdObj = net accounts | Select-string 'password history'; "
            "$HisPwdStr = $HisPwdObj.ToString(); "
            "$HisPwdStr -match '\\d{1,2}' | out-null; "
            f"$HisPwdStr -match '{expected}' | out-null; "
            "$LO_HisPwd = $matches[0]; Write-Output $LO_HisPwd;"
        )
    if ptype == 'PASSWORD_POLICY' and pname == 'MAXIMUM_PASSWORD_AGE':
        return (
            "$MaxPwdObj = net accounts | Select-string 'Maximum password age'; "
            "$MaxPwdStr = $MaxPwdObj.ToString(); "
            "$MaxPwdStr -match '\\d{1,3}' | out-null; "
            f"$MaxPwdStr -match '{expected}' | out-null; "
            "$LO_MaxPwd = $matches[0]; Write-Output $LO_MaxPwd;"
        )
    if ptype == 'PASSWORD_POLICY' and pname == 'MINIMUM_PASSWORD_AGE':
        return (
            "$MinAgeObj = net accounts | Select-string 'Minimum password age'; "
            "$MinAgeStr = $MinAgeObj.ToString(); "
            "$MinAgeStr -match '\\d{1,3}' | out-null; "
            f"$MinAgeStr -match '{expected}' | out-null; "
            "$LO_MinAge = $matches[0]; Write-Output $LO_MinAge;"
        )
    if ptype == 'PASSWORD_POLICY' and pname == 'MINIMUM_PASSWORD_LENGTH':
        return (
            "$MinLenObj = net accounts | Select-string 'password length'; "
            "$MinLenStr = $MinLenObj.ToString(); "
            "$MinLenStr -match '\\d{1,3}' | out-null; "
            f"$MinLenStr -match '{expected}' | out-null; "
            "$LO_MinLen = $matches[0]; Write-Output $LO_MinLen;"
        )
    if ptype == 'PASSWORD_POLICY' and pname == 'COMPLEXITY_REQUIREMENTS':
        return (
            "$secfile = [System.IO.Path]::GetTempFileName(); "
            "secedit /export /cfg $secfile /areas SECURITYPOLICY /quiet | Out-Null; "
            "$complexity = Select-String -Path $secfile -Pattern 'PasswordComplexity\\s*=\\s*(\\d)'; "
            "Remove-Item $secfile -ErrorAction SilentlyContinue; "
            "if ($complexity -and $complexity.Matches.Groups[1].Value -eq '1') { Write-Output 'Enabled' } else { Write-Output 'Disabled' }"
        )
    if ptype == 'PASSWORD_POLICY' and pname == 'REVERSIBLE_ENCRYPTION':
        return (
            "$secfile = [System.IO.Path]::GetTempFileName(); "
            "secedit /export /cfg $secfile /areas SECURITYPOLICY /quiet | Out-Null; "
            "$line = Select-String -Path $secfile -Pattern 'ClearTextPassword\\s*=\\s*(\\d)' | Select-Object -First 1; "
            "Remove-Item $secfile -ErrorAction SilentlyContinue; "
            "if($line -and $line.Matches.Count -gt 0){ "
            "if($line.Matches[0].Groups[1].Value -eq '1'){ Write-Output 'Enabled' } else { Write-Output 'Disabled' } "
            "} else { Write-Output 'Not Found' }"
        )
    if ptype == 'PASSWORD_POLICY' and pname == 'LOCKOUT_ADMINS':
        return (
            "$secfile = [System.IO.Path]::GetTempFileName(); "
            "secedit /export /cfg $secfile /areas SECURITYPOLICY /quiet | Out-Null; "
            "$line = Select-String -Path $secfile -Pattern 'AllowAdministratorLockout\\s*=\\s*(\\d)' | Select-Object -First 1; "
            "Remove-Item $secfile -ErrorAction SilentlyContinue; "
            "if($line -and $line.Matches.Count -gt 0){ "
            "if($line.Matches[0].Groups[1].Value -eq '1'){ Write-Output 'Enabled' } else { Write-Output 'Disabled' } "
            "} else { Write-Output 'Not Found' }"
        )
    if ptype == 'PASSWORD_POLICY' and pname == 'FORCE_LOGOFF':
        return (
            "$LogoffObj = net accounts | Select-string 'Force user logoff'; "
            "$LogoffStr = $LogoffObj.ToString(); "
            "$LogoffStr -match '\\d{1,4}|Never' | out-null; "
            "$LO_Logoff = $matches[0]; Write-Output $LO_Logoff;"
        )
    if ptype == 'LOCKOUT_POLICY' and pname == 'LOCKOUT_DURATION':
        return (
            "$LoDurObj = net accounts | Select-string 'Lockout duration'; "
            "$LoDurStr = $LoDurObj.ToString(); "
            "$LoDurStr -match '\\d{1,4}' | out-null; "
            f"$LoDurStr -match '{expected}' | out-null; "
            "$LO_LoDur = $matches[0]; Write-Output $LO_LoDur;"
        )
    if ptype == 'LOCKOUT_POLICY' and pname == 'LOCKOUT_THRESHOLD':
        return (
            "$LoThrObj = net accounts | Select-string 'threshold'; "
            "$LoThrStr = $LoThrObj.ToString(); "
            "$LoThrStr -match '\\d{1,3}' | out-null; "
            f"$LoThrStr -match '{expected}' | out-null; "
            "$LO_LoThr = $matches[0]; Write-Output $LO_LoThr;"
        )
    if ptype == 'LOCKOUT_POLICY' and pname == 'LOCKOUT_RESET':
        return (
            "$LoObsObj = net accounts | Select-string 'Lockout observation'; "
            "$LoObsStr = $LoObsObj.ToString(); "
            "$LoObsStr -match '\\d{1,4}' | out-null; "
            f"$LoObsStr -match '{expected}' | out-null; "
            "$LO_LoObs = $matches[0]; Write-Output $LO_LoObs;"
        )
    return ''


def _derive_audit_policy_expected(value_data):
    val = _norm_upper_token(_strip_outer_quotes_once(value_data))
    if 'SUCCESS' in val and 'FAILURE' in val:
        return 'Success and Failure'
    if 'SUCCESS' in val:
        return 'Success'
    if 'FAILURE' in val:
        return 'Failure'
    return _strip_outer_quotes_once(value_data)


def _convert_policy_item_to_audit_powershell(fields):
    ptype = _norm_upper_token(fields.get('type', ''))
    if ptype == 'AUDIT_POLICY_SUBCATEGORY':
        subcat = _strip_outer_quotes_once(fields.get('audit_policy_subcategory', ''))
        expected = _derive_audit_policy_expected(fields.get('value_data', ''))
        if not subcat or not expected:
            return fields

        ps_script = (
            f"$AP = auditpol /get /Subcategory:'{subcat}' |Select-String '{subcat}'; "
            "$APStr=$AP.ToString(); "
            f"$APStr -match '{expected}' | out-null; "
            "$LOAP = $matches[0]; Write-Output $LOAP;"
        )

        converted = OrderedDict()
        converted['type'] = 'AUDIT_POWERSHELL'
        for key in ('description', 'info', 'solution', 'reference', 'see_also'):
            if key in fields:
                converted[key] = fields.get(key)
        converted['value_type'] = 'POLICY_TEXT'
        converted['value_data'] = f'"{expected}"'
        converted['powershell_args'] = '-NoProfile -ExecutionPolicy Bypass -Command'
        converted['powershell_script'] = ps_script
        converted['expect'] = expected
        return converted

    if ptype not in {'PASSWORD_POLICY', 'LOCKOUT_POLICY'}:
        return fields

    policy_name_key = 'password_policy' if ptype == 'PASSWORD_POLICY' else 'lockout_policy'
    policy_name = _norm_upper_token(fields.get(policy_name_key, ''))
    expected = _derive_expected_scalar(fields)
    ps_script = _build_policy_powershell_args(ptype, policy_name, expected)
    if not ps_script:
        return fields

    converted = OrderedDict()
    converted['type'] = 'AUDIT_POWERSHELL'

    # Preserve existing primary metadata in a stable order.
    for key in ('description', 'info', 'solution', 'reference', 'see_also'):
        if key in fields:
            converted[key] = fields.get(key)

    converted['value_type'] = 'POLICY_TEXT'
    converted['value_data'] = f'"{expected}"' if expected else fields.get('value_data', '')
    converted['powershell_args'] = '-NoProfile -ExecutionPolicy Bypass -Command'
    converted['powershell_script'] = ps_script
    converted['expect'] = expected if expected else 'NON_COMPLIANT'

    return converted


def _ps_single_quote_escape(value):
    return str(value or '').replace("'", "''")


def _rewrite_auditpol_powershell_script(script_text):
    script = _strip_outer_quotes_once(script_text)
    if not script or 'auditpol' not in script.lower():
        return script

    subcat_match = re.search(
        r"/subcategory\s*:\s*['\"]([^'\"]+)['\"]",
        script,
        flags=re.IGNORECASE,
    )
    if not subcat_match:
        return script

    subcat_raw = subcat_match.group(1).strip()
    subcat = _ps_single_quote_escape(subcat_raw)
    subcat_regex = re.escape(subcat_raw)
    return (
        f"$rows = auditpol /get /Subcategory:'{subcat}' /r 2>$null | ConvertFrom-Csv; "
        f"$row = if($rows){{ $rows | Where-Object {{ $_.Subcategory -eq '{subcat}' }} | Select-Object -First 1 }} else {{ $null }}; "
        "$inc = if($row){ [string]$row.'Inclusion Setting' } else { '' }; "
        "if(-not [string]::IsNullOrWhiteSpace($inc)){ Write-Output $inc.Trim(); return }; "
        f"$AP = auditpol /get /Subcategory:'{subcat}' 2>$null | ForEach-Object {{ $_.ToString() }}; "
        f"$line = $AP | Where-Object {{ $_ -match '^\\s*{subcat_regex}\\s{{2,}}' }} | Select-Object -First 1; "
        "if(-not $line){ Write-Output 'Not Found'; return }; "
        "if($line -match 'Success and Failure'){ Write-Output 'Success and Failure'; return }; "
        "if($line -match 'No Auditing'){ Write-Output 'No Auditing'; return }; "
        "if($line -match 'Success'){ Write-Output 'Success'; return }; "
        "if($line -match 'Failure'){ Write-Output 'Failure'; return }; "
        "Write-Output $line;"
    )


def _rewrite_ntlm_outgoing_traffic_script(script_text):
    script = _strip_outer_quotes_once(script_text)
    if 'RestrictSendingNTLMTraffic' not in script:
        return script

    return (
        "if(Test-Path -Path 'HKLM:\\System\\CurrentControlSet\\Control\\Lsa\\MSV1_0'){"
        "$noutput = (Get-ItemProperty -Path 'HKLM:\\System\\CurrentControlSet\\Control\\Lsa\\MSV1_0').RestrictSendingNTLMTraffic;"
        "if($null -eq $noutput){ Write-Output 'Compliant'; return };"
        "try { if([int]$noutput -ge 1){ Write-Output 'Compliant' } else { Write-Output 'Non-Compliant' } }"
        "catch { Write-Output 'Non-Compliant' }"
        "}else { Write-Output 'Non-Compliant' }"
    )


def _rewrite_netlogon_parameters_default_script(script_text):
    script = _strip_outer_quotes_once(script_text)

    # Repair prior over-normalization variants like Parametersssss.
    script = re.sub(
        r'(?i)Services\\Netlogon\\Parameters+',
        r'Services\\Netlogon\\Parameters',
        script,
    )

    if not re.search(r'(?i)Services\\Netlogon\\Parameters', script):
        return script

    prop_match = re.search(
        r"Get-ItemProperty\s*-Path\s*'[^']*Netlogon\\Parameters+'\)\.(\w+)",
        script,
        flags=re.IGNORECASE,
    )
    if not prop_match:
        return script

    prop = prop_match.group(1)
    defaults = {
        'DisablePasswordChange': '0',
        'MaximumPasswordAge': '30',
        'RequireSignOrSeal': '1',
        'SealSecureChannel': '1',
        'SignSecureChannel': '1',
        'RequireStrongKey': '1',
        'RefusePasswordChange': '0',
    }
    default_value = defaults.get(prop)
    if default_value is None:
        return script

    return (
        "$p='HKLM:\\System\\CurrentControlSet\\Services\\Netlogon\\Parameters'; "
        f"$k='{prop}'; $d='{default_value}'; "
        "if(-not (Test-Path -Path $p)){ Write-Output $d; return }; "
        "$r=Get-ItemProperty -Path $p -ErrorAction SilentlyContinue; "
        "$pr=$r.PSObject.Properties[$k]; "
        "if($null -eq $pr -or $null -eq $pr.Value){ Write-Output $d } else { Write-Output $pr.Value }"
    )


def _rewrite_password_complexity_script(script_text):
    script = _strip_outer_quotes_once(script_text)
    s_upper = _norm_upper_token(script)
    if 'NET ACCOUNTS' not in s_upper or 'PASSWORD COMPLEXITY' not in s_upper:
        return script

    return (
        "$secfile = [System.IO.Path]::GetTempFileName(); "
        "secedit /export /cfg $secfile /areas SECURITYPOLICY /quiet | Out-Null; "
        "$complexity = Select-String -Path $secfile -Pattern 'PasswordComplexity\\s*=\\s*(\\d)'; "
        "Remove-Item $secfile -ErrorAction SilentlyContinue; "
        "if ($complexity -and $complexity.Matches.Groups[1].Value -eq '1') { Write-Output 'Enabled' } else { Write-Output 'Disabled' }"
    )


def _normalize_legacy_net_accounts_expected(script_text, raw_value_data):
    script = _strip_outer_quotes_once(script_text)
    expected = _strip_outer_quotes_once(raw_value_data)
    if not script or not expected:
        return raw_value_data

    s_upper = _norm_upper_token(script)
    if 'NET ACCOUNTS' not in s_upper:
        return raw_value_data
    if not (
        'LOCKOUT OBSERVATION' in s_upper
        or 'LOCKOUT DURATION' in s_upper
        or 'THRESHOLD' in s_upper
    ):
        return raw_value_data

    first_num = _first_int(expected)
    if first_num is None:
        return raw_value_data
    return f'"{first_num}"'


def _best_effort_expected_text(fields):
    for key in ('value_data', 'expect', 'sql_expect', 'regex', 'not_expect', 'show_output'):
        raw_val = fields.get(key, '')
        raw_text = str(raw_val or '').strip()
        if ('&&' in raw_text or '||' in raw_text) and raw_text.count('"') >= 2:
            raw = raw_text
        else:
            raw = _strip_outer_quotes_once(raw_val)
        if raw:
            return raw
    return '.+'


def _format_converted_value_data(expected, check_type=''):
    raw = str(expected or '').strip()
    if not raw:
        return '""'

    semantic_boolean = _normalize_boolean_value_data_expression(raw)
    if semantic_boolean:
        return f'"{semantic_boolean}"'

    stripped = _strip_outer_quotes_once(raw)
    return f'"{stripped}"'


def _split_top_level(text, delimiter):
    parts = []
    buf = []
    depth = 0
    in_quote = False
    i = 0
    while i < len(text):
        ch = text[i]
        if ch == '"' and (i == 0 or text[i - 1] != '\\'):
            in_quote = not in_quote
            buf.append(ch)
            i += 1
            continue

        if not in_quote:
            if ch == '(':
                depth += 1
            elif ch == ')' and depth > 0:
                depth -= 1
            elif depth == 0 and text.startswith(delimiter, i):
                parts.append(''.join(buf).strip())
                buf = []
                i += len(delimiter)
                continue

        buf.append(ch)
        i += 1

    tail = ''.join(buf).strip()
    if tail:
        parts.append(tail)
    return parts


def _canonicalize_user_right_member(value):
    member = _strip_outer_quotes_once(value)
    member = member.split('\\')[-1]
    member = re.sub(r'\s+', ' ', member).strip()
    if not member:
        return ''
    if member == member.lower():
        member = ' '.join(part.capitalize() for part in member.split())
    return member


def _strip_balanced_outer_parens(text):
    s = str(text or '').strip()
    while s.startswith('(') and s.endswith(')'):
        depth = 0
        balanced = True
        for idx, ch in enumerate(s):
            if ch == '(':
                depth += 1
            elif ch == ')':
                depth -= 1
                if depth == 0 and idx != len(s) - 1:
                    balanced = False
                    break
            if depth < 0:
                balanced = False
                break
        if not balanced or depth != 0:
            break
        s = s[1:-1].strip()
    return s


def _regex_escape_user_right_member(value):
    return re.escape(value).replace(r'\ ', ' ')


def _looks_like_quoted_boolean_expression(text):
    raw = str(text or '').strip()
    return ('&&' in raw or '||' in raw or r'\|\|' in raw) and raw.count('"') >= 2


def _prepare_boolean_expression_text(text):
    raw = str(text or '').strip()
    # Normalize source strings like \"A\" \|\| \"B\" before tokenization.
    raw = raw.replace(r'\"', '"')
    if raw.startswith('"') and raw.endswith('"') and len(raw) > 2 and raw.count('"') > 2 and raw[1] in {'^', '('}:
        raw = raw[1:-1].strip()
    if raw.startswith('^') and raw.endswith('$'):
        raw = raw[1:-1].strip()
    raw = raw.replace(r'\|\|', '||')
    raw = raw.replace(r'\&\&', '&&')
    raw = raw.replace(r'\(', '(').replace(r'\)', ')')
    return raw


def _regex_escape_expected_literal(value):
    return re.escape(str(value or ''))


def _normalize_no_members_token_regex(pattern_text):
    pattern = str(pattern_text or '')
    if "'?NO_MEMBERS'?" in pattern:
        return pattern
    return re.sub(r'\bNO_MEMBERS\b', "'?NO_MEMBERS'?", pattern)


def _allow_no_members_alternative(pattern_text):
    pattern = str(pattern_text or '').strip()
    if not pattern or 'NO_MEMBERS' in pattern:
        return pattern

    no_members_alt = "'?NO_MEMBERS'?"
    inline_flags = re.match(r'^\(\?([a-zA-Z]+)\)\^(.*)\$$', pattern, flags=re.DOTALL)
    if inline_flags:
        flags = inline_flags.group(1)
        inner = inline_flags.group(2)
        return f'(?{flags})^(?:{inner}|{no_members_alt})$'

    if pattern.startswith('^') and pattern.endswith('$') and len(pattern) >= 2:
        inner = pattern[1:-1]
        return f'^(?:{inner}|{no_members_alt})$'

    return f'(?s)^(?:{pattern}|{no_members_alt})$'


def _build_registry_compliance_script(registry_path, expected_paths):
    path_checks = ' -and '.join(
        f"$paths -contains '{_ps_single_quote_escape(path)}'"
        for path in expected_paths
        if str(path or '').strip()
    )
    if not path_checks:
        return ''

    registry_path_ps = _ps_single_quote_escape(registry_path)
    return (
        f"if(Test-Path -Path '{registry_path_ps}'){{"
        f"$paths = (Get-ItemProperty -Path '{registry_path_ps}').Machine;"
        f"if([string]::IsNullOrWhiteSpace([string]$paths)){{ Write-Output '{NON_COMPLIANT_OUTPUT}'; return }};"
        f"if({path_checks}) {{ Write-Output 'Compliant' }} else {{ Write-Output 'Non-Compliant' }}"
        f"}}else {{ Write-Output '{NON_COMPLIANT_OUTPUT}' }}"
    )


def _build_generic_boolean_expected(raw_expected):
    expected = _prepare_boolean_expression_text(raw_expected)
    if '&&' not in expected and '||' not in expected:
        return _strip_outer_quotes_once(expected)

    expr = _strip_balanced_outer_parens(expected)
    alt_terms = []
    for alt in _split_top_level(expr, '||'):
        alt = _strip_balanced_outer_parens(alt)
        terms = []
        for term in _split_top_level(alt, '&&'):
            token = _strip_outer_quotes_once(_strip_balanced_outer_parens(term).strip())
            if token:
                terms.append(token)
        if terms:
            alt_terms.append(terms)

    if not alt_terms:
        return expected

    regex_alts = []
    for terms in alt_terms:
        if len(terms) == 1:
            regex_alts.append(_regex_escape_expected_literal(terms[0]))
            continue
        lookaheads = ''.join(
            f'(?=.*{_regex_escape_expected_literal(term)})'
            for term in terms
        )
        regex_alts.append(f'{lookaheads}.*')

    if len(regex_alts) == 1:
        only = regex_alts[0]
        if only.endswith('.*') and only.startswith('(?='):
            return f'(?s)^{only}$'
        return f'^{only}$'

    return '(?s)^(?:' + '|'.join(regex_alts) + ')$'


def _normalize_boolean_value_data_expression(raw_value_data):
    raw = str(raw_value_data or '').strip()
    if not raw:
        return ''

    normalized = raw.replace(r'\&\&', '&&').replace(r'\|\|', '||')
    if '&&' not in normalized and '||' not in normalized:
        return ''

    semantic = _build_generic_boolean_expected(normalized)
    if not semantic:
        return ''

    try:
        re.compile(semantic)
    except re.error:
        return ''

    return semantic


def _build_applocker_xml_expected_regex(raw_value_data):
    raw = _strip_outer_quotes_once(raw_value_data)
    if not raw:
        return ''

    compact = str(raw).replace('\\r\\n', '').replace('\\r', '').replace('\\n', '')
    compact = compact.replace('\r\n', '').replace('\r', '').replace('\n', '').strip()
    if not (compact.startswith('<FilePublisherRule') or compact.startswith('<FilePathRule')):
        return ''

    pattern = re.escape(compact)
    pattern = pattern.replace(r'LowSection="\*"', r'LowSection=".*"')
    pattern = pattern.replace(r'HighSection="\*"', r'HighSection=".*"')
    return f'(?s)^{pattern}.*$'


def _build_user_rights_expected(fields, fallback_expected):
    raw_expected = _prepare_boolean_expression_text(fallback_expected)
    if _looks_like_quoted_boolean_expression(raw_expected):
        expected = raw_expected
    else:
        expected = _strip_outer_quotes_once(fallback_expected).strip()
    desc_upper = _norm_upper_token(fields.get('description', ''))
    if 'NO ONE' in desc_upper and _norm_upper_token(expected) in {'.+', '^.+$', 'NO ONE'}:
        return '^NO_MEMBERS$'

    if expected.startswith('^'):
        return expected

    expected = _strip_balanced_outer_parens(expected)

    alt_sets = []
    for alt in _split_top_level(expected, '||'):
        members = []
        for raw_member in re.findall(r'"([^"]+)"', alt):
            member = _canonicalize_user_right_member(raw_member)
            if member and member not in members:
                members.append(member)
        if not members:
            continue
        normalized = tuple(sorted(members, key=lambda item: item.upper()))
        if normalized not in alt_sets:
            alt_sets.append(normalized)

    if not alt_sets:
        # Prevent emitting operator-only patterns when source expressions are malformed.
        if re.fullmatch(r'\^?\\?\|\\?\|\$?', expected):
            return '^.+$'
        return expected

    alt_sets.sort(key=lambda item: (len(item), [part.upper() for part in item]))

    if len(alt_sets) == 1:
        return '^' + ','.join(_regex_escape_user_right_member(part) for part in alt_sets[0]) + '$'

    chain = True
    for idx in range(1, len(alt_sets)):
        prev = alt_sets[idx - 1]
        curr = alt_sets[idx]
        if len(curr) != len(prev) + 1 or curr[:len(prev)] != prev:
            chain = False
            break

    if chain:
        pattern = '^' + _regex_escape_user_right_member(alt_sets[0][0])
        for idx in range(1, len(alt_sets[0])):
            pattern += f',{_regex_escape_user_right_member(alt_sets[0][idx])}'
        for idx in range(1, len(alt_sets)):
            added = alt_sets[idx][-1]
            pattern += f'(,{_regex_escape_user_right_member(added)})?'
        return pattern + '$'

    alternates = []
    for alt in alt_sets:
        alternates.append(','.join(_regex_escape_user_right_member(part) for part in alt))
    return '^(?:' + '|'.join(alternates) + ')$'


def _fallback_user_rights_expected_from_description(fields):
    desc = _strip_outer_quotes_once(fields.get('description', ''))
    if not desc:
        return ''

    m = re.search(r'\bis set to\b\s*(.+)$', desc, flags=re.IGNORECASE)
    if not m:
        return ''

    raw_tail = m.group(1).strip().rstrip('.')
    if not raw_tail:
        return ''

    members = []
    for part in raw_tail.split(','):
        member = _canonicalize_user_right_member(part)
        if member and member not in members:
            members.append(member)

    if not members:
        return ''

    members = sorted(members, key=lambda item: item.upper())
    return '^' + ','.join(_regex_escape_user_right_member(part) for part in members) + '$'


def _build_user_rights_powershell_script(right_type):
    right_ps = _ps_single_quote_escape(right_type)
    right_line_pattern = _ps_single_quote_escape(rf'^\s*{re.escape(right_type)}\s*=')
    return (
        "$t=[IO.Path]::GetTempFileName(); "
        "$ok=$false; "
        "try { secedit /export /cfg $t /areas USER_RIGHTS /mergedpolicy /quiet | Out-Null; $ok=$true } catch {} ; "
        "if(-not $ok){ secedit /export /cfg $t /areas USER_RIGHTS /quiet | Out-Null }; "
        f"$l=(Get-Content $t | Where-Object {{ $_ -match '{right_line_pattern}' }} | Select-Object -Last 1); "
        "$m=if($l){ ($l -replace '.*=', '').Split(',') | ForEach-Object { $s=$_.Trim(); $s=$s -replace '^\\*', ''; "
        "if($s -match '^S-1-5-'){ try{([Security.Principal.SecurityIdentifier]$s).Translate([Security.Principal.NTAccount]).Value.Split('\\\\')[-1]}catch{$s} } else { if($s -match '\\\\'){ $s.Split('\\\\')[-1] } else { $s } } } | Where-Object { $_ } | Sort-Object -Unique }else{ @() }; "
        "Remove-Item $t -Force -ErrorAction SilentlyContinue; if($m){ $m -join ',' } else { 'NO_MEMBERS' }"
    )


def _rewrite_user_rights_powershell_script(script_text):
    script = _strip_outer_quotes_once(script_text)
    if 'secedit /export /cfg $t /areas USER_RIGHTS' not in script:
        return script

    right_match = re.search(r"-match\s+'([^']+)'", script)
    if not right_match:
        return script

    right_expr = right_match.group(1)
    right_name_match = re.search(r'(Se[A-Za-z0-9_]+)', right_expr)
    if not right_name_match:
        return script

    return _build_user_rights_powershell_script(right_name_match.group(1))


def _build_sid_account_name_powershell_script(sid_suffix):
    return (
        "$A = Get-CimInstance -ClassName Win32_UserAccount -EA SilentlyContinue | "
        f"Where-Object {{ $_.SID -match '-{sid_suffix}$' }} | Select-Object -ExpandProperty Name -First 1; "
        f"if([string]::IsNullOrEmpty($A)){{ Write-Output '{NON_COMPLIANT_OUTPUT}' }} else {{ Write-Output $A }}"
    )


def _build_sid_account_status_powershell_script(sid_suffix):
    return (
        "$A = Get-CimInstance -ClassName Win32_UserAccount -EA SilentlyContinue | "
        f"Where-Object {{ $_.SID -match '-{sid_suffix}$' }} | Select-Object -First 1; "
        "if($A -and $A.Disabled -ne $null){ if($A.Disabled){ Write-Output 'Disabled' } else { Write-Output 'Enabled' } }"
        f" elseif($A -and -not [string]::IsNullOrEmpty($A.Name)){{ Write-Output $A.Name }} else {{ Write-Output '{NON_COMPLIANT_OUTPUT}' }}"
    )


def _apply_powershell_arg_corrections(desc_upper, script_text):
    script = str(script_text or '')

    # Canonicalize legacy lockout-admin scripts that output 1/0 or rely on
    # positional match state. Emit explicit status values and deterministic
    # non-compliant output when no data is returned.
    if 'ALLOWADMINISTRATORLOCKOUT' in _norm_upper_token(script):
        return (
            "$secfile = [System.IO.Path]::GetTempFileName(); "
            "secedit /export /cfg $secfile /areas SECURITYPOLICY /quiet | Out-Null; "
            "$line = Select-String -Path $secfile -Pattern 'AllowAdministratorLockout\\s*=\\s*(\\d)' | Select-Object -First 1; "
            "Remove-Item $secfile -ErrorAction SilentlyContinue; "
            "if($line -and $line.Matches.Count -gt 0){ "
            "if($line.Matches[0].Groups[1].Value -eq '1'){ Write-Output 'Enabled' } else { Write-Output 'Disabled' } "
            f"}} else {{ Write-Output '{NON_COMPLIANT_OUTPUT}' }}"
        )

    if 'REG_CHECK|DEFAULTPASSWORD|HKLM\\SOFTWARE\\MICROSOFT\\WINDOWS NT\\CURRENTVERSION\\WINLOGON' in _norm_upper_token(script):
        return (
            "$p='HKLM:\\Software\\Microsoft\\Windows NT\\CurrentVersion\\Winlogon'; "
            "if(Test-Path $p){ "
            "$v=(Get-ItemProperty $p -Name 'DefaultPassword' -EA SilentlyContinue).DefaultPassword; "
            "if([string]::IsNullOrEmpty($v)){ 'NOT_EXIST' }else{ 'EXIST' } "
            "}else{ 'NOT_EXIST' }"
        )

    # This control must explicitly check value existence, not emit a marker string.
    if 'ENSURE DEFAULTPASSWORD DOES NOT EXIST' in desc_upper:
        return (
            "$p='HKLM:\\Software\\Microsoft\\Windows NT\\CurrentVersion\\Winlogon'; "
            "if(Test-Path $p){ "
            "$v=(Get-ItemProperty $p -Name 'DefaultPassword' -EA SilentlyContinue).DefaultPassword; "
            "if([string]::IsNullOrEmpty($v)){ 'NOT_EXIST' }else{ 'EXIST' } "
            "}else{ 'NOT_EXIST' }"
        )

    # This policy is machine-scoped and should read from HKLM.
    if 'ALWAYS INSTALL WITH ELEVATED PRIVILEGES' in desc_upper:
        script = script.replace(
            'Registry::HKU\\Software\\Policies\\Microsoft\\Windows\\Installer',
            'Registry::HKLM\\Software\\Policies\\Microsoft\\Windows\\Installer',
        )

    return script


def _ensure_non_compliant_on_empty_output(script_text):
    script = _strip_outer_quotes_once(script_text)
    if not str(script or '').strip():
        return f"Write-Output '{NON_COMPLIANT_OUTPUT}'"

    # Avoid wrapping the same script multiple times across repeated runs.
    if '$__pysc_result' in script:
        return script

    return (
        "$__pysc_result = (& { "
        + script
        + " } | Out-String); "
        + f"if([string]::IsNullOrWhiteSpace([string]$__pysc_result)){{ Write-Output '{NON_COMPLIANT_OUTPUT}' }} "
        + "else { Write-Output ($__pysc_result.Trim()) }"
    )


def _build_check_account_powershell_script(fields):
    acct = _norm_upper_token(fields.get('account_type', ''))
    desc = _norm_upper_token(fields.get('description', ''))
    expected = _norm_upper_token(_best_effort_expected_text(fields))
    check_type = _norm_upper_token(fields.get('check_type', ''))

    if acct == 'ADMINISTRATOR_ACCOUNT':
        if 'RENAME' in desc or check_type in {'CHECK_NOT_REGEX', 'CHECK_NOT_EQUAL'}:
            return _build_sid_account_name_powershell_script('500')
        return _build_sid_account_status_powershell_script('500')

    if acct == 'GUEST_ACCOUNT':
        if 'RENAME' in desc or check_type == 'CHECK_NOT_EQUAL' or expected == 'GUEST':
            return _build_sid_account_name_powershell_script('501')
        return _build_sid_account_status_powershell_script('501')

    acct_ps = _ps_single_quote_escape(_strip_outer_quotes_once(fields.get('account_type', '')))
    return (
        "$A = Get-CimInstance -ClassName Win32_UserAccount -ErrorAction SilentlyContinue | "
        f"Where-Object {{ $_.Name -eq '{acct_ps}' }} | Select-Object -First 1; "
        f"if($A){{ Write-Output $A.Name }} else {{ Write-Output '{NON_COMPLIANT_OUTPUT}' }};"
    )


def _best_effort_powershell_script(fields):
    ctype = _norm_upper_token(fields.get('type', ''))
    expected = _best_effort_expected_text(fields)

    reg_key = _strip_outer_quotes_once(fields.get('reg_key', ''))
    reg_item = _strip_outer_quotes_once(fields.get('reg_item', fields.get('key_item', '')))
    if ctype in {'REGISTRY_SETTING', 'REG_CHECK', 'BANNER_CHECK'} and reg_key and reg_item:
        reg_key_ps = _ps_single_quote_escape(reg_key)
        reg_item_ps = _ps_single_quote_escape(reg_item)
        return (
            f"$P = 'Registry::{reg_key_ps}'; "
            f"$K = '{reg_item_ps}'; "
            f"if(-not (Test-Path -Path $P)){{ Write-Output '{NON_COMPLIANT_OUTPUT}'; return }}; "
            "$R = Get-ItemProperty -Path $P -ErrorAction SilentlyContinue; "
            f"if($null -eq $R){{ Write-Output '{NON_COMPLIANT_OUTPUT}'; return }}; "
            "$prop = $R.PSObject.Properties[$K]; "
            f"if($null -eq $prop -or $null -eq $prop.Value){{ Write-Output '{NON_COMPLIANT_OUTPUT}'; return }}; "
            "$V = $prop.Value; "
            f"if($null -eq $V -or [string]::IsNullOrWhiteSpace([string]$V)){{ Write-Output '{NON_COMPLIANT_OUTPUT}' }} "
            "elseif($V -is [System.Array]){ "
            "$joined = ($V -join [Environment]::NewLine); "
            f"if([string]::IsNullOrWhiteSpace([string]$joined)){{ Write-Output '{NON_COMPLIANT_OUTPUT}' }} else {{ Write-Output $joined }} }} "
            "else { Write-Output $V; }"
        )

    if ctype == 'USER_RIGHTS_POLICY':
        right = _strip_outer_quotes_once(fields.get('right_type', ''))
        if right:
            return _build_user_rights_powershell_script(right)

    if ctype == 'CHECK_ACCOUNT':
        acct = _strip_outer_quotes_once(fields.get('account_type', ''))
        if acct:
            return _build_check_account_powershell_script(fields)

    if ctype == 'WMI_POLICY':
        wmi_key = _strip_outer_quotes_once(fields.get('wmi_key', ''))
        wmi_ns = _strip_outer_quotes_once(fields.get('wmi_namespace', ''))
        wmi_req = _strip_outer_quotes_once(fields.get('wmi_request', ''))
        wmi_attr = _strip_outer_quotes_once(fields.get('wmi_attribute', ''))
        if wmi_req:
            wmi_req_ps = _ps_single_quote_escape(wmi_req)
            wmi_ns_ps = _ps_single_quote_escape(wmi_ns or 'root\\cimv2')
            if wmi_attr:
                wmi_attr_ps = _ps_single_quote_escape(wmi_attr)
                return (
                    f"$W = Get-WmiObject -Namespace '{wmi_ns_ps}' -Query '{wmi_req_ps}' -ErrorAction SilentlyContinue; "
                    f"if($null -eq $W){{ Write-Output '{NON_COMPLIANT_OUTPUT}'; return }}; "
                    f"$V = $W.{wmi_attr_ps}; "
                    f"if($null -eq $V -or [string]::IsNullOrWhiteSpace([string]$V)){{ Write-Output '{NON_COMPLIANT_OUTPUT}' }} else {{ Write-Output $V }};"
                )
            return (
                f"$W = Get-WmiObject -Namespace '{wmi_ns_ps}' -Query '{wmi_req_ps}' -ErrorAction SilentlyContinue; "
                f"if($null -eq $W){{ Write-Output '{NON_COMPLIANT_OUTPUT}' }} else {{ Write-Output $W }};"
            )
        if wmi_key:
            wmi_key_ps = _ps_single_quote_escape(wmi_key)
            return (
                f"$W = Get-WmiObject -Class '{wmi_key_ps}' -ErrorAction SilentlyContinue; "
                f"if($null -eq $W){{ Write-Output '{NON_COMPLIANT_OUTPUT}' }} else {{ Write-Output $W }};"
            )

    if ctype == 'CMD_EXEC':
        cmd = _strip_outer_quotes_once(fields.get('cmd', ''))
        if cmd:
            cmd_ps = _ps_single_quote_escape(cmd)
            return f"$C = cmd /c '{cmd_ps}'; Write-Output $C;"

    if ctype in {'FILE_CONTENT_CHECK', 'FILE_CONTENT_CHECK_NOT', 'FILE_CHECK', 'FILE_CHECK_NOT', 'OFFLINE_CONFIG_CHECK', 'OFFLINE_BANNER_CHECK'}:
        item = _strip_outer_quotes_once(fields.get('item', ''))
        if item:
            item_ps = _ps_single_quote_escape(item)
            return (
                f"if (Test-Path '{item_ps}') {{ Get-Content -Path '{item_ps}' -ErrorAction SilentlyContinue }} "
                f"else {{ Write-Output '{NON_COMPLIANT_OUTPUT}' }}"
            )

    marker_parts = [ctype]
    key_hint = _strip_outer_quotes_once(
        fields.get('reg_item')
        or fields.get('key_item')
        or fields.get('right_type')
        or fields.get('account_type')
        or fields.get('audit_policy_subcategory')
        or fields.get('item')
        or fields.get('cmd')
        or ''
    )
    if key_hint:
        marker_parts.append(key_hint)
    if expected:
        marker_parts.append(expected)
    marker = _ps_single_quote_escape('|'.join(marker_parts))
    return f"Write-Output '{marker}';"


def _convert_remaining_item_to_audit_powershell(fields):
    ctype = _norm_upper_token(fields.get('type', ''))
    if not ctype or ctype == 'AUDIT_POWERSHELL':
        return fields

    expected = _best_effort_expected_text(fields)
    original_check_type = str(fields.get('check_type', '') or '').strip()
    original_check_type_upper = _norm_upper_token(original_check_type)
    if ctype == 'USER_RIGHTS_POLICY' and original_check_type_upper not in {'CHECK_SUPERSET', 'CHECK_SUBSET', 'CHECK_EQUAL'}:
        expected = _build_user_rights_expected(fields, expected)
    script = _best_effort_powershell_script(fields)

    converted = OrderedDict()
    converted['type'] = 'AUDIT_POWERSHELL'
    for key in ('description', 'info', 'solution', 'reference', 'see_also'):
        if key in fields:
            converted[key] = fields.get(key)
    converted['value_type'] = 'POLICY_TEXT'
    converted['value_data'] = _format_converted_value_data(expected, original_check_type)
    converted['powershell_args'] = '-NoProfile -ExecutionPolicy Bypass -Command'
    converted['powershell_script'] = script
    converted['expect'] = _strip_outer_quotes_once(_format_converted_value_data(expected, original_check_type))
    return converted


def _should_force_convert_non_powershell(platform_hint, fields):
    hint = _norm_upper_token(platform_hint)
    if hint in {'MSSRV', 'MSWRK'}:
        return True

    desc_platform = _norm_upper_token(detect_platform(fields.get('description', '')))
    return desc_platform in {'MSSRV', 'MSWRK'}


def _is_windows_audit_platform(platform_hint='', fields=None):
    hint = _norm_upper_token(platform_hint)
    if hint in {'MSSRV', 'MSWRK'}:
        return True

    desc_platform = _norm_upper_token(detect_platform((fields or {}).get('description', '')))
    return desc_platform in {'MSSRV', 'MSWRK'}


def _ensure_audit_powershell_metadata_fields(fields):
    if _norm_upper_token(fields.get('type', '')) != 'AUDIT_POWERSHELL':
        return fields

    ensured = OrderedDict(fields)
    raw_expected = str(ensured.get('value_data', '') or '').strip()
    if raw_expected:
        script = _strip_outer_quotes_once(ensured.get('powershell_args', ''))
        check_type_upper = _norm_upper_token(ensured.get('check_type', ''))
        if (
            'secedit /export /cfg $t /areas USER_RIGHTS' in script
            and _looks_like_quoted_boolean_expression(raw_expected)
        ):
            normalized_expected = _build_user_rights_expected(ensured, raw_expected)
            ensured['value_data'] = f'"{normalized_expected}"'
        else:
            semantic_boolean = _normalize_boolean_value_data_expression(raw_expected)
            if semantic_boolean:
                ensured['value_data'] = f'"{semantic_boolean}"'
            else:
                normalized_expected = _build_bracket_range_expected(raw_expected)
                if normalized_expected:
                    ensured['value_data'] = f'"{normalized_expected}"'
    if 'solution' not in ensured:
        ensured['solution'] = '""'
    if 'reference' not in ensured:
        ensured['reference'] = '""'
    if 'see_also' not in ensured:
        ensured['see_also'] = f'"{SEE_ALSO_REPLACEMENT}"'
    return ensured


def _finalize_audit_powershell_fields_for_emit(fields, platform_hint=''):
    if _norm_upper_token(fields.get('type', '')) != 'AUDIT_POWERSHELL':
        return OrderedDict(fields)

    finalized = OrderedDict()
    for key, value in fields.items():
        if key in {'powershell_args', 'powershell_script', 'powershell_option'}:
            continue
        if key == 'expect' and _is_windows_audit_platform(platform_hint, fields):
            continue
        finalized[key] = value

    ps_args = fields.get('powershell_script', fields.get('powershell_args', ''))

    # Normalize known value-data drift that causes false mismatches.
    raw_value_data = str(finalized.get('value_data', '') or '')
    if 'Sysmonlog' in raw_value_data:
        finalized['value_data'] = raw_value_data.replace('Sysmonlog', 'SysmonLog')

    desc_upper = _norm_upper_token(_strip_outer_quotes_once(finalized.get('description', '')))
    forced_ps_args = None
    if (
        'ACCOUNT LOCKOUT RESET' in desc_upper
        or 'ACCOUNT LOCKOUT DURATION' in desc_upper
        or 'ACCOUNT LOCKOUT THRESHOLD' in desc_upper
    ):
        scalar = _first_int(_strip_outer_quotes_once(finalized.get('value_data', '')))
        if scalar is not None:
            finalized['value_data'] = f'"{scalar}"'

    # Keep expected values aligned with canonical lockout-admin script output.
    if 'ALLOW ADMINISTRATOR ACCOUNT LOCKOUT' in desc_upper:
        expected_inner = _norm_upper_token(_strip_outer_quotes_once(finalized.get('value_data', '')))
        if expected_inner in {'1', 'ENABLED', 'TRUE'}:
            finalized['value_data'] = '"(?i)^Enabled$"'
            finalized['check_type'] = 'CHECK_REGEX'
        elif expected_inner in {'0', 'DISABLED', 'FALSE'}:
            finalized['value_data'] = '"(?i)^Disabled$"'
            finalized['check_type'] = 'CHECK_REGEX'

    # For password expiry warning, enforce a minimum threshold instead of hard-capping
    # so values above the documented minimum still pass.
    if 'PROMPT USER TO CHANGE PASSWORD BEFORE EXPIRATION' in desc_upper:
        min_days = _first_int(_strip_outer_quotes_once(finalized.get('value_data', '')))
        if min_days is not None:
            finalized['value_data'] = f'"{_build_ge_integer_regex(min_days)}"'

    if ps_args:
        script_text = _strip_outer_quotes_once(ps_args)
        script_text = _normalize_embedded_powershell_script(script_text)

        # Repair broken USER_RIGHTS expected patterns that collapse to only
        # operators/anchors after escaping (for example "^\^\|\|\$$").
        value_data_inner = _strip_outer_quotes_once(finalized.get('value_data', ''))
        if (
            'secedit /export /cfg $t /areas USER_RIGHTS' in script_text
            and value_data_inner
            and re.fullmatch(r'[\^$\\|()\s]+', value_data_inner)
        ):
            fallback_expected = _fallback_user_rights_expected_from_description(fields)
            if fallback_expected:
                finalized['value_data'] = f'"{fallback_expected}"'
                value_data_inner = fallback_expected

        # In merged MSSRV output, many environments return NO_MEMBERS for
        # USER_RIGHTS checks where the baseline text expects explicit members.
        # Allow that runtime value while still preserving the canonical list match.
        if (
            'secedit /export /cfg $t /areas USER_RIGHTS' in script_text
            and 'NO ONE' not in desc_upper
            and 'DENY LOG ON LOCALLY TO INCLUDE GUESTS' not in desc_upper
        ):
            expected_inner = _strip_outer_quotes_once(finalized.get('value_data', ''))
            expected_inner = _normalize_user_rights_expected_for_tenable(fields, expected_inner)
            if expected_inner and 'NO_MEMBERS' not in expected_inner:
                finalized['value_data'] = f'"{_allow_no_members_alternative(expected_inner)}"'
            elif expected_inner and 'NO_MEMBERS' in expected_inner:
                normalized_expected = _normalize_no_members_token_regex(expected_inner)
                finalized['value_data'] = f'"{normalized_expected}"'

            # Tenable parser safety: this control is frequently emitted with an
            # over-escaped alternate expression for WdiServiceHost that can
            # produce unknown token parse failures (for example "\\\ ").
            if 'PROFILE SYSTEM PERFORMANCE' in desc_upper and 'WDISERVICEHOST' in desc_upper:
                finalized['value_data'] = '"^(?:Administrators,(?:NT SERVICE\\\\)?WdiServiceHost|\'?NO_MEMBERS\'?)$"'

        # Keep this control strict: custom groups should fail unless Guests is present as expected.
        if 'DENY LOG ON LOCALLY TO INCLUDE GUESTS' in desc_upper:
            finalized['value_data'] = '"^Guests$"'

        # This USER_RIGHTS check is emitted as a regex pattern and must be
        # evaluated with regex semantics to allow Administrators or NO_MEMBERS.
        if 'PERFORM VOLUME MAINTENANCE TASKS IS SET TO ADMINISTRATORS' in desc_upper:
            finalized['check_type'] = 'CHECK_REGEX'

        # Baseline-aligned expected lists for remotely accessible registry paths.
        if 'NETWORK ACCESS - REMOTELY ACCESSIBLE REGISTRY PATHS AND SUB-PATHS ALLOWED EXACT PATHS IS CONFIGURED' in desc_upper:
            finalized['info'] = (
                '"This audit is written to dynamically identify if all paths are present in any order. '
                'This policy setting determines which registry paths will be accessible over the network, '
                'regardless of the users or groups listed in the access control list (ACL) of the winreg registry key."'
            )
            finalized['value_data'] = '"^Compliant$"'
            ps_args = _build_registry_compliance_script(
                r'HKLM:\System\CurrentControlSet\Control\SecurePipeServers\Winreg\AllowedExactPaths',
                [
                    'System\\CurrentControlSet\\Control\\ProductOptions',
                    'System\\CurrentControlSet\\Control\\Server Applications',
                    'Software\\Microsoft\\Windows NT\\CurrentVersion',
                ],
            )
            if ps_args:
                forced_ps_args = f'"{ps_args}"'
            finalized['check_type'] = 'CHECK_REGEX'
        if 'NETWORK ACCESS - REMOTELY ACCESSIBLE REGISTRY PATHS AND SUB-PATHS ALLOWED PATHS IS CONFIGURED' in desc_upper:
            finalized['info'] = (
                '"This audit is written to dynamically identify if all paths are present in any order. '
                'This policy setting determines which registry paths and sub-paths will be accessible over the network, '
                'regardless of the users or groups listed in the access control list (ACL) of the winreg registry key."'
            )
            finalized['value_data'] = '"^Compliant$"'
            ps_args = _build_registry_compliance_script(
                r'HKLM:\System\CurrentControlSet\Control\SecurePipeServers\Winreg\AllowedPaths',
                [
                    'System\\CurrentControlSet\\Control\\Print\\Printers',
                    'System\\CurrentControlSet\\Services\\Eventlog',
                    'Software\\Microsoft\\OLAP Server',
                    'Software\\Microsoft\\Windows NT\\CurrentVersion\\Print',
                    'Software\\Microsoft\\Windows NT\\CurrentVersion\\Windows',
                    'System\\CurrentControlSet\\Control\\ContentIndex',
                    'System\\CurrentControlSet\\Control\\Terminal Server',
                    'System\\CurrentControlSet\\Control\\Terminal Server\\UserConfig',
                    'System\\CurrentControlSet\\Control\\Terminal Server\\DefaultUserConfiguration',
                    'Software\\Microsoft\\Windows NT\\CurrentVersion\\Perflib',
                    'System\\CurrentControlSet\\Services\\SysmonLog',
                ],
            )
            if ps_args:
                forced_ps_args = f'"{ps_args}"'
            finalized['check_type'] = 'CHECK_REGEX'

        if 'SERVICES - RAPID7 INSIGHT AGENT SERVICE' in desc_upper:
            finalized['value_data'] = '"Running"'
            finalized['check_type'] = 'CHECK_REGEX'
            ps_args = '"$noutput = (Get-wmiobject win32_service | where-object {$_.name -eq \'ir_agent\'}).state; if ($noutput -eq $Null) {write-host \'Service Not Found\'} else {$noutput}"'

        # Normalize known script path typos observed in source audits.
        script_text = re.sub(
            r"(?i)Services\\Netlogon\\Parameter(?!s)(['\\])",
            r"Services\\Netlogon\\Parameters\1",
            script_text,
        )
        script_text = re.sub(
            r'(?i)Services\\Netlogon\\Parameters+',
            r'Services\\Netlogon\\Parameters',
            script_text,
        )
        script_text = re.sub(
            r"(?i)PublicProfile\\Loggin\\Software\\Policies\\Microsoft\\WindowsFirewall\\PublicProfile\\Logging",
            r"PublicProfile\\Logging",
            script_text,
        )

        # Persist normalized script text by default; conditional rewrites below
        # may still replace this with more specialized scripts.
        script_text = _apply_powershell_arg_corrections(desc_upper, script_text)
        script_text = re.sub(
            r"(?i)Write-Output\s+['\"](?:Value Not Found|Path Not Found|NOT_FOUND|Not Found)['\"]",
            f"Write-Output '{NON_COMPLIANT_OUTPUT}'",
            script_text,
        )
        ps_args = f'"{script_text}"'
        if forced_ps_args is not None:
            ps_args = forced_ps_args

        if 'PasswordExpiryWarning' in script_text:
            min_days = _first_int(_strip_outer_quotes_once(finalized.get('value_data', '')))
            if min_days is not None:
                finalized['value_data'] = f'"{_build_ge_integer_regex(min_days)}"'

        normalized_expected = _normalize_legacy_net_accounts_expected(script_text, finalized.get('value_data', ''))
        if normalized_expected != finalized.get('value_data', ''):
            finalized['value_data'] = normalized_expected

        # AppLocker XML payloads can differ by trailing line endings; normalize
        # these checks to regex mode and tolerate wildcard version ranges.
        applocker_expected = _build_applocker_xml_expected_regex(finalized.get('value_data', ''))
        if applocker_expected:
            finalized['check_type'] = 'CHECK_REGEX'
            finalized['value_data'] = f'"{applocker_expected}"'
            if not _is_windows_audit_platform(platform_hint, fields):
                finalized['expect'] = f'"{applocker_expected}"'

        if "net user 'ADMINISTRATOR_ACCOUNT'" in script_text:
            ps_args = f'"{_build_sid_account_name_powershell_script("500")}"'
        elif "net user 'GUEST_ACCOUNT'" in script_text:
            ps_args = f'"{_build_check_account_powershell_script(fields | {"account_type": "GUEST_ACCOUNT"})}"'
        elif 'secedit /export /cfg $t /areas USER_RIGHTS' in script_text:
            ps_args = f'"{_rewrite_user_rights_powershell_script(script_text)}"'
        elif re.search(r'\bnet\s+accounts\b', script_text, flags=re.IGNORECASE) and re.search(r'password\s+complexity', script_text, flags=re.IGNORECASE):
            ps_args = f'"{_rewrite_password_complexity_script(script_text)}"'
        elif 'RestrictSendingNTLMTraffic' in script_text:
            ps_args = f'"{_rewrite_ntlm_outgoing_traffic_script(script_text)}"'
        elif re.search(r'(?i)Services\\Netlogon\\Parameters', script_text):
            ps_args = f'"{_rewrite_netlogon_parameters_default_script(script_text)}"'
        elif re.search(r'\bauditpol\b', script_text, flags=re.IGNORECASE):
            ps_args = f'"{_rewrite_auditpol_powershell_script(script_text)}"'

        # Universal guard: if any control script returns no data, force
        # deterministic non-compliant output.
        ps_args = f'"{_ensure_non_compliant_on_empty_output(ps_args)}"'

        # Keep the canonical Tenable field name for PowerShell checks.
        finalized['powershell_args'] = _render_powershell_args_value(ps_args)

        # Emit option only when explicitly non-default to keep output clean.
        ps_option = str(fields.get('powershell_option', '') or '').strip()
        if ps_option and _norm_upper_token(ps_option) != 'CAN_BE_EQUAL':
            finalized['powershell_option'] = ps_option

        # Final sticky override after all rewrites: local admin must be disabled.
        if 'ACCOUNT - LOCAL ADMIN ACCOUNT' in desc_upper:
            desc_text = _strip_outer_quotes_once(finalized.get('description', ''))
            desc_text = re.sub(
                r'(?i)Account\s*-\s*Local\s+Admin\s+Account\s+Enabled',
                'Account - Local Admin Account Disabled',
                desc_text,
            )
            finalized['description'] = f'"{desc_text}"'
            finalized['value_data'] = '"Disabled"'
            finalized['powershell_args'] = _render_powershell_args_value(_build_sid_account_status_powershell_script("500"))

    return finalized


def _strip_outer_quotes_once(value):
    if value is None:
        return ''
    s = str(value).strip()
    if len(s) >= 2 and s[0] == s[-1] and s[0] in ('"', "'"):
        return s[1:-1]
    return s


def _strip_bom_prefix(value):
    if value is None:
        return ''
    text = str(value).replace('\ufeff', '').replace('ï»¿', '')
    return text.lstrip('\ufeff').lstrip('ï»¿')


def _sanitize_audit_lines(lines):
    cleaned = []
    for index, line in enumerate(lines or []):
        text = _strip_bom_prefix(line)
        if index == 0:
            text = text.lstrip('\ufeff').lstrip('ï»¿')
        cleaned.append(text)
    return cleaned


def _assert_no_encoding_markers(text, context):
    issues = []
    if '\ufeff' in text:
        issues.append('U+FEFF')
    if 'ï»¿' in text:
        issues.append('mojibake-UTF8-BOM')
    if issues:
        raise ValueError(f"{context}: forbidden encoding markers found ({', '.join(issues)})")


def _norm_token(value):
    if value is None:
        return ''
    s = str(value).strip().strip('"').strip("'")
    s = re.sub(r'\s+', ' ', s)
    return s


def _norm_upper_token(value):
    return _norm_token(value).upper()


def _normalize_user_rights_expected_for_tenable(fields, expected_text):
    expected = _strip_outer_quotes_once(expected_text)
    if not expected:
        return ''

    # Repair over-escaped regex fragments that Tenable can reject,
    # then rebuild a clean expected pattern from the control description.
    if r'\(\?' in expected or r'\|\|' in expected or r'\ "' in expected or '\\ ' in expected:
        fallback = _fallback_user_rights_expected_from_description(fields)
        if fallback:
            return _allow_no_members_alternative(fallback)

    # Source variants may include quoted alternatives like
    # "NT SERVICE\\WdiServiceHost" || "WdiServiceHost".
    # Collapse those to a deterministic comma-delimited expectation.
    if '||' in expected and 'WDISERVICEHOST' in _norm_upper_token(expected):
        return _allow_no_members_alternative('^Administrators,(NT SERVICE\\\\)?WdiServiceHost$')

    return expected


def enforce_solution_reference_adjacency(pairs):
    """Ensure solution and reference are adjacent in output ordering."""
    keys = [k for k, _ in pairs]
    if "solution" not in keys or "reference" not in keys:
        return pairs

    solution_idx = keys.index("solution")
    solution_val = pairs[solution_idx][1]
    reference_val = pairs[keys.index("reference")][1]

    remainder = [(k, v) for (k, v) in pairs if k not in ("solution", "reference")]
    insert_at = min(solution_idx, len(remainder))
    return (
        remainder[:insert_at]
        + [("solution", solution_val), ("reference", reference_val)]
        + remainder[insert_at:]
    )


CUSTOM_ITEM_FIELD_ORDER = [
    'type',
    'description',
    'info',
    'solution',
    'reference',
    'see_also',
    'value_type',
    'value_data',
    'powershell_option',
    'powershell_args',
    'expect',
    'check_type',
]


def order_custom_item_pairs(pairs):
    """Order custom_item fields in a consistent sequence for all emitted blocks."""
    by_key = OrderedDict()
    for k, v in pairs:
        if k not in by_key:
            by_key[k] = v

    ordered = []
    for key in CUSTOM_ITEM_FIELD_ORDER:
        if key in by_key:
            ordered.append((key, by_key.pop(key)))

    for key, val in by_key.items():
        ordered.append((key, val))

    return ordered


PLATFORM_RE = re.compile(r"\b([A-Z]{2,6})\b")


def detect_platform(description):
    if not description:
        return 'UNKNOWN'
    d = description.strip().strip('"').strip("'")
    m = re.search(r"-\s*([A-Z]{2,6})\s*-", d)
    if m:
        return m.group(1)
    m2 = PLATFORM_RE.search(d)
    if m2:
        return m2.group(1)
    return 'UNKNOWN'


def determine_platform_from_filename(path):
    name = os.path.basename(path).lower()
    # New rules: 'vmware' -> VMware, 'enterprise' + linux/rhel -> RHEL
    if 'vmware' in name:
        return 'VMware'
    if 'mssrv' in name:
        return 'MSSRV'
    if 'mswrk' in name:
        return 'MSWRK'
    if 'enterprise' in name and ('linux' in name or 'rhel' in name):
        return 'RHEL'
    if 'enterprise' in name:
        return 'MSWRK'
    if 'sql' in name:
        return 'SQL'
    if 'server' in name and 'sql' not in name:
        return 'MSSRV'
    if 'ios' in name:
        return 'IOS'
    if 'palo alto' in name or ('palo' in name and 'alto' in name):
        return 'PAFW'
    if 'pafw' in name:
        return 'PAFW'
    if 'nx' in name:
        return 'NX-OS'
    if 'f5' in name:
        return 'F5'
    if 'azure' in name:
        return 'MSAZ'
    if 'asa' in name:
        return 'ASA'
    if 'amazon' in name or 'aws' in name:
        return 'Amazon'
    return 'UNKNOWN'


WINDOWS_PARSE_POLICY = {'allow_repair_fallback': False, 'repair_pipeline': ('quoted_fields',)}


RHEL_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('wrapper', 'quoted_fields', 'value_data')}


SQL_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('quoted_fields', 'value_data', 'powershell', 'wrapper')}


IOS_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('wrapper', 'quoted_fields')}


PAFW_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('quoted_fields', 'value_data', 'wrapper')}


NXOS_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('wrapper', 'quoted_fields', 'value_data')}


F5_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('quoted_fields', 'powershell', 'wrapper')}


MSAZ_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('quoted_fields', 'value_data', 'powershell')}


ASA_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('wrapper', 'quoted_fields', 'value_data', 'powershell')}


AMAZON_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('quoted_fields', 'value_data', 'wrapper', 'powershell')}


VMWARE_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('quoted_fields', 'powershell', 'value_data', 'wrapper')}


UNKNOWN_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('quoted_fields', 'value_data')}


PLATFORM_PARSE_PROFILES = {
    # Windows baselines: shared strict parser-only policy.
    'MSSRV': WINDOWS_PARSE_POLICY,
    'MSWRK': WINDOWS_PARSE_POLICY,

    # Platform-specific parser profiles (tunable independently).
    'RHEL': RHEL_PARSE_POLICY,
    'SQL': SQL_PARSE_POLICY,
    'IOS': IOS_PARSE_POLICY,
    'PAFW': PAFW_PARSE_POLICY,
    'NX-OS': NXOS_PARSE_POLICY,
    'F5': F5_PARSE_POLICY,
    'MSAZ': MSAZ_PARSE_POLICY,
    'ASA': ASA_PARSE_POLICY,
    'AMAZON': AMAZON_PARSE_POLICY,
    'VMWARE': VMWARE_PARSE_POLICY,
    'UNKNOWN': UNKNOWN_PARSE_POLICY,
}


def _normalized_platform_key(platform_hint):
    platform = _norm_upper_token(platform_hint)
    if platform == 'VMWARE':
        return 'VMWARE'
    return platform


def _platform_parse_profile(platform_hint):
    platform_key = _normalized_platform_key(platform_hint)
    default_profile = UNKNOWN_PARSE_POLICY
    return platform_key, PLATFORM_PARSE_PROFILES.get(platform_key, default_profile)


def _apply_parse_repair_pipeline(text, repair_pipeline):
    current = str(text or '')
    for step in repair_pipeline:
        if step == 'value_data':
            current, _ = _repair_malformed_value_data_lines(current)
        elif step == 'powershell':
            current, _ = _repair_malformed_powershell_lines(current)
        elif step == 'quoted_fields':
            current, _ = _repair_malformed_quoted_field_lines(current)
        elif step == 'wrapper':
            current = _normalize_tenable_audit_wrapper(current)
    return current


def _save_workbook_with_lock_fallback(wb, outpath):
    temp_outpath = f"{outpath}.{os.getpid()}.tmp"
    wb.save(temp_outpath)

    # Windows can hold transient locks (Explorer previews, AV, sync clients), so retry briefly.
    for _ in range(8):
        try:
            os.replace(temp_outpath, outpath)
            return outpath
        except PermissionError:
            time.sleep(0.25)

    base, ext = os.path.splitext(outpath)
    fallback_outpath = _timestamped_output_path(f"{base}.new{ext or '.xlsx'}")
    wb.save(fallback_outpath)

    try:
        os.remove(temp_outpath)
    except Exception:
        pass

    print(f"Could not replace locked workbook '{outpath}'. Written to '{fallback_outpath}' instead.")
    return fallback_outpath


def _sanitize_sheet_title(title, existing_titles):
    cleaned = re.sub(r'[\\/\?\*\[\]:]', '_', str(title)).strip()
    if not cleaned:
        cleaned = 'Sheet'
    cleaned = cleaned[:31]

    base = cleaned
    suffix = 1
    while cleaned in existing_titles:
        extra = f'_{suffix}'
        cleaned = f'{base[:31 - len(extra)]}{extra}'
        suffix += 1

    existing_titles.add(cleaned)
    return cleaned


def _timestamped_output_path(path):
    base, ext = os.path.splitext(path)
    if _TS_SUFFIX_RE.search(base):
        return path
    return f'{base}_{RUN_TIMESTAMP}{ext}'


def _parsing_results_output_path(folder_path):
    return _timestamped_output_path(os.path.join(folder_path, PARSING_RESULTS_FILENAME))


def _write_parsing_results_workbook(records, outpath):
    if Workbook is None:
        raise RuntimeError('openpyxl is required to write Excel workbook; pip install openpyxl')

    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet(title='Parsing Log')
    summary_headers = [
        'audit_file',
        'source_folder',
        'status',
        'node_count',
        'custom_item_count',
        'report_count',
        'text_count',
        'unknown_key_count',
        'unknown_keys',
        'normalized_output',
        'notes',
    ]
    summary_ws.append(summary_headers)

    existing_titles = {'Parsing Log'}
    for record in records:
        summary_ws.append([record.get(h, '') for h in summary_headers])

        sheet_title = _sanitize_sheet_title(record.get('sheet_name') or record.get('audit_file') or 'Audit', existing_titles)
        ws = wb.create_sheet(title=sheet_title)
        ws.append(['row_type', 'label', 'value'])
        ws.append(['audit_file', 'value', record.get('audit_file', '')])
        ws.append(['source_folder', 'value', record.get('source_folder', '')])
        ws.append(['status', 'value', record.get('status', '')])
        ws.append(['node_count', 'value', record.get('node_count', 0)])
        ws.append(['custom_item_count', 'value', record.get('custom_item_count', 0)])
        ws.append(['report_count', 'value', record.get('report_count', 0)])
        ws.append(['text_count', 'value', record.get('text_count', 0)])
        ws.append(['unknown_key_count', 'value', record.get('unknown_key_count', 0)])
        ws.append(['unknown_keys', 'value', record.get('unknown_keys', '')])
        ws.append(['normalized_output', 'value', record.get('normalized_output', '')])
        ws.append(['notes', 'value', record.get('notes', '')])

        ws.append([])
        ws.append(['node_index', 'node_type', 'details'])
        for node in record.get('nodes', []):
            ws.append([
                node.get('node_index', ''),
                node.get('node_type', ''),
                json.dumps(node, ensure_ascii=False),
            ])

    final_outpath = _save_workbook_with_lock_fallback(wb, outpath)
    print(f'Wrote parsing results workbook: {final_outpath}')
    return final_outpath


def _write_parsing_results_for_folder(folder_path):
    folder_path = os.path.abspath(folder_path)
    records = PARSING_RESULTS_BY_FOLDER.get(folder_path, [])
    if not records:
        return None
    return _write_parsing_results_workbook(records, _parsing_results_output_path(folder_path))


def _strip_redundant_platform_description_prefix(text):
    inner = _strip_outer_quotes_once(text)
    if not inner:
        return ''

    while True:
        cleaned = re.sub(
            r'^[\-\u2013\u2014]?\s*(?P<platform>MSWRK|MSSRV|WINSRV|WINWRK)\s*-\s*[A-Za-z]{2}\s*-\s*(?P=platform)\s*-\s*',
            '',
            inner,
            flags=re.IGNORECASE,
        ).strip()
        if cleaned == inner:
            break
        inner = cleaned

    inner = re.sub(r'^[\-\u2013\u2014]?\s*(?:MSWRK|MSSRV|WINSRV|WINWRK)\s*[\-\u2013\u2014]\s*', '', inner, flags=re.IGNORECASE).strip()
    inner = re.sub(r'^[\-\u2013\u2014]?\s*[A-Za-z]{2}\s*[\-\u2013\u2014]\s*', '', inner).strip()
    return inner


def _prefixed_description(desc_value, platform, ordinal):
    desc_norm = normalize_description(desc_value)
    if not desc_norm:
        desc_norm = '"Control"'

    inner = _strip_redundant_platform_description_prefix(desc_norm)
    inner = re.sub(r'^\d+\.\d{4}\s*-\s*[A-Za-z0-9_-]+\s*-\s*', '', inner).strip()
    inner = re.sub(r'^\d+(?:\.\d+)+\s*', '', inner).strip()
    inner = re.sub(r'^[\-\u2013\u2014]\s*[A-Za-z0-9_-]+\s*[\-\u2013\u2014]\s*', '', inner).strip()
    inner = re.sub(r'^(MSWRK|MSSRV|WINSRV|WINWRK)\s*-\s*', '', inner, flags=re.IGNORECASE).strip()
    inner = re.sub(r'^[\-\u2013\u2014]?\s*' + re.escape(platform) + r'\s*[\-\u2013\u2014]\s*', '', inner, flags=re.IGNORECASE).strip()

    # Remove control-family/platform prefixes from source descriptions.
    # Example: "AC - MSSRV - Perform volume maintenance tasks..." -> "Perform volume maintenance tasks..."
    while True:
        cleaned = re.sub(
            r'^(AC|AU|CM|IA|SC)\s*-\s*(MSWRK|MSSRV)\s*-\s*',
            '',
            inner,
            flags=re.IGNORECASE,
        ).strip()
        if cleaned == inner:
            break
        inner = cleaned

    number = 1.0 + (ordinal / 10000.0)
    return f'"{number:.4f} - {platform} - {inner}"'


def _is_target_os_installed_description(desc_text):
    desc = _norm_upper_token(_strip_outer_quotes_once(desc_text or ''))
    return 'WINDOWS SERVER IS INSTALLED' in desc or 'WINDOWS WORKSTATION IS INSTALLED' in desc


def _drop_target_os_passed_reports(lines):
    cleaned = []
    i = 0
    while i < len(lines):
        if re.match(r'^\s*<report\b[^>]*>\s*$', lines[i], flags=re.IGNORECASE):
            j = i + 1
            block = [lines[i]]
            while j < len(lines):
                block.append(lines[j])
                if re.match(r'^\s*</report>\s*$', lines[j], flags=re.IGNORECASE):
                    break
                j += 1

            block_text = '\n'.join(block)
            if re.search(r'PASSED\s*-\s*TARGET\s+OS\s+MATCHES\s+BASELINE', block_text, flags=re.IGNORECASE):
                i = j + 1
                continue

            cleaned.extend(block)
            i = j + 1
            continue

        cleaned.append(lines[i])
        i += 1

    return cleaned


def _drop_target_os_custom_items(lines):
    cleaned = []
    i = 0
    while i < len(lines):
        if re.match(r'^\s*<custom_item>\s*$', lines[i], flags=re.IGNORECASE):
            j = i + 1
            block = [lines[i]]
            while j < len(lines):
                block.append(lines[j])
                if re.match(r'^\s*</custom_item>\s*$', lines[j], flags=re.IGNORECASE):
                    break
                j += 1

            desc_text = ''
            for line in block:
                m = re.match(r'^\s*description\s*:\s*(.+?)\s*$', line, flags=re.IGNORECASE)
                if m:
                    desc_text = m.group(1)
                    break

            if _is_target_os_installed_description(desc_text):
                i = j + 1
                continue

            cleaned.extend(block)
            i = j + 1
            continue

        cleaned.append(lines[i])
        i += 1

    return cleaned


def _drop_rapid7_service_item_blocks(lines):
    cleaned = []
    i = 0
    while i < len(lines):
        if re.match(r'^\s*<(custom_item|item)>\s*$', lines[i], flags=re.IGNORECASE):
            m_open = re.match(r'^\s*<(custom_item|item)>\s*$', lines[i], flags=re.IGNORECASE)
            tag_name = m_open.group(1).lower() if m_open else 'custom_item'
            close_re = rf'^\s*</{tag_name}>\s*$'

            j = i + 1
            block = [lines[i]]
            while j < len(lines):
                block.append(lines[j])
                if re.match(close_re, lines[j], flags=re.IGNORECASE):
                    break
                j += 1

            block_text_upper = _norm_upper_token('\n'.join(block))
            if 'RAPID7' in block_text_upper and 'SERVICE' in block_text_upper:
                i = j + 1
                continue

            cleaned.extend(block)
            i = j + 1
            continue

        cleaned.append(lines[i])
        i += 1

    return cleaned


def _strip_target_os_applicability_wrappers(lines):
    out = []
    i = 0
    while i < len(lines):
        if not re.match(r'^\s*<if>\s*$', lines[i], flags=re.IGNORECASE):
            out.append(lines[i])
            i += 1
            continue

        depth = 0
        j = i
        while j < len(lines):
            if re.match(r'^\s*<if>\s*$', lines[j], flags=re.IGNORECASE):
                depth += 1
            elif re.match(r'^\s*</if>\s*$', lines[j], flags=re.IGNORECASE):
                depth -= 1
                if depth == 0:
                    break
            j += 1

        if j >= len(lines):
            out.append(lines[i])
            i += 1
            continue

        block = lines[i:j + 1]

        has_target_gate = any(
            re.match(r'^\s*description\s*:\s*.+$', line, flags=re.IGNORECASE)
            and _is_target_os_installed_description(line.split(':', 1)[1].strip())
            for line in block
        )

        if not has_target_gate:
            out.extend(block)
            i = j + 1
            continue

        then_start = None
        then_end = None
        depth = 0
        for k, line in enumerate(block):
            if re.match(r'^\s*<if>\s*$', line, flags=re.IGNORECASE):
                depth += 1
            elif re.match(r'^\s*</if>\s*$', line, flags=re.IGNORECASE):
                depth -= 1
            elif depth == 1 and re.match(r'^\s*<then>\s*$', line, flags=re.IGNORECASE):
                then_start = k
            elif depth == 1 and re.match(r'^\s*</then>\s*$', line, flags=re.IGNORECASE):
                then_end = k

        if then_start is None or then_end is None or then_end <= then_start:
            out.extend(block)
            i = j + 1
            continue

        then_body = block[then_start + 1:then_end]
        then_body = _drop_target_os_passed_reports(then_body)
        then_body = _drop_target_os_custom_items(then_body)
        then_body = _drop_rapid7_service_item_blocks(then_body)
        out.extend(then_body)
        i = j + 1

    return out


def _cleanup_normalized_output_lines(lines):
    cleaned = _sanitize_audit_lines(lines)
    cleaned = _strip_target_os_applicability_wrappers(cleaned)
    cleaned = _drop_target_os_custom_items(cleaned)
    cleaned = _drop_rapid7_service_item_blocks(cleaned)
    cleaned = _drop_target_os_passed_reports(cleaned)
    return cleaned


def _docker_desktop_executable_candidates():
    candidates = []
    for base in (
        os.environ.get('ProgramFiles', ''),
        os.environ.get('ProgramFiles(x86)', ''),
        os.environ.get('LocalAppData', ''),
    ):
        if not base:
            continue
        candidates.append(os.path.join(base, 'Docker', 'Docker', 'Docker Desktop.exe'))
    return candidates


def _start_docker_desktop_if_available():
    for candidate in _docker_desktop_executable_candidates():
        if not os.path.isfile(candidate):
            continue
        try:
            subprocess.Popen(
                [candidate],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                stdin=subprocess.DEVNULL,
            )
            return True, f'Started Docker Desktop: {candidate}'
        except Exception as exc:
            return False, f'Could not start Docker Desktop: {exc}'
    return False, 'Docker Desktop executable was not found'


def _wait_for_docker_daemon(timeout_seconds=120, poll_seconds=3):
    deadline = time.time() + timeout_seconds
    last_output = ''

    while time.time() < deadline:
        try:
            probe = subprocess.run(['docker', 'info'], capture_output=True, text=True)
        except FileNotFoundError:
            return False, 'docker is not installed or not on PATH'
        except Exception as exc:
            return False, str(exc)

        combined = (probe.stdout or '').strip()
        if probe.stderr:
            combined = (combined + '\n' + probe.stderr.strip()).strip() if combined else probe.stderr.strip()

        if probe.returncode == 0:
            return True, ''

        last_output = combined
        if any(
            marker in combined.lower()
            for marker in (
                'error during connect',
                'dockerdesktoplinuxengine',
                'the system cannot find the file specified',
                'is the docker daemon running',
                'cannot connect to the docker daemon',
            )
        ):
            time.sleep(poll_seconds)
            continue

        time.sleep(poll_seconds)

    return False, last_output or 'docker daemon did not become ready in time'


def _ensure_docker_ready():
    ready, message = _wait_for_docker_daemon(timeout_seconds=2, poll_seconds=1)
    if ready:
        return True, ''

    start_ok, start_message = _start_docker_desktop_if_available()
    if not start_ok:
        return False, message or start_message

    ready, message = _wait_for_docker_daemon(timeout_seconds=120, poll_seconds=3)
    if ready:
        return True, ''

    return False, message or start_message


def _run_check_audit_in_docker(audit_path):
    audit_path = os.path.abspath(audit_path)
    if not os.path.isfile(audit_path):
        return 1, f'Audit file does not exist: {audit_path}'

    ready, startup_message = _ensure_docker_ready()
    if not ready:
        return 127, startup_message

    mount_dir = os.path.dirname(audit_path).replace('\\', '/')
    docker_cmd = [
        'docker', 'run', '--rm',
        '-v', f'{mount_dir}:/audit',
        'tenable/audit-utils',
        'check_audit',
        f'/audit/{os.path.basename(audit_path)}',
    ]

    try:
        proc = subprocess.run(docker_cmd, capture_output=True, text=True)
    except FileNotFoundError:
        return 127, 'docker is not installed or not on PATH'
    except Exception as exc:
        return 1, str(exc)

    combined = (proc.stdout or '').strip()
    if proc.stderr:
        combined = (combined + '\n' + proc.stderr.strip()).strip() if combined else proc.stderr.strip()

    docker_unavailable_patterns = (
        'error during connect',
        'dockerdesktoplinuxengine',
        'the system cannot find the file specified',
        'is the docker daemon running',
        'cannot connect to the docker daemon',
    )
    combined_lower = combined.lower()
    if any(pat in combined_lower for pat in docker_unavailable_patterns):
        return 127, 'docker daemon is unavailable; skipping check_audit validation'

    return proc.returncode, combined


def _normalize_tenable_audit_wrapper(text, check_type_name=None, check_type_version=None, group_policy_name=None):
    lines = text.splitlines()
    normalized = []
    opened_check_type = False
    opened_group_policy = False

    for line in lines:
        stripped = line.strip()
        m_check = re.match(r'^<check_type\s*:\s*"([^"]+)"(?:\s+version\s*:\s*"([^"]+)")?\s*>$', stripped, flags=re.IGNORECASE)
        if m_check:
            if not opened_check_type:
                if check_type_name is None and check_type_version is None:
                    normalized.append(line.rstrip())
                else:
                    original_name = (m_check.group(1) or '').strip()
                    original_version = (m_check.group(2) or '').strip()
                    out_name = (check_type_name or original_name or 'Windows').strip()
                    out_version = original_version if check_type_version is None else str(check_type_version).strip()
                    if out_version:
                        normalized.append(f'<check_type:"{out_name}" version:"{out_version}">')
                    else:
                        normalized.append(f'<check_type:"{out_name}">')
                opened_check_type = True
            continue
        if re.match(r'^<group_policy(?:\s*:\s*"[^"]+")?\s*>$', stripped, flags=re.IGNORECASE):
            if not opened_group_policy:
                if group_policy_name is None:
                    normalized.append(line.rstrip())
                elif group_policy_name:
                    normalized.append(f'<group_policy:"{group_policy_name}">')
                else:
                    normalized.append('<group_policy>')
                opened_group_policy = True
            continue
        if re.match(r'^</group_policy>\s*$', stripped, flags=re.IGNORECASE):
            normalized.append('</group_policy>')
            continue
        if re.match(r'^</check_type>\s*$', stripped, flags=re.IGNORECASE):
            normalized.append('</check_type>')
            continue
        normalized.append(line.rstrip())

    return re.sub(r'\n{3,}', '\n\n', '\n'.join(normalized).rstrip() + '\n')


def _count_unescaped_double_quotes(text):
    count = 0
    escaped = False
    for ch in str(text or ''):
        if escaped:
            escaped = False
            continue
        if ch == '\\':
            escaped = True
            continue
        if ch == '"':
            count += 1
    return count


def _repair_value_data_expression(value):
    raw = str(value or '').strip()
    if not raw:
        return value

    normalized = raw.replace(r'\&\&', '&&').replace(r'\|\|', '||')

    # Numeric POLICY_DWORD selections such as `2 || 3` or `4 || 5` are
    # scanner-accepted as-is and must not be rewritten into regex form.
    if re.fullmatch(r'"?\d+"?\s*\|\|\s*"?\d+"?', normalized):
        return value

    semantic_boolean = _normalize_boolean_value_data_expression(normalized)
    if semantic_boolean:
        return f'"{semantic_boolean}"'

    return value


def _normalize_embedded_powershell_script(script_text):
    script = str(script_text or '')
    if not script:
        return script

    # Avoid nested raw double quotes inside audit-quoted powershell fields.
    script = script.replace('-join "`r`n"', '-join [Environment]::NewLine')
    script = script.replace('-join "`n"', '-join [Environment]::NewLine')

    # Tenable runtime evaluation requires pipeline output.
    # Write-Host writes to host stream and can produce POWERSHELL_NO_RESULT.
    script = re.sub(r'(?i)\bwrite-host\b', 'Write-Output', script)

    return script


def _repair_malformed_powershell_lines(text):
    changed = False
    out_lines = []
    for line in str(text or '').splitlines():
        match = re.match(r'^(\s*(?:powershell_args|powershell_script)\s*:\s*)(.+?)\s*$', line)
        if not match:
            out_lines.append(line)
            continue

        prefix = match.group(1)
        value = match.group(2).strip()
        if not (len(value) >= 2 and value[0] == '"' and value[-1] == '"'):
            out_lines.append(line)
            continue

        inner = value[1:-1]
        fixed_inner = _normalize_embedded_powershell_script(inner)
        fixed_value = f'"{fixed_inner}"'

        if fixed_value != value:
            changed = True
            out_lines.append(f'{prefix}{fixed_value}')
        else:
            out_lines.append(line)

    repaired = '\n'.join(out_lines).rstrip() + '\n'
    return repaired, changed


def _escape_unescaped_double_quotes(value):
    return re.sub(r'(?<!\\)"', r'\\"', str(value or ''))


AUDIT_QUOTED_STRING_FIELDS = {
    'cmd',
    'expect',
    'f5_command',
    'file_required',
    'interface_name',
    'item',
    'json_transform',
    'not_expect',
    'policy_arn',
    'powershell_args',
    'powershell_script',
    'regex',
    'request',
    'show_output',
    'shared_key',
    'sql_expect',
    'tmsh',
    'value_data',
    'wmi_attribute',
    'wmi_key',
    'wmi_namespace',
    'wmi_request',
    'where',
    'xsl_stmt',
}


def _render_audit_string_value(value):
    raw = _strip_outer_quotes_once(value)
    escaped = _escape_unescaped_double_quotes(raw)
    return f'"{escaped}"'


def _render_powershell_args_value(script_text, launcher_args='-NoProfile -ExecutionPolicy Bypass -Command'):
    script = _normalize_embedded_powershell_script(_strip_outer_quotes_once(script_text))
    escaped_script = script.replace("'", "''")
    if escaped_script:
        return f'"{launcher_args} \'{escaped_script}\'"'
    return f'"{launcher_args}"'


def _repair_malformed_quoted_field_lines(text, target_keys=None):
    changed = False
    out_lines = []
    # Azure and other non-Windows audits frequently carry jq/json snippets
    # in request/json_transform values that contain embedded quotes.
    # Include the standard quoted fields in auto-repair so malformed or
    # multiline quoted values can be normalized instead of aborting the
    # entire normalization run.
    keys = {k.lower() for k in (target_keys or {'description', 'info', 'solution', 'reference', 'see_also', 'value_type', 'value_data', 'expect', 'sql_expect', 'xsl_stmt', 'request', 'json_transform'})}

    lines = str(text or '').splitlines()
    i = 0
    while i < len(lines):
        line = lines[i]
        match = re.match(r'^(\s*([A-Za-z_][A-Za-z0-9_]*)\s*:\s*)(.+?)\s*$', line)
        if not match:
            out_lines.append(line)
            i += 1
            continue

        prefix = match.group(1)
        key = match.group(2).lower()
        value = match.group(3).strip()

        if key not in keys or not value.startswith('"'):
            out_lines.append(line)
            i += 1
            continue

        if value.endswith('"') and _count_unescaped_double_quotes(value) % 2 == 0:
            out_lines.append(line)
            i += 1
            continue

        collected = [value[1:] if value.startswith('"') else value]
        j = i + 1
        closed = value.endswith('"') and _count_unescaped_double_quotes(value) % 2 != 0
        while j < len(lines):
            segment = lines[j]
            if segment.endswith('"') and _count_unescaped_double_quotes(segment) % 2 != 0:
                collected.append(segment[:-1].strip())
                closed = True
                break
            collected.append(segment.strip())
            j += 1

        if not closed:
            inner = value[1:] if value.startswith('"') else value
            inner = inner[:-1] if inner.endswith('"') else inner
        else:
            inner = ' '.join(part for part in collected if part).strip()

        fixed_inner = _escape_unescaped_double_quotes(inner)
        if fixed_inner.endswith('\\'):
            # Prevent escaping the final quote of the rendered audit value.
            fixed_inner += '\\'
        fixed_value = f'"{fixed_inner}"'

        if fixed_value != value:
            changed = True
            out_lines.append(f'{prefix}{fixed_value}')
        else:
            out_lines.append(line)
        i = j + 1 if closed else i + 1

    repaired = '\n'.join(out_lines).rstrip() + '\n'
    return repaired, changed


def _repair_malformed_value_data_lines(text):
    changed = False
    out_lines = []
    for line in str(text or '').splitlines():
        match = re.match(r'^(\s*value_data\s*:\s*)(.+?)\s*$', line)
        if not match:
            out_lines.append(line)
            continue

        prefix = match.group(1)
        value = match.group(2)
        fixed = _repair_value_data_expression(value)
        if str(fixed) != str(value):
            changed = True
            out_lines.append(f'{prefix}{fixed}')
        else:
            out_lines.append(line)

    repaired = '\n'.join(out_lines).rstrip() + '\n'
    return repaired, changed


def _line_level_parse_errors(text):
    errors = []
    in_custom_item = False

    for line_no, line in enumerate(str(text or '').splitlines(), start=1):
        stripped = line.strip()

        if not stripped:
            continue

        if stripped.startswith('#'):
            continue

        if stripped.lower() == '<custom_item>':
            in_custom_item = True
            continue
        if stripped.lower() == '</custom_item>':
            in_custom_item = False
            continue

        if stripped.startswith('<') and stripped.endswith('>'):
            continue

        if not in_custom_item:
            continue

        kv = re.match(r'^\s*([A-Za-z_][A-Za-z0-9_]*)\s*:\s*(.+?)\s*$', line)
        if not kv:
            errors.append(f'line {line_no}: invalid custom_item field syntax')
            continue

        key = kv.group(1).lower()
        value = kv.group(2)
        key_upper = key.upper()

        value_stripped = value.strip()
        if key in REAL_KEYS and value_stripped.startswith('"'):
            if (not value_stripped.endswith('"')) or (_count_unescaped_double_quotes(value) % 2 != 0):
                errors.append(f'line {line_no}: unbalanced double quotes in {key}')

        if key in {'powershell_args', 'powershell_script'} and value_stripped.startswith('"') and value_stripped.endswith('"'):
            inner = value_stripped[1:-1]
            if inner and re.search(r'(?<!\\)"', inner):
                errors.append(f'line {line_no}: unescaped embedded double quote in {key}')

        if key == 'value_data' and key in REAL_KEYS:
            raw = value_stripped
            normalized = raw.replace(r'\&\&', '&&').replace(r'\|\|', '||')
            if ('&&' in normalized or '||' in normalized) and normalized.startswith('"') and normalized.endswith('"'):
                inner = normalized[1:-1]
                if re.search(r'(?<!\\)"', inner):
                    errors.append(f'line {line_no}: malformed value_data boolean expression')

    return errors


def _windows_expect_field_errors(text, platform_hint='UNKNOWN'):
    if not _is_windows_audit_platform(platform_hint):
        return []

    errors = []
    try:
        document, _, _ = _parse_document_for_platform(str(text or '').splitlines(), platform_hint)
    except Exception as exc:
        return [f'Could not parse document for Windows expect validation: {exc}']

    for idx, node in enumerate(document, start=1):
        if node.get('type') != 'custom_item':
            continue

        fields = node.get('fields', {})
        if _norm_upper_token(fields.get('type', '')) != 'AUDIT_POWERSHELL':
            continue
        if 'expect' not in fields:
            continue

        desc = _strip_outer_quotes_once(fields.get('description', ''))
        label = desc or f'custom_item #{idx}'
        errors.append(f'{label}: Windows AUDIT_POWERSHELL blocks must not emit expect')

    return errors


def _value_data_semantic_errors(text, platform_hint='UNKNOWN'):
    errors = []
    try:
        document, _, _ = _parse_document_for_platform(str(text or '').splitlines(), platform_hint)
    except Exception as exc:
        return [f'Could not parse document for semantic validation: {exc}']

    for idx, node in enumerate(document, start=1):
        if node.get('type') != 'custom_item':
            continue

        fields = node.get('fields', {})
        raw_value = str(fields.get('value_data', '') or '').strip()
        if not raw_value:
            continue

        check_type = _norm_upper_token(fields.get('check_type', ''))
        semantic_value = _normalize_boolean_value_data_expression(raw_value)
        candidate = semantic_value or _strip_outer_quotes_once(raw_value)

        # Many source audits use CHECK_REGEX with literal value_data paths.
        # Validate those as escaped literals to avoid false positives while
        # still compiling intentional regex patterns as-is.
        is_intentional_regex = bool(
            re.search(
                r'\^|\$|\[|\]|\(\?:|\(\?=|\(\?!|\|\||&&|\||\.\*|\.\+|\{\d+(,\d*)?\}',
                candidate,
            )
        )
        should_compile = bool(semantic_value or is_intentional_regex or check_type == 'CHECK_REGEX')
        if should_compile:
            pattern = candidate if (semantic_value or is_intentional_regex) else re.escape(candidate)
            try:
                re.compile(pattern)
            except re.error as exc:
                desc = _strip_outer_quotes_once(fields.get('description', ''))
                label = desc or f'custom_item #{idx}'
                errors.append(f'{label}: invalid value_data regex ({exc})')

    return errors


def _document_parse_errors(text, platform_hint='UNKNOWN'):
    try:
        _parse_document_for_platform(str(text or '').splitlines(), platform_hint)
    except Exception as exc:
        return [f'Could not fully parse document: {exc}']
    return []


def _repair_and_scan_audit_text(text, *, check_type_name=None, check_type_version=None, group_policy_name=None, platform_hint='UNKNOWN'):
    current = str(text or '')
    changed = False

    _, profile = _platform_parse_profile(platform_hint)
    repaired_text = _apply_parse_repair_pipeline(current, profile.get('repair_pipeline', ()))
    if repaired_text != current:
        current = repaired_text
        changed = True

    normalized = _normalize_tenable_audit_wrapper(
        current,
        check_type_name=check_type_name,
        check_type_version=check_type_version,
        group_policy_name=group_policy_name,
    )
    if normalized != current:
        current = normalized
        changed = True

    parse_errors = _line_level_parse_errors(current)
    if parse_errors:
        repaired_text = _apply_parse_repair_pipeline(current, profile.get('repair_pipeline', ()))
        if repaired_text != current:
            current = repaired_text
            changed = True
            parse_errors = _line_level_parse_errors(current)
    if parse_errors:
        return current, changed, 'line-level parse validation failed', parse_errors

    document_errors = _document_parse_errors(current, platform_hint=platform_hint)
    if document_errors:
        return current, changed, 'document parse validation failed', document_errors

    windows_expect_errors = _windows_expect_field_errors(current, platform_hint=platform_hint)
    if windows_expect_errors:
        return current, changed, 'windows audit structural validation failed', windows_expect_errors

    semantic_errors = _value_data_semantic_errors(current, platform_hint=platform_hint)
    if semantic_errors:
        return current, changed, 'value_data semantic validation failed', semantic_errors

    return current, changed, '', []


def validate_and_repair_audit_file(audit_path, *, check_type_name=None, check_type_version=None, group_policy_name=None, platform_hint=None):
    platform_for_parse = platform_hint or determine_platform_from_filename(audit_path)
    try:
        with open(audit_path, encoding='utf-8', errors='replace') as fh:
            current = _strip_bom_prefix(fh.read())
    except Exception:
        _record_validation_result(audit_path, 'failed', 'Could not read file for validation.')
        return False

    current, changed, error_label, scan_errors = _repair_and_scan_audit_text(
        current,
        check_type_name=check_type_name,
        check_type_version=check_type_version,
        group_policy_name=group_policy_name,
        platform_hint=platform_for_parse,
    )

    if changed:
        current = _strip_bom_prefix(current)
        with open(audit_path, 'w', encoding='utf-8', newline='') as fh:
            fh.write(current.replace('\r\n', '\n').replace('\r', '\n').replace('\n', '\r\n'))

    if scan_errors:
        detail = '\n'.join(scan_errors[:25])
        print(f'{error_label} for {audit_path}')
        print(detail)
        _record_validation_result(audit_path, 'failed', detail)
        return False

    code, output = _run_check_audit_in_docker(audit_path)
    if code == 127:
        print(f'Skipping check_audit for {audit_path}: {output}')
        _record_validation_result(audit_path, 'skipped', output)
        return False
    if code == 0:
        _record_validation_result(audit_path, 'passed', '')
        return True

    if output:
        print(f'check_audit failed for {audit_path}')
        print(output)

    repaired, changed, error_label, scan_errors = _repair_and_scan_audit_text(
        current,
        check_type_name=check_type_name,
        check_type_version=check_type_version,
        group_policy_name=group_policy_name,
        platform_hint=platform_for_parse,
    )

    if changed:
        with open(audit_path, 'w', encoding='utf-8', newline='') as fh:
            fh.write(repaired.replace('\r\n', '\n').replace('\r', '\n').replace('\n', '\r\n'))
        if scan_errors:
            detail = '\n'.join(scan_errors[:25])
            print(f'{error_label} for {audit_path}')
            print(detail)
            _record_validation_result(audit_path, 'failed', detail)
            return False
        code, output = _run_check_audit_in_docker(audit_path)
        if code == 0:
            print(f'check_audit passed after repair: {audit_path}')
            _record_validation_result(audit_path, 'passed', '')
            return True
        if output:
            print(f'check_audit still failed after repair for {audit_path}')
            print(output)

    _record_validation_result(audit_path, 'failed', output)
    return False


def _parse_document_for_platform(lines, platform_hint, source_path=''):
    platform_key, profile = _platform_parse_profile(platform_hint)
    strict_mode_label = 'windows-strict' if platform_key in {'MSSRV', 'MSWRK'} else f'{platform_key.lower()}-strict'
    lines = _sanitize_audit_lines(lines)
    source_text = ''.join(lines)

    try:
        document = parse_document(lines)
        return document, strict_mode_label, ''
    except Exception as strict_exc:
        if not profile.get('allow_repair_fallback', False):
            raise

        repaired_text = _apply_parse_repair_pipeline(source_text, profile.get('repair_pipeline', ()))
        repaired_lines = repaired_text.splitlines()

        document = parse_document(repaired_lines)
        note = f'strict parse failed; {platform_key} repaired fallback used ({strict_exc})'
        if source_path:
            print(f'WARNING: {source_path}: {note}')
        else:
            print(f'WARNING: {note}')
        repaired_mode_label = 'windows-repaired' if platform_key in {'MSSRV', 'MSWRK'} else f'{platform_key.lower()}-repaired'
        return document, repaired_mode_label, note


def _record_validation_result(audit_path, status, details=''):
    if not audit_path:
        return

    key = os.path.abspath(audit_path)
    existing = VALIDATION_RESULTS.get(key)
    rank = {'passed': 0, 'skipped': 1, 'failed': 2}
    new_rank = rank.get(status, 0)
    old_rank = rank.get((existing or {}).get('status', 'passed'), 0)

    if existing is None or new_rank >= old_rank:
        VALIDATION_RESULTS[key] = {
            'status': status,
            'details': (details or '').strip(),
        }


def _reset_validation_summary():
    VALIDATION_RESULTS.clear()


def _print_validation_summary():
    if not VALIDATION_RESULTS:
        print('\nValidation summary: no check_audit validations were executed in this run.')
        return

    totals = {'passed': 0, 'failed': 0, 'skipped': 0}
    for rec in VALIDATION_RESULTS.values():
        status = rec.get('status', 'passed')
        if status not in totals:
            continue
        totals[status] += 1

    print('\nValidation summary')
    print(f"  Passed : {totals['passed']}")
    print(f"  Failed : {totals['failed']}")
    print(f"  Skipped: {totals['skipped']}")

    if totals['failed']:
        print('\nFailed .audit files:')
        for path, rec in VALIDATION_RESULTS.items():
            if rec.get('status') != 'failed':
                continue
            print(f'  - {path}')
            details = rec.get('details', '')
            if details:
                first_line = details.splitlines()[0].strip()
                if first_line:
                    print(f'      reason: {first_line}')


def _extract_check_type_meta(audit_path):
    if not audit_path or not os.path.isfile(audit_path):
        return '', '', ''
    try:
        with open(audit_path, encoding='utf-8', errors='ignore') as fh:
            text = _strip_bom_prefix(fh.read(20000))
    except Exception:
        return '', '', ''

    check_type_name = ''
    check_type_version = ''
    group_policy_name = ''

    m_ct = re.search(r'<check_type\s*:\s*"([^"]+)"(?:\s+version\s*:\s*"([^"]+)")?\s*>', text, flags=re.IGNORECASE)
    if m_ct:
        check_type_name = (m_ct.group(1) or '').strip()
        check_type_version = (m_ct.group(2) or '').strip()

    m_gp = re.search(r'<group_policy\s*:\s*"([^"]+)"\s*>', text, flags=re.IGNORECASE)
    if m_gp:
        group_policy_name = (m_gp.group(1) or '').strip()

    return check_type_name, check_type_version, group_policy_name


def _run_vscode_command(args):
    """Best-effort wrapper around VS Code CLI; never raises to callers."""
    code_exe = shutil.which('code')
    if not code_exe:
        return False, 'VS Code CLI not found in PATH.'

    try:
        proc = subprocess.run([code_exe] + args, capture_output=True, text=True)
    except Exception as exc:
        return False, str(exc)

    output = f"{proc.stdout or ''}\n{proc.stderr or ''}".strip()
    if proc.returncode != 0:
        return False, output or f'code exited with status {proc.returncode}'
    return True, output


def open_audit_in_vscode_for_lint(audit_path):
    ok, err = _run_vscode_command(['--reuse-window', audit_path])
    if not ok:
        print(f'Could not open in VS Code: {audit_path} ({err})')
        return False

    print(f'Opened in VS Code for AuditLang diagnostics: {audit_path}')
    print('Use View -> Problems to review parse errors with line numbers.')
    return True


def emit(document, variables, platform_hint=''):
    output = []
    rendered_blocks = []
    all_keys = []

    def _emit_description_platform_label(raw_desc):
        plat = _norm_upper_token(platform_hint)
        if not plat or plat == 'UNKNOWN':
            plat = _norm_upper_token(detect_platform(raw_desc))
        if plat == 'MSAZ':
            return 'AZURE'
        if not plat or plat == 'UNKNOWN':
            return 'CONTROL'
        return plat

    desc_counter = 0
    unknown_keys = set()

    # Build rendered blocks
    for node in document:
        if node["type"] not in ("custom_item", "report-warning", "report-passed"):
            continue

        if node["type"] == "report-passed":
            pairs = []
            # if "solution" in node["fields"]:
            #     print("\nSOLUTION VALUE:")
            #     print(repr(node["fields"]["solution"]))

            for k, v in node["fields"].items():
                if k in IGNORED_KEYS:
                    continue
                if k not in REAL_KEYS:
                    unknown_keys.add(k)
                    continue
                if k == "see_also":
                    pairs.append((k, f"\"{SEE_ALSO_REPLACEMENT}\""))
                elif k == "info":
                    info = normalize_info(v, _info_sentence_limit(platform_hint))
                    if info:
                        pairs.append((k, info))
                elif k == "reference":
                    raw_ref = resolve_variables(v, variables)
                    pairs.append((k, normalize_reference_or_passthrough(raw_ref)))
                elif k == "solution":
                    sol = normalize_solution(resolve_variables(v, variables))
                    if sol:
                        pairs.append((k, sol))
                elif k in AUDIT_QUOTED_STRING_FIELDS:
                    pairs.append((k, _render_audit_string_value(resolve_variables(v, variables))))
                else:
                    pairs.append((k, resolve_variables(v, variables)))
            pairs = enforce_solution_reference_adjacency(pairs)
            rendered_blocks.append(pairs)
            all_keys.extend(k for k, _ in pairs)
            continue

        pairs = []

        custom_fields = _convert_policy_item_to_audit_powershell(node["fields"])
        if _should_force_convert_non_powershell(platform_hint, node["fields"]):
            custom_fields = _convert_remaining_item_to_audit_powershell(custom_fields)
        custom_fields = _ensure_audit_powershell_metadata_fields(custom_fields)
        custom_fields = _finalize_audit_powershell_fields_for_emit(custom_fields, platform_hint)

        for k, v in custom_fields.items():
            if k in IGNORED_KEYS:
                continue
            if k not in REAL_KEYS:
                unknown_keys.add(k)
                continue

            if k == "see_also":
                pairs.append((k, f"\"{SEE_ALSO_REPLACEMENT}\""))

            elif k == "info":
                info = normalize_info(v, _info_sentence_limit(platform_hint))
                if info:
                    pairs.append((k, info))

            elif k == "reference":
                raw_ref = resolve_variables(v, variables)
                pairs.append((k, normalize_reference_or_passthrough(raw_ref)))

            elif k == "solution":
                sol = normalize_solution(resolve_variables(v, variables))
                if sol:
                    pairs.append((k, sol))

            elif k in AUDIT_QUOTED_STRING_FIELDS:
                pairs.append((k, _render_audit_string_value(resolve_variables(v, variables))))

            elif k == "description":
                desc_value = resolve_variables(v, variables)
                platform_label = _emit_description_platform_label(desc_value)
                desc = _prefixed_description(desc_value, platform_label, desc_counter)
                desc_counter += 1
                pairs.append((k, desc))

            else:
                pairs.append((k, resolve_variables(v, variables)))

        pairs = order_custom_item_pairs(pairs)
        rendered_blocks.append(pairs)
        all_keys.extend(k for k, _ in pairs)

    width = max(len(k) for k in all_keys) if all_keys else 0
    block_idx = 0

    # Emit final output
    for node in document:
        if node["type"] == "text":
            if not node["text"].lstrip().startswith("#"):
                output.append(resolve_variables(node["text"], variables))

        elif node["type"] == "report-passed":
            output.append('<report type:"PASSED">')
            for k, v in rendered_blocks[block_idx]:
                output.append(f"  {k.ljust(width)} : {v}")
            output.append("</report>")
            block_idx += 1
            continue

        elif node["type"] == "report-warning":
            output.append('<report type:"WARNING">')
            for k, v in rendered_blocks[block_idx]:
                output.append(f"  {k.ljust(width)} : {v}")
            output.append("</report>")
            block_idx += 1

        elif node["type"] == "custom_item":
            output.append("<custom_item>")
            for k, v in rendered_blocks[block_idx]:
                output.append(f"  {k.ljust(width)} : {v}")
            output.append("</custom_item>")
            block_idx += 1

    return output, unknown_keys


def _persist_key(key, set_name):
    """Add key to the named set in this script file for future runs."""
    script = os.path.abspath(__file__)
    with open(script, encoding="utf-8") as f:
        content = f.read()
    # Find the set block and insert the new key before the closing }
    pattern = rf'({set_name}\s*=\s*\{{)(.*?)(\}})'
    m = re.search(pattern, content, flags=re.DOTALL)
    if not m:
        print(f"  Could not find {set_name} â€” add '{key}' manually.")
        return
    prefix, body, closing = m.group(1), m.group(2), m.group(3)
    # Ensure trailing comma on last entry
    stripped = body.rstrip()
    if stripped and not stripped.endswith(','):
        body = stripped + ',\n'
    else:
        body = body.rstrip('\n') + '\n'
    new_block = f'{prefix}{body}    "{key}",\n{closing}'
    new_content = content[:m.start()] + new_block + content[m.end():]
    with open(script, "w", encoding="utf-8") as f:
        f.write(new_content)
    print(f"  '{key}' added to {set_name} in {os.path.basename(script)}.")


def process_file(infile, open_in_vscode=False, strict_mode=False):
    if not os.path.isfile(infile):
        print(f"ERROR: Input file does not exist: {infile}")
        return False

    base = os.path.splitext(os.path.basename(infile))[0]

    input_folder = os.path.dirname(infile)
    normalized_folder = os.path.join(input_folder, "Normalized")
    os.makedirs(normalized_folder, exist_ok=True)

    outfile = _timestamped_output_path(os.path.join(
        normalized_folder,
        f"{base}.audit"
    ))

    with open(infile, encoding="utf-8") as f:
        lines = _sanitize_audit_lines(f.readlines())

    source_check_type_name, source_check_type_version, source_group_policy_name = _extract_check_type_meta(infile)
    platform_hint = determine_platform_from_filename(infile)

    variables = extract_variables(lines)
    try:
        document, parse_mode, parse_note = _parse_document_for_platform(lines, platform_hint, infile)
    except Exception as exc:
        detail = f'parse failed for platform {platform_hint}: {exc}'
        print(f'ERROR: {detail}')
        _record_validation_result(infile, 'failed', detail)
        return False

    output, unknown_keys = emit(document, variables, platform_hint)
    output = _cleanup_normalized_output_lines(output)

    preflight_text, preflight_changed, preflight_error_label, preflight_errors = _repair_and_scan_audit_text(
        "\n".join(output) + "\n",
        check_type_name=source_check_type_name or None,
        check_type_version=source_check_type_version or None,
        group_policy_name=source_group_policy_name or None,
        platform_hint=platform_hint,
    )
    if preflight_changed:
        output = preflight_text.rstrip("\n").splitlines()
    if preflight_errors:
        print(f"ERROR: {preflight_error_label} for generated output: {infile}")
        for err in preflight_errors[:25]:
            print(err)
        _record_validation_result(infile, 'failed', '\n'.join(preflight_errors[:25]))
        return False

    node_type_counts = {}
    node_rows = []
    for idx, node in enumerate(document, start=1):
        node_type = node.get('type', 'unknown')
        node_type_counts[node_type] = node_type_counts.get(node_type, 0) + 1
        details = {}
        if node_type in ('custom_item', 'report-warning', 'report-passed'):
            details['description'] = node.get('fields', {}).get('description', '')
            details['field_count'] = len(node.get('fields', {}))
        elif node_type == 'text':
            details['text'] = node.get('text', '')[:200]
        node_rows.append({
            'node_index': idx,
            'node_type': node_type,
            'details': details,
        })

    with open(outfile, "w", encoding="utf-8", newline="") as f:
        safe_text = _strip_bom_prefix("\n".join(output) + "\n")
        _assert_no_encoding_markers(safe_text, outfile)
        f.write(safe_text.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "\r\n"))

    validate_and_repair_audit_file(
        outfile,
        check_type_name=source_check_type_name or None,
        check_type_version=source_check_type_version or None,
        group_policy_name=source_group_policy_name or None,
        platform_hint=platform_hint,
    )

    print("\nNormalized audit written to:")
    print(f"  {outfile}")

    if open_in_vscode:
        open_audit_in_vscode_for_lint(outfile)

    folder_key = os.path.abspath(os.path.dirname(infile))
    PARSING_RESULTS_BY_FOLDER.setdefault(folder_key, []).append({
        'audit_file': os.path.basename(infile),
        'sheet_name': os.path.splitext(os.path.basename(infile))[0],
        'source_folder': os.path.dirname(infile),
        'status': 'parsed',
        'node_count': len(document),
        'custom_item_count': node_type_counts.get('custom_item', 0),
        'report_count': node_type_counts.get('report-passed', 0) + node_type_counts.get('report-warning', 0),
        'text_count': node_type_counts.get('text', 0),
        'unknown_key_count': len(unknown_keys),
        'unknown_keys': '; '.join(sorted(unknown_keys)),
        'normalized_output': outfile,
        'notes': (
            f'Normalized with Tenable Audit Lang parser ({parse_mode})'
            + (f'; {parse_note}' if parse_note else '')
        ),
        'nodes': node_rows,
    })

    if unknown_keys:
        print("\nUnrecognized keys found â€” classify each:")
        reclassified = False

        for k in sorted(unknown_keys):
            while True:
                ans = input(f"  '{k}': (R)eal, (I)gnored, (S)kip? ").strip().upper()
                if ans in ('R', 'I', 'S'):
                    break

            if ans == 'R':
                REAL_KEYS.add(k)
                _persist_key(k, 'REAL_KEYS')
                reclassified = True

            elif ans == 'I':
                IGNORED_KEYS.add(k)
                _persist_key(k, 'IGNORED_KEYS')
                reclassified = True

        if reclassified:
            output, _ = emit(document, variables, platform_hint)
            output = _cleanup_normalized_output_lines(output)

            preflight_text, preflight_changed, preflight_error_label, preflight_errors = _repair_and_scan_audit_text(
                "\n".join(output) + "\n",
                check_type_name=source_check_type_name or None,
                check_type_version=source_check_type_version or None,
                group_policy_name=source_group_policy_name or None,
                platform_hint=platform_hint,
            )
            if preflight_changed:
                output = preflight_text.rstrip("\n").splitlines()
            if preflight_errors:
                print(f"ERROR: {preflight_error_label} for generated output: {infile}")
                for err in preflight_errors[:25]:
                    print(err)
                _record_validation_result(infile, 'failed', '\n'.join(preflight_errors[:25]))
                return False

            with open(outfile, "w", encoding="utf-8", newline="") as f:
                safe_text = _strip_bom_prefix("\n".join(output) + "\n")
                _assert_no_encoding_markers(safe_text, outfile)
                f.write(safe_text.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "\r\n"))

            validate_and_repair_audit_file(
                outfile,
                check_type_name=source_check_type_name or None,
                check_type_version=source_check_type_version or None,
                group_policy_name=source_group_policy_name or None,
                platform_hint=platform_hint,
            )

            print("\nRe-processed with updated key classifications.")

    return True


def process_folder(folder, open_in_vscode=False, strict_mode=False):
    audit_files = sorted(
        f for f in os.listdir(folder)
        if f.lower().endswith(".audit")
    )

    if not audit_files:
        print("No .audit files found.")
        return True

    print(f"\nFound {len(audit_files)} audit files.\n")

    failed_files = []

    for fname in audit_files:
        infile = os.path.join(folder, fname)

        print("-" * 60)
        print(f"Processing: {fname}")

        ok = process_file(infile, open_in_vscode=open_in_vscode, strict_mode=strict_mode)
        if not ok:
            failed_files.append(infile)

    if failed_files:
        print("\nPreflight/normalization failures:")
        for path in failed_files:
            print(f"  {path}")
        if strict_mode:
            print("\nERROR: strict mode enabled and one or more files failed preflight/normalization.")
            return False

    return True

