
#!/usr/bin/env python3
"""
This script is essentially an Audit File Normalizer and Standardizer for Nessus/Tenable .audit files.
This focuses on cleaning, restructuring, normalizing, and standardizing the contents of individual audit files.
Its goal is to create a consistent Tenable parsable audit format suitable for enterprise use,
converting vendor/CIS/STIG-generated audits into an internally standardized audit baseline.

High-Level Workflow
For each .audit file:

Extract variable definitions.
Parse the audit structure.
Identify report and custom-item blocks.
Normalize descriptions, references, and info fields.
Replace variables with actual values.
Remove unwanted fields.
Renumber descriptions.
Standardize formatting.
Write a new -normalized.audit file.
Learn previously unknown key field names and update itself for future runs.
"""
import os
import re
import sys
import subprocess
import time
import shutil
import importlib.util
import urllib.request
from collections import OrderedDict, defaultdict
from datetime import datetime
import json
import csv

REPO_ROOT = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "..").replace("\\", "/")
)
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

SCRIPT_DIR = os.path.abspath(os.path.dirname(__file__))
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

AUDIT_INPUTS_ROOT = os.path.join(SCRIPT_DIR, 'audit_inputs')

_PARSER_START_BLOCK_RE = re.compile(r"^\s*<\s*(custom_item|report)\b([^>]*)>\s*$", re.IGNORECASE)
_PARSER_END_CUSTOM_RE = re.compile(r"^\s*</\s*custom_item\s*>\s*$", re.IGNORECASE)
_PARSER_END_REPORT_RE = re.compile(r"^\s*</\s*report\s*>\s*$", re.IGNORECASE)
_PARSER_FIELD_RE = re.compile(r"^\s*([A-Za-z0-9_]+)\s*:\s*(.*)$")
_PARSER_REPORT_TYPE_RE = re.compile(r"\btype\s*:\s*\"?([A-Za-z-]+)\"?", re.IGNORECASE)
_PARSER_VAR_LINE_RE = re.compile(r"^\s*([A-Za-z_][A-Za-z0-9_]*)\s*[:=]\s*(.*?)\s*$")


def _parser_lines(text_or_lines):
    if isinstance(text_or_lines, str):
        return text_or_lines.splitlines()
    return [str(line).rstrip("\n") for line in text_or_lines]


def _parser_unquote(value):
    value = value.strip()
    if len(value) >= 2 and value[0] == value[-1] and value[0] in ('"', "'"):
        return value[1:-1]
    return value


def extract_variables(text_or_lines):
    """Extract top-level and commented XML-style audit variable definitions."""
    lines = _parser_lines(text_or_lines)
    variables = {}

    for line in lines:
        if _PARSER_START_BLOCK_RE.match(line):
            break
        match = _PARSER_VAR_LINE_RE.match(line)
        if not match:
            continue
        key, value = match.group(1), match.group(2)
        if key.lower() not in {'type', 'description', 'info', 'reference'} and not (value.startswith('<') and value.endswith('>')):
            variables[key] = _parser_unquote(value)

    in_variable = False
    variable_name = ''
    variable_default = ''
    for line in lines:
        cleaned = re.sub(r"^\s*#\s*", "", line).strip()
        if cleaned.lower() == '<variable>':
            in_variable, variable_name, variable_default = True, '', ''
        elif in_variable and cleaned.lower() == '</variable>':
            if variable_name and variable_default:
                variables[variable_name] = _parser_unquote(variable_default)
            in_variable = False
        elif in_variable:
            name_match = re.match(r"^<name>(.*?)</name>$", cleaned, flags=re.IGNORECASE)
            default_match = re.match(r"^<default>(.*?)</default>$", cleaned, flags=re.IGNORECASE)
            if name_match:
                variable_name = name_match.group(1).strip()
            elif default_match:
                variable_default = default_match.group(1).strip()
    return variables


def parse_document(text_or_lines):
    """Parse an audit file into text, custom-item, and report nodes."""
    document = []
    in_block = False
    block_type = ''
    block_fields = OrderedDict()
    last_key = ''

    for line in _parser_lines(text_or_lines):
        if not in_block:
            start_match = _PARSER_START_BLOCK_RE.match(line)
            if not start_match:
                document.append({'type': 'text', 'text': line})
                continue
            tag, attributes = start_match.group(1).lower(), start_match.group(2) or ''
            if tag == 'report':
                type_match = _PARSER_REPORT_TYPE_RE.search(attributes)
                report_type = type_match.group(1).strip().lower() if type_match else 'warning'
                block_type = f"report-{'passed' if report_type == 'passed' else 'warning'}"
            else:
                block_type = 'custom_item'
            block_fields, last_key, in_block = OrderedDict(), '', True
            continue

        is_end = (block_type.startswith('report') and _PARSER_END_REPORT_RE.match(line)) or (block_type == 'custom_item' and _PARSER_END_CUSTOM_RE.match(line))
        if is_end:
            document.append({'type': block_type, 'fields': block_fields})
            in_block, block_type, block_fields, last_key = False, '', OrderedDict(), ''
            continue

        field_match = _PARSER_FIELD_RE.match(line)
        if field_match:
            last_key = field_match.group(1).strip()
            block_fields[last_key] = field_match.group(2).strip()
        elif last_key and line.strip():
            block_fields[last_key] = (block_fields[last_key] + ' ' + line.strip()).strip()

    if in_block and block_type:
        document.append({'type': block_type, 'fields': block_fields})
    return document

try:
    import openpyxl
    from openpyxl import Workbook
except Exception:
    openpyxl = None
    Workbook = None


PARSING_RESULTS_FILENAME = 'Parsing Results.xlsx'
PARSING_RESULTS_BY_FOLDER = OrderedDict()
VALIDATION_RESULTS = OrderedDict()
RUN_TIMESTAMP = time.strftime('%y%m%d%H', time.localtime())
_TS_SUFFIX_RE = re.compile(r'_\d{8}$')
NON_COMPLIANT_OUTPUT = '__NON_COMPLIANT__'
THREAT_INTEL_CACHE_PATH = os.path.join(SCRIPT_DIR, 'threat_intel_cache.json')
THREAT_INTEL_FEED_URL = os.environ.get('PYSC_THREAT_INTEL_FEED_URL', '').strip()
THREAT_INTEL_CACHE_MAX_AGE_HOURS = max(1, int(os.environ.get('PYSC_THREAT_INTEL_CACHE_MAX_AGE_HOURS', '168') or 168))
THREAT_INTEL_FORCE_REFRESH = os.environ.get('PYSC_THREAT_INTEL_REFRESH', '').strip().lower() in {'1', 'true', 'yes'}
THREAT_INTEL_DATA = {}


# =============================================================================
# CONFIGURATION
# =============================================================================

REAL_KEYS = {
    "type", "description", "info", "reference", "see_also", "solution",
    "api_request_type", "request", "xsl_stmt", "not_expect", "show_output",
    "powershell_args", "key_item",
    "value_type", "value_data", "reg_key", "reg_item", "reg_option",
    "audit_policy_subcategory", "right_type", "reg_include_hku_users",
    "check_type", "account_type", "password_policy", "lockout_policy",
    "regex", "expect", "severity",
    "wmi_key",
    "wmi_namespace",
    "wmi_request",
    "wmi_attribute",
    "f5_command",
    "item",
    "json_transform",
    "match_all",
    "sql_expect",
    "sql_request",
    "sql_types",
    "cmd",
    "is_substring",
    "mask",
    "min_occurrences",
    "rpm",
    "string_required",
    "timeout",
    "policy_arn",
    "interface_name",
    "shared_key",
    "where",
    "tmsh",
    "dont_echo_cmd",
    "file_required",
    "powershell_option",
    "operator",
    "only_show_cmd_output",
    "system",
    "reg_ignore_hku_users",
    "reg_type",
}

IGNORED_KEYS = {
    "Impact",
    "Note",
    "4944",
    "4945",
    "4946",
    "4947",
    "4948",
    "4949",
    "4950",
    "4951",
    "4952",
    "4953",
    "4954",
    "4956",
    "4957",
    "4958",
    "5063",
    "5064",
    "5065",
    "5066",
    "5067",
    "5068",
    "5069",
    "5070",
    "6145",
    "Caution",
    "Disabled",
    "Enabled",
    "Important",
    "Warning",
    "Example",
    "MinimumPasswordLength",
    "PasswordReusePrevention",
    "aws_action",
    "NOTE",
    "content",
    "context",
    "Rationale",
    "Steps",
    "WARNING",
    "https",
    "1",
    "2",
    "Source",
    "or",
    "levels",
    "service",
    "status",
    "Notes",
    "Satisfies",
    "days",
    "name",
    "optionally",
    "file",
    "group",
    "owner",
    "required",
    "ignore",
    "IMPORTANT",
    "Run",
    "Default",
    "Mitigated",
    "Name",
    "None",
    "Vulnerable",
    "Run",
}

SEE_ALSO_REPLACEMENT = "See HTH Policies and Standards"
BASELINE_MSSRV_CSV_PATH = os.environ.get(
    'PYSC_BASELINE_MSSRV_CSV', os.path.join(SCRIPT_DIR, 'Baseline_-_MSSRV.csv')
)
MERGED_MSSRV_CSV_PATH = os.environ.get(
    'PYSC_MERGED_MSSRV_CSV', os.path.join(SCRIPT_DIR, 'Merged_2607.csv')
)
DESCRIPTION_MATCH_XLSX_PATH = os.environ.get(
    'PYSC_DESCRIPTION_MATCH_XLSX', os.path.join(SCRIPT_DIR, 'Baseline_vs_Merged_MSSRV_Description_Matches.xlsx')
)
PRODUCTION_AUDIT_ROOT = os.path.abspath(os.environ.get(
    'PYSC_PRODUCTION_AUDIT_ROOT', os.path.join(SCRIPT_DIR, 'actual_audit_inputs')
))
PRODUCTION_NORMALIZED_ROOT = os.path.abspath(os.environ.get(
    'PYSC_PRODUCTION_NORMALIZED_ROOT', os.path.join(PRODUCTION_AUDIT_ROOT, 'Normalized')
))
PRODUCTION_GAP_ROOT = os.path.abspath(os.environ.get(
    'PYSC_PRODUCTION_GAP_ROOT', os.path.join(PRODUCTION_AUDIT_ROOT, 'For_Gap')
))
PRODUCTION_GAP_OUTPUT_ROOT = os.path.abspath(os.environ.get(
    'PYSC_PRODUCTION_GAP_OUTPUT_ROOT', PRODUCTION_AUDIT_ROOT
))
BASELINE_PLUGIN_ID_FILTER = ''

# =============================================================================
# HELPERS
# =============================================================================

def resolve_variables(text, variables):
    for k, v in variables.items():
        text = text.replace(f"@{k}@", v)
    return text

def normalize_info(raw):
    if not raw:
        return None
    s = raw.strip()
    s = re.sub(r'^[\'"]+', '', s)
    s = re.sub(r'[\'"]+$', '', s)
    s = re.sub(r'\s+', ' ', s)
    if s.startswith('This audit is written to dynamically identify if all paths are present in any order.'):
        s = s.rstrip('.') + '.'
        return f'"{s}"'
    sentence = s.split('.')[0].strip()
    if not sentence:
        return None
    return f"\"{sentence}.\""

def normalize_reference(raw):
    if not raw:
        return None
    flat = _strip_outer_quotes_once(raw)
    flat = re.sub(r"\s+", " ", flat)
    parts = [p.strip() for p in flat.split(",")]
    controls = []
    for p in parts:
        p = p.strip().strip('"').strip("'")
        m = re.match(r"^(?:NIST\s+)?800-53r5\|(.+)$", p, flags=re.IGNORECASE)
        if m:
            controls.append(m.group(1).strip().strip('"').strip("'"))
    if not controls:
        return None
    seen = set()
    unique = []
    for c in controls:
        if c not in seen:
            seen.add(c)
            unique.append(c)
    return "\"NIST 800-53r5|" + " ".join(unique) + "\""


def normalize_reference_or_passthrough(raw):
    """Prefer normalized NIST references, but retain non-NIST references instead of dropping them."""
    ref = normalize_reference(raw)
    if ref:
        return ref

    cleaned = _strip_outer_quotes_once(raw)
    if cleaned is None:
        cleaned = ''
    return f'"{cleaned}"'


def normalize_description(raw):
    if not raw:
        return None
    s = raw.strip()
    s = s.strip('"')
    s = re.sub(r"^\d+(\.\d+)+\s*", "", s)
    s = re.sub(r"'([^']+)'", r"\1", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = _escape_unescaped_double_quotes(s)
    return f'"{s}"'


def normalize_solution(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None

    # Normalize to a double-quoted value and guarantee a trailing quote.
    if s.startswith("'") and s.endswith("'") and len(s) >= 2:
        s = s[1:-1]
    elif s.startswith("'"):
        s = s[1:]
    elif s.startswith('"') and s.endswith('"') and len(s) >= 2:
        s = s[1:-1]
    elif s.startswith('"'):
        s = s[1:]

    s = s.rstrip("'\"").rstrip()
    s = re.sub(r'\s+', ' ', s).strip()
    s = _escape_unescaped_double_quotes(s)
    if s.endswith('\\'):
        # Prevent escaping the final quote of the rendered audit value.
        s += '\\'
    s = f'"{s}"'
    return s


def _first_int(text):
    if text is None:
        return None
    m = re.search(r'\d+', str(text))
    if not m:
        return None
    try:
        return int(m.group(0))
    except ValueError:
        return None


def _derive_expected_scalar(fields):
    raw = _strip_outer_quotes_once(fields.get('value_data', ''))
    if not raw:
        raw = ''
    else:
        # Range style values like [24..MAX] -> 24.
        m_range = re.search(r'\[(\d+)\s*\.\.', raw)
        if m_range:
            return m_range.group(1)

        # Alternate values like "1" || "2" -> first value.
        m_first = re.search(r'\d+', raw)
        if m_first:
            return m_first.group(0)

    # Fall back to description numbers and avoid control-id prefixes like 1.2.3.
    desc = str(fields.get('description', ''))
    nums = re.findall(r'\d+', desc)
    if nums:
        return nums[-1]

    return raw


def _build_ge_integer_regex(min_value):
    digits = str(int(min_value))
    parts = []

    def rec(prefix, index, tight):
        if index == len(digits):
            parts.append(prefix)
            return

        current = int(digits[index])
        remaining = len(digits) - index - 1

        if tight:
            rec(prefix + str(current), index + 1, True)
            for candidate in range(current + 1, 10):
                if remaining:
                    parts.append(prefix + str(candidate) + rf'\d{{{remaining}}}')
                else:
                    parts.append(prefix + str(candidate))
        else:
            if len(digits) - index > 0:
                parts.append(prefix + rf'\d{{{len(digits) - index}}}')

    rec('', 0, True)

    if len(digits) == 1:
        longer = r'[1-9]\d+'
    else:
        longer = rf'[1-9]\d{{{len(digits)},}}'
    parts.append(longer)

    unique = []
    seen = set()
    for part in parts:
        if part not in seen:
            seen.add(part)
            unique.append(part)
    return '^(?:' + '|'.join(unique) + ')$'


def _build_bracket_range_expected(raw_expected):
    raw = _strip_outer_quotes_once(raw_expected).strip()
    match = re.fullmatch(r'\[(MIN|\d+)\s*\.\.\.?\s*(MAX|\d+)\]', raw, flags=re.IGNORECASE)
    if not match:
        return None

    lower = match.group(1).upper()
    upper = match.group(2).upper()

    lower_num = None if lower == 'MIN' else int(lower)
    upper_num = None if upper == 'MAX' else int(upper)

    if lower_num is not None and upper_num is not None:
        if upper_num < lower_num:
            return raw
        values = [str(i) for i in range(lower_num, upper_num + 1)]
        return '^(?:' + '|'.join(values) + ')$'

    if lower_num is None and upper_num is not None:
        values = [str(i) for i in range(0, upper_num + 1)]
        return '^(?:' + '|'.join(values) + ')$'

    if lower_num is not None and upper_num is None:
        return _build_ge_integer_regex(lower_num)

    return r'^\d+$'


def _build_policy_powershell_args(policy_type, policy_name, expected_value):
    ptype = _norm_upper_token(policy_type)
    pname = _norm_upper_token(policy_name)
    expected = str(expected_value or '').strip()

    if ptype == 'PASSWORD_POLICY' and pname == 'ENFORCE_PASSWORD_HISTORY':
        return (
            "$HisPwdObj = net accounts | Select-string 'password history'; "
            "$HisPwdStr = $HisPwdObj.ToString(); "
            "$HisPwdStr -match '\\d{1,2}' | out-null; "
            f"$HisPwdStr -match '{expected}' | out-null; "
            "$LO_HisPwd = $matches[0]; Write-Output $LO_HisPwd;"
        )
    if ptype == 'PASSWORD_POLICY' and pname == 'MAXIMUM_PASSWORD_AGE':
        return (
            "$MaxPwdObj = net accounts | Select-string 'Maximum password age'; "
            "$MaxPwdStr = $MaxPwdObj.ToString(); "
            "$MaxPwdStr -match '\\d{1,3}' | out-null; "
            f"$MaxPwdStr -match '{expected}' | out-null; "
            "$LO_MaxPwd = $matches[0]; Write-Output $LO_MaxPwd;"
        )
    if ptype == 'PASSWORD_POLICY' and pname == 'MINIMUM_PASSWORD_AGE':
        return (
            "$MinAgeObj = net accounts | Select-string 'Minimum password age'; "
            "$MinAgeStr = $MinAgeObj.ToString(); "
            "$MinAgeStr -match '\\d{1,3}' | out-null; "
            f"$MinAgeStr -match '{expected}' | out-null; "
            "$LO_MinAge = $matches[0]; Write-Output $LO_MinAge;"
        )
    if ptype == 'PASSWORD_POLICY' and pname == 'MINIMUM_PASSWORD_LENGTH':
        return (
            "$MinLenObj = net accounts | Select-string 'password length'; "
            "$MinLenStr = $MinLenObj.ToString(); "
            "$MinLenStr -match '\\d{1,3}' | out-null; "
            f"$MinLenStr -match '{expected}' | out-null; "
            "$LO_MinLen = $matches[0]; Write-Output $LO_MinLen;"
        )
    if ptype == 'PASSWORD_POLICY' and pname == 'COMPLEXITY_REQUIREMENTS':
        return (
            "$secfile = [System.IO.Path]::GetTempFileName(); "
            "secedit /export /cfg $secfile /areas SECURITYPOLICY /quiet | Out-Null; "
            "$complexity = Select-String -Path $secfile -Pattern 'PasswordComplexity\\s*=\\s*(\\d)'; "
            "Remove-Item $secfile -ErrorAction SilentlyContinue; "
            "if ($complexity -and $complexity.Matches.Groups[1].Value -eq '1') { Write-Output 'Enabled' } else { Write-Output 'Disabled' }"
        )
    if ptype == 'PASSWORD_POLICY' and pname == 'REVERSIBLE_ENCRYPTION':
        return (
            "$secfile = [System.IO.Path]::GetTempFileName(); "
            "secedit /export /cfg $secfile /areas SECURITYPOLICY /quiet | Out-Null; "
            "$line = Select-String -Path $secfile -Pattern 'ClearTextPassword\\s*=\\s*(\\d)' | Select-Object -First 1; "
            "Remove-Item $secfile -ErrorAction SilentlyContinue; "
            "if($line -and $line.Matches.Count -gt 0){ "
            "if($line.Matches[0].Groups[1].Value -eq '1'){ Write-Output 'Enabled' } else { Write-Output 'Disabled' } "
            "} else { Write-Output 'Not Found' }"
        )
    if ptype == 'PASSWORD_POLICY' and pname == 'LOCKOUT_ADMINS':
        return (
            "$secfile = [System.IO.Path]::GetTempFileName(); "
            "secedit /export /cfg $secfile /areas SECURITYPOLICY /quiet | Out-Null; "
            "$line = Select-String -Path $secfile -Pattern 'AllowAdministratorLockout\\s*=\\s*(\\d)' | Select-Object -First 1; "
            "Remove-Item $secfile -ErrorAction SilentlyContinue; "
            "if($line -and $line.Matches.Count -gt 0){ "
            "if($line.Matches[0].Groups[1].Value -eq '1'){ Write-Output 'Enabled' } else { Write-Output 'Disabled' } "
            "} else { Write-Output 'Not Found' }"
        )
    if ptype == 'PASSWORD_POLICY' and pname == 'FORCE_LOGOFF':
        return (
            "$LogoffObj = net accounts | Select-string 'Force user logoff'; "
            "$LogoffStr = $LogoffObj.ToString(); "
            "$LogoffStr -match '\\d{1,4}|Never' | out-null; "
            "$LO_Logoff = $matches[0]; Write-Output $LO_Logoff;"
        )
    if ptype == 'LOCKOUT_POLICY' and pname == 'LOCKOUT_DURATION':
        return (
            "$LoDurObj = net accounts | Select-string 'Lockout duration'; "
            "$LoDurStr = $LoDurObj.ToString(); "
            "$LoDurStr -match '\\d{1,4}' | out-null; "
            f"$LoDurStr -match '{expected}' | out-null; "
            "$LO_LoDur = $matches[0]; Write-Output $LO_LoDur;"
        )
    if ptype == 'LOCKOUT_POLICY' and pname == 'LOCKOUT_THRESHOLD':
        return (
            "$LoThrObj = net accounts | Select-string 'threshold'; "
            "$LoThrStr = $LoThrObj.ToString(); "
            "$LoThrStr -match '\\d{1,3}' | out-null; "
            f"$LoThrStr -match '{expected}' | out-null; "
            "$LO_LoThr = $matches[0]; Write-Output $LO_LoThr;"
        )
    if ptype == 'LOCKOUT_POLICY' and pname == 'LOCKOUT_RESET':
        return (
            "$LoObsObj = net accounts | Select-string 'Lockout observation'; "
            "$LoObsStr = $LoObsObj.ToString(); "
            "$LoObsStr -match '\\d{1,4}' | out-null; "
            f"$LoObsStr -match '{expected}' | out-null; "
            "$LO_LoObs = $matches[0]; Write-Output $LO_LoObs;"
        )
    return ''


def _derive_audit_policy_expected(value_data):
    val = _norm_upper_token(_strip_outer_quotes_once(value_data))
    if 'SUCCESS' in val and 'FAILURE' in val:
        return 'Success and Failure'
    if 'SUCCESS' in val:
        return 'Success'
    if 'FAILURE' in val:
        return 'Failure'
    return _strip_outer_quotes_once(value_data)


def _convert_policy_item_to_audit_powershell(fields):
    ptype = _norm_upper_token(fields.get('type', ''))
    if ptype == 'AUDIT_POLICY_SUBCATEGORY':
        subcat = _strip_outer_quotes_once(fields.get('audit_policy_subcategory', ''))
        expected = _derive_audit_policy_expected(fields.get('value_data', ''))
        if not subcat or not expected:
            return fields

        ps_script = (
            f"$AP = auditpol /get /Subcategory:'{subcat}' |Select-String '{subcat}'; "
            "$APStr=$AP.ToString(); "
            f"$APStr -match '{expected}' | out-null; "
            "$LOAP = $matches[0]; Write-Output $LOAP;"
        )

        converted = OrderedDict()
        converted['type'] = 'AUDIT_POWERSHELL'
        for key in ('description', 'info', 'solution', 'reference', 'see_also'):
            if key in fields:
                converted[key] = fields.get(key)
        converted['value_type'] = 'POLICY_TEXT'
        converted['value_data'] = f'"{expected}"'
        converted['powershell_args'] = '-NoProfile -ExecutionPolicy Bypass -Command'
        converted['powershell_script'] = ps_script
        converted['expect'] = expected
        return converted

    if ptype not in {'PASSWORD_POLICY', 'LOCKOUT_POLICY'}:
        return fields

    policy_name_key = 'password_policy' if ptype == 'PASSWORD_POLICY' else 'lockout_policy'
    policy_name = _norm_upper_token(fields.get(policy_name_key, ''))
    expected = _derive_expected_scalar(fields)
    ps_script = _build_policy_powershell_args(ptype, policy_name, expected)
    if not ps_script:
        return fields

    converted = OrderedDict()
    converted['type'] = 'AUDIT_POWERSHELL'

    # Preserve existing primary metadata in a stable order.
    for key in ('description', 'info', 'solution', 'reference', 'see_also'):
        if key in fields:
            converted[key] = fields.get(key)

    converted['value_type'] = 'POLICY_TEXT'
    converted['value_data'] = f'"{expected}"' if expected else fields.get('value_data', '')
    converted['powershell_args'] = '-NoProfile -ExecutionPolicy Bypass -Command'
    converted['powershell_script'] = ps_script
    converted['expect'] = expected if expected else 'NON_COMPLIANT'

    return converted


def _ps_single_quote_escape(value):
    return str(value or '').replace("'", "''")


def _rewrite_auditpol_powershell_script(script_text):
    script = _strip_outer_quotes_once(script_text)
    if not script or 'auditpol' not in script.lower():
        return script

    subcat_match = re.search(
        r"/subcategory\s*:\s*['\"]([^'\"]+)['\"]",
        script,
        flags=re.IGNORECASE,
    )
    if not subcat_match:
        return script

    subcat_raw = subcat_match.group(1).strip()
    subcat = _ps_single_quote_escape(subcat_raw)
    subcat_regex = re.escape(subcat_raw)
    return (
        f"$rows = auditpol /get /Subcategory:'{subcat}' /r 2>$null | ConvertFrom-Csv; "
        f"$row = if($rows){{ $rows | Where-Object {{ $_.Subcategory -eq '{subcat}' }} | Select-Object -First 1 }} else {{ $null }}; "
        "$inc = if($row){ [string]$row.'Inclusion Setting' } else { '' }; "
        "if(-not [string]::IsNullOrWhiteSpace($inc)){ Write-Output $inc.Trim(); return }; "
        f"$AP = auditpol /get /Subcategory:'{subcat}' 2>$null | ForEach-Object {{ $_.ToString() }}; "
        f"$line = $AP | Where-Object {{ $_ -match '^\\s*{subcat_regex}\\s{{2,}}' }} | Select-Object -First 1; "
        "if(-not $line){ Write-Output 'Not Found'; return }; "
        "if($line -match 'Success and Failure'){ Write-Output 'Success and Failure'; return }; "
        "if($line -match 'No Auditing'){ Write-Output 'No Auditing'; return }; "
        "if($line -match 'Success'){ Write-Output 'Success'; return }; "
        "if($line -match 'Failure'){ Write-Output 'Failure'; return }; "
        "Write-Output $line;"
    )


def _rewrite_ntlm_outgoing_traffic_script(script_text):
    script = _strip_outer_quotes_once(script_text)
    if 'RestrictSendingNTLMTraffic' not in script:
        return script

    return (
        "if(Test-Path -Path 'HKLM:\\System\\CurrentControlSet\\Control\\Lsa\\MSV1_0'){"
        "$noutput = (Get-ItemProperty -Path 'HKLM:\\System\\CurrentControlSet\\Control\\Lsa\\MSV1_0').RestrictSendingNTLMTraffic;"
        "if($null -eq $noutput){ Write-Output 'Compliant'; return };"
        "try { if([int]$noutput -ge 1){ Write-Output 'Compliant' } else { Write-Output 'Non-Compliant' } }"
        "catch { Write-Output 'Non-Compliant' }"
        "}else { Write-Output 'Non-Compliant' }"
    )


def _rewrite_netlogon_parameters_default_script(script_text):
    script = _strip_outer_quotes_once(script_text)

    # Repair prior over-normalization variants like Parametersssss.
    script = re.sub(
        r'(?i)Services\\Netlogon\\Parameters+',
        r'Services\\Netlogon\\Parameters',
        script,
    )

    if not re.search(r'(?i)Services\\Netlogon\\Parameters', script):
        return script

    prop_match = re.search(
        r"Get-ItemProperty\s*-Path\s*'[^']*Netlogon\\Parameters+'\)\.(\w+)",
        script,
        flags=re.IGNORECASE,
    )
    if not prop_match:
        return script

    prop = prop_match.group(1)
    defaults = {
        'DisablePasswordChange': '0',
        'MaximumPasswordAge': '30',
        'RequireSignOrSeal': '1',
        'SealSecureChannel': '1',
        'SignSecureChannel': '1',
        'RequireStrongKey': '1',
        'RefusePasswordChange': '0',
    }
    default_value = defaults.get(prop)
    if default_value is None:
        return script

    return (
        "$p='HKLM:\\System\\CurrentControlSet\\Services\\Netlogon\\Parameters'; "
        f"$k='{prop}'; $d='{default_value}'; "
        "if(-not (Test-Path -Path $p)){ Write-Output $d; return }; "
        "$r=Get-ItemProperty -Path $p -ErrorAction SilentlyContinue; "
        "$pr=$r.PSObject.Properties[$k]; "
        "if($null -eq $pr -or $null -eq $pr.Value){ Write-Output $d } else { Write-Output $pr.Value }"
    )


def _rewrite_password_complexity_script(script_text):
    script = _strip_outer_quotes_once(script_text)
    s_upper = _norm_upper_token(script)
    if 'NET ACCOUNTS' not in s_upper or 'PASSWORD COMPLEXITY' not in s_upper:
        return script

    return (
        "$secfile = [System.IO.Path]::GetTempFileName(); "
        "secedit /export /cfg $secfile /areas SECURITYPOLICY /quiet | Out-Null; "
        "$complexity = Select-String -Path $secfile -Pattern 'PasswordComplexity\\s*=\\s*(\\d)'; "
        "Remove-Item $secfile -ErrorAction SilentlyContinue; "
        "if ($complexity -and $complexity.Matches.Groups[1].Value -eq '1') { Write-Output 'Enabled' } else { Write-Output 'Disabled' }"
    )


def _normalize_legacy_net_accounts_expected(script_text, raw_value_data):
    script = _strip_outer_quotes_once(script_text)
    expected = _strip_outer_quotes_once(raw_value_data)
    if not script or not expected:
        return raw_value_data

    s_upper = _norm_upper_token(script)
    if 'NET ACCOUNTS' not in s_upper:
        return raw_value_data
    if not (
        'LOCKOUT OBSERVATION' in s_upper
        or 'LOCKOUT DURATION' in s_upper
        or 'THRESHOLD' in s_upper
    ):
        return raw_value_data

    first_num = _first_int(expected)
    if first_num is None:
        return raw_value_data
    return f'"{first_num}"'


def _best_effort_expected_text(fields):
    for key in ('value_data', 'expect', 'sql_expect', 'regex', 'not_expect', 'show_output'):
        raw_val = fields.get(key, '')
        raw_text = str(raw_val or '').strip()
        if ('&&' in raw_text or '||' in raw_text) and raw_text.count('"') >= 2:
            raw = raw_text
        else:
            raw = _strip_outer_quotes_once(raw_val)
        if raw:
            return raw
    return '.+'


def _format_converted_value_data(expected, check_type=''):
    raw = str(expected or '').strip()
    if not raw:
        return '""'

    semantic_boolean = _normalize_boolean_value_data_expression(raw)
    if semantic_boolean:
        return f'"{semantic_boolean}"'

    stripped = _strip_outer_quotes_once(raw)
    return f'"{stripped}"'


def _split_top_level(text, delimiter):
    parts = []
    buf = []
    depth = 0
    in_quote = False
    i = 0
    while i < len(text):
        ch = text[i]
        if ch == '"' and (i == 0 or text[i - 1] != '\\'):
            in_quote = not in_quote
            buf.append(ch)
            i += 1
            continue

        if not in_quote:
            if ch == '(':
                depth += 1
            elif ch == ')' and depth > 0:
                depth -= 1
            elif depth == 0 and text.startswith(delimiter, i):
                parts.append(''.join(buf).strip())
                buf = []
                i += len(delimiter)
                continue

        buf.append(ch)
        i += 1

    tail = ''.join(buf).strip()
    if tail:
        parts.append(tail)
    return parts


def _canonicalize_user_right_member(value):
    member = _strip_outer_quotes_once(value)
    member = member.split('\\')[-1]
    member = re.sub(r'\s+', ' ', member).strip()
    if not member:
        return ''
    if member == member.lower():
        member = ' '.join(part.capitalize() for part in member.split())
    return member


def _strip_balanced_outer_parens(text):
    s = str(text or '').strip()
    while s.startswith('(') and s.endswith(')'):
        depth = 0
        balanced = True
        for idx, ch in enumerate(s):
            if ch == '(':
                depth += 1
            elif ch == ')':
                depth -= 1
                if depth == 0 and idx != len(s) - 1:
                    balanced = False
                    break
            if depth < 0:
                balanced = False
                break
        if not balanced or depth != 0:
            break
        s = s[1:-1].strip()
    return s


def _regex_escape_user_right_member(value):
    return re.escape(value).replace(r'\ ', ' ')


def _looks_like_quoted_boolean_expression(text):
    raw = str(text or '').strip()
    return ('&&' in raw or '||' in raw or r'\|\|' in raw) and raw.count('"') >= 2


def _prepare_boolean_expression_text(text):
    raw = str(text or '').strip()
    # Normalize source strings like \"A\" \|\| \"B\" before tokenization.
    raw = raw.replace(r'\"', '"')
    if raw.startswith('"') and raw.endswith('"') and len(raw) > 2 and raw.count('"') > 2 and raw[1] in {'^', '('}:
        raw = raw[1:-1].strip()
    if raw.startswith('^') and raw.endswith('$'):
        raw = raw[1:-1].strip()
    raw = raw.replace(r'\|\|', '||')
    raw = raw.replace(r'\&\&', '&&')
    raw = raw.replace(r'\(', '(').replace(r'\)', ')')
    return raw


def _regex_escape_expected_literal(value):
    return re.escape(str(value or ''))


def _normalize_no_members_token_regex(pattern_text):
    pattern = str(pattern_text or '')
    if "'?NO_MEMBERS'?" in pattern:
        return pattern
    return re.sub(r'\bNO_MEMBERS\b', "'?NO_MEMBERS'?", pattern)


def _allow_no_members_alternative(pattern_text):
    pattern = str(pattern_text or '').strip()
    if not pattern or 'NO_MEMBERS' in pattern:
        return pattern

    no_members_alt = "'?NO_MEMBERS'?"
    inline_flags = re.match(r'^\(\?([a-zA-Z]+)\)\^(.*)\$$', pattern, flags=re.DOTALL)
    if inline_flags:
        flags = inline_flags.group(1)
        inner = inline_flags.group(2)
        return f'(?{flags})^(?:{inner}|{no_members_alt})$'

    if pattern.startswith('^') and pattern.endswith('$') and len(pattern) >= 2:
        inner = pattern[1:-1]
        return f'^(?:{inner}|{no_members_alt})$'

    return f'(?s)^(?:{pattern}|{no_members_alt})$'


def _build_contains_all_regex(parts):
    escaped_parts = []
    for part in parts:
        raw = str(part or '').strip()
        if not raw:
            continue
        tokens = [tok for tok in re.split(r'\s+', raw) if tok]
        if not tokens:
            continue
        escaped_parts.append(r'\s+'.join(re.escape(tok) for tok in tokens))
    if not escaped_parts:
        return ''
    # Tenable reports these list-valued policy outputs wrapped in single quotes.
    # Make the regex tolerant to those outer quotes while still requiring every path.
    return "(?is)^'?(?:" + ''.join(f'(?=.*{part})' for part in escaped_parts) + ").*'?$"


def _build_registry_compliance_script(registry_path, expected_paths):
    path_checks = ' -and '.join(
        f"$paths -contains '{_ps_single_quote_escape(path)}'"
        for path in expected_paths
        if str(path or '').strip()
    )
    if not path_checks:
        return ''

    registry_path_ps = _ps_single_quote_escape(registry_path)
    return (
        f"if(Test-Path -Path '{registry_path_ps}'){{"
        f"$paths = (Get-ItemProperty -Path '{registry_path_ps}').Machine;"
        f"if([string]::IsNullOrWhiteSpace([string]$paths)){{ Write-Output '{NON_COMPLIANT_OUTPUT}'; return }};"
        f"if({path_checks}) {{ Write-Output 'Compliant' }} else {{ Write-Output 'Non-Compliant' }}"
        f"}}else {{ Write-Output '{NON_COMPLIANT_OUTPUT}' }}"
    )


def _build_generic_boolean_expected(raw_expected):
    expected = _prepare_boolean_expression_text(raw_expected)
    if '&&' not in expected and '||' not in expected:
        return _strip_outer_quotes_once(expected)

    expr = _strip_balanced_outer_parens(expected)
    alt_terms = []
    for alt in _split_top_level(expr, '||'):
        alt = _strip_balanced_outer_parens(alt)
        terms = []
        for term in _split_top_level(alt, '&&'):
            token = _strip_outer_quotes_once(_strip_balanced_outer_parens(term).strip())
            if token:
                terms.append(token)
        if terms:
            alt_terms.append(terms)

    if not alt_terms:
        return expected

    regex_alts = []
    for terms in alt_terms:
        if len(terms) == 1:
            regex_alts.append(_regex_escape_expected_literal(terms[0]))
            continue
        lookaheads = ''.join(
            f'(?=.*{_regex_escape_expected_literal(term)})'
            for term in terms
        )
        regex_alts.append(f'{lookaheads}.*')

    if len(regex_alts) == 1:
        only = regex_alts[0]
        if only.endswith('.*') and only.startswith('(?='):
            return f'(?s)^{only}$'
        return f'^{only}$'

    return '(?s)^(?:' + '|'.join(regex_alts) + ')$'


def _normalize_boolean_value_data_expression(raw_value_data):
    raw = str(raw_value_data or '').strip()
    if not raw:
        return ''

    normalized = raw.replace(r'\&\&', '&&').replace(r'\|\|', '||')
    if '&&' not in normalized and '||' not in normalized:
        return ''

    semantic = _build_generic_boolean_expected(normalized)
    if not semantic:
        return ''

    try:
        re.compile(semantic)
    except re.error:
        return ''

    return semantic


def _build_applocker_xml_expected_regex(raw_value_data):
    raw = _strip_outer_quotes_once(raw_value_data)
    if not raw:
        return ''

    compact = str(raw).replace('\\r\\n', '').replace('\\r', '').replace('\\n', '')
    compact = compact.replace('\r\n', '').replace('\r', '').replace('\n', '').strip()
    if not (compact.startswith('<FilePublisherRule') or compact.startswith('<FilePathRule')):
        return ''

    pattern = re.escape(compact)
    pattern = pattern.replace(r'LowSection="\*"', r'LowSection=".*"')
    pattern = pattern.replace(r'HighSection="\*"', r'HighSection=".*"')
    return f'(?s)^{pattern}.*$'


def _build_user_rights_expected(fields, fallback_expected):
    raw_expected = _prepare_boolean_expression_text(fallback_expected)
    if _looks_like_quoted_boolean_expression(raw_expected):
        expected = raw_expected
    else:
        expected = _strip_outer_quotes_once(fallback_expected).strip()
    desc_upper = _norm_upper_token(fields.get('description', ''))
    if 'NO ONE' in desc_upper and _norm_upper_token(expected) in {'.+', '^.+$', 'NO ONE'}:
        return '^NO_MEMBERS$'

    if expected.startswith('^'):
        return expected

    expected = _strip_balanced_outer_parens(expected)

    alt_sets = []
    for alt in _split_top_level(expected, '||'):
        members = []
        for raw_member in re.findall(r'"([^"]+)"', alt):
            member = _canonicalize_user_right_member(raw_member)
            if member and member not in members:
                members.append(member)
        if not members:
            continue
        normalized = tuple(sorted(members, key=lambda item: item.upper()))
        if normalized not in alt_sets:
            alt_sets.append(normalized)

    if not alt_sets:
        # Prevent emitting operator-only patterns when source expressions are malformed.
        if re.fullmatch(r'\^?\\?\|\\?\|\$?', expected):
            return '^.+$'
        return expected

    alt_sets.sort(key=lambda item: (len(item), [part.upper() for part in item]))

    if len(alt_sets) == 1:
        return '^' + ','.join(_regex_escape_user_right_member(part) for part in alt_sets[0]) + '$'

    chain = True
    for idx in range(1, len(alt_sets)):
        prev = alt_sets[idx - 1]
        curr = alt_sets[idx]
        if len(curr) != len(prev) + 1 or curr[:len(prev)] != prev:
            chain = False
            break

    if chain:
        pattern = '^' + _regex_escape_user_right_member(alt_sets[0][0])
        for idx in range(1, len(alt_sets[0])):
            pattern += f',{_regex_escape_user_right_member(alt_sets[0][idx])}'
        for idx in range(1, len(alt_sets)):
            added = alt_sets[idx][-1]
            pattern += f'(,{_regex_escape_user_right_member(added)})?'
        return pattern + '$'

    alternates = []
    for alt in alt_sets:
        alternates.append(','.join(_regex_escape_user_right_member(part) for part in alt))
    return '^(?:' + '|'.join(alternates) + ')$'


def _fallback_user_rights_expected_from_description(fields):
    desc = _strip_outer_quotes_once(fields.get('description', ''))
    if not desc:
        return ''

    m = re.search(r'\bis set to\b\s*(.+)$', desc, flags=re.IGNORECASE)
    if not m:
        return ''

    raw_tail = m.group(1).strip().rstrip('.')
    if not raw_tail:
        return ''

    members = []
    for part in raw_tail.split(','):
        member = _canonicalize_user_right_member(part)
        if member and member not in members:
            members.append(member)

    if not members:
        return ''

    members = sorted(members, key=lambda item: item.upper())
    return '^' + ','.join(_regex_escape_user_right_member(part) for part in members) + '$'


def _build_user_rights_powershell_script(right_type):
    right_ps = _ps_single_quote_escape(right_type)
    right_line_pattern = _ps_single_quote_escape(rf'^\s*{re.escape(right_type)}\s*=')
    return (
        "$t=[IO.Path]::GetTempFileName(); "
        "$ok=$false; "
        "try { secedit /export /cfg $t /areas USER_RIGHTS /mergedpolicy /quiet | Out-Null; $ok=$true } catch {} ; "
        "if(-not $ok){ secedit /export /cfg $t /areas USER_RIGHTS /quiet | Out-Null }; "
        f"$l=(Get-Content $t | Where-Object {{ $_ -match '{right_line_pattern}' }} | Select-Object -Last 1); "
        "$m=if($l){ ($l -replace '.*=', '').Split(',') | ForEach-Object { $s=$_.Trim(); $s=$s -replace '^\\*', ''; "
        "if($s -match '^S-1-5-'){ try{([Security.Principal.SecurityIdentifier]$s).Translate([Security.Principal.NTAccount]).Value.Split('\\\\')[-1]}catch{$s} } else { if($s -match '\\\\'){ $s.Split('\\\\')[-1] } else { $s } } } | Where-Object { $_ } | Sort-Object -Unique }else{ @() }; "
        "Remove-Item $t -Force -ErrorAction SilentlyContinue; if($m){ $m -join ',' } else { 'NO_MEMBERS' }"
    )


def _rewrite_user_rights_powershell_script(script_text):
    script = _strip_outer_quotes_once(script_text)
    if 'secedit /export /cfg $t /areas USER_RIGHTS' not in script:
        return script

    right_match = re.search(r"-match\s+'([^']+)'", script)
    if not right_match:
        return script

    right_expr = right_match.group(1)
    right_name_match = re.search(r'(Se[A-Za-z0-9_]+)', right_expr)
    if not right_name_match:
        return script

    return _build_user_rights_powershell_script(right_name_match.group(1))


def _build_sid_account_name_powershell_script(sid_suffix):
    return (
        "$A = Get-CimInstance -ClassName Win32_UserAccount -EA SilentlyContinue | "
        f"Where-Object {{ $_.SID -match '-{sid_suffix}$' }} | Select-Object -ExpandProperty Name -First 1; "
        f"if([string]::IsNullOrEmpty($A)){{ Write-Output '{NON_COMPLIANT_OUTPUT}' }} else {{ Write-Output $A }}"
    )


def _build_sid_account_status_powershell_script(sid_suffix):
    return (
        "$A = Get-CimInstance -ClassName Win32_UserAccount -EA SilentlyContinue | "
        f"Where-Object {{ $_.SID -match '-{sid_suffix}$' }} | Select-Object -First 1; "
        "if($A -and $A.Disabled -ne $null){ if($A.Disabled){ Write-Output 'Disabled' } else { Write-Output 'Enabled' } }"
        f" elseif($A -and -not [string]::IsNullOrEmpty($A.Name)){{ Write-Output $A.Name }} else {{ Write-Output '{NON_COMPLIANT_OUTPUT}' }}"
    )


def _apply_powershell_arg_corrections(desc_upper, script_text):
    script = str(script_text or '')

    # Canonicalize legacy lockout-admin scripts that output 1/0 or rely on
    # positional match state. Emit explicit status values and deterministic
    # non-compliant output when no data is returned.
    if 'ALLOWADMINISTRATORLOCKOUT' in _norm_upper_token(script):
        return (
            "$secfile = [System.IO.Path]::GetTempFileName(); "
            "secedit /export /cfg $secfile /areas SECURITYPOLICY /quiet | Out-Null; "
            "$line = Select-String -Path $secfile -Pattern 'AllowAdministratorLockout\\s*=\\s*(\\d)' | Select-Object -First 1; "
            "Remove-Item $secfile -ErrorAction SilentlyContinue; "
            "if($line -and $line.Matches.Count -gt 0){ "
            "if($line.Matches[0].Groups[1].Value -eq '1'){ Write-Output 'Enabled' } else { Write-Output 'Disabled' } "
            f"}} else {{ Write-Output '{NON_COMPLIANT_OUTPUT}' }}"
        )

    if 'REG_CHECK|DEFAULTPASSWORD|HKLM\\SOFTWARE\\MICROSOFT\\WINDOWS NT\\CURRENTVERSION\\WINLOGON' in _norm_upper_token(script):
        return (
            "$p='HKLM:\\Software\\Microsoft\\Windows NT\\CurrentVersion\\Winlogon'; "
            "if(Test-Path $p){ "
            "$v=(Get-ItemProperty $p -Name 'DefaultPassword' -EA SilentlyContinue).DefaultPassword; "
            "if([string]::IsNullOrEmpty($v)){ 'NOT_EXIST' }else{ 'EXIST' } "
            "}else{ 'NOT_EXIST' }"
        )

    # This control must explicitly check value existence, not emit a marker string.
    if 'ENSURE DEFAULTPASSWORD DOES NOT EXIST' in desc_upper:
        return (
            "$p='HKLM:\\Software\\Microsoft\\Windows NT\\CurrentVersion\\Winlogon'; "
            "if(Test-Path $p){ "
            "$v=(Get-ItemProperty $p -Name 'DefaultPassword' -EA SilentlyContinue).DefaultPassword; "
            "if([string]::IsNullOrEmpty($v)){ 'NOT_EXIST' }else{ 'EXIST' } "
            "}else{ 'NOT_EXIST' }"
        )

    # This policy is machine-scoped and should read from HKLM.
    if 'ALWAYS INSTALL WITH ELEVATED PRIVILEGES' in desc_upper:
        script = script.replace(
            'Registry::HKU\\Software\\Policies\\Microsoft\\Windows\\Installer',
            'Registry::HKLM\\Software\\Policies\\Microsoft\\Windows\\Installer',
        )

    return script


def _ensure_non_compliant_on_empty_output(script_text):
    script = _strip_outer_quotes_once(script_text)
    if not str(script or '').strip():
        return f"Write-Output '{NON_COMPLIANT_OUTPUT}'"

    # Avoid wrapping the same script multiple times across repeated runs.
    if '$__pysc_result' in script:
        return script

    return (
        "$__pysc_result = (& { "
        + script
        + " } | Out-String); "
        + f"if([string]::IsNullOrWhiteSpace([string]$__pysc_result)){{ Write-Output '{NON_COMPLIANT_OUTPUT}' }} "
        + "else { Write-Output ($__pysc_result.Trim()) }"
    )


def _build_check_account_powershell_script(fields):
    acct = _norm_upper_token(fields.get('account_type', ''))
    desc = _norm_upper_token(fields.get('description', ''))
    expected = _norm_upper_token(_best_effort_expected_text(fields))
    check_type = _norm_upper_token(fields.get('check_type', ''))

    if acct == 'ADMINISTRATOR_ACCOUNT':
        if 'RENAME' in desc or check_type in {'CHECK_NOT_REGEX', 'CHECK_NOT_EQUAL'}:
            return _build_sid_account_name_powershell_script('500')
        return _build_sid_account_status_powershell_script('500')

    if acct == 'GUEST_ACCOUNT':
        if 'RENAME' in desc or check_type == 'CHECK_NOT_EQUAL' or expected == 'GUEST':
            return _build_sid_account_name_powershell_script('501')
        return _build_sid_account_status_powershell_script('501')

    acct_ps = _ps_single_quote_escape(_strip_outer_quotes_once(fields.get('account_type', '')))
    return (
        "$A = Get-CimInstance -ClassName Win32_UserAccount -ErrorAction SilentlyContinue | "
        f"Where-Object {{ $_.Name -eq '{acct_ps}' }} | Select-Object -First 1; "
        f"if($A){{ Write-Output $A.Name }} else {{ Write-Output '{NON_COMPLIANT_OUTPUT}' }};"
    )


def _best_effort_powershell_script(fields):
    ctype = _norm_upper_token(fields.get('type', ''))
    expected = _best_effort_expected_text(fields)

    reg_key = _strip_outer_quotes_once(fields.get('reg_key', ''))
    reg_item = _strip_outer_quotes_once(fields.get('reg_item', fields.get('key_item', '')))
    if ctype in {'REGISTRY_SETTING', 'REG_CHECK', 'BANNER_CHECK'} and reg_key and reg_item:
        reg_key_ps = _ps_single_quote_escape(reg_key)
        reg_item_ps = _ps_single_quote_escape(reg_item)
        return (
            f"$P = 'Registry::{reg_key_ps}'; "
            f"$K = '{reg_item_ps}'; "
            f"if(-not (Test-Path -Path $P)){{ Write-Output '{NON_COMPLIANT_OUTPUT}'; return }}; "
            "$R = Get-ItemProperty -Path $P -ErrorAction SilentlyContinue; "
            f"if($null -eq $R){{ Write-Output '{NON_COMPLIANT_OUTPUT}'; return }}; "
            "$prop = $R.PSObject.Properties[$K]; "
            f"if($null -eq $prop -or $null -eq $prop.Value){{ Write-Output '{NON_COMPLIANT_OUTPUT}'; return }}; "
            "$V = $prop.Value; "
            f"if($null -eq $V -or [string]::IsNullOrWhiteSpace([string]$V)){{ Write-Output '{NON_COMPLIANT_OUTPUT}' }} "
            "elseif($V -is [System.Array]){ "
            "$joined = ($V -join [Environment]::NewLine); "
            f"if([string]::IsNullOrWhiteSpace([string]$joined)){{ Write-Output '{NON_COMPLIANT_OUTPUT}' }} else {{ Write-Output $joined }} }} "
            "else { Write-Output $V; }"
        )

    if ctype == 'USER_RIGHTS_POLICY':
        right = _strip_outer_quotes_once(fields.get('right_type', ''))
        if right:
            return _build_user_rights_powershell_script(right)

    if ctype == 'CHECK_ACCOUNT':
        acct = _strip_outer_quotes_once(fields.get('account_type', ''))
        if acct:
            return _build_check_account_powershell_script(fields)

    if ctype == 'WMI_POLICY':
        wmi_key = _strip_outer_quotes_once(fields.get('wmi_key', ''))
        wmi_ns = _strip_outer_quotes_once(fields.get('wmi_namespace', ''))
        wmi_req = _strip_outer_quotes_once(fields.get('wmi_request', ''))
        wmi_attr = _strip_outer_quotes_once(fields.get('wmi_attribute', ''))
        if wmi_req:
            wmi_req_ps = _ps_single_quote_escape(wmi_req)
            wmi_ns_ps = _ps_single_quote_escape(wmi_ns or 'root\\cimv2')
            if wmi_attr:
                wmi_attr_ps = _ps_single_quote_escape(wmi_attr)
                return (
                    f"$W = Get-WmiObject -Namespace '{wmi_ns_ps}' -Query '{wmi_req_ps}' -ErrorAction SilentlyContinue; "
                    f"if($null -eq $W){{ Write-Output '{NON_COMPLIANT_OUTPUT}'; return }}; "
                    f"$V = $W.{wmi_attr_ps}; "
                    f"if($null -eq $V -or [string]::IsNullOrWhiteSpace([string]$V)){{ Write-Output '{NON_COMPLIANT_OUTPUT}' }} else {{ Write-Output $V }};"
                )
            return (
                f"$W = Get-WmiObject -Namespace '{wmi_ns_ps}' -Query '{wmi_req_ps}' -ErrorAction SilentlyContinue; "
                f"if($null -eq $W){{ Write-Output '{NON_COMPLIANT_OUTPUT}' }} else {{ Write-Output $W }};"
            )
        if wmi_key:
            wmi_key_ps = _ps_single_quote_escape(wmi_key)
            return (
                f"$W = Get-WmiObject -Class '{wmi_key_ps}' -ErrorAction SilentlyContinue; "
                f"if($null -eq $W){{ Write-Output '{NON_COMPLIANT_OUTPUT}' }} else {{ Write-Output $W }};"
            )

    if ctype == 'CMD_EXEC':
        cmd = _strip_outer_quotes_once(fields.get('cmd', ''))
        if cmd:
            cmd_ps = _ps_single_quote_escape(cmd)
            return f"$C = cmd /c '{cmd_ps}'; Write-Output $C;"

    if ctype in {'FILE_CONTENT_CHECK', 'FILE_CONTENT_CHECK_NOT', 'FILE_CHECK', 'FILE_CHECK_NOT', 'OFFLINE_CONFIG_CHECK', 'OFFLINE_BANNER_CHECK'}:
        item = _strip_outer_quotes_once(fields.get('item', ''))
        if item:
            item_ps = _ps_single_quote_escape(item)
            return (
                f"if (Test-Path '{item_ps}') {{ Get-Content -Path '{item_ps}' -ErrorAction SilentlyContinue }} "
                f"else {{ Write-Output '{NON_COMPLIANT_OUTPUT}' }}"
            )

    marker_parts = [ctype]
    key_hint = _strip_outer_quotes_once(
        fields.get('reg_item')
        or fields.get('key_item')
        or fields.get('right_type')
        or fields.get('account_type')
        or fields.get('audit_policy_subcategory')
        or fields.get('item')
        or fields.get('cmd')
        or ''
    )
    if key_hint:
        marker_parts.append(key_hint)
    if expected:
        marker_parts.append(expected)
    marker = _ps_single_quote_escape('|'.join(marker_parts))
    return f"Write-Output '{marker}';"


def _convert_remaining_item_to_audit_powershell(fields):
    ctype = _norm_upper_token(fields.get('type', ''))
    if not ctype or ctype == 'AUDIT_POWERSHELL':
        return fields

    expected = _best_effort_expected_text(fields)
    original_check_type = str(fields.get('check_type', '') or '').strip()
    original_check_type_upper = _norm_upper_token(original_check_type)
    if ctype == 'USER_RIGHTS_POLICY' and original_check_type_upper not in {'CHECK_SUPERSET', 'CHECK_SUBSET', 'CHECK_EQUAL'}:
        expected = _build_user_rights_expected(fields, expected)
    script = _best_effort_powershell_script(fields)

    converted = OrderedDict()
    converted['type'] = 'AUDIT_POWERSHELL'
    for key in ('description', 'info', 'solution', 'reference', 'see_also'):
        if key in fields:
            converted[key] = fields.get(key)
    converted['value_type'] = 'POLICY_TEXT'
    converted['value_data'] = _format_converted_value_data(expected, original_check_type)
    converted['powershell_args'] = '-NoProfile -ExecutionPolicy Bypass -Command'
    converted['powershell_script'] = script
    converted['expect'] = _strip_outer_quotes_once(_format_converted_value_data(expected, original_check_type))
    return converted


def _should_force_convert_non_powershell(platform_hint, fields):
    hint = _norm_upper_token(platform_hint)
    if hint in {'MSSRV', 'MSWRK'}:
        return True

    desc_platform = _norm_upper_token(detect_platform(fields.get('description', '')))
    return desc_platform in {'MSSRV', 'MSWRK'}


def _is_windows_audit_platform(platform_hint='', fields=None):
    hint = _norm_upper_token(platform_hint)
    if hint in {'MSSRV', 'MSWRK'}:
        return True

    desc_platform = _norm_upper_token(detect_platform((fields or {}).get('description', '')))
    return desc_platform in {'MSSRV', 'MSWRK'}


def _ensure_audit_powershell_metadata_fields(fields):
    if _norm_upper_token(fields.get('type', '')) != 'AUDIT_POWERSHELL':
        return fields

    ensured = OrderedDict(fields)
    raw_expected = str(ensured.get('value_data', '') or '').strip()
    if raw_expected:
        script = _strip_outer_quotes_once(ensured.get('powershell_args', ''))
        check_type_upper = _norm_upper_token(ensured.get('check_type', ''))
        if (
            'secedit /export /cfg $t /areas USER_RIGHTS' in script
            and _looks_like_quoted_boolean_expression(raw_expected)
        ):
            normalized_expected = _build_user_rights_expected(ensured, raw_expected)
            ensured['value_data'] = f'"{normalized_expected}"'
        else:
            semantic_boolean = _normalize_boolean_value_data_expression(raw_expected)
            if semantic_boolean:
                ensured['value_data'] = f'"{semantic_boolean}"'
            else:
                normalized_expected = _build_bracket_range_expected(raw_expected)
                if normalized_expected:
                    ensured['value_data'] = f'"{normalized_expected}"'
    if 'solution' not in ensured:
        ensured['solution'] = '""'
    if 'reference' not in ensured:
        ensured['reference'] = '""'
    if 'see_also' not in ensured:
        ensured['see_also'] = f'"{SEE_ALSO_REPLACEMENT}"'
    return ensured


def _finalize_audit_powershell_fields_for_emit(fields, platform_hint=''):
    if _norm_upper_token(fields.get('type', '')) != 'AUDIT_POWERSHELL':
        return OrderedDict(fields)

    finalized = OrderedDict()
    for key, value in fields.items():
        if key in {'powershell_args', 'powershell_script', 'powershell_option'}:
            continue
        if key == 'expect' and _is_windows_audit_platform(platform_hint, fields):
            continue
        finalized[key] = value

    ps_args = fields.get('powershell_script', fields.get('powershell_args', ''))

    # Normalize known value-data drift that causes false mismatches.
    raw_value_data = str(finalized.get('value_data', '') or '')
    if 'Sysmonlog' in raw_value_data:
        finalized['value_data'] = raw_value_data.replace('Sysmonlog', 'SysmonLog')

    desc_upper = _norm_upper_token(_strip_outer_quotes_once(finalized.get('description', '')))
    forced_ps_args = None
    if (
        'ACCOUNT LOCKOUT RESET' in desc_upper
        or 'ACCOUNT LOCKOUT DURATION' in desc_upper
        or 'ACCOUNT LOCKOUT THRESHOLD' in desc_upper
    ):
        scalar = _first_int(_strip_outer_quotes_once(finalized.get('value_data', '')))
        if scalar is not None:
            finalized['value_data'] = f'"{scalar}"'

    # Keep expected values aligned with canonical lockout-admin script output.
    if 'ALLOW ADMINISTRATOR ACCOUNT LOCKOUT' in desc_upper:
        expected_inner = _norm_upper_token(_strip_outer_quotes_once(finalized.get('value_data', '')))
        if expected_inner in {'1', 'ENABLED', 'TRUE'}:
            finalized['value_data'] = '"(?i)^Enabled$"'
            finalized['check_type'] = 'CHECK_REGEX'
        elif expected_inner in {'0', 'DISABLED', 'FALSE'}:
            finalized['value_data'] = '"(?i)^Disabled$"'
            finalized['check_type'] = 'CHECK_REGEX'

    # For password expiry warning, enforce a minimum threshold instead of hard-capping
    # so values above the documented minimum still pass.
    if 'PROMPT USER TO CHANGE PASSWORD BEFORE EXPIRATION' in desc_upper:
        min_days = _first_int(_strip_outer_quotes_once(finalized.get('value_data', '')))
        if min_days is not None:
            finalized['value_data'] = f'"{_build_ge_integer_regex(min_days)}"'

    if ps_args:
        script_text = _strip_outer_quotes_once(ps_args)
        script_text = _normalize_embedded_powershell_script(script_text)

        # Repair broken USER_RIGHTS expected patterns that collapse to only
        # operators/anchors after escaping (for example "^\^\|\|\$$").
        value_data_inner = _strip_outer_quotes_once(finalized.get('value_data', ''))
        if (
            'secedit /export /cfg $t /areas USER_RIGHTS' in script_text
            and value_data_inner
            and re.fullmatch(r'[\^$\\|()\s]+', value_data_inner)
        ):
            fallback_expected = _fallback_user_rights_expected_from_description(fields)
            if fallback_expected:
                finalized['value_data'] = f'"{fallback_expected}"'
                value_data_inner = fallback_expected

        # In merged MSSRV output, many environments return NO_MEMBERS for
        # USER_RIGHTS checks where the baseline text expects explicit members.
        # Allow that runtime value while still preserving the canonical list match.
        if (
            'secedit /export /cfg $t /areas USER_RIGHTS' in script_text
            and 'NO ONE' not in desc_upper
            and 'DENY LOG ON LOCALLY TO INCLUDE GUESTS' not in desc_upper
        ):
            expected_inner = _strip_outer_quotes_once(finalized.get('value_data', ''))
            expected_inner = _normalize_user_rights_expected_for_tenable(fields, expected_inner)
            if expected_inner and 'NO_MEMBERS' not in expected_inner:
                finalized['value_data'] = f'"{_allow_no_members_alternative(expected_inner)}"'
            elif expected_inner and 'NO_MEMBERS' in expected_inner:
                normalized_expected = _normalize_no_members_token_regex(expected_inner)
                finalized['value_data'] = f'"{normalized_expected}"'

            # Tenable parser safety: this control is frequently emitted with an
            # over-escaped alternate expression for WdiServiceHost that can
            # produce unknown token parse failures (for example "\\\ ").
            if 'PROFILE SYSTEM PERFORMANCE' in desc_upper and 'WDISERVICEHOST' in desc_upper:
                finalized['value_data'] = '"^(?:Administrators,(?:NT SERVICE\\\\)?WdiServiceHost|\'?NO_MEMBERS\'?)$"'

        # Keep this control strict: custom groups should fail unless Guests is present as expected.
        if 'DENY LOG ON LOCALLY TO INCLUDE GUESTS' in desc_upper:
            finalized['value_data'] = '"^Guests$"'

        # This USER_RIGHTS check is emitted as a regex pattern and must be
        # evaluated with regex semantics to allow Administrators or NO_MEMBERS.
        if 'PERFORM VOLUME MAINTENANCE TASKS IS SET TO ADMINISTRATORS' in desc_upper:
            finalized['check_type'] = 'CHECK_REGEX'

        # Baseline-aligned expected lists for remotely accessible registry paths.
        if 'NETWORK ACCESS - REMOTELY ACCESSIBLE REGISTRY PATHS AND SUB-PATHS ALLOWED EXACT PATHS IS CONFIGURED' in desc_upper:
            finalized['info'] = (
                '"This audit is written to dynamically identify if all paths are present in any order. '
                'This policy setting determines which registry paths will be accessible over the network, '
                'regardless of the users or groups listed in the access control list (ACL) of the winreg registry key."'
            )
            finalized['value_data'] = '"^Compliant$"'
            ps_args = _build_registry_compliance_script(
                r'HKLM:\System\CurrentControlSet\Control\SecurePipeServers\Winreg\AllowedExactPaths',
                [
                    'System\\CurrentControlSet\\Control\\ProductOptions',
                    'System\\CurrentControlSet\\Control\\Server Applications',
                    'Software\\Microsoft\\Windows NT\\CurrentVersion',
                ],
            )
            if ps_args:
                forced_ps_args = f'"{ps_args}"'
            finalized['check_type'] = 'CHECK_REGEX'
        if 'NETWORK ACCESS - REMOTELY ACCESSIBLE REGISTRY PATHS AND SUB-PATHS ALLOWED PATHS IS CONFIGURED' in desc_upper:
            finalized['info'] = (
                '"This audit is written to dynamically identify if all paths are present in any order. '
                'This policy setting determines which registry paths and sub-paths will be accessible over the network, '
                'regardless of the users or groups listed in the access control list (ACL) of the winreg registry key."'
            )
            finalized['value_data'] = '"^Compliant$"'
            ps_args = _build_registry_compliance_script(
                r'HKLM:\System\CurrentControlSet\Control\SecurePipeServers\Winreg\AllowedPaths',
                [
                    'System\\CurrentControlSet\\Control\\Print\\Printers',
                    'System\\CurrentControlSet\\Services\\Eventlog',
                    'Software\\Microsoft\\OLAP Server',
                    'Software\\Microsoft\\Windows NT\\CurrentVersion\\Print',
                    'Software\\Microsoft\\Windows NT\\CurrentVersion\\Windows',
                    'System\\CurrentControlSet\\Control\\ContentIndex',
                    'System\\CurrentControlSet\\Control\\Terminal Server',
                    'System\\CurrentControlSet\\Control\\Terminal Server\\UserConfig',
                    'System\\CurrentControlSet\\Control\\Terminal Server\\DefaultUserConfiguration',
                    'Software\\Microsoft\\Windows NT\\CurrentVersion\\Perflib',
                    'System\\CurrentControlSet\\Services\\SysmonLog',
                ],
            )
            if ps_args:
                forced_ps_args = f'"{ps_args}"'
            finalized['check_type'] = 'CHECK_REGEX'

        if 'SERVICES - RAPID7 INSIGHT AGENT SERVICE' in desc_upper:
            finalized['value_data'] = '"Running"'
            finalized['check_type'] = 'CHECK_REGEX'
            ps_args = '"$noutput = (Get-wmiobject win32_service | where-object {$_.name -eq \'ir_agent\'}).state; if ($noutput -eq $Null) {write-host \'Service Not Found\'} else {$noutput}"'

        # Normalize known script path typos observed in source audits.
        script_text = re.sub(
            r"(?i)Services\\Netlogon\\Parameter(?!s)(['\\])",
            r"Services\\Netlogon\\Parameters\1",
            script_text,
        )
        script_text = re.sub(
            r'(?i)Services\\Netlogon\\Parameters+',
            r'Services\\Netlogon\\Parameters',
            script_text,
        )
        script_text = re.sub(
            r"(?i)PublicProfile\\Loggin\\Software\\Policies\\Microsoft\\WindowsFirewall\\PublicProfile\\Logging",
            r"PublicProfile\\Logging",
            script_text,
        )

        # Persist normalized script text by default; conditional rewrites below
        # may still replace this with more specialized scripts.
        script_text = _apply_powershell_arg_corrections(desc_upper, script_text)
        script_text = re.sub(
            r"(?i)Write-Output\s+['\"](?:Value Not Found|Path Not Found|NOT_FOUND|Not Found)['\"]",
            f"Write-Output '{NON_COMPLIANT_OUTPUT}'",
            script_text,
        )
        ps_args = f'"{script_text}"'
        if forced_ps_args is not None:
            ps_args = forced_ps_args

        if 'PasswordExpiryWarning' in script_text:
            min_days = _first_int(_strip_outer_quotes_once(finalized.get('value_data', '')))
            if min_days is not None:
                finalized['value_data'] = f'"{_build_ge_integer_regex(min_days)}"'

        normalized_expected = _normalize_legacy_net_accounts_expected(script_text, finalized.get('value_data', ''))
        if normalized_expected != finalized.get('value_data', ''):
            finalized['value_data'] = normalized_expected

        # AppLocker XML payloads can differ by trailing line endings; normalize
        # these checks to regex mode and tolerate wildcard version ranges.
        applocker_expected = _build_applocker_xml_expected_regex(finalized.get('value_data', ''))
        if applocker_expected:
            finalized['check_type'] = 'CHECK_REGEX'
            finalized['value_data'] = f'"{applocker_expected}"'
            if not _is_windows_audit_platform(platform_hint, fields):
                finalized['expect'] = f'"{applocker_expected}"'

        if "net user 'ADMINISTRATOR_ACCOUNT'" in script_text:
            ps_args = f'"{_build_sid_account_name_powershell_script("500")}"'
        elif "net user 'GUEST_ACCOUNT'" in script_text:
            ps_args = f'"{_build_check_account_powershell_script(fields | {"account_type": "GUEST_ACCOUNT"})}"'
        elif 'secedit /export /cfg $t /areas USER_RIGHTS' in script_text:
            ps_args = f'"{_rewrite_user_rights_powershell_script(script_text)}"'
        elif re.search(r'\bnet\s+accounts\b', script_text, flags=re.IGNORECASE) and re.search(r'password\s+complexity', script_text, flags=re.IGNORECASE):
            ps_args = f'"{_rewrite_password_complexity_script(script_text)}"'
        elif 'RestrictSendingNTLMTraffic' in script_text:
            ps_args = f'"{_rewrite_ntlm_outgoing_traffic_script(script_text)}"'
        elif re.search(r'(?i)Services\\Netlogon\\Parameters', script_text):
            ps_args = f'"{_rewrite_netlogon_parameters_default_script(script_text)}"'
        elif re.search(r'\bauditpol\b', script_text, flags=re.IGNORECASE):
            ps_args = f'"{_rewrite_auditpol_powershell_script(script_text)}"'

        # Universal guard: if any control script returns no data, force
        # deterministic non-compliant output.
        ps_args = f'"{_ensure_non_compliant_on_empty_output(ps_args)}"'

        # Keep the canonical Tenable field name for PowerShell checks.
        finalized['powershell_args'] = _render_powershell_args_value(ps_args)

        # Emit option only when explicitly non-default to keep output clean.
        ps_option = str(fields.get('powershell_option', '') or '').strip()
        if ps_option and _norm_upper_token(ps_option) != 'CAN_BE_EQUAL':
            finalized['powershell_option'] = ps_option

        # Final sticky override after all rewrites: local admin must be disabled.
        if 'ACCOUNT - LOCAL ADMIN ACCOUNT' in desc_upper:
            desc_text = _strip_outer_quotes_once(finalized.get('description', ''))
            desc_text = re.sub(
                r'(?i)Account\s*-\s*Local\s+Admin\s+Account\s+Enabled',
                'Account - Local Admin Account Disabled',
                desc_text,
            )
            finalized['description'] = f'"{desc_text}"'
            finalized['value_data'] = '"Disabled"'
            finalized['powershell_args'] = _render_powershell_args_value(_build_sid_account_status_powershell_script("500"))

    return finalized


CONTROL_KEY_FIELDS_BY_TYPE = {
    'REGISTRY_SETTING': ['reg_key', 'reg_item', 'reg_option', 'check_type'],
    'AUDIT_POLICY_SUBCATEGORY': ['audit_policy_subcategory', 'check_type'],
    'AUDIT_POWERSHELL': ['powershell_script', 'powershell_args'],
    'BANNER_CHECK': ['reg_key', 'reg_item', 'check_type'],
    'CHECK_ACCOUNT': ['account_type', 'check_type'],
    'LOCKOUT_POLICY': ['lockout_policy', 'check_type'],
    'PASSWORD_POLICY': ['password_policy', 'check_type'],
    'REG_CHECK': ['reg_option', 'key_item', 'check_type'],
    'USER_RIGHTS_POLICY': ['right_type', 'check_type'],
}

CONTROL_KEY_FALLBACK_FIELDS = [
    'reg_key',
    'reg_item',
    'key_item',
    'audit_policy_subcategory',
    'account_type',
    'lockout_policy',
    'password_policy',
    'right_type',
    'powershell_args',
    'cmd',
    'item',
    'wmi_key',
    'interface_name',
    'policy_arn',
]


def _strip_outer_quotes_once(value):
    if value is None:
        return ''
    s = str(value).strip()
    if len(s) >= 2 and s[0] == s[-1] and s[0] in ('"', "'"):
        return s[1:-1]
    return s


def _strip_bom_prefix(value):
    if value is None:
        return ''
    text = str(value).replace('\ufeff', '').replace('ï»¿', '')
    return text.lstrip('\ufeff').lstrip('ï»¿')


def _write_audit_text(path, text):
    """Write audit text with CRLF line endings regardless of host platform.

    Tenable .audit files use Windows-style CRLF endings; opening in text mode
    without an explicit newline would translate '\\n' to the host OS's
    os.linesep (LF on Linux), diverging from the CRLF golden snapshots.
    """
    normalized = text.replace('\r\n', '\n').replace('\r', '\n').replace('\n', '\r\n')
    with open(path, 'w', encoding='utf-8', newline='') as fh:
        fh.write(normalized)


def _sanitize_audit_lines(lines):
    cleaned = []
    for index, line in enumerate(lines or []):
        text = _strip_bom_prefix(line)
        if index == 0:
            text = text.lstrip('\ufeff').lstrip('ï»¿')
        cleaned.append(text)
    return cleaned


def _assert_no_encoding_markers(text, context):
    issues = []
    if '\ufeff' in text:
        issues.append('U+FEFF')
    if 'ï»¿' in text:
        issues.append('mojibake-UTF8-BOM')
    if issues:
        raise ValueError(f"{context}: forbidden encoding markers found ({', '.join(issues)})")


def derive_expected_value(fields):
    raw = fields.get('value_data', '')
    cleaned = _strip_outer_quotes_once(raw)
    if re.fullmatch(r'-?\d+', cleaned):
        try:
            return int(cleaned)
        except ValueError:
            return cleaned
    if re.fullmatch(r'-?\d+\.\d+', cleaned):
        try:
            return float(cleaned)
        except ValueError:
            return cleaned
    return cleaned


def summarize_powershell_audit_target(powershell_args):
    script = _strip_outer_quotes_once(powershell_args)
    if not script:
        return ''

    m_wmi_service = re.search(
        r"get-wmiobject\s+win32_service[\s\S]*?\.name\s*-eq\s*['\"]([^'\"]+)['\"]",
        script,
        flags=re.IGNORECASE,
    )
    if m_wmi_service:
        return m_wmi_service.group(1)

    m_net_accounts_quoted = re.search(
        r"\bnet\s+accounts\b[^\n;]*?Select-string\s+['\"]([^'\"]+)['\"]",
        script,
        flags=re.IGNORECASE,
    )
    if m_net_accounts_quoted:
        return f"net accounts {m_net_accounts_quoted.group(1)}"

    m_net_accounts_word = re.search(
        r"\bnet\s+accounts\b[^\n;]*?Select-string\s+([A-Za-z][A-Za-z0-9_-]*)",
        script,
        flags=re.IGNORECASE,
    )
    if m_net_accounts_word:
        return f"net accounts {m_net_accounts_word.group(1)}"

    m_optional_feature = re.search(
        r"Get-WindowsOptionalFeature\b[^\n;]*?-FeatureName\s+(['\"]?)([A-Za-z0-9_.-]+)\1",
        script,
        flags=re.IGNORECASE,
    )
    if m_optional_feature:
        return m_optional_feature.group(2)

    m_reg_itemprop = re.search(
        r"Get-ItemProperty\b[^\n;)]*?(?:-Path\s+)?['\"](?:registry::)?([^'\"]+)['\"]\)\.([A-Za-z0-9_]+)",
        script,
        flags=re.IGNORECASE,
    )
    if m_reg_itemprop:
        return f"{m_reg_itemprop.group(1)} {m_reg_itemprop.group(2)}"

    m_get_service = re.search(
        r"Get-Service\b[^\n;]*?-Name\s+(['\"]?)([A-Za-z0-9_.-]+)\1",
        script,
        flags=re.IGNORECASE,
    )
    if m_get_service:
        return f"$s={m_get_service.group(2)}"

    m_cim_class = re.search(r"Get-CimInstance\s+([A-Za-z0-9_]+)", script, flags=re.IGNORECASE)
    m_cim_name = re.search(r"\.Name\s*-eq\s*['\"]([^'\"]+)['\"]", script, flags=re.IGNORECASE)
    if m_cim_class and m_cim_name:
        return f"$s={m_cim_class.group(1)} '{m_cim_name.group(1)}'"

    m_auditpol = re.search(
        r"\bauditpol\b[^\n;]*?/subcategory\s*:\s*['\"]([^'\"]+)['\"]",
        script,
        flags=re.IGNORECASE,
    )
    if m_auditpol:
        return f"auditpol {m_auditpol.group(1).strip()}"

    if re.search(r'\bsecedit\b', script, flags=re.IGNORECASE):
        m_policy = re.search(
            r"-match\s+['\"]\^?(?:\s*|\\+s\*)*([A-Za-z0-9_]+)(?:\s*|\\+s\*)*=",
            script,
            flags=re.IGNORECASE,
        )
        if m_policy:
            return f"secedit {m_policy.group(1)}"
        return 'secedit'

    m_registry_prop = re.search(r"PSObject\.Properties\[['\"]([^'\"]+)['\"]\]", script, flags=re.IGNORECASE)
    if m_registry_prop:
        return f"registry {m_registry_prop.group(1)}"

    m_registry_path = re.search(r"-Path\s+['\"](?:Registry::)?([^'\"]+)['\"]", script, flags=re.IGNORECASE)
    if m_registry_path:
        return f"registry {m_registry_path.group(1)}"

    return re.sub(r'\s+', ' ', script).strip()


def _norm_token(value):
    if value is None:
        return ''
    s = str(value).strip().strip('"').strip("'")
    s = re.sub(r'\s+', ' ', s)
    return s


def _norm_upper_token(value):
    return _norm_token(value).upper()


def _normalize_registry_root(path_text):
    path_norm = _norm_token(path_text)
    path_upper = path_norm.upper()
    if path_upper.startswith('HKCU:\\'):
        return 'HKU\\' + path_norm[6:]
    if path_upper.startswith('HKLM:\\'):
        return 'HKLM\\' + path_norm[6:]
    if path_upper.startswith('HKEY_CURRENT_USER\\'):
        return 'HKU\\' + path_norm[len('HKEY_CURRENT_USER\\'):]
    if path_upper.startswith('HKEY_LOCAL_MACHINE\\'):
        return 'HKLM\\' + path_norm[len('HKEY_LOCAL_MACHINE\\'):]
    return path_norm


def _normalize_registry_path_aliases(path_text):
    """Normalize known path aliases/typos to improve cross-source key matching."""
    path_norm = _norm_token(path_text)
    if not path_norm:
        return ''

    # Common singular/plural drift seen in Netlogon policy checks.
    path_norm = re.sub(
        r'(?i)\\Services\\Netlogon\\Parameter(\\|$)',
        r'\\Services\\Netlogon\\Parameters\1',
        path_norm,
    )

    # Collapse duplicated Windows Firewall PublicProfile Logging paths.
    path_norm = re.sub(
        r'(?i)\\PublicProfile\\Loggin\\Software\\Policies\\Microsoft\\WindowsFirewall\\PublicProfile\\Logging(\\|$)',
        r'\\PublicProfile\\Logging\1',
        path_norm,
    )

    return path_norm


def _registry_key_from_powershell_target(ps_target):
    target = _norm_token(ps_target)
    if not target:
        return ''

    match = re.match(r'((?:HKCU|HKLM):\\.+)\s+([A-Za-z0-9_]+)$', target, flags=re.IGNORECASE)
    if not match:
        return ''

    reg_key_path = _normalize_registry_root(match.group(1))
    reg_key_path = _normalize_registry_path_aliases(reg_key_path)
    reg_key = _norm_upper_token(reg_key_path)
    reg_item = _norm_upper_token(match.group(2))
    if reg_key and reg_item:
        return f'REGISTRY:{reg_key}|{reg_item}'
    return ''


def _normalize_user_rights_expected_for_tenable(fields, expected_text):
    expected = _strip_outer_quotes_once(expected_text)
    if not expected:
        return ''

    # Repair over-escaped regex fragments that Tenable can reject,
    # then rebuild a clean expected pattern from the control description.
    if r'\(\?' in expected or r'\|\|' in expected or r'\ "' in expected or '\\ ' in expected:
        fallback = _fallback_user_rights_expected_from_description(fields)
        if fallback:
            return _allow_no_members_alternative(fallback)

    # Source variants may include quoted alternatives like
    # "NT SERVICE\\WdiServiceHost" || "WdiServiceHost".
    # Collapse those to a deterministic comma-delimited expectation.
    if '||' in expected and 'WDISERVICEHOST' in _norm_upper_token(expected):
        return _allow_no_members_alternative('^Administrators,(NT SERVICE\\\\)?WdiServiceHost$')

    return expected


def _classify_account_rule(fields):
    acct = _norm_upper_token(fields.get('account_type', ''))
    desc = _norm_upper_token(fields.get('description', ''))
    expected = _norm_upper_token(_best_effort_expected_text(fields))
    check_type = _norm_upper_token(fields.get('check_type', ''))

    if 'RENAME' in desc or check_type in {'CHECK_NOT_EQUAL', 'CHECK_NOT_REGEX'}:
        return acct, 'RENAME'

    if expected in {'ENABLED', 'DISABLED', 'TRUE', 'FALSE'}:
        return acct, 'STATUS'

    return acct, 'GENERIC'


def _classify_account_powershell_rule(fields):
    script = _norm_upper_token(_strip_outer_quotes_once(fields.get('powershell_script', fields.get('powershell_args', ''))))
    desc = _norm_upper_token(fields.get('description', ''))

    acct = ''
    if '-500$' in script:
        acct = 'ADMINISTRATOR_ACCOUNT'
    elif '-501$' in script:
        acct = 'GUEST_ACCOUNT'
    else:
        return '', ''

    if 'EXPANDPROPERTY NAME' in script or 'RENAME' in desc:
        return acct, 'RENAME'
    if '.DISABLED' in script or "'DISABLED'" in script or "'ENABLED'" in script:
        return acct, 'STATUS'
    return acct, 'GENERIC'


def derive_evaluated_item_key(fields):
    """Build a canonical key representing what control item is being evaluated."""
    ctype = _norm_upper_token(fields.get('type', fields.get('control_type', '')))

    if ctype == 'PASSWORD_POLICY':
        pp = _norm_upper_token(fields.get('password_policy', ''))
        return f'PASSWORD_POLICY:{pp}' if pp else 'PASSWORD_POLICY'

    if ctype == 'LOCKOUT_POLICY':
        lp = _norm_upper_token(fields.get('lockout_policy', ''))
        return f'LOCKOUT_POLICY:{lp}' if lp else 'LOCKOUT_POLICY'

    if ctype == 'AUDIT_POLICY_SUBCATEGORY':
        sub = _norm_upper_token(fields.get('audit_policy_subcategory', ''))
        return f'AUDIT_POLICY_SUBCATEGORY:{sub}' if sub else 'AUDIT_POLICY_SUBCATEGORY'

    if ctype in {'REGISTRY_SETTING', 'REG_CHECK', 'BANNER_CHECK'}:
        reg_key = _norm_upper_token(fields.get('reg_key', ''))
        reg_item = _norm_upper_token(fields.get('reg_item', fields.get('key_item', '')))
        if reg_key or reg_item:
            return f'REGISTRY:{reg_key}|{reg_item}'
        return 'REGISTRY'

    if ctype == 'CHECK_ACCOUNT':
        acct, semantic = _classify_account_rule(fields)
        if acct and semantic:
            return f'CHECK_ACCOUNT:{acct}:{semantic}'
        return f'CHECK_ACCOUNT:{acct}' if acct else 'CHECK_ACCOUNT'

    if ctype == 'USER_RIGHTS_POLICY':
        right_type = _norm_upper_token(fields.get('right_type', ''))
        return f'USER_RIGHTS_POLICY:{right_type}' if right_type else 'USER_RIGHTS_POLICY'

    if ctype == 'AUDIT_POWERSHELL':
        ps_raw = _norm_token(fields.get('powershell_script', fields.get('powershell_args', '')))
        ps_target = _norm_token(summarize_powershell_audit_target(ps_raw))
        ps_target_upper = ps_target.upper()
        ps_raw_upper = ps_raw.upper()

        acct, semantic = _classify_account_powershell_rule(fields)
        if acct and semantic:
            return f'CHECK_ACCOUNT:{acct}:{semantic}'

        reg_key = _registry_key_from_powershell_target(ps_target)
        if reg_key:
            return reg_key

        if ps_target_upper.startswith('AUDITPOL '):
            return f"AUDIT_POLICY_SUBCATEGORY:{ps_target_upper[len('AUDITPOL '):].strip()}"

        if ps_target_upper == 'WINRM':
            return 'REGISTRY:HKLM\\SOFTWARE\\POLICIES\\MICROSOFT\\WINDOWS\\WINRM\\SERVICE|ALLOWBASIC'

        if ps_target_upper == 'NET ACCOUNTS PASSWORD HISTORY':
            return 'PASSWORD_POLICY:ENFORCE_PASSWORD_HISTORY'
        if ps_target_upper == 'NET ACCOUNTS MAXIMUM PASSWORD AGE':
            return 'PASSWORD_POLICY:MAXIMUM_PASSWORD_AGE'
        if ps_target_upper == 'NET ACCOUNTS MAXIMUM PASSWORD':
            return 'PASSWORD_POLICY:MAXIMUM_PASSWORD_AGE'
        if ps_target_upper == 'NET ACCOUNTS MINIMUM PASSWORD AGE':
            return 'PASSWORD_POLICY:MINIMUM_PASSWORD_AGE'
        if ps_target_upper == 'NET ACCOUNTS PASSWORD LENGTH':
            return 'PASSWORD_POLICY:MINIMUM_PASSWORD_LENGTH'
        if ps_target_upper == 'NET ACCOUNTS LOCKOUT DURATION':
            return 'LOCKOUT_POLICY:LOCKOUT_DURATION'
        if ps_target_upper == 'NET ACCOUNTS LOCKOUT THRESHOLD':
            return 'LOCKOUT_POLICY:LOCKOUT_THRESHOLD'
        if ps_target_upper == 'NET ACCOUNTS THRESHOLD':
            return 'LOCKOUT_POLICY:LOCKOUT_THRESHOLD'
        if ps_target_upper == 'NET ACCOUNTS LOCKOUT OBSERVATION':
            return 'LOCKOUT_POLICY:LOCKOUT_RESET'
        if 'NET ACCOUNTS' in ps_target_upper and 'PASSWORD HISTORY' in ps_target_upper:
            return 'PASSWORD_POLICY:ENFORCE_PASSWORD_HISTORY'
        if 'NET ACCOUNTS' in ps_target_upper and 'MAXIMUM PASSWORD AGE' in ps_target_upper:
            return 'PASSWORD_POLICY:MAXIMUM_PASSWORD_AGE'
        if 'NET ACCOUNTS' in ps_target_upper and 'MAXIMUM PASSWORD' in ps_target_upper:
            return 'PASSWORD_POLICY:MAXIMUM_PASSWORD_AGE'
        if 'NET ACCOUNTS' in ps_target_upper and 'MINIMUM PASSWORD AGE' in ps_target_upper:
            return 'PASSWORD_POLICY:MINIMUM_PASSWORD_AGE'
        if 'NET ACCOUNTS' in ps_target_upper and 'PASSWORD LENGTH' in ps_target_upper:
            return 'PASSWORD_POLICY:MINIMUM_PASSWORD_LENGTH'
        if 'NET ACCOUNTS' in ps_target_upper and 'LOCKOUT DURATION' in ps_target_upper:
            return 'LOCKOUT_POLICY:LOCKOUT_DURATION'
        if 'NET ACCOUNTS' in ps_target_upper and 'LOCKOUT THRESHOLD' in ps_target_upper:
            return 'LOCKOUT_POLICY:LOCKOUT_THRESHOLD'
        if 'NET ACCOUNTS' in ps_target_upper and ' THRESHOLD' in ps_target_upper:
            return 'LOCKOUT_POLICY:LOCKOUT_THRESHOLD'
        if 'NET ACCOUNTS' in ps_target_upper and 'LOCKOUT OBSERVATION' in ps_target_upper:
            return 'LOCKOUT_POLICY:LOCKOUT_RESET'

        if ps_target_upper in {'$S=WINS', 'WINS'}:
            return 'AUDIT_POWERSHELL:WINS'
        if 'GET-WINDOWSFEATURE' in ps_target_upper and "NAME -EQ 'WINS'" in ps_target_upper:
            return 'AUDIT_POWERSHELL:WINS'

        if 'PASSWORD HISTORY' in ps_raw_upper:
            return 'PASSWORD_POLICY:ENFORCE_PASSWORD_HISTORY'
        if 'MAXIMUMPASSWORDAGE' in ps_raw_upper or 'MAXIMUM PASSWORD AGE' in ps_raw_upper:
            return 'PASSWORD_POLICY:MAXIMUM_PASSWORD_AGE'
        if 'MINIMUMPASSWORDAGE' in ps_raw_upper or 'MINIMUM PASSWORD AGE' in ps_raw_upper:
            return 'PASSWORD_POLICY:MINIMUM_PASSWORD_AGE'
        if 'MINIMUMPASSWORDLENGTH' in ps_raw_upper or 'MINIMUM PASSWORD LENGTH' in ps_raw_upper:
            return 'PASSWORD_POLICY:MINIMUM_PASSWORD_LENGTH'
        if 'PASSWORDCOMPLEXITY' in ps_raw_upper:
            return 'PASSWORD_POLICY:COMPLEXITY_REQUIREMENTS'
        if 'CLEARTEXTPASSWORD' in ps_raw_upper:
            return 'PASSWORD_POLICY:REVERSIBLE_ENCRYPTION'
        if 'ALLOWADMINISTRATORLOCKOUT' in ps_raw_upper:
            return 'PASSWORD_POLICY:LOCKOUT_ADMINS'
        if 'LOCKOUTDURATION' in ps_raw_upper:
            return 'LOCKOUT_POLICY:LOCKOUT_DURATION'
        if 'LOCKOUTBADCOUNT' in ps_raw_upper:
            return 'LOCKOUT_POLICY:LOCKOUT_THRESHOLD'
        if 'RESETLOCKOUTCOUNT' in ps_raw_upper:
            return 'LOCKOUT_POLICY:LOCKOUT_RESET'

        if ps_target:
            return f'AUDIT_POWERSHELL:{ps_target_upper}'
        return 'AUDIT_POWERSHELL'

    return f'{ctype}:{_norm_upper_token(fields.get("control_key", ""))}' if ctype else _norm_upper_token(fields.get('control_key', ''))


def derive_control_key_parts(fields):
    control_type = str(fields.get('type', '')).strip().upper()
    if control_type == 'AUDIT_POWERSHELL':
        parts = []
        ps_val = summarize_powershell_audit_target(fields.get('powershell_script', fields.get('powershell_args', '')))
        if ps_val:
            parts.append(('powershell_script', ps_val))
        return control_type, parts

    selected = CONTROL_KEY_FIELDS_BY_TYPE.get(control_type, CONTROL_KEY_FALLBACK_FIELDS)

    parts = []
    for name in selected:
        if name == 'powershell_args':
            val = summarize_powershell_audit_target(fields.get(name, ''))
        else:
            val = _strip_outer_quotes_once(fields.get(name, ''))
        if val:
            parts.append((name, val))

    if not parts:
        for name in CONTROL_KEY_FALLBACK_FIELDS:
            if name == 'powershell_args':
                val = summarize_powershell_audit_target(fields.get(name, ''))
            else:
                val = _strip_outer_quotes_once(fields.get(name, ''))
            if val:
                parts.append((name, val))
                break

    return control_type, parts


def derive_control_key(fields):
    control_type, parts = derive_control_key_parts(fields)
    if not control_type:
        control_type = 'UNKNOWN'
    if not parts:
        return control_type
    return control_type + '|' + '|'.join(f'{k}={v}' for k, v in parts)


def derive_control_keyword(fields):
    _, parts = derive_control_key_parts(fields)
    if not parts:
        return ''
    return ' | '.join(v for _, v in parts)


def enforce_solution_reference_adjacency(pairs):
    """Ensure solution and reference are adjacent in output ordering."""
    keys = [k for k, _ in pairs]
    if "solution" not in keys or "reference" not in keys:
        return pairs

    solution_idx = keys.index("solution")
    solution_val = pairs[solution_idx][1]
    reference_val = pairs[keys.index("reference")][1]

    remainder = [(k, v) for (k, v) in pairs if k not in ("solution", "reference")]
    insert_at = min(solution_idx, len(remainder))
    return (
        remainder[:insert_at]
        + [("solution", solution_val), ("reference", reference_val)]
        + remainder[insert_at:]
    )


CUSTOM_ITEM_FIELD_ORDER = [
    'type',
    'description',
    'info',
    'solution',
    'reference',
    'see_also',
    'value_type',
    'value_data',
    'powershell_option',
    'powershell_args',
    'expect',
    'check_type',
]


def order_custom_item_pairs(pairs):
    """Order custom_item fields in a consistent sequence for all emitted blocks."""
    by_key = OrderedDict()
    for k, v in pairs:
        if k not in by_key:
            by_key[k] = v

    ordered = []
    for key in CUSTOM_ITEM_FIELD_ORDER:
        if key in by_key:
            ordered.append((key, by_key.pop(key)))

    for key, val in by_key.items():
        ordered.append((key, val))

    return ordered


# -----------------------------------------------------------------------------
# Catalog / Excel export helpers (embedded from tools/catalog_controls.py)
# -----------------------------------------------------------------------------

FIELD_RE = re.compile(r"^\s*([A-Za-z0-9_]+)\s*:\s*(.*)$")
PLATFORM_RE = re.compile(r"\b([A-Z]{2,6})\b")


def parse_custom_items(text):
    items = []
    parts = re.split(r"<custom_item>|</custom_item>", text, flags=re.IGNORECASE)
    for i in range(1, len(parts), 2):
        body = parts[i]
        fields = {}
        for line in body.splitlines():
            m = FIELD_RE.match(line)
            if m:
                key = m.group(1).lower()
                val = m.group(2).strip()
                fields[key] = val
        items.append(fields)
    return items


def detect_platform(description):
    if not description:
        return 'UNKNOWN'
    d = description.strip().strip('"').strip("'")
    m = re.search(r"-\s*([A-Z]{2,6})\s*-", d)
    if m:
        return m.group(1)
    m2 = PLATFORM_RE.search(d)
    if m2:
        return m2.group(1)
    return 'UNKNOWN'


def determine_platform_from_filename(path):
    name = os.path.basename(path).lower()
    # New rules: 'vmware' -> VMware, 'enterprise' + linux/rhel -> RHEL
    if 'vmware' in name:
        return 'VMware'
    if 'mssrv' in name:
        return 'MSSRV'
    if 'mswrk' in name:
        return 'MSWRK'
    if 'enterprise' in name and ('linux' in name or 'rhel' in name):
        return 'RHEL'
    if 'enterprise' in name:
        return 'MSWRK'
    if 'sql' in name:
        return 'SQL'
    if 'server' in name and 'sql' not in name:
        return 'MSSRV'
    if 'ios' in name:
        return 'IOS'
    if 'palo alto' in name or ('palo' in name and 'alto' in name):
        return 'PAFW'
    if 'pafw' in name:
        return 'PAFW'
    if 'nx' in name:
        return 'NX-OS'
    if 'f5' in name:
        return 'F5'
    if 'azure' in name:
        return 'MSAZ'
    if 'asa' in name:
        return 'ASA'
    if 'amazon' in name or 'aws' in name:
        return 'Amazon'
    return 'UNKNOWN'


WINDOWS_PARSE_POLICY = {'allow_repair_fallback': False, 'repair_pipeline': ('quoted_fields',)}


RHEL_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('wrapper', 'quoted_fields', 'value_data')}
SQL_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('quoted_fields', 'value_data', 'powershell', 'wrapper')}
IOS_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('wrapper', 'quoted_fields')}
PAFW_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('quoted_fields', 'value_data', 'wrapper')}
NXOS_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('wrapper', 'quoted_fields', 'value_data')}
F5_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('quoted_fields', 'powershell', 'wrapper')}
MSAZ_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('quoted_fields', 'value_data', 'powershell')}
ASA_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('wrapper', 'quoted_fields', 'value_data', 'powershell')}
AMAZON_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('quoted_fields', 'value_data', 'wrapper', 'powershell')}
VMWARE_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('quoted_fields', 'powershell', 'value_data', 'wrapper')}
UNKNOWN_PARSE_POLICY = {'allow_repair_fallback': True, 'repair_pipeline': ('quoted_fields', 'value_data')}


# Operational risk prioritization model used for gap analysis staging.
# Priority Score = Family Weight * Attack Prevalence * Asset Exposure * Exploitability Modifier
SECURITY_FAMILY_WEIGHTS = {
    'PRIVILEGED_ACCESS': 1.40,
    'CREDENTIAL_ACCESS': 1.35,
    'AUTHENTICATION': 1.25,
    'NETWORK_BOUNDARY': 1.20,
    'LOGGING_DETECTION': 1.15,
    'SYSTEM_HARDENING': 1.00,
    'OTHER': 0.95,
}


DEFAULT_CONTROL_TAXONOMY = [
    {
        'id': 'CTRL-WIN-AUTH-LOCKOUT',
        'family': 'AUTHENTICATION',
        'match_any': ['LOCKOUT', 'LOCKOUT_THRESHOLD', 'LOCKOUT_DURATION', 'LOCKOUT_RESET'],
        'attack_prevalence': 1.35,
        'asset_exposure': 1.20,
        'exploitability_modifier': 1.10,
        'threat_weight': 1.05,
        'source': 'builtin',
    },
    {
        'id': 'CTRL-WIN-CRED-NTLM',
        'family': 'CREDENTIAL_ACCESS',
        'match_any': ['NTLM', 'MSV1_0', 'RESTRICTSENDINGNTLMTRAFFIC'],
        'attack_prevalence': 1.45,
        'asset_exposure': 1.25,
        'exploitability_modifier': 1.10,
        'threat_weight': 1.10,
        'source': 'builtin',
    },
    {
        'id': 'CTRL-WIN-PRIV-USER-RIGHTS',
        'family': 'PRIVILEGED_ACCESS',
        'match_any': ['USER_RIGHTS', 'SEDEBUGPRIVILEGE', 'SETAKEOWNERSHIPPRIVILEGE', 'ADMINISTRATOR'],
        'attack_prevalence': 1.35,
        'asset_exposure': 1.30,
        'exploitability_modifier': 1.05,
        'threat_weight': 1.10,
        'source': 'builtin',
    },
    {
        'id': 'CTRL-WIN-NET-REMOTE-MGMT',
        'family': 'NETWORK_BOUNDARY',
        'match_any': ['RDP', 'WINRM', 'SMB', 'REMOTE DESKTOP', 'NETWORK ACCESS'],
        'attack_prevalence': 1.40,
        'asset_exposure': 1.30,
        'exploitability_modifier': 1.10,
        'threat_weight': 1.10,
        'source': 'builtin',
    },
    {
        'id': 'CTRL-WIN-DET-AUDITPOL',
        'family': 'LOGGING_DETECTION',
        'match_any': ['AUDITPOL', 'AUDIT POLICY', 'EVENT LOG', 'POWERSHELL TRANSCRIPTION'],
        'attack_prevalence': 1.20,
        'asset_exposure': 1.10,
        'exploitability_modifier': 1.00,
        'threat_weight': 1.00,
        'source': 'builtin',
    },
    {
        'id': 'CTRL-WIN-CRED-PASSWORD-POLICY',
        'family': 'CREDENTIAL_ACCESS',
        'match_any': ['PASSWORD_POLICY', 'PASSWORD LENGTH', 'PASSWORD HISTORY', 'PASSWORDCOMPLEXITY', 'CLEARTEXTPASSWORD'],
        'attack_prevalence': 1.35,
        'asset_exposure': 1.15,
        'exploitability_modifier': 1.05,
        'threat_weight': 1.05,
        'source': 'builtin',
    },
]


PLATFORM_PARSE_PROFILES = {
    # Windows baselines: shared strict parser-only policy.
    'MSSRV': WINDOWS_PARSE_POLICY,
    'MSWRK': WINDOWS_PARSE_POLICY,

    # Platform-specific parser profiles (tunable independently).
    'RHEL': RHEL_PARSE_POLICY,
    'SQL': SQL_PARSE_POLICY,
    'IOS': IOS_PARSE_POLICY,
    'PAFW': PAFW_PARSE_POLICY,
    'NX-OS': NXOS_PARSE_POLICY,
    'F5': F5_PARSE_POLICY,
    'MSAZ': MSAZ_PARSE_POLICY,
    'ASA': ASA_PARSE_POLICY,
    'AMAZON': AMAZON_PARSE_POLICY,
    'VMWARE': VMWARE_PARSE_POLICY,
    'UNKNOWN': UNKNOWN_PARSE_POLICY,
}


def _normalized_platform_key(platform_hint):
    platform = _norm_upper_token(platform_hint)
    if platform == 'VMWARE':
        return 'VMWARE'
    return platform


def _platform_parse_profile(platform_hint):
    platform_key = _normalized_platform_key(platform_hint)
    default_profile = UNKNOWN_PARSE_POLICY
    return platform_key, PLATFORM_PARSE_PROFILES.get(platform_key, default_profile)


def _apply_parse_repair_pipeline(text, repair_pipeline):
    current = str(text or '')
    for step in repair_pipeline:
        if step == 'value_data':
            current, _ = _repair_malformed_value_data_lines(current)
        elif step == 'powershell':
            current, _ = _repair_malformed_powershell_lines(current)
        elif step == 'quoted_fields':
            current, _ = _repair_malformed_quoted_field_lines(current)
        elif step == 'wrapper':
            current = _normalize_tenable_audit_wrapper(current)
    return current


def extract_sequence_rows(text):
    """Return rows in source order, interleaving structural tags and custom items.

    Row types:
      - {'kind': 'tag', 'custom_tag': '<custom_item>'|'</custom_item>'}
      - {'kind': 'tag', 'conditional_tag': '<if>'|'</if>'|'<condition ...>'|'</condition>'|'<report ...>'|'</report>'}
      - {'kind': 'item', 'fields': {...}, 'condition_type': str, 'report_fields': {...}}
    """
    rows = []
    lines = text.splitlines()
    i = 0

    if_stack = []
    condition_stack = []

    while i < len(lines):
        raw = lines[i]
        s = raw.strip()

        if re.match(r'^<if>\s*$', s, flags=re.IGNORECASE):
            rows.append({'kind': 'tag', 'conditional_tag': '<if>'})
            if_stack.append({'report_fields': {}})
            i += 1
            continue

        if re.match(r'^</if>\s*$', s, flags=re.IGNORECASE):
            rows.append({'kind': 'tag', 'conditional_tag': '</if>'})
            if if_stack:
                if_stack.pop()
            i += 1
            continue

        if re.match(r'^<condition\b[^>]*>\s*$', s, flags=re.IGNORECASE):
            rows.append({'kind': 'tag', 'conditional_tag': s})
            condition_stack.append(s)
            i += 1
            continue

        if re.match(r'^</condition>\s*$', s, flags=re.IGNORECASE):
            rows.append({'kind': 'tag', 'conditional_tag': '</condition>'})
            if condition_stack:
                condition_stack.pop()
            i += 1
            continue

        if re.match(r'^<report\b[^>]*>\s*$', s, flags=re.IGNORECASE):
            rows.append({'kind': 'tag', 'conditional_tag': s})
            report_fields = {}
            i += 1
            while i < len(lines):
                inner = lines[i].strip()
                if re.match(r'^</report>\s*$', inner, flags=re.IGNORECASE):
                    rows.append({'kind': 'tag', 'conditional_tag': '</report>'})
                    break
                fm = FIELD_RE.match(lines[i])
                if fm:
                    report_fields[fm.group(1).lower()] = fm.group(2).strip()
                i += 1
            if if_stack:
                if_stack[-1]['report_fields'] = report_fields
            i += 1
            continue

        if re.match(r'^<custom_item>\s*$', s, flags=re.IGNORECASE):
            rows.append({'kind': 'tag', 'custom_tag': '<custom_item>'})
            fields = {}
            i += 1
            while i < len(lines):
                inner = lines[i].strip()
                if re.match(r'^</custom_item>\s*$', inner, flags=re.IGNORECASE):
                    break
                fm = FIELD_RE.match(lines[i])
                if fm:
                    fields[fm.group(1).lower()] = fm.group(2).strip()
                i += 1

            cond_type = ''
            if condition_stack:
                m = re.search(r'type\s*:\s*"([^\"]+)"', condition_stack[-1], flags=re.IGNORECASE)
                if m:
                    cond_type = m.group(1)

            report_fields = if_stack[-1]['report_fields'] if if_stack else {}
            rows.append({
                'kind': 'item',
                'fields': fields,
                'condition_type': cond_type,
                'report_fields': report_fields,
            })
            rows.append({'kind': 'tag', 'custom_tag': '</custom_item>'})

            if i < len(lines) and re.match(r'^</custom_item>\s*$', lines[i].strip(), flags=re.IGNORECASE):
                i += 1
            continue

        # Some benchmarks (for example Cisco ASA) use <item> instead of
        # <custom_item>. Treat these as control rows for catalog generation.
        if re.match(r'^<item>\s*$', s, flags=re.IGNORECASE):
            rows.append({'kind': 'tag', 'conditional_tag': '<item>'})
            fields = {}
            i += 1
            while i < len(lines):
                inner = lines[i].strip()
                if re.match(r'^</item>\s*$', inner, flags=re.IGNORECASE):
                    break
                fm = FIELD_RE.match(lines[i])
                if fm:
                    fields[fm.group(1).lower()] = fm.group(2).strip()
                i += 1

            cond_type = ''
            if condition_stack:
                m = re.search(r'type\s*:\s*"([^\"]+)"', condition_stack[-1], flags=re.IGNORECASE)
                if m:
                    cond_type = m.group(1)

            report_fields = if_stack[-1]['report_fields'] if if_stack else {}
            rows.append({
                'kind': 'item',
                'fields': fields,
                'condition_type': cond_type,
                'report_fields': report_fields,
            })
            rows.append({'kind': 'tag', 'conditional_tag': '</item>'})

            if i < len(lines) and re.match(r'^</item>\s*$', lines[i].strip(), flags=re.IGNORECASE):
                i += 1
            continue

        i += 1

    return rows


def _sanitize_for_excel(v):
    if v is None:
        return ''
    if not isinstance(v, str):
        v = str(v)
    return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', v)


def _save_workbook_with_lock_fallback(wb, outpath):
    temp_outpath = f"{outpath}.{os.getpid()}.tmp"
    wb.save(temp_outpath)

    # Windows can hold transient locks (Explorer previews, AV, sync clients), so retry briefly.
    for _ in range(8):
        try:
            os.replace(temp_outpath, outpath)
            return outpath
        except PermissionError:
            time.sleep(0.25)

    base, ext = os.path.splitext(outpath)
    fallback_outpath = _timestamped_output_path(f"{base}.new{ext or '.xlsx'}")
    wb.save(fallback_outpath)

    try:
        os.remove(temp_outpath)
    except Exception:
        pass

    print(f"Could not replace locked workbook '{outpath}'. Written to '{fallback_outpath}' instead.")
    return fallback_outpath


def _write_catalog_workbook(rows_by_platform, ordered_real_headers, outpath):
    if Workbook is None:
        raise RuntimeError('openpyxl is required to write Excel workbook; pip install openpyxl')

    wb = Workbook()
    wb.remove(wb.active)
    for plat, rows in sorted(rows_by_platform.items()):
        safe_name = plat[:31]
        ws = wb.create_sheet(title=safe_name)
        all_keys = set()
        for r in rows:
            all_keys.update(r.keys())
        meta_keys = [
            'control_type', 'control_key', 'control_keyword', 'expected_value',
            'description', 'info', 'reference', 'condition_type',
            'report_type', 'report_description', 'raw_fields', 'source_count',
            'source_files', 'source_file', 'Custom item', 'Conditional'
        ]
        headers = [k for k in meta_keys if k in all_keys]
        headers += [k for k in ordered_real_headers if k not in headers]
        headers += sorted(k for k in all_keys if k not in headers)
        ws.append(headers)
        for r in rows:
            ws.append([_sanitize_for_excel(r.get(h, '')) for h in headers])

    return _save_workbook_with_lock_fallback(wb, outpath)


def _sanitize_sheet_title(title, existing_titles):
    cleaned = re.sub(r'[\\/\?\*\[\]:]', '_', str(title)).strip()
    if not cleaned:
        cleaned = 'Sheet'
    cleaned = cleaned[:31]

    base = cleaned
    suffix = 1
    while cleaned in existing_titles:
        extra = f'_{suffix}'
        cleaned = f'{base[:31 - len(extra)]}{extra}'
        suffix += 1

    existing_titles.add(cleaned)
    return cleaned


def _timestamped_output_path(path):
    base, ext = os.path.splitext(path)
    if _TS_SUFFIX_RE.search(base):
        return path
    return f'{base}_{RUN_TIMESTAMP}{ext}'


def _resolve_existing_or_latest_timestamped_path(base_path):
    preferred = _timestamped_output_path(base_path)
    if os.path.isfile(preferred):
        return preferred

    base_dir = os.path.dirname(base_path) or '.'
    stem = os.path.splitext(os.path.basename(base_path))[0]
    ext = os.path.splitext(base_path)[1]
    pattern = re.compile(rf'^{re.escape(stem)}_(\d{{8}}){re.escape(ext)}$')

    candidates = []
    try:
        for name in os.listdir(base_dir):
            if not pattern.match(name):
                continue
            full = os.path.join(base_dir, name)
            if os.path.isfile(full):
                candidates.append(full)
    except Exception:
        return preferred

    if not candidates:
        return preferred
    return max(candidates, key=os.path.getmtime)


def _parsing_results_output_path(folder_path):
    return _timestamped_output_path(os.path.join(folder_path, PARSING_RESULTS_FILENAME))


def _write_parsing_results_workbook(records, outpath):
    if Workbook is None:
        raise RuntimeError('openpyxl is required to write Excel workbook; pip install openpyxl')

    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet(title='Parsing Log')
    summary_headers = [
        'audit_file',
        'source_folder',
        'status',
        'node_count',
        'custom_item_count',
        'report_count',
        'text_count',
        'unknown_key_count',
        'unknown_keys',
        'normalized_output',
        'notes',
    ]
    summary_ws.append(summary_headers)

    existing_titles = {'Parsing Log'}
    for record in records:
        summary_ws.append([record.get(h, '') for h in summary_headers])

        sheet_title = _sanitize_sheet_title(record.get('sheet_name') or record.get('audit_file') or 'Audit', existing_titles)
        ws = wb.create_sheet(title=sheet_title)
        ws.append(['row_type', 'label', 'value'])
        ws.append(['audit_file', 'value', record.get('audit_file', '')])
        ws.append(['source_folder', 'value', record.get('source_folder', '')])
        ws.append(['status', 'value', record.get('status', '')])
        ws.append(['node_count', 'value', record.get('node_count', 0)])
        ws.append(['custom_item_count', 'value', record.get('custom_item_count', 0)])
        ws.append(['report_count', 'value', record.get('report_count', 0)])
        ws.append(['text_count', 'value', record.get('text_count', 0)])
        ws.append(['unknown_key_count', 'value', record.get('unknown_key_count', 0)])
        ws.append(['unknown_keys', 'value', record.get('unknown_keys', '')])
        ws.append(['normalized_output', 'value', record.get('normalized_output', '')])
        ws.append(['notes', 'value', record.get('notes', '')])

        ws.append([])
        ws.append(['node_index', 'node_type', 'details'])
        for node in record.get('nodes', []):
            ws.append([
                node.get('node_index', ''),
                node.get('node_type', ''),
                json.dumps(node, ensure_ascii=False),
            ])

    final_outpath = _save_workbook_with_lock_fallback(wb, outpath)
    print(f'Wrote parsing results workbook: {final_outpath}')
    return final_outpath


def _write_parsing_results_for_folder(folder_path):
    folder_path = os.path.abspath(folder_path)
    records = PARSING_RESULTS_BY_FOLDER.get(folder_path, [])
    if not records:
        return None
    return _write_parsing_results_workbook(records, _parsing_results_output_path(folder_path))


def generate_catalog(input_folder, output_file=None, output_dir=None):
    records_by_platform = {}
    unique_records_by_platform = {}
    real_key_counts = {k: 0 for k in REAL_KEYS}
    real_key_first_seen = {k: 10**9 for k in REAL_KEYS}
    seen_seq = 0

    files = []
    for root, dirnames, filenames in os.walk(input_folder):
        # Only consider source .audit controls; skip generated normalized outputs.
        dirnames[:] = [d for d in dirnames if d.lower() != 'normalized']
        for f in filenames:
            if f.lower().endswith('.audit'):
                files.append(os.path.join(root, f))

    for path in sorted(files):
        with open(path, encoding='utf-8') as fh:
            txt = _strip_bom_prefix(fh.read())

        variables = extract_variables(txt.splitlines())
        entries = extract_sequence_rows(txt)
        platform = determine_platform_from_filename(path)

        for entry in entries:
            if entry['kind'] == 'item':
                it = {
                    key: resolve_variables(value, variables) if isinstance(value, str) else value
                    for key, value in entry['fields'].items()
                }
                cond_type = resolve_variables(entry.get('condition_type', ''), variables)
                report_fields = {
                    key: resolve_variables(value, variables) if isinstance(value, str) else value
                    for key, value in entry.get('report_fields', {}).items()
                }
            else:
                it = {}
                cond_type = ''
                report_fields = {}

            desc = it.get('description', '')
            plat = platform or detect_platform(desc)

            for k in it.keys():
                if k in REAL_KEYS:
                    seen_seq += 1
                    if real_key_first_seen[k] == 10**9:
                        real_key_first_seen[k] = seen_seq
                    if str(it.get(k, '')).strip():
                        real_key_counts[k] += 1

            real_key_values = {k: it.get(k, '') for k in REAL_KEYS}
            rec = {
                **real_key_values,
                'source_file': os.path.relpath(path),
                'control_type': it.get('type', ''),
                'control_key': derive_control_key(it) if entry['kind'] == 'item' else '',
                'control_keyword': derive_control_keyword(it) if entry['kind'] == 'item' else '',
                'expected_value': derive_expected_value(it) if entry['kind'] == 'item' else '',
                'description': desc,
                'info': it.get('info', ''),
                'reference': it.get('reference', ''),
                'raw_fields': json.dumps(it, ensure_ascii=False),
                'condition_type': cond_type,
                'report_type': report_fields.get('type', ''),
                'report_description': report_fields.get('description', ''),
                'Custom item': entry.get('custom_tag', ''),
                'Conditional': entry.get('conditional_tag', ''),
            }
            records_by_platform.setdefault(plat, []).append(rec)

            if entry['kind'] == 'item':
                unique_key = str(rec.get('control_key', '')).strip()
                if not unique_key:
                    unique_key = json.dumps(
                        {
                            **real_key_values,
                            'condition_type': cond_type,
                            'report_type': report_fields.get('type', ''),
                            'report_description': report_fields.get('description', ''),
                        },
                        sort_keys=True,
                        ensure_ascii=False,
                    )
                unique_bucket = unique_records_by_platform.setdefault(plat, OrderedDict())
                if unique_key not in unique_bucket:
                    unique_bucket[unique_key] = dict(rec)
                else:
                    existing = unique_bucket[unique_key]
                    src = rec.get('source_file', '')
                    existing_files = existing.get('source_files', '')
                    existing_set = {s for s in existing_files.split(';') if s}
                    if src and src not in existing_set:
                        existing_set.add(src)
                        ordered_sources = sorted(existing_set)
                        existing['source_files'] = ';'.join(ordered_sources)
                        existing['source_count'] = len(ordered_sources)

    rows_by_platform = {}
    for plat, rows in records_by_platform.items():
        out_rows = []
        for r in rows:
            rr = dict(r)
            rr['source_count'] = 1 if rr.get('source_file') else 0
            rr['source_files'] = rr.get('source_file', '')
            out_rows.append(rr)
        rows_by_platform[plat] = out_rows

    unique_rows_by_platform = {}
    for plat, rows_map in unique_records_by_platform.items():
        out_rows = []
        for rec in rows_map.values():
            rr = dict(rec)
            src_files = rr.get('source_files') or rr.get('source_file', '')
            rr['source_files'] = src_files
            rr['source_count'] = len([s for s in src_files.split(';') if s]) if src_files else 0
            out_rows.append(rr)
        unique_rows_by_platform[plat] = out_rows

    ordered_real_headers = sorted(
        REAL_KEYS,
        key=lambda k: (
            -real_key_counts.get(k, 0),
            real_key_first_seen.get(k, 10**9),
            k,
        ),
    )

    if output_file:
        all_catalog_outpath = _timestamped_output_path(output_file)
        outdir = output_dir or os.path.dirname(all_catalog_outpath) or input_folder
    else:
        normalized_dir = output_dir or os.path.join(input_folder, 'Normalized')
        os.makedirs(normalized_dir, exist_ok=True)
        outdir = normalized_dir
        all_catalog_outpath = _timestamped_output_path(os.path.join(outdir, 'All_Controls_Catalog.xlsx'))

    unique_catalog_outpath = _timestamped_output_path(os.path.join(outdir, 'Unique_Controls_Catalog.xlsx'))

    final_outpath = _write_catalog_workbook(rows_by_platform, ordered_real_headers, all_catalog_outpath)
    print(f'Wrote workbook: {final_outpath}')

    final_unique_outpath = _write_catalog_workbook(
        unique_rows_by_platform,
        ordered_real_headers,
        unique_catalog_outpath,
    )
    print(f'Wrote workbook: {final_unique_outpath}')

    return final_outpath


def find_duplicates_in_workbook(wb_path):
    if openpyxl is None:
        raise RuntimeError('openpyxl is required; pip install openpyxl')
    wb = openpyxl.load_workbook(wb_path, read_only=True)
    out = {}

    def _norm(s):
        if s is None:
            return ''
        s = str(s).strip().strip('"').strip("'")
        s = re.sub(r'\s+', ' ', s)
        return s.lower()

    for name in wb.sheetnames:
        ws = wb[name]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            out[name] = []
            continue
        hidx = { (h.lower() if h else ''): i for i, h in enumerate(headers) }
        desc_i = hidx.get('description')
        type_i = hidx.get('control_type')
        src_i = hidx.get('source_file')
        counts = {}
        for r in rows:
            desc = _norm(r[desc_i]) if desc_i is not None and desc_i < len(r) else ''
            ctype = _norm(r[type_i]) if type_i is not None and type_i < len(r) else ''
            src = r[src_i] if src_i is not None and src_i < len(r) else ''
            key = (ctype, desc)
            counts.setdefault(key, []).append(src)
        dups = [
            {'control_type': k[0], 'description': k[1], 'count': len(v), 'examples': v[:5]}
            for k, v in counts.items() if len(v) > 1
        ]
        out[name] = sorted(dups, key=lambda x: -x['count'])
    return out


def export_duplicates_csvs(wb_path, out_dir=None):
    if openpyxl is None:
        raise RuntimeError('openpyxl is required; pip install openpyxl')
    if out_dir is None:
        out_dir = os.path.dirname(wb_path)
    wb = openpyxl.load_workbook(wb_path, read_only=True)

    def _norm(s):
        if s is None:
            return ''
        s = str(s).strip().strip('"').strip("'")
        s = re.sub(r'\s+', ' ', s)
        return s

    written = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [h.lower() if h else '' for h in next(rows)]
        except StopIteration:
            continue
        hidx = {h: i for i, h in enumerate(headers)}
        desc_i = hidx.get('description')
        type_i = hidx.get('control_type')
        src_i = hidx.get('source_file')
        info_i = hidx.get('info')
        ref_i = hidx.get('reference')
        see_i = hidx.get('see_also')
        show_i = hidx.get('show_output')
        cond_i = hidx.get('condition_type') or hidx.get('condition')
        rpt_type_i = hidx.get('report_type')
        rpt_desc_i = hidx.get('report_description')
        raw_i = hidx.get('raw_fields')

        counts = {}
        rows_list = list(rows)
        for r in rows_list:
            desc = _norm(r[desc_i]) if desc_i is not None and desc_i < len(r) else ''
            ctype = _norm(r[type_i]) if type_i is not None and type_i < len(r) else ''
            key = (ctype.lower(), desc.lower())
            counts.setdefault(key, []).append(r)

        duplicates = {k: v for k, v in counts.items() if len(v) > 1}
        if not duplicates:
            continue

        out_csv = _timestamped_output_path(os.path.join(out_dir, f'duplicates_{name}.csv'))
        with open(out_csv, 'w', newline='', encoding='utf-8') as fh:
            writer = csv.writer(fh)
            hdr = ['sheet','control_type','description','info','reference','see_also','show_output','condition_type','report_type','report_description','raw_fields','count','examples']
            writer.writerow(hdr)
            for (ctype, desc), rows_list in sorted(duplicates.items(), key=lambda x: -len(x[1])):
                first = rows_list[0]
                info = first[info_i] if info_i is not None and info_i < len(first) else ''
                ref = first[ref_i] if ref_i is not None and ref_i < len(first) else ''
                see = first[see_i] if see_i is not None and see_i < len(first) else ''
                show = first[show_i] if show_i is not None and show_i < len(first) else ''
                cond = first[cond_i] if cond_i is not None and cond_i < len(first) else ''
                rpt_type = first[rpt_type_i] if rpt_type_i is not None and rpt_type_i < len(first) else ''
                rpt_desc = first[rpt_desc_i] if rpt_desc_i is not None and rpt_desc_i < len(first) else ''
                raw = first[raw_i] if raw_i is not None and raw_i < len(first) else ''
                examples = []
                for r in rows_list[:10]:
                    src = r[src_i] if src_i is not None and src_i < len(r) else ''
                    examples.append(src)
                writer.writerow([name, ctype, desc, info, ref, see, show, cond, rpt_type, rpt_desc, raw, len(rows_list), ';'.join(examples)])

        written.append(out_csv)
    return written


def _iter_workbook_rows(wb_path):
    if openpyxl is None:
        raise RuntimeError('openpyxl is required; pip install openpyxl')

    wb = openpyxl.load_workbook(wb_path, read_only=True)
    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        try:
            headers = next(it)
        except StopIteration:
            continue

        hdrs = [str(h).strip() if h is not None else '' for h in headers]
        for r in it:
            rec = {'sheet': sheet}
            for i, h in enumerate(hdrs):
                if not h:
                    continue
                rec[h] = r[i] if i < len(r) else ''
            rows.append(rec)
    return rows


def match_unique_catalogs(path_a, path_b, output_xlsx=None):
    """Match controls across two unique catalogs by evaluated item key.

    Writes an Excel workbook with two sheets:
      - Matched
      - Unmatched
    """
    rows_a = _iter_workbook_rows(path_a)
    rows_b = _iter_workbook_rows(path_b)

    buckets_a = {}
    buckets_b = {}

    for rec in rows_a:
        key = derive_evaluated_item_key(rec)
        buckets_a.setdefault(key, []).append(rec)
    for rec in rows_b:
        key = derive_evaluated_item_key(rec)
        buckets_b.setdefault(key, []).append(rec)

    # Match only controls present in the left (actual_audit_inputs) catalog.
    all_keys = sorted(set(buckets_a))
    matched = []

    for key in all_keys:
        left = buckets_a.get(key, [])
        right = buckets_b.get(key, [])
        status = 'matched' if right else 'left_only'

        left_sheets = sorted({str(x.get('sheet', '')).strip().upper() for x in left if str(x.get('sheet', '')).strip()})
        platform = ''
        for candidate in ('MSWRK', 'MSSRV'):
            if candidate in left_sheets:
                platform = candidate
                break

        left_keys = '; '.join(sorted({str(x.get('control_key', '')).strip() for x in left if str(x.get('control_key', '')).strip()}))
        right_keys = '; '.join(sorted({str(x.get('control_key', '')).strip() for x in right if str(x.get('control_key', '')).strip()}))
        left_desc = '; '.join(sorted({str(x.get('description', '')).strip() for x in left if str(x.get('description', '')).strip()}))
        right_desc = '; '.join(sorted({str(x.get('description', '')).strip() for x in right if str(x.get('description', '')).strip()}))

        matched.append({
            'match_status': status,
            'platform': platform,
            'evaluated_item_key': key,
            'left_count': len(left),
            'right_count': len(right),
            'left_control_keys': left_keys,
            'right_control_keys': right_keys,
            'left_descriptions': left_desc,
            'right_descriptions': right_desc,
        })

    if output_xlsx is None:
        output_xlsx = os.path.join(os.path.dirname(path_a), 'Matched_Controls_Crosswalk.xlsx')
    output_xlsx = _timestamped_output_path(output_xlsx)

    fieldnames = [
        'match_status',
        'platform',
        'evaluated_item_key',
        'left_count',
        'right_count',
        'left_control_keys',
        'right_control_keys',
        'left_descriptions',
        'right_descriptions',
    ]

    if Workbook is None:
        raise RuntimeError('openpyxl is required to write Excel workbook; pip install openpyxl')

    wb = Workbook()
    wb.remove(wb.active)

    sheet_map = {
        ('MSWRK', 'matched'): wb.create_sheet(title='MSWRK_Matched'),
        ('MSWRK', 'left_only'): wb.create_sheet(title='MSWRK_Unmatched'),
        ('MSSRV', 'matched'): wb.create_sheet(title='MSSRV_Matched'),
        ('MSSRV', 'left_only'): wb.create_sheet(title='MSSRV_Unmatched'),
    }
    for ws in sheet_map.values():
        ws.append(fieldnames)

    for row in matched:
        platform = row.get('platform', '')
        if platform not in {'MSWRK', 'MSSRV'}:
            continue
        values = [_sanitize_for_excel(row.get(k, '')) for k in fieldnames]
        ws = sheet_map.get((platform, row.get('match_status')))
        if ws is not None:
            ws.append(values)

    final_out = _save_workbook_with_lock_fallback(wb, _timestamped_output_path(output_xlsx))

    total = len(matched)
    matched_n = sum(1 for m in matched if m['match_status'] == 'matched')
    left_only_n = sum(1 for m in matched if m['match_status'] == 'left_only')
    right_only_n = 0

    print(f'Matched crosswalk written: {final_out}')
    print(f'Total keys: {total} | matched: {matched_n} | left_only: {left_only_n} | right_only: {right_only_n}')
    return final_out


def _normalize_platform_sheet_name(sheet_name):
    name = _norm_upper_token(sheet_name)
    if name in {'MSWRK', 'MSSRV'}:
        return name
    return ''


def _extract_first_description_line(description_text):
    if description_text is None:
        return ''
    lines = re.split(r'\r?\n', str(description_text))
    if not lines:
        return ''
    return lines[0].strip().strip('"').strip()


def _normalize_scan_description_key(description_text):
    line = _extract_first_description_line(description_text)
    line = re.sub(r'^\d+\.\d+\s*-\s*', '', line)
    line = re.sub(r'^MSSRV\s*-\s*', '', line, flags=re.IGNORECASE)
    line = re.sub(r'\s*:\s*\[(PASSED|FAILED)\]\s*$', '', line, flags=re.IGNORECASE)
    line = re.sub(r'\s+', ' ', line).strip()
    return line.upper()


def _extract_scan_state(description_text):
    match = re.search(r'\[(PASSED|FAILED)\]', str(description_text or ''), flags=re.IGNORECASE)
    return (match.group(1).upper() if match else '')


def _load_scan_description_index(csv_path, plugin_id='21156'):
    by_key = OrderedDict()
    duplicate_rows = 0

    if not os.path.isfile(csv_path):
        return by_key, duplicate_rows

    with open(csv_path, 'r', encoding='utf-8-sig', errors='ignore', newline='') as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            pid = str(row.get('Plugin ID', '')).strip()
            if plugin_id and pid != str(plugin_id):
                continue

            description = row.get('Description', '')
            key = _normalize_scan_description_key(description)
            if not key:
                continue

            item = {
                'key': key,
                'description_first_line': _extract_first_description_line(description),
                'state': _extract_scan_state(description),
            }

            if key in by_key:
                duplicate_rows += 1
                continue
            by_key[key] = item

    return by_key, duplicate_rows


def _filter_mssrv_rows_to_baseline_descriptions(rows, baseline_keys):
    if not baseline_keys:
        return rows, set()

    selected = OrderedDict()
    for row in rows:
        key = _normalize_scan_description_key(row.get('fields', {}).get('description', ''))
        if key not in baseline_keys:
            continue

        existing = selected.get(key)
        if existing is None:
            selected[key] = row
            continue

        # Prefer active rows when duplicates resolve to the same baseline key.
        if (not existing.get('active')) and row.get('active'):
            selected[key] = row

    missing = set(baseline_keys) - set(selected.keys())
    ordered_rows = []
    for key in baseline_keys:
        row = selected.get(key)
        if row is not None:
            ordered_rows.append(row)

    return ordered_rows, missing


def _is_known_non_control_baseline_key(key):
    text = _norm_upper_token(key)
    if not text:
        return False

    # Known scan/meta descriptions that should not be treated as hardening controls.
    known_prefixes = (
        'PASSED - TARGET OS MATCHES BASELINE',
        'THIS PLUGIN DISPLAYS, FOR EACH TESTED HOST, INFORMATION ABOUT THE SCAN ITSELF',
        'USING THE WMI INTERFACE, NESSUS WAS ABLE TO RUN',
    )
    if text.startswith(known_prefixes):
        return True

    if 'SCAN ITSELF' in text:
        return True
    if 'NETSTAT' in text and 'ENUMERATE THE OPEN PORTS' in text:
        return True

    return False


def _partition_missing_baseline_keys(missing_keys):
    known_meta = []
    actionable = []
    for key in sorted(missing_keys):
        if _is_known_non_control_baseline_key(key):
            known_meta.append(key)
        else:
            actionable.append(key)
    return known_meta, actionable


def write_description_match_workbook(baseline_csv, merged_csv, output_xlsx):
    if Workbook is None:
        raise RuntimeError('openpyxl is required to write Excel workbook; pip install openpyxl')

    baseline_idx, baseline_dupes = _load_scan_description_index(baseline_csv, plugin_id=BASELINE_PLUGIN_ID_FILTER)
    merged_idx, merged_dupes = _load_scan_description_index(merged_csv, plugin_id=BASELINE_PLUGIN_ID_FILTER)

    baseline_keys = set(baseline_idx.keys())
    merged_keys = set(merged_idx.keys())

    matching_keys = sorted(baseline_keys & merged_keys)
    baseline_only_keys = sorted(baseline_keys - merged_keys)
    merged_only_keys = sorted(merged_keys - baseline_keys)

    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet(title='Summary')
    summary_ws.append(['metric', 'value'])
    summary_ws.append(['baseline_csv', baseline_csv])
    summary_ws.append(['merged_csv', merged_csv])
    summary_ws.append(['baseline_unique_descriptions', len(baseline_keys)])
    summary_ws.append(['baseline_duplicate_rows_skipped', baseline_dupes])
    summary_ws.append(['merged_unique_descriptions', len(merged_keys)])
    summary_ws.append(['merged_duplicate_rows_skipped', merged_dupes])
    summary_ws.append(['matching_descriptions', len(matching_keys)])
    summary_ws.append(['baseline_only_descriptions', len(baseline_only_keys)])
    summary_ws.append(['merged_only_descriptions', len(merged_only_keys)])

    matching_ws = wb.create_sheet(title='Matching_Descriptions')
    matching_ws.append(['description_key', 'baseline_first_line', 'baseline_state', 'merged_first_line', 'merged_state'])
    for key in matching_keys:
        b = baseline_idx[key]
        m = merged_idx[key]
        matching_ws.append([
            _sanitize_for_excel(key),
            _sanitize_for_excel(b.get('description_first_line', '')),
            _sanitize_for_excel(b.get('state', '')),
            _sanitize_for_excel(m.get('description_first_line', '')),
            _sanitize_for_excel(m.get('state', '')),
        ])

    baseline_only_ws = wb.create_sheet(title='Baseline_Only')
    baseline_only_ws.append(['description_key', 'baseline_first_line', 'baseline_state'])
    for key in baseline_only_keys:
        b = baseline_idx[key]
        baseline_only_ws.append([
            _sanitize_for_excel(key),
            _sanitize_for_excel(b.get('description_first_line', '')),
            _sanitize_for_excel(b.get('state', '')),
        ])

    merged_only_ws = wb.create_sheet(title='Merged_Only')
    merged_only_ws.append(['description_key', 'merged_first_line', 'merged_state'])
    for key in merged_only_keys:
        m = merged_idx[key]
        merged_only_ws.append([
            _sanitize_for_excel(key),
            _sanitize_for_excel(m.get('description_first_line', '')),
            _sanitize_for_excel(m.get('state', '')),
        ])

    final_out = _save_workbook_with_lock_fallback(wb, _timestamped_output_path(output_xlsx))
    print(f'Wrote baseline/merged description match workbook: {final_out}')
    print(
        f'Description match summary | baseline_unique={len(baseline_keys)} | '
        f'merged_unique={len(merged_keys)} | matching={len(matching_keys)} | '
        f'baseline_only={len(baseline_only_keys)} | merged_only={len(merged_only_keys)}'
    )
    return final_out


def _parse_raw_fields_cell(raw_cell):
    if raw_cell is None:
        return {}
    text = str(raw_cell).strip()
    if not text:
        return {}
    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            return {str(k).lower(): v for k, v in parsed.items()}
    except Exception:
        return {}
    return {}


def _load_unique_catalog_index(wb_path):
    rows = _iter_workbook_rows(wb_path)
    by_platform = {'MSWRK': OrderedDict(), 'MSSRV': OrderedDict()}

    def _role_scope_suffix(fields):
        desc = _norm_upper_token(_strip_outer_quotes_once(fields.get('description', '')))
        if 'MSSRV.DC' in desc or '(DC ONLY)' in desc or ' DC ONLY' in desc:
            return 'DC_ONLY'
        if '(MS ONLY)' in desc or ' MS ONLY' in desc:
            return 'MS_ONLY'
        return ''

    for rec in rows:
        platform = _normalize_platform_sheet_name(rec.get('sheet', ''))
        if platform not in by_platform:
            continue

        fields = _parse_raw_fields_cell(rec.get('raw_fields', ''))
        if not fields:
            continue

        key = derive_evaluated_item_key(fields)
        if not key:
            key = str(rec.get('control_key', '')).strip()
        if not key:
            continue

        # Keep DC-only and MS-only variants distinct so role-specific controls
        # do not overwrite each other during unique-catalog indexing.
        scope_suffix = _role_scope_suffix(fields)
        if scope_suffix:
            key = f'{key}|SCOPE:{scope_suffix}'

        if key not in by_platform[platform]:
            by_platform[platform][key] = {
                'fields': fields,
                'condition_type': rec.get('condition_type', ''),
                'report_type': rec.get('report_type', ''),
                'report_description': rec.get('report_description', ''),
                'control_key': rec.get('control_key', ''),
            }

    return by_platform


def _strip_redundant_platform_description_prefix(text):
    inner = _strip_outer_quotes_once(text)
    if not inner:
        return ''

    while True:
        cleaned = re.sub(
            r'^[\-\u2013\u2014]?\s*(?P<platform>MSWRK|MSSRV|WINSRV|WINWRK)\s*-\s*[A-Za-z]{2}\s*-\s*(?P=platform)\s*-\s*',
            '',
            inner,
            flags=re.IGNORECASE,
        ).strip()
        if cleaned == inner:
            break
        inner = cleaned

    inner = re.sub(r'^[\-\u2013\u2014]?\s*(?:MSWRK|MSSRV|WINSRV|WINWRK)\s*[\-\u2013\u2014]\s*', '', inner, flags=re.IGNORECASE).strip()
    inner = re.sub(r'^[\-\u2013\u2014]?\s*[A-Za-z]{2}\s*[\-\u2013\u2014]\s*', '', inner).strip()
    return inner


def _prefixed_description(desc_value, platform, ordinal):
    desc_norm = normalize_description(desc_value)
    if not desc_norm:
        desc_norm = '"Control"'

    inner = _strip_redundant_platform_description_prefix(desc_norm)
    inner = re.sub(r'^\d+\.\d{4}\s*-\s*[A-Za-z0-9_-]+\s*-\s*', '', inner).strip()
    inner = re.sub(r'^\d+(?:\.\d+)+\s*', '', inner).strip()
    inner = re.sub(r'^[\-\u2013\u2014]\s*[A-Za-z0-9_-]+\s*[\-\u2013\u2014]\s*', '', inner).strip()
    inner = re.sub(r'^(MSWRK|MSSRV|WINSRV|WINWRK)\s*-\s*', '', inner, flags=re.IGNORECASE).strip()
    inner = re.sub(r'^[\-\u2013\u2014]?\s*' + re.escape(platform) + r'\s*[\-\u2013\u2014]\s*', '', inner, flags=re.IGNORECASE).strip()

    # Remove control-family/platform prefixes from source descriptions.
    # Example: "AC - MSSRV - Perform volume maintenance tasks..." -> "Perform volume maintenance tasks..."
    while True:
        cleaned = re.sub(
            r'^(AC|AU|CM|IA|SC)\s*-\s*(MSWRK|MSSRV)\s*-\s*',
            '',
            inner,
            flags=re.IGNORECASE,
        ).strip()
        if cleaned == inner:
            break
        inner = cleaned

    number = 1.0 + (ordinal / 10000.0)
    return f'"{number:.4f} - {platform} - {inner}"'


def _render_custom_item_block(fields, platform_hint, ordinal, apply_prefix=True):
    merged_fields = _convert_policy_item_to_audit_powershell(fields)
    if _should_force_convert_non_powershell(platform_hint, fields):
        merged_fields = _convert_remaining_item_to_audit_powershell(merged_fields)
    merged_fields = _ensure_audit_powershell_metadata_fields(merged_fields)
    merged_fields = _finalize_audit_powershell_fields_for_emit(merged_fields, platform_hint)
    merged_fields = OrderedDict(merged_fields)
    if apply_prefix:
        merged_fields['description'] = _prefixed_description(merged_fields.get('description', ''), platform_hint, ordinal)

    # Final render-time alignment for MSSRV compliance output parity.
    desc_inner = _strip_outer_quotes_once(merged_fields.get('description', ''))
    desc_upper = _norm_upper_token(desc_inner)
    ps_script = _strip_outer_quotes_once(merged_fields.get('powershell_script', merged_fields.get('powershell_args', '')))
    ps_upper = _norm_upper_token(ps_script)
    id_match = re.match(r'^(\d+\.\d+)\s*-', desc_inner)
    control_id = id_match.group(1) if id_match else ''

    if 'USER_RIGHTS' in ps_upper and 'NO ONE' not in desc_upper:
        expected_inner = _strip_outer_quotes_once(merged_fields.get('value_data', ''))
        if expected_inner and 'NO_MEMBERS' not in expected_inner and 'DENY LOG ON LOCALLY TO INCLUDE GUESTS' not in desc_upper:
            merged_fields['value_data'] = f'"{_allow_no_members_alternative(expected_inner)}"'
        elif expected_inner and 'NO_MEMBERS' in expected_inner and 'DENY LOG ON LOCALLY TO INCLUDE GUESTS' not in desc_upper:
            normalized_expected = _normalize_no_members_token_regex(expected_inner)
            if normalized_expected != expected_inner:
                merged_fields['value_data'] = f'"{normalized_expected}"'

    pairs = []
    for k, v in merged_fields.items():
        if k in IGNORED_KEYS or k not in REAL_KEYS:
            continue
        if k == 'see_also':
            pairs.append((k, f'"{SEE_ALSO_REPLACEMENT}"'))
        elif k == 'info':
            info = normalize_info(v)
            if info:
                pairs.append((k, info))
        elif k == 'reference':
            pairs.append((k, normalize_reference_or_passthrough(v)))
        elif k == 'solution':
            sol = normalize_solution(v)
            if sol:
                pairs.append((k, sol))
        elif k in AUDIT_QUOTED_STRING_FIELDS:
            pairs.append((k, _render_audit_string_value(v)))
        else:
            pairs.append((k, str(v)))

    pairs = order_custom_item_pairs(pairs)
    width = max(len(k) for k, _ in pairs) if pairs else 0
    lines = ['<custom_item>']
    for k, v in pairs:
        lines.append(f'  {k.ljust(width)} : {v}')
    lines.append('</custom_item>')
    return lines


def _comment_block_lines(lines):
    return [f'# {line}' for line in lines]


def _target_check_block(platform):
    if platform == 'MSSRV':
        value_data = '"^[Ww][Ii][Nn][Dd][Oo][Ww][Ss].+[Ss][Ee][Rr][Vv][Ee][Rr].+$"'
        description = '"Windows Server is installed"'
    else:
        value_data = '"^[Ww][Ii][Nn][Dd][Oo][Ww][Ss](?!.*[Ss][Ee][Rr][Vv][Ee][Rr]).+$"'
        description = '"Windows Workstation is installed"'

    lines = []
    lines.append('<custom_item>')
    lines.append('  type        : REGISTRY_SETTING')
    lines.append(f'  description : {description}')
    lines.append('  value_type  : POLICY_TEXT')
    lines.append(f'  value_data  : {value_data}')
    lines.append('  reg_key     : "HKLM\\Software\\Microsoft\\Windows Nt\\Currentversion"')
    lines.append('  reg_item    : "ProductName"')
    lines.append('  check_type  : CHECK_REGEX')
    lines.append('</custom_item>')
    return lines


def _domain_role_check_block(domain_roles, description):
    value_data = ' || '.join(str(role) for role in domain_roles)

    lines = []
    lines.append('<custom_item>')
    lines.append('  type          : WMI_POLICY')
    lines.append(f'  description   : "{description}"')
    lines.append('  value_type    : POLICY_DWORD')
    lines.append(f'  value_data    : {value_data}')
    lines.append('  wmi_namespace : "root/CIMV2"')
    lines.append('  wmi_request   : "select DomainRole from Win32_ComputerSystem"')
    lines.append('  wmi_attribute : "DomainRole"')
    lines.append('  wmi_key       : "DomainRole"')
    lines.append('</custom_item>')
    return lines


def _dc_target_check_block():
    lines = []
    lines.append('<custom_item>')
    lines.append('  type          : WMI_POLICY')
    lines.append('  description   : "Check if server is Domain Controller"')
    lines.append('  value_type    : POLICY_DWORD')
    lines.append('  value_data    : 4 || 5')
    lines.append('  wmi_namespace : "root/CIMV2"')
    lines.append('  wmi_request   : "select DomainRole from Win32_ComputerSystem"')
    lines.append('  wmi_attribute : "DomainRole"')
    lines.append('  wmi_key       : "DomainRole"')
    lines.append('</custom_item>')
    return lines


def _is_dc_only_control(fields):
    desc = _norm_upper_token(_strip_outer_quotes_once(fields.get('description', '')))
    return 'MSSRV.DC' in desc or '(DC ONLY)' in desc or ' DC ONLY' in desc


def _is_ms_only_control(fields):
    desc = _norm_upper_token(_strip_outer_quotes_once(fields.get('description', '')))
    return '(MS ONLY)' in desc or ' MS ONLY' in desc


def _is_target_os_applicability_control(fields):
    desc = _norm_upper_token(_strip_outer_quotes_once(fields.get('description', '')))
    info = _norm_upper_token(_strip_outer_quotes_once(fields.get('info', '')))
    powershell = _norm_upper_token(_strip_outer_quotes_once(fields.get('powershell_script', fields.get('powershell_args', ''))))
    reg_key = _norm_upper_token(_strip_outer_quotes_once(fields.get('reg_key', '')))
    reg_item = _norm_upper_token(_strip_outer_quotes_once(fields.get('reg_item', '')))
    is_target_desc = (
        'WINDOWS SERVER IS INSTALLED' in desc
        or 'WINDOWS WORKSTATION IS INSTALLED' in desc
    )
    is_product_name_probe = 'CURRENTVERSION' in powershell and 'PRODUCTNAME' in powershell
    is_registry_product_name_probe = 'WINDOWS NT\\CURRENTVERSION' in reg_key and reg_item == 'PRODUCTNAME'
    return is_target_desc and (
        'TARGET OS VALIDATION FOR BASELINE APPLICABILITY.' in info
        or is_product_name_probe
        or is_registry_product_name_probe
    )


def _is_domain_controller_probe_control(fields):
    desc = _norm_upper_token(_strip_outer_quotes_once(fields.get('description', '')))
    return 'CHECK IF SERVER IS DOMAIN CONTROLLER' in desc


def _is_rapid7_insight_agent_control(fields):
    # Unique filter rule: remove any control block that refers to a Rapid7 service.
    probe_keys = (
        'description',
        'info',
        'powershell_args',
        'powershell_script',
        'request',
        'cmd',
        'item',
        'reg_item',
    )
    joined = ' '.join(
        _norm_upper_token(_strip_outer_quotes_once(fields.get(key, '')))
        for key in probe_keys
    )
    return 'RAPID7' in joined and 'SERVICE' in joined


def _is_for_gap_target_applicability_control(fields):
    if _is_target_os_applicability_control(fields):
        return True
    desc = _norm_upper_token(_strip_outer_quotes_once(fields.get('description', '')))
    return (
        'TARGET OS MATCHES BASELINE' in desc
        or 'TARGET OS DOES NOT MATCH BASELINE' in desc
    )


def _is_for_gap_target_gate_control(fields):
    if _is_for_gap_target_applicability_control(fields):
        return True

    desc = _norm_upper_token(_strip_outer_quotes_once(fields.get('description', '')))
    req = _norm_upper_token(_strip_outer_quotes_once(fields.get('request', '')))
    xsl = _norm_upper_token(_strip_outer_quotes_once(fields.get('xsl_stmt', '')))

    if 'CHECK FOR PALO ALTO VERSION' in desc:
        return True

    if 'PANORAMA MODEL' in desc or 'PANORAMA SYSTEM-MODE' in desc:
        return True

    is_system_info_probe = '<SHOW><SYSTEM><INFO>' in req or '/RESPONSE/RESULT/SYSTEM/SW-VERSION' in xsl
    if is_system_info_probe and 'CHECK FOR' in desc and 'VERSION' in desc:
        return True

    return False


def _is_preconditional_inventory_control(fields):
    desc = _norm_upper_token(_strip_outer_quotes_once(fields.get('description', '')))
    return (
        'NETWORK INFORMATION' in desc
        or 'SYSTEM/DOMAIN IDENTIFICATION' in desc
    )


def _is_adcs_authorized_validation_control(fields):
    desc = _norm_upper_token(_strip_outer_quotes_once(fields.get('description', '')))
    return (
        'ACTIVE DIRECTORY CERTIFICATE SERVICES ROLE AUTHORIZED VALIDATION' in desc
        or 'CERTIFICATE SERVICES ROLE AUTHORIZED VALIDATION' in desc
    )


def _append_condition_block(out_lines, *, platform, active_rows, inactive_rows, warning_description, ordinal_start, dc_only=False, include_target_gate=True):
    if include_target_gate:
        out_lines.append('<if>')
        out_lines.append('  <condition type:"AND">')
        if dc_only:
            out_lines.extend(_dc_target_check_block())
        else:
            out_lines.extend(_target_check_block(platform))
        out_lines.append('  </condition>')
        out_lines.append('')
        out_lines.append('  <then>')
        out_lines.append('<report type:"PASSED">')
        passed_suffix = 'MSSRV.DC' if dc_only else platform
        out_lines.append(f'  description : "PASSED - TARGET OS MATCHES BASELINE - {passed_suffix}"')
        out_lines.append(f'  see_also    : "{SEE_ALSO_REPLACEMENT}"')
        out_lines.append('</report>')
        out_lines.append('')

    ordinal = ordinal_start
    for row in active_rows:
        block_lines = _render_custom_item_block(row['fields'], platform, ordinal)
        ordinal += 1
        out_lines.extend(block_lines)
        out_lines.append('')

    for row in inactive_rows:
        block_lines = _render_custom_item_block(row['fields'], platform, ordinal)
        ordinal += 1
        out_lines.extend(_comment_block_lines(block_lines))
        out_lines.append('')

    if include_target_gate:
        out_lines.append('  </then>')
        out_lines.append('')
        out_lines.append('  <else>')
        out_lines.append('    <report type:"WARNING">')
        out_lines.append(f'      description : {warning_description}')
        out_lines.append(f'      see_also    : "{SEE_ALSO_REPLACEMENT}"')
        out_lines.append('    </report>')
        out_lines.append('  </else>')
        out_lines.append('</if>')
        out_lines.append('')

    return ordinal


def _append_domain_role_block(out_lines, *, title, domain_roles, description, active_rows, inactive_rows, warning_description, ordinal_start, passed_suffix):
    if title:
        out_lines.append('#########################################################################################################################################')
        out_lines.append(f'########################################################### {title} ####################################################################')
        out_lines.append('#########################################################################################################################################')
        out_lines.append('')

    out_lines.append('<if>')
    out_lines.append('  <condition type:"AND">')
    out_lines.extend(_domain_role_check_block(domain_roles, description))
    out_lines.append('  </condition>')
    out_lines.append('')
    out_lines.append('  <then>')

    ordinal = ordinal_start
    for row in active_rows:
        block_lines = _render_custom_item_block(row['fields'], 'MSSRV', ordinal)
        ordinal += 1
        out_lines.extend(block_lines)
        out_lines.append('')

    for row in inactive_rows:
        block_lines = _render_custom_item_block(row['fields'], 'MSSRV', ordinal)
        ordinal += 1
        out_lines.extend(_comment_block_lines(block_lines))
        out_lines.append('')

    out_lines.append('  </then>')
    out_lines.append('')
    out_lines.append('  <else>')
    out_lines.append('  </else>')
    out_lines.append('</if>')
    out_lines.append('')

    return ordinal


def _is_target_os_installed_description(desc_text):
    desc = _norm_upper_token(_strip_outer_quotes_once(desc_text or ''))
    return 'WINDOWS SERVER IS INSTALLED' in desc or 'WINDOWS WORKSTATION IS INSTALLED' in desc


def _drop_target_os_passed_reports(lines):
    cleaned = []
    i = 0
    while i < len(lines):
        if re.match(r'^\s*<report\b[^>]*>\s*$', lines[i], flags=re.IGNORECASE):
            j = i + 1
            block = [lines[i]]
            while j < len(lines):
                block.append(lines[j])
                if re.match(r'^\s*</report>\s*$', lines[j], flags=re.IGNORECASE):
                    break
                j += 1

            block_text = '\n'.join(block)
            if re.search(r'PASSED\s*-\s*TARGET\s+OS\s+MATCHES\s+BASELINE', block_text, flags=re.IGNORECASE):
                i = j + 1
                continue

            cleaned.extend(block)
            i = j + 1
            continue

        cleaned.append(lines[i])
        i += 1

    return cleaned


def _drop_target_os_custom_items(lines):
    cleaned = []
    i = 0
    while i < len(lines):
        if re.match(r'^\s*<custom_item>\s*$', lines[i], flags=re.IGNORECASE):
            j = i + 1
            block = [lines[i]]
            while j < len(lines):
                block.append(lines[j])
                if re.match(r'^\s*</custom_item>\s*$', lines[j], flags=re.IGNORECASE):
                    break
                j += 1

            desc_text = ''
            for line in block:
                m = re.match(r'^\s*description\s*:\s*(.+?)\s*$', line, flags=re.IGNORECASE)
                if m:
                    desc_text = m.group(1)
                    break

            if _is_target_os_installed_description(desc_text):
                i = j + 1
                continue

            cleaned.extend(block)
            i = j + 1
            continue

        cleaned.append(lines[i])
        i += 1

    return cleaned


def _drop_rapid7_service_item_blocks(lines):
    cleaned = []
    i = 0
    while i < len(lines):
        if re.match(r'^\s*<(custom_item|item)>\s*$', lines[i], flags=re.IGNORECASE):
            m_open = re.match(r'^\s*<(custom_item|item)>\s*$', lines[i], flags=re.IGNORECASE)
            tag_name = m_open.group(1).lower() if m_open else 'custom_item'
            close_re = rf'^\s*</{tag_name}>\s*$'

            j = i + 1
            block = [lines[i]]
            while j < len(lines):
                block.append(lines[j])
                if re.match(close_re, lines[j], flags=re.IGNORECASE):
                    break
                j += 1

            block_text_upper = _norm_upper_token('\n'.join(block))
            if 'RAPID7' in block_text_upper and 'SERVICE' in block_text_upper:
                i = j + 1
                continue

            cleaned.extend(block)
            i = j + 1
            continue

        cleaned.append(lines[i])
        i += 1

    return cleaned


def _strip_target_os_applicability_wrappers(lines):
    out = []
    i = 0
    while i < len(lines):
        if not re.match(r'^\s*<if>\s*$', lines[i], flags=re.IGNORECASE):
            out.append(lines[i])
            i += 1
            continue

        depth = 0
        j = i
        while j < len(lines):
            if re.match(r'^\s*<if>\s*$', lines[j], flags=re.IGNORECASE):
                depth += 1
            elif re.match(r'^\s*</if>\s*$', lines[j], flags=re.IGNORECASE):
                depth -= 1
                if depth == 0:
                    break
            j += 1

        if j >= len(lines):
            out.append(lines[i])
            i += 1
            continue

        block = lines[i:j + 1]

        has_target_gate = any(
            re.match(r'^\s*description\s*:\s*.+$', line, flags=re.IGNORECASE)
            and _is_target_os_installed_description(line.split(':', 1)[1].strip())
            for line in block
        )

        if not has_target_gate:
            out.extend(block)
            i = j + 1
            continue

        then_start = None
        then_end = None
        depth = 0
        for k, line in enumerate(block):
            if re.match(r'^\s*<if>\s*$', line, flags=re.IGNORECASE):
                depth += 1
            elif re.match(r'^\s*</if>\s*$', line, flags=re.IGNORECASE):
                depth -= 1
            elif depth == 1 and re.match(r'^\s*<then>\s*$', line, flags=re.IGNORECASE):
                then_start = k
            elif depth == 1 and re.match(r'^\s*</then>\s*$', line, flags=re.IGNORECASE):
                then_end = k

        if then_start is None or then_end is None or then_end <= then_start:
            out.extend(block)
            i = j + 1
            continue

        then_body = block[then_start + 1:then_end]
        then_body = _drop_target_os_passed_reports(then_body)
        then_body = _drop_target_os_custom_items(then_body)
        then_body = _drop_rapid7_service_item_blocks(then_body)
        out.extend(then_body)
        i = j + 1

    return out


def _cleanup_normalized_output_lines(lines):
    cleaned = _sanitize_audit_lines(lines)
    cleaned = _strip_target_os_applicability_wrappers(cleaned)
    cleaned = _drop_target_os_custom_items(cleaned)
    cleaned = _drop_rapid7_service_item_blocks(cleaned)
    cleaned = _drop_target_os_passed_reports(cleaned)
    return cleaned


def _is_blank_field(value):
    raw = _strip_outer_quotes_once(value)
    return not str(raw).strip()


def _merge_control_fields(master_fields, other_fields):
    merged = dict(master_fields or {})
    fallback = dict(other_fields or {})

    # Backfill missing metadata from the matched counterpart control.
    for key in ('info', 'solution', 'reference', 'see_also'):
        if _is_blank_field(merged.get(key, '')) and not _is_blank_field(fallback.get(key, '')):
            merged[key] = fallback.get(key)

    # Ensure all requested metadata fields exist with sane defaults.
    if _is_blank_field(merged.get('info', '')):
        merged['info'] = '"No additional information provided."'
    if _is_blank_field(merged.get('solution', '')):
        merged['solution'] = '"Configure this setting to the expected baseline value shown in the control description and value_data."'
    if _is_blank_field(merged.get('reference', '')):
        merged['reference'] = '""'
    if _is_blank_field(merged.get('see_also', '')):
        merged['see_also'] = f'"{SEE_ALSO_REPLACEMENT}"'

    return merged


def _to_float_or_default(value, default):
    try:
        return float(value)
    except Exception:
        return float(default)


def _default_threat_intel_payload():
    return {
        'version': '1',
        'updated_at': time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime()),
        'source': 'builtin',
        'control_taxonomy': list(DEFAULT_CONTROL_TAXONOMY),
        # Optional external join payload keyed by taxonomy control ID.
        # Example: {"CTRL-WIN-CRED-NTLM": {"attack_prevalence": 1.7, "threat_weight": 1.2, "source": "cisa_kev_snapshot"}}
        'threat_by_control_id': {},
    }


def _safe_read_json(path):
    try:
        with open(path, 'r', encoding='utf-8') as fh:
            return json.load(fh)
    except Exception:
        return {}


def _safe_write_json(path, payload):
    try:
        parent = os.path.dirname(path) or '.'
        os.makedirs(parent, exist_ok=True)
        temp_path = f'{path}.{os.getpid()}.tmp'
        with open(temp_path, 'w', encoding='utf-8') as fh:
            json.dump(payload, fh, indent=2, sort_keys=True)
            fh.write('\n')
        os.replace(temp_path, path)
        return True
    except Exception:
        return False


def _merge_threat_intel_payload(base_payload, overlay_payload):
    base = dict(base_payload or {})
    overlay = dict(overlay_payload or {})

    base_map = {str(entry.get('id', '')).strip(): dict(entry) for entry in base.get('control_taxonomy', []) if str(entry.get('id', '')).strip()}
    for entry in overlay.get('control_taxonomy', []) or []:
        cid = str(entry.get('id', '')).strip()
        if not cid:
            continue
        merged = dict(base_map.get(cid, {}))
        merged.update(dict(entry))
        base_map[cid] = merged

    merged_threat_by_id = dict(base.get('threat_by_control_id', {}) or {})
    merged_threat_by_id.update(dict(overlay.get('threat_by_control_id', {}) or {}))

    merged = dict(base)
    merged.update({k: v for k, v in overlay.items() if k not in {'control_taxonomy', 'threat_by_control_id'}})
    merged['control_taxonomy'] = sorted(base_map.values(), key=lambda e: str(e.get('id', '')))
    merged['threat_by_control_id'] = merged_threat_by_id
    return merged


def _fetch_remote_threat_intel(url):
    if not url:
        return {}
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'PySC-ThreatIntel/1.0'})
        with urllib.request.urlopen(req, timeout=20) as resp:
            if getattr(resp, 'status', 200) >= 400:
                return {}
            content_type = (resp.headers.get('Content-Type') or '').lower()
            body = resp.read().decode('utf-8', errors='replace')
            if 'application/json' not in content_type and not body.lstrip().startswith('{'):
                return {}
            return json.loads(body)
    except Exception:
        return {}


def _cache_is_stale(path, max_age_hours):
    if not os.path.isfile(path):
        return True
    age_seconds = max(0, time.time() - os.path.getmtime(path))
    return age_seconds > (float(max_age_hours) * 3600.0)


def _load_threat_intel_cache(force_refresh=False):
    global THREAT_INTEL_DATA

    merged = _default_threat_intel_payload()
    local_payload = _safe_read_json(THREAT_INTEL_CACHE_PATH)
    if local_payload:
        merged = _merge_threat_intel_payload(merged, local_payload)

    should_refresh = bool(force_refresh or THREAT_INTEL_FORCE_REFRESH)
    if not should_refresh and THREAT_INTEL_FEED_URL:
        should_refresh = _cache_is_stale(THREAT_INTEL_CACHE_PATH, THREAT_INTEL_CACHE_MAX_AGE_HOURS)

    if should_refresh and THREAT_INTEL_FEED_URL:
        remote_payload = _fetch_remote_threat_intel(THREAT_INTEL_FEED_URL)
        if remote_payload:
            merged = _merge_threat_intel_payload(merged, remote_payload)
            merged['updated_at'] = time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime())
            merged['source'] = str(remote_payload.get('source', THREAT_INTEL_FEED_URL))
            _safe_write_json(THREAT_INTEL_CACHE_PATH, merged)

    if not os.path.isfile(THREAT_INTEL_CACHE_PATH):
        _safe_write_json(THREAT_INTEL_CACHE_PATH, merged)

    THREAT_INTEL_DATA = merged
    return merged


def _taxonomy_entries_from_intel_data():
    data = THREAT_INTEL_DATA or _load_threat_intel_cache(force_refresh=False)
    entries = []
    for raw in data.get('control_taxonomy', []) or []:
        cid = str(raw.get('id', '')).strip()
        if not cid:
            continue
        entry = dict(raw)
        entry['id'] = cid
        entry['family'] = _norm_upper_token(raw.get('family', 'OTHER')) or 'OTHER'
        tokens = raw.get('match_any', []) or []
        entry['match_any'] = [_norm_upper_token(tok) for tok in tokens if _norm_upper_token(tok)]
        entries.append(entry)
    return entries


def _match_control_to_taxonomy(fields):
    text = _joined_control_text_upper(fields)
    best_entry = None
    best_score = -1

    for entry in _taxonomy_entries_from_intel_data():
        tokens = entry.get('match_any', [])
        if not tokens:
            continue
        matched = sum(1 for tok in tokens if tok in text)
        if matched <= 0:
            continue
        if matched > best_score:
            best_score = matched
            best_entry = entry

    if not best_entry:
        return '', None
    return best_entry.get('id', ''), best_entry


def _external_intel_for_control_id(control_id):
    if not control_id:
        return {}
    data = THREAT_INTEL_DATA or _load_threat_intel_cache(force_refresh=False)
    by_id = data.get('threat_by_control_id', {}) or {}
    payload = by_id.get(control_id, {})
    return dict(payload) if isinstance(payload, dict) else {}


def _joined_control_text_upper(fields):
    probe_keys = (
        'description',
        'info',
        'type',
        'value_data',
        'check_type',
        'reg_key',
        'reg_item',
        'key_item',
        'right_type',
        'account_type',
        'audit_policy_subcategory',
        'password_policy',
        'lockout_policy',
        'powershell_args',
        'powershell_script',
        'cmd',
        'item',
    )
    return ' '.join(
        _norm_upper_token(_strip_outer_quotes_once(fields.get(key, '')))
        for key in probe_keys
    )


def _classify_security_family(fields):
    text = _joined_control_text_upper(fields)

    if any(token in text for token in ('SEDEBUGPRIVILEGE', 'SETAKEOWNERSHIPPRIVILEGE', 'ADMINISTRATOR', 'LOCAL ADMIN', 'PRIVILEGE')):
        return 'PRIVILEGED_ACCESS'
    if any(token in text for token in ('NTLM', 'KERBEROS', 'PASSWORD', 'CREDENTIAL', 'LSA', 'CLEARTEXTPASSWORD', 'PASSWORDCOMPLEXITY')):
        return 'CREDENTIAL_ACCESS'
    if any(token in text for token in ('ACCOUNT LOCKOUT', 'LOCKOUT', 'LOGON', 'SIGN-IN', 'AUTHENTICATION', 'WINLOGON')):
        return 'AUTHENTICATION'
    if any(token in text for token in ('FIREWALL', 'RDP', 'SMB', 'WINRM', 'SSH', 'TLS', 'REMOTE', 'NETWORK ACCESS')):
        return 'NETWORK_BOUNDARY'
    if any(token in text for token in ('AUDITPOL', 'AUDIT POLICY', 'EVENT LOG', 'LOGGING', 'DETECTION', 'POWERSHELL TRANSCRIPTION')):
        return 'LOGGING_DETECTION'
    if any(token in text for token in ('REGISTRY', 'SERVICE', 'UAC', 'DEFENDER', 'BITLOCKER', 'PATCH', 'CIS')):
        return 'SYSTEM_HARDENING'
    return 'OTHER'


def _estimate_attack_prevalence(fields):
    text = _joined_control_text_upper(fields)
    score = 1.00

    high_signal_tokens = (
        'RDP', 'SMB', 'NTLM', 'KERBEROS', 'WINRM', 'LSA',
        'ADMINISTRATOR', 'GUEST ACCOUNT', 'PASSWORD', 'LOCKOUT',
    )
    medium_signal_tokens = (
        'FIREWALL', 'AUDIT POLICY', 'EVENT LOG', 'POWERSHELL', 'REGISTRY',
    )

    if any(tok in text for tok in high_signal_tokens):
        score = max(score, 1.35)
    if any(tok in text for tok in medium_signal_tokens):
        score = max(score, 1.15)

    return score


def _estimate_asset_exposure_multiplier(fields, platform_hint=''):
    text = _joined_control_text_upper(fields)
    platform = _norm_upper_token(platform_hint)

    score = 1.00
    if any(tok in text for tok in ('TIER-0', 'TIER 0', 'DOMAIN CONTROLLER', 'MSSRV.DC', '(DC ONLY)')):
        score = max(score, 1.50)
    elif any(tok in text for tok in ('INTERNET', 'PUBLIC', 'EXTERNAL', 'REMOTE', 'RDP', 'WINRM', 'SMB')):
        score = max(score, 1.30)

    if platform in {'PAFW', 'F5', 'ASA', 'NXOS'}:
        score = max(score, 1.25)

    return score


def _estimate_exploitability_modifier(fields):
    text = _joined_control_text_upper(fields)

    # Easier bypass -> larger modifier (higher operational risk priority).
    if any(tok in text for tok in ('CHECK_EQUAL', 'CHECK_NOT_EQUAL', 'EXIST', 'NOT_EXIST', 'SERVICE')):
        return 1.20
    if any(tok in text for tok in ('CHECK_REGEX', 'CHECK_SUPERSET', 'CHECK_SUBSET', 'AUDITPOL', 'USER_RIGHTS')):
        return 1.00
    return 0.95


def _compute_operational_risk_score(fields, platform_hint=''):
    control_id, taxonomy_entry = _match_control_to_taxonomy(fields)
    family = (taxonomy_entry or {}).get('family') or _classify_security_family(fields)
    family_weight = SECURITY_FAMILY_WEIGHTS.get(family, SECURITY_FAMILY_WEIGHTS['OTHER'])
    prevalence = _estimate_attack_prevalence(fields)
    exposure = _estimate_asset_exposure_multiplier(fields, platform_hint=platform_hint)
    exploitability = _estimate_exploitability_modifier(fields)

    taxonomy_source = ''
    if taxonomy_entry:
        prevalence = max(prevalence, _to_float_or_default(taxonomy_entry.get('attack_prevalence'), prevalence))
        exposure = max(exposure, _to_float_or_default(taxonomy_entry.get('asset_exposure'), exposure))
        exploitability = max(exploitability, _to_float_or_default(taxonomy_entry.get('exploitability_modifier'), exploitability))
        taxonomy_source = str(taxonomy_entry.get('source', '') or '')

    external_intel = _external_intel_for_control_id(control_id)
    threat_weight = 1.0
    intel_source = taxonomy_source
    if external_intel:
        prevalence = max(prevalence, _to_float_or_default(external_intel.get('attack_prevalence'), prevalence))
        exposure = max(exposure, _to_float_or_default(external_intel.get('asset_exposure'), exposure))
        exploitability = max(exploitability, _to_float_or_default(external_intel.get('exploitability_modifier'), exploitability))
        threat_weight = max(1.0, _to_float_or_default(external_intel.get('threat_weight'), 1.0))
        intel_source = str(external_intel.get('source', '') or intel_source)

    score = family_weight * prevalence * exposure * exploitability * threat_weight
    return {
        'control_id': control_id,
        'family': family,
        'family_weight': family_weight,
        'attack_prevalence': prevalence,
        'asset_exposure': exposure,
        'exploitability_modifier': exploitability,
        'threat_weight': threat_weight,
        'intel_source': intel_source,
        'priority_score': score,
    }


def _description_core_for_dedupe(description):
    text = _strip_outer_quotes_once(description)
    text = re.sub(r'^\d+\.\d+\s*-\s*(MSWRK|MSSRV)\s*-\s*', '', text, flags=re.IGNORECASE)
    text = _strip_redundant_platform_description_prefix(text)
    return _norm_upper_token(text)


def _merged_row_dedupe_signature(fields):
    # Collapse duplicate controls by logical description across all merge buckets.
    description_core = _description_core_for_dedupe(fields.get('description', ''))
    return description_core


def _row_priority_score(row):
    risk = row.get('risk_meta', {}) or {}
    try:
        return float(risk.get('priority_score', 0.0) or 0.0)
    except Exception:
        return 0.0


def _prefer_row_for_gap(existing, candidate):
    existing_active = bool(existing.get('active'))
    candidate_active = bool(candidate.get('active'))
    if existing_active != candidate_active:
        return candidate_active

    existing_score = _row_priority_score(existing)
    candidate_score = _row_priority_score(candidate)
    if abs(candidate_score - existing_score) > 1e-12:
        return candidate_score > existing_score

    existing_rank = int(existing.get('source_rank', 99))
    candidate_rank = int(candidate.get('source_rank', 99))
    if existing_rank != candidate_rank:
        return candidate_rank < existing_rank

    return str(candidate.get('key', '')) < str(existing.get('key', ''))


def _dedupe_merged_rows(merged_rows):
    deduped = []
    seen = {}
    dropped = 0

    for row in merged_rows:
        sig = _merged_row_dedupe_signature(row.get('fields', {}))
        existing_index = seen.get(sig)
        if existing_index is None:
            seen[sig] = len(deduped)
            deduped.append(row)
            continue

        dropped += 1
        existing = deduped[existing_index]
        # Tie-break duplicate descriptions by activity, risk score, and source rank.
        if _prefer_row_for_gap(existing, row):
            deduped[existing_index] = row

    return deduped, dropped


def _resolve_source_path_from_catalog(catalog_path, source_file_value):
    if not source_file_value:
        return ''
    src = str(source_file_value).strip().strip('"').strip("'")
    if not src:
        return ''
    if os.path.isabs(src) and os.path.isfile(src):
        return src
    candidate = os.path.normpath(os.path.join(os.path.dirname(catalog_path), src))
    if os.path.isfile(candidate):
        return candidate
    return ''


def _write_uncommented_all_copy(src_path):
    if not src_path or not os.path.isfile(src_path):
        return ''

    dst_path = os.path.join(os.path.dirname(src_path), f'ALL_{os.path.basename(src_path)}')
    try:
        with open(src_path, encoding='utf-8', errors='replace') as fh:
            lines = _sanitize_audit_lines(fh.readlines())
    except Exception:
        return ''

    out_lines = []
    in_custom_item_block = False

    for line in lines:
        if re.match(r'^#\s*<custom_item>\s*$', line):
            in_custom_item_block = True
            out_lines.append(re.sub(r'^#\s?', '', line))
            continue

        if in_custom_item_block:
            out_lines.append(re.sub(r'^#\s?', '', line))
            if re.match(r'^#\s*</custom_item>\s*$', line):
                in_custom_item_block = False
            continue

        out_lines.append(line)

    with open(dst_path, 'w', encoding='utf-8', newline='\n') as fh:
        cleaned_lines = _sanitize_audit_lines(out_lines)
        cleaned_text = ''.join(cleaned_lines)
        _assert_no_encoding_markers(cleaned_text, dst_path)
        fh.writelines(cleaned_lines)

    return dst_path


def _logical_audit_key(path):
    base = os.path.basename(path)
    stem, ext = os.path.splitext(base)
    stem = _TS_SUFFIX_RE.sub('', stem)
    return f'{stem.lower()}{ext.lower()}'


def _latest_audit_files_by_key(folder, include_prefixes=None, exclude_prefixes=None):
    chosen = {}
    if not os.path.isdir(folder):
        return []

    include_prefixes = tuple(include_prefixes or ())
    exclude_prefixes = tuple(exclude_prefixes or ())

    for name in sorted(os.listdir(folder), key=str.lower):
        if not name.lower().endswith('.audit'):
            continue
        if name.upper().startswith('ALL_'):
            continue
        if include_prefixes and not name.startswith(include_prefixes):
            continue
        if exclude_prefixes and name.startswith(exclude_prefixes):
            continue

        path = os.path.join(folder, name)
        key = _logical_audit_key(path)
        current = chosen.get(key)
        if current is None:
            chosen[key] = path
            continue

        current_base = os.path.basename(current)
        current_stem = os.path.splitext(current_base)[0]
        new_stem = os.path.splitext(name)[0]
        current_stamp = _TS_SUFFIX_RE.search(current_stem)
        new_stamp = _TS_SUFFIX_RE.search(new_stem)
        current_val = current_stamp.group(0).lstrip('_') if current_stamp else ''
        new_val = new_stamp.group(0).lstrip('_') if new_stamp else ''
        if new_val > current_val or (new_val == current_val and os.path.getmtime(path) > os.path.getmtime(current)):
            chosen[key] = path

    return sorted(chosen.values(), key=lambda p: os.path.basename(p).lower())


def _copy_file_to_folder(src_path, dst_folder):
    if not src_path or not os.path.isfile(src_path):
        return ''
    os.makedirs(dst_folder, exist_ok=True)
    dst_path = os.path.join(dst_folder, os.path.basename(src_path))
    shutil.copy2(src_path, dst_path)
    return dst_path


def _gap_type_label_from_path(path):
    platform = determine_platform_from_filename(path)
    platform = (platform or 'UNKNOWN').upper()
    return {
        'MSAZ': 'AZURE',
        'NX-OS': 'NXOS',
        'VMWARE': 'VMWARE',
    }.get(platform, platform)


def _gap_source_rank(path):
    norm = str(path).replace('/', '\\').lower()
    if '\\actual_audit_inputs\\normalized\\' in norm:
        return 0
    if '\\actual_audit_inputs\\merged\\' in norm:
        return 0
    if '\\actual_audit_inputs\\' in norm and '\\for_gap\\' not in norm:
        return 0
    if '\\audit_inputs\\normalized\\' in norm:
        return 1
    return 2


def _iter_custom_item_blocks(audit_path):
    if not audit_path or not os.path.isfile(audit_path):
        return

    field_re = re.compile(r'^\s*([A-Za-z0-9_]+)\s*:\s*(.*?)\s*$', re.IGNORECASE)

    try:
        with open(audit_path, encoding='utf-8', errors='replace') as fh:
            lines = _sanitize_audit_lines(fh.readlines())
    except Exception:
        return

    in_block = False
    block_lines = []
    block_commented = False

    for raw_line in lines:
        stripped = raw_line.lstrip()
        was_comment = stripped.startswith('#')
        content_line = stripped[1:].lstrip() if was_comment else raw_line

        if re.match(r'^\s*<(?:custom_item|item)>\s*$', content_line, flags=re.IGNORECASE):
            in_block = True
            block_lines = [raw_line]
            block_commented = was_comment
            continue

        if in_block:
            block_lines.append(raw_line)
            if re.match(r'^\s*</(?:custom_item|item)>\s*$', content_line, flags=re.IGNORECASE):
                fields = {}
                for line in block_lines:
                    line_stripped = line.lstrip()
                    parsed_line = line_stripped[1:].lstrip() if line_stripped.startswith('#') else line
                    field_match = field_re.match(parsed_line)
                    if field_match:
                        fields[field_match.group(1).lower()] = re.sub(r'\s+', ' ', field_match.group(2)).strip()

                yield {
                    'active': not block_commented,
                    'fields': fields,
                }
                in_block = False
                block_lines = []
                block_commented = False


def _iter_target_report_blocks(audit_path):
    if not audit_path or not os.path.isfile(audit_path):
        return

    try:
        with open(audit_path, encoding='utf-8', errors='replace') as fh:
            lines = _sanitize_audit_lines(fh.readlines())
    except Exception:
        return

    i = 0
    while i < len(lines):
        raw = lines[i]
        stripped = raw.lstrip()
        is_commented = stripped.startswith('#')
        content = stripped[1:].lstrip() if is_commented else raw

        if not re.match(r'^\s*<report\b[^>]*>\s*$', content, flags=re.IGNORECASE):
            i += 1
            continue

        report_fields = {}
        i += 1
        while i < len(lines):
            current = lines[i]
            current_stripped = current.lstrip()
            current_commented = current_stripped.startswith('#')
            current_content = current_stripped[1:].lstrip() if current_commented else current

            if re.match(r'^\s*</report>\s*$', current_content, flags=re.IGNORECASE):
                break

            fm = FIELD_RE.match(current_content)
            if fm:
                report_fields[fm.group(1).lower()] = fm.group(2).strip()
            i += 1

        desc_inner = _norm_upper_token(_strip_outer_quotes_once(report_fields.get('description', '')))
        if 'TARGET OS MATCHES BASELINE' in desc_inner or 'TARGET OS DOES NOT MATCH BASELINE' in desc_inner:
            fields = {
                'type': report_fields.get('type', 'FILE_CONTENT_CHECK'),
                'description': report_fields.get('description', ''),
                'see_also': report_fields.get('see_also', ''),
                'info': report_fields.get('info', ''),
                'reference': report_fields.get('reference', ''),
                'solution': report_fields.get('solution', ''),
                'item': report_fields.get('item', ''),
            }
            yield {
                'active': not is_commented,
                'fields': fields,
            }

        i += 1


def _write_gap_combined_audit(out_path, source_files, blocks, check_type_name, check_type_version, group_policy_name, platform_label, renumber=True):
    out_lines = []
    out_lines.append('# generated_by: for-gap-combined-dedupe')
    if check_type_name and check_type_version:
        out_lines.append(f'<check_type:"{check_type_name}" version:"{check_type_version}">')
    elif check_type_name:
        out_lines.append(f'<check_type:"{check_type_name}">')
    else:
        out_lines.append('<check_type:"Windows" version:"2">')

    out_lines.append(f'<group_policy:"{group_policy_name or "For Gap Combined"}">')
    out_lines.append('')

    normalized_blocks = []
    ordinal = 0
    for block in blocks:
        fields = dict(block.get('fields', {}))
        is_target_applicability = _is_for_gap_target_applicability_control(fields)
        is_target_gate = _is_for_gap_target_gate_control(fields)
        is_unnumbered_target = is_target_applicability or is_target_gate
        if is_unnumbered_target:
            desc_inner = _strip_outer_quotes_once(fields.get('description', ''))
            desc_inner = re.sub(r'^\d+\.\d{4}\s*-\s*[A-Za-z0-9_-]+\s*-\s*', '', desc_inner).strip()
            if desc_inner:
                fields['description'] = f'"{desc_inner}"'

        apply_prefix_for_block = renumber and (not is_unnumbered_target)
        rendered = _render_custom_item_block(
            fields,
            platform_label or check_type_name or 'Windows',
            ordinal,
            apply_prefix=apply_prefix_for_block,
        )
        normalized_blocks.append({
            'active': bool(block.get('active')),
            'is_target_gate': is_target_gate,
            'rendered': rendered,
        })
        ordinal += 1

    active_target_gates = [b for b in normalized_blocks if b.get('is_target_gate') and b.get('active')]

    deduped_active_target_gates = []
    seen_target_gate_keys = set()
    for gate in active_target_gates:
        rendered = gate.get('rendered', [])
        gate_key = ''
        for line in rendered:
            if 'description' not in line:
                continue
            desc_inner = _norm_upper_token(_strip_outer_quotes_once(line.split(':', 1)[1].strip()))
            if 'CHECK FOR PALO ALTO VERSION 10' in desc_inner:
                gate_key = 'pafw_version_10'
            elif 'CHECK FOR PALO ALTO VERSION 11' in desc_inner:
                gate_key = 'pafw_version_11'
            elif 'PANORAMA MODEL' in desc_inner:
                gate_key = 'pafw_panorama_model'
            elif 'PANORAMA SYSTEM-MODE' in desc_inner:
                gate_key = 'pafw_panorama_system_mode'
            else:
                gate_key = desc_inner
            break

        if not gate_key or gate_key in seen_target_gate_keys:
            continue
        seen_target_gate_keys.add(gate_key)
        deduped_active_target_gates.append(gate)

    use_target_gate_wrapper = (platform_label or '').upper() == 'PAFW' and bool(deduped_active_target_gates)
    if use_target_gate_wrapper:
        out_lines.append('<if>')
        out_lines.append('  <condition type:"OR">')
        for gate in deduped_active_target_gates:
            for line in gate['rendered']:
                out_lines.append(f'    {line}')
            out_lines.append('')
        out_lines.append('  </condition>')
        out_lines.append('')
        out_lines.append('  <then>')
        out_lines.append('    <report type:"PASSED">')
        out_lines.append(f'      description : "{os.path.basename(out_path)} target checks matched baseline"')
        out_lines.append('      info        : "Target platform gates passed for this merged baseline."')
        out_lines.append('      see_also    : "See HTH Policies and Standards"')
        out_lines.append('    </report>')
        out_lines.append('')

    for block in normalized_blocks:
        if use_target_gate_wrapper and block.get('is_target_gate') and block.get('active'):
            continue
        rendered = block['rendered']
        if not block.get('active'):
            rendered = _comment_block_lines(rendered)
        out_lines.extend(rendered)
        out_lines.append('')

    if use_target_gate_wrapper:
        out_lines.append('  </then>')
        out_lines.append('</if>')
        out_lines.append('')

    out_lines.append('</group_policy>')
    out_lines.append('</check_type>')

    out_text = '\n'.join(out_lines).rstrip() + '\n'
    with open(out_path, 'w', encoding='utf-8', newline='\n') as fh:
        fh.write(out_text)

    validate_and_repair_audit_file(
        out_path,
        check_type_name='Windows',
        check_type_version='2',
        group_policy_name='Windows Security Hardening Through Group Policy and Registry Settings',
        platform_hint=platform_label,
    )

    print(f'Wrote gap combined audit: {out_path} ({len(source_files)} source file(s), {len(blocks)} unique checks)')
    return out_path


def _stage_gap_analysis_files():
    gap_root = os.path.join(SCRIPT_DIR, 'actual_audit_inputs', 'For_Gap')
    source_roots = [
        os.path.join(SCRIPT_DIR, 'actual_audit_inputs'),
        os.path.join(SCRIPT_DIR, 'actual_audit_inputs', 'Normalized'),
        os.path.join(SCRIPT_DIR, 'actual_audit_inputs', 'Merged'),
        os.path.join(SCRIPT_DIR, 'audit_inputs', 'Normalized'),
    ]
    os.makedirs(gap_root, exist_ok=True)

    source_files = []
    for root in source_roots:
        if os.path.isdir(root):
            source_files.extend(_latest_audit_files_by_key(root, exclude_prefixes=('ALL_',)))

    grouped_paths = {}
    for src_path in source_files:
        type_label = _gap_type_label_from_path(src_path)
        if type_label == 'UNKNOWN':
            continue
        grouped_paths.setdefault(type_label, []).append((_gap_source_rank(src_path), src_path))

    print(f'Gap analysis folder ready: {gap_root}')

    for name in os.listdir(gap_root):
        existing_path = os.path.join(gap_root, name)
        try:
            if os.path.isdir(existing_path):
                shutil.rmtree(existing_path)
            else:
                os.remove(existing_path)
        except Exception:
            pass

    written = []
    merged_root = os.path.join(SCRIPT_DIR, 'actual_audit_inputs', 'Merged')
    for type_label in sorted(grouped_paths):
        ranked_paths = sorted(grouped_paths[type_label], key=lambda item: (item[0], os.path.basename(item[1]).lower()))

        if type_label in {'MSSRV', 'MSWRK'}:
            merged_paths = _latest_audit_files_by_key(
                merged_root,
                include_prefixes=(f'Merged_{type_label}',),
            )
            if merged_paths:
                dst_path = os.path.join(gap_root, f'{type_label}.audit')
                shutil.copy2(merged_paths[0], dst_path)
                validate_and_repair_audit_file(
                    dst_path,
                    check_type_name='Windows',
                    check_type_version='2',
                    group_policy_name='Windows Security Hardening Through Group Policy and Registry Settings',
                    platform_hint=type_label,
                )
                written.append(dst_path)
                print(f'Wrote gap exact copy: {dst_path} <- {merged_paths[0]}')
            continue

        merged_rows = []
        for source_rank, path in ranked_paths:
            for block in _iter_custom_item_blocks(path):
                normalized_fields = _merge_control_fields(block.get('fields', {}), {})
                if type_label == 'RHEL' and _is_blank_field(normalized_fields.get('system', '')):
                    normalized_fields['system'] = '"Linux"'
                description_sort = _description_core_for_dedupe(normalized_fields.get('description', ''))
                risk_meta = _compute_operational_risk_score(normalized_fields, platform_hint=type_label)
                # Only controls from actual_audit_inputs are considered active defaults.
                # Controls sourced only from audit_inputs are carried as inactive candidates.
                is_master_source = source_rank == 0
                is_target_gate = _is_for_gap_target_gate_control(normalized_fields)
                merged_rows.append({
                    'key': f'{source_rank}:{os.path.basename(path)}:{len(merged_rows)}',
                    'source_rank': source_rank,
                    'active': bool(block.get('active')) and (is_master_source or is_target_gate),
                    'fields': normalized_fields,
                    'description_sort': description_sort,
                    'risk_meta': risk_meta,
                })

            for block in _iter_target_report_blocks(path):
                normalized_fields = _merge_control_fields(block.get('fields', {}), {})
                if type_label == 'RHEL' and _is_blank_field(normalized_fields.get('system', '')):
                    normalized_fields['system'] = '"Linux"'
                description_sort = _description_core_for_dedupe(normalized_fields.get('description', ''))
                risk_meta = _compute_operational_risk_score(normalized_fields, platform_hint=type_label)
                is_master_source = source_rank == 0
                merged_rows.append({
                    'key': f'{source_rank}:{os.path.basename(path)}:target:{len(merged_rows)}',
                    'source_rank': source_rank,
                    'active': bool(block.get('active')) and is_master_source,
                    'fields': normalized_fields,
                    'description_sort': description_sort,
                    'risk_meta': risk_meta,
                })

        merged_rows.sort(
            key=lambda r: (
                -_row_priority_score(r),
                r.get('source_rank', 99),
                r['description_sort'],
                r['key'],
            )
        )
        deduped_rows, dropped_dupes = _dedupe_merged_rows(merged_rows)
        deduped_rows.sort(
            key=lambda r: (
                -_row_priority_score(r),
                r.get('source_rank', 99),
                r.get('description_sort', ''),
                str(r.get('key', '')),
            )
        )
        deduped_blocks = [
            {'active': row.get('active', True), 'fields': row.get('fields', {})}
            for row in deduped_rows
        ]

        if not deduped_blocks:
            continue

        template_path = ranked_paths[0][1]
        source_file_paths = [item[1] for item in ranked_paths]
        check_type_name, check_type_version, group_policy_name = _extract_check_type_meta(template_path)
        group_policy_inner = _norm_upper_token(_strip_outer_quotes_once(group_policy_name or ''))
        if not group_policy_inner or group_policy_inner == 'FOR GAP COMBINED':
            group_policy_name = type_label
        out_path = os.path.join(gap_root, f'{type_label}.audit')
        written.append(
            _write_gap_combined_audit(
                out_path,
                source_file_paths,
                deduped_blocks,
                check_type_name,
                check_type_version,
                group_policy_name,
                type_label,
                renumber=True,
            )
        )
        print(f'Gap merge summary for {type_label}: input_rows={len(merged_rows)} | deduped={dropped_dupes}')

    print(f'Staged {len(written)} combined type file(s) for gap analysis in {gap_root}')
    return gap_root


_PRODUCTION_REFERENCE_RE = re.compile(r'\b([A-Z]{2,3}-\d+(?:\(\d+\))?)\b', re.IGNORECASE)
_PRODUCTION_DESCRIPTION_RE = re.compile(r'^\s*description\s*:\s*"?(.*?)"?\s*$', re.IGNORECASE)
_PRODUCTION_REFERENCE_LINE_RE = re.compile(r'^\s*reference\s*:\s*"?(.*?)"?\s*$', re.IGNORECASE)
_PRODUCTION_TYPE_RE = re.compile(r'^\s*type\s*:\s*(.+?)\s*$', re.IGNORECASE)
_PRODUCTION_TIMESTAMP_RE = re.compile(r'_\d{8}(?=\.audit$)', re.IGNORECASE)


def _production_normalize_reference(reference):
    match = re.match(r'^([A-Z]{2,3})-0*(\d+)(?:\(0*(\d+)\))?$', str(reference).upper())
    if not match:
        return str(reference).upper()
    family, number, enhancement = match.groups()
    return f'{family}-{int(number)}' + (f'({int(enhancement)})' if enhancement else '')


def _production_audit_files(root):
    if not os.path.isdir(root):
        return []
    latest = {}
    for name in sorted(os.listdir(root), key=str.lower):
        if not name.lower().endswith('.audit') or name.upper().startswith('ALL_'):
            continue
        key = _PRODUCTION_TIMESTAMP_RE.sub('', name).lower()
        path = os.path.join(root, name)
        current = latest.get(key)
        if current is None or os.path.getmtime(path) >= os.path.getmtime(current):
            latest[key] = path
    return sorted(latest.values(), key=str.lower)


def _production_type_name(path, description):
    text = f'{os.path.basename(path)} {description}'.upper()
    for token, label in (
        ('AZURE', 'AZURE'), ('NX-OS', 'NETNXOS'), ('NXOS', 'NETNXOS'),
        ('PALO', 'NETPAFW'), ('PAFW', 'NETPAFW'), ('ASA', 'NETASA'), ('F5', 'NETF5'),
        ('IOS', 'IOS'), ('RHEL', 'RHEL'), ('SQL', 'SQL'), ('MSSRV', 'MSSRV'),
        ('MSWRK', 'MSWRK'), ('SERVER', 'WINSRV'), ('WINDOWS', 'WINWRK'),
    ):
        if token in text:
            return label
    return 'UNKNOWN'


def _production_records(root, source):
    records = []
    for path in _production_audit_files(root):
        try:
            lines = open(path, encoding='utf-8', errors='ignore').read().splitlines()
        except OSError:
            continue
        in_block = False
        inactive = False
        fields = {}
        for raw_line in lines + ['</custom_item>']:
            stripped = raw_line.lstrip()
            commented = stripped.startswith('#')
            line = re.sub(r'^\s*#\s*', '', raw_line).strip()
            if re.match(r'^<custom_item>\s*$', line, re.IGNORECASE):
                in_block, inactive, fields = True, commented, {}
                continue
            if in_block and re.match(r'^</custom_item>\s*$', line, re.IGNORECASE):
                description = fields.get('description', '').strip().strip('"')
                reference = fields.get('reference', '').strip().strip('"')
                tokens = sorted({_production_normalize_reference(token) for token in _PRODUCTION_REFERENCE_RE.findall(reference)})
                if description and tokens:
                    records.append({
                        'source': source,
                        'file': os.path.basename(path),
                        'type': _production_type_name(path, description),
                        'description': description,
                        'key': re.sub(r'\s+', ' ', description).upper(),
                        'reference': reference,
                        'tokens': tokens,
                        'inactive': inactive,
                    })
                in_block, fields = False, {}
                continue
            if not in_block:
                continue
            for field_name, pattern in (
                ('description', _PRODUCTION_DESCRIPTION_RE),
                ('reference', _PRODUCTION_REFERENCE_LINE_RE),
                ('check_type', _PRODUCTION_TYPE_RE),
            ):
                match = pattern.match(line)
                if match:
                    fields[field_name] = match.group(1).strip()
                    break
    return records


def _production_lines(records):
    return '\n'.join(sorted({f'[{record["file"]}] {record["description"]}' for record in records}))


def _production_autosize(worksheet):
    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
    for column in worksheet.columns:
        letter = column[0].column_letter
        width = max((max((len(part) for part in str(cell.value or '').splitlines()), default=0) for cell in column), default=10)
        worksheet.column_dimensions[letter].width = min(max(width + 2, 12), 72)
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')


def run_production_gap_analysis():
    """Compare current normalized audits to staged merged audits and write an Excel gap report."""
    if openpyxl is None or Workbook is None:
        raise RuntimeError('openpyxl is required for production gap analysis; pip install openpyxl')
    normalized = _production_records(PRODUCTION_NORMALIZED_ROOT, 'Normalized')
    merged = _production_records(PRODUCTION_GAP_ROOT, 'Merged')
    if not normalized and not merged:
        raise RuntimeError(f'No referenced custom items found in {PRODUCTION_NORMALIZED_ROOT} or {PRODUCTION_GAP_ROOT}')

    normalized_by_token = defaultdict(list)
    merged_by_token = defaultdict(list)
    for record in normalized:
        for token in record['tokens']:
            normalized_by_token[token].append(record)
    for record in merged:
        for token in record['tokens']:
            merged_by_token[token].append(record)

    rows = []
    for token in sorted(set(normalized_by_token) | set(merged_by_token)):
        expected = normalized_by_token[token]
        actual = merged_by_token[token]
        expected_active = [record for record in expected if not record['inactive']]
        actual_active = [record for record in actual if not record['inactive']]
        expected_keys = {record['key'] for record in expected_active}
        actual_keys = {record['key'] for record in actual_active}
        missing = [record for record in expected_active if record['key'] not in actual_keys]
        added = [record for record in actual_active if record['key'] not in expected_keys]
        inactive = [record for record in expected + actual if record['inactive']]
        if missing and added:
            status = 'Changed'
        elif missing:
            status = 'Missing in Merged'
        elif added:
            status = 'Added in Merged'
        elif inactive:
            status = 'Matched with Inactive Checks'
        else:
            status = 'Matched'
        coverage = round(100 * (len(expected_keys & actual_keys) / len(expected_keys)), 1) if expected_keys else None
        priority = len({record['key'] for record in missing}) * (3 if token.split('-', 1)[0] in {'AC', 'IA', 'AU', 'SC', 'SI'} else 1)
        rows.append({
            'reference': token, 'family': token.split('-', 1)[0], 'status': status,
            'coverage': coverage, 'priority': priority, 'normalized': expected,
            'merged': actual, 'missing': missing, 'added': added, 'inactive': inactive,
        })

    os.makedirs(PRODUCTION_GAP_OUTPUT_ROOT, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = 'Executive_Summary'
    summary.append(['Production NIST Reference Gap Analysis', datetime.now().isoformat(timespec='seconds')])
    summary.append([])
    summary.append(['Metric', 'Count'])
    summary.append(['Normalized checks', len(normalized)])
    summary.append(['Merged checks', len(merged)])
    summary.append(['NIST references', len(rows)])
    for status in ('Missing in Merged', 'Changed', 'Added in Merged', 'Matched with Inactive Checks', 'Matched'):
        summary.append([status, sum(row['status'] == status for row in rows)])
    _production_autosize(summary)

    all_sheet = workbook.create_sheet('All_Combined')
    headers = ['Reference', 'Family', 'Gap Status', 'Coverage %', 'Priority Score', 'Normalized Checks', 'Merged Checks', 'Missing in Merged', 'Added in Merged', 'Inactive Checks']
    all_sheet.append(headers)
    for row in rows:
        all_sheet.append([
            row['reference'], row['family'], row['status'], row['coverage'] if row['coverage'] is not None else '', row['priority'],
            _production_lines(row['normalized']), _production_lines(row['merged']), _production_lines(row['missing']),
            _production_lines(row['added']), _production_lines(row['inactive']),
        ])
    _production_autosize(all_sheet)

    priority_sheet = workbook.create_sheet('Priority_All')
    priority_sheet.append(['Rank', 'Priority Score', 'Reference', 'Family', 'Gap Status', 'Coverage %', 'Missing Checks'])
    actionable = [row for row in rows if row['status'] in {'Missing in Merged', 'Changed', 'Matched with Inactive Checks'}]
    for rank, row in enumerate(sorted(actionable, key=lambda item: (-item['priority'], item['reference'])), start=1):
        priority_sheet.append([rank, row['priority'], row['reference'], row['family'], row['status'], row['coverage'] if row['coverage'] is not None else '', _production_lines(row['missing'] or row['inactive'])])
    _production_autosize(priority_sheet)

    for title, statuses in (('Missing_In_Merged', {'Missing in Merged', 'Changed'}), ('Added_In_Merged', {'Added in Merged', 'Changed'})):
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        for row in rows:
            if row['status'] in statuses:
                sheet.append([row['reference'], row['family'], row['status'], row['coverage'] if row['coverage'] is not None else '', row['priority'], _production_lines(row['normalized']), _production_lines(row['merged']), _production_lines(row['missing']), _production_lines(row['added']), _production_lines(row['inactive'])])
        _production_autosize(sheet)

    output_path = os.path.join(PRODUCTION_GAP_OUTPUT_ROOT, f'Production_NIST_Reference_Gap_Analysis_{datetime.now():%y%m%d%H%M%S}.xlsx')
    workbook.save(output_path)
    print(f'Wrote production reference gap analysis: {output_path}')
    print(f'Normalized checks={len(normalized)} | Merged checks={len(merged)} | References={len(rows)}')
    return output_path


def _run_production_reference_gap_analysis():
    try:
        print('\nGenerating production reference gap analysis workbook...')
        run_production_gap_analysis()
        return True
    except Exception as exc:
        print(f'WARNING: Gap analysis generation failed: {exc}')
        return False


def _docker_desktop_executable_candidates():
    candidates = []
    for base in (
        os.environ.get('ProgramFiles', ''),
        os.environ.get('ProgramFiles(x86)', ''),
        os.environ.get('LocalAppData', ''),
    ):
        if not base:
            continue
        candidates.append(os.path.join(base, 'Docker', 'Docker', 'Docker Desktop.exe'))
    return candidates


def _start_docker_desktop_if_available():
    for candidate in _docker_desktop_executable_candidates():
        if not os.path.isfile(candidate):
            continue
        try:
            subprocess.Popen(
                [candidate],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                stdin=subprocess.DEVNULL,
            )
            return True, f'Started Docker Desktop: {candidate}'
        except Exception as exc:
            return False, f'Could not start Docker Desktop: {exc}'
    return False, 'Docker Desktop executable was not found'


def _wait_for_docker_daemon(timeout_seconds=120, poll_seconds=3):
    deadline = time.time() + timeout_seconds
    last_output = ''

    while time.time() < deadline:
        try:
            probe = subprocess.run(['docker', 'info'], capture_output=True, text=True)
        except FileNotFoundError:
            return False, 'docker is not installed or not on PATH'
        except Exception as exc:
            return False, str(exc)

        combined = (probe.stdout or '').strip()
        if probe.stderr:
            combined = (combined + '\n' + probe.stderr.strip()).strip() if combined else probe.stderr.strip()

        if probe.returncode == 0:
            return True, ''

        last_output = combined
        if any(
            marker in combined.lower()
            for marker in (
                'error during connect',
                'dockerdesktoplinuxengine',
                'the system cannot find the file specified',
                'is the docker daemon running',
                'cannot connect to the docker daemon',
            )
        ):
            time.sleep(poll_seconds)
            continue

        time.sleep(poll_seconds)

    return False, last_output or 'docker daemon did not become ready in time'


def _ensure_docker_ready():
    ready, message = _wait_for_docker_daemon(timeout_seconds=2, poll_seconds=1)
    if ready:
        return True, ''

    start_ok, start_message = _start_docker_desktop_if_available()
    if not start_ok:
        return False, message or start_message

    ready, message = _wait_for_docker_daemon(timeout_seconds=120, poll_seconds=3)
    if ready:
        return True, ''

    return False, message or start_message


def _run_check_audit_in_docker(audit_path):
    audit_path = os.path.abspath(audit_path)
    if not os.path.isfile(audit_path):
        return 1, f'Audit file does not exist: {audit_path}'

    ready, startup_message = _ensure_docker_ready()
    if not ready:
        return 127, startup_message

    mount_dir = os.path.dirname(audit_path).replace('\\', '/')
    docker_cmd = [
        'docker', 'run', '--rm',
        '-v', f'{mount_dir}:/audit',
        'tenable/audit-utils',
        'check_audit',
        f'/audit/{os.path.basename(audit_path)}',
    ]

    try:
        proc = subprocess.run(docker_cmd, capture_output=True, text=True)
    except FileNotFoundError:
        return 127, 'docker is not installed or not on PATH'
    except Exception as exc:
        return 1, str(exc)

    combined = (proc.stdout or '').strip()
    if proc.stderr:
        combined = (combined + '\n' + proc.stderr.strip()).strip() if combined else proc.stderr.strip()

    docker_unavailable_patterns = (
        'error during connect',
        'dockerdesktoplinuxengine',
        'the system cannot find the file specified',
        'is the docker daemon running',
        'cannot connect to the docker daemon',
    )
    combined_lower = combined.lower()
    if any(pat in combined_lower for pat in docker_unavailable_patterns):
        return 127, 'docker daemon is unavailable; skipping check_audit validation'

    return proc.returncode, combined


def _normalize_tenable_audit_wrapper(text, check_type_name=None, check_type_version=None, group_policy_name=None):
    lines = text.splitlines()
    normalized = []
    opened_check_type = False
    opened_group_policy = False

    for line in lines:
        stripped = line.strip()
        m_check = re.match(r'^<check_type\s*:\s*"([^"]+)"(?:\s+version\s*:\s*"([^"]+)")?\s*>$', stripped, flags=re.IGNORECASE)
        if m_check:
            if not opened_check_type:
                if check_type_name is None and check_type_version is None:
                    normalized.append(line.rstrip())
                else:
                    original_name = (m_check.group(1) or '').strip()
                    original_version = (m_check.group(2) or '').strip()
                    out_name = (check_type_name or original_name or 'Windows').strip()
                    out_version = original_version if check_type_version is None else str(check_type_version).strip()
                    if out_version:
                        normalized.append(f'<check_type:"{out_name}" version:"{out_version}">')
                    else:
                        normalized.append(f'<check_type:"{out_name}">')
                opened_check_type = True
            continue
        if re.match(r'^<group_policy(?:\s*:\s*"[^"]+")?\s*>$', stripped, flags=re.IGNORECASE):
            if not opened_group_policy:
                if group_policy_name is None:
                    normalized.append(line.rstrip())
                elif group_policy_name:
                    normalized.append(f'<group_policy:"{group_policy_name}">')
                else:
                    normalized.append('<group_policy>')
                opened_group_policy = True
            continue
        if re.match(r'^</group_policy>\s*$', stripped, flags=re.IGNORECASE):
            normalized.append('</group_policy>')
            continue
        if re.match(r'^</check_type>\s*$', stripped, flags=re.IGNORECASE):
            normalized.append('</check_type>')
            continue
        normalized.append(line.rstrip())

    return re.sub(r'\n{3,}', '\n\n', '\n'.join(normalized).rstrip() + '\n')


def _count_unescaped_double_quotes(text):
    count = 0
    escaped = False
    for ch in str(text or ''):
        if escaped:
            escaped = False
            continue
        if ch == '\\':
            escaped = True
            continue
        if ch == '"':
            count += 1
    return count


def _repair_value_data_expression(value):
    raw = str(value or '').strip()
    if not raw:
        return value

    normalized = raw.replace(r'\&\&', '&&').replace(r'\|\|', '||')

    # Numeric POLICY_DWORD selections such as `2 || 3` or `4 || 5` are
    # scanner-accepted as-is and must not be rewritten into regex form.
    if re.fullmatch(r'"?\d+"?\s*\|\|\s*"?\d+"?', normalized):
        return value

    semantic_boolean = _normalize_boolean_value_data_expression(normalized)
    if semantic_boolean:
        return f'"{semantic_boolean}"'

    return value


def _normalize_embedded_powershell_script(script_text):
    script = str(script_text or '')
    if not script:
        return script

    # Avoid nested raw double quotes inside audit-quoted powershell fields.
    script = script.replace('-join "`r`n"', '-join [Environment]::NewLine')
    script = script.replace('-join "`n"', '-join [Environment]::NewLine')

    # Tenable runtime evaluation requires pipeline output.
    # Write-Host writes to host stream and can produce POWERSHELL_NO_RESULT.
    script = re.sub(r'(?i)\bwrite-host\b', 'Write-Output', script)

    return script


def _repair_malformed_powershell_lines(text):
    changed = False
    out_lines = []
    for line in str(text or '').splitlines():
        match = re.match(r'^(\s*(?:powershell_args|powershell_script)\s*:\s*)(.+?)\s*$', line)
        if not match:
            out_lines.append(line)
            continue

        prefix = match.group(1)
        value = match.group(2).strip()
        if not (len(value) >= 2 and value[0] == '"' and value[-1] == '"'):
            out_lines.append(line)
            continue

        inner = value[1:-1]
        fixed_inner = _normalize_embedded_powershell_script(inner)
        fixed_value = f'"{fixed_inner}"'

        if fixed_value != value:
            changed = True
            out_lines.append(f'{prefix}{fixed_value}')
        else:
            out_lines.append(line)

    repaired = '\n'.join(out_lines).rstrip() + '\n'
    return repaired, changed


def _escape_unescaped_double_quotes(value):
    return re.sub(r'(?<!\\)"', r'\\"', str(value or ''))


AUDIT_QUOTED_STRING_FIELDS = {
    'cmd',
    'expect',
    'f5_command',
    'file_required',
    'interface_name',
    'item',
    'json_transform',
    'not_expect',
    'policy_arn',
    'powershell_args',
    'powershell_script',
    'regex',
    'request',
    'show_output',
    'shared_key',
    'sql_expect',
    'tmsh',
    'value_data',
    'wmi_attribute',
    'wmi_key',
    'wmi_namespace',
    'wmi_request',
    'where',
    'xsl_stmt',
}


def _render_audit_string_value(value):
    raw = _strip_outer_quotes_once(value)
    escaped = _escape_unescaped_double_quotes(raw)
    return f'"{escaped}"'


def _render_powershell_args_value(script_text, launcher_args='-NoProfile -ExecutionPolicy Bypass -Command'):
    script = _normalize_embedded_powershell_script(_strip_outer_quotes_once(script_text))
    escaped_script = script.replace("'", "''")
    if escaped_script:
        return f'"{launcher_args} \'{escaped_script}\'"'
    return f'"{launcher_args}"'


def _repair_malformed_quoted_field_lines(text, target_keys=None):
    changed = False
    out_lines = []
    # Azure and other non-Windows audits frequently carry jq/json snippets
    # in request/json_transform values that contain embedded quotes.
    # Include the standard quoted fields in auto-repair so malformed or
    # multiline quoted values can be normalized instead of aborting the
    # entire normalization run.
    keys = {k.lower() for k in (target_keys or {'description', 'info', 'solution', 'reference', 'see_also', 'value_type', 'value_data', 'expect', 'sql_expect', 'xsl_stmt', 'request', 'json_transform'})}

    lines = str(text or '').splitlines()
    i = 0
    while i < len(lines):
        line = lines[i]
        match = re.match(r'^(\s*([A-Za-z_][A-Za-z0-9_]*)\s*:\s*)(.+?)\s*$', line)
        if not match:
            out_lines.append(line)
            i += 1
            continue

        prefix = match.group(1)
        key = match.group(2).lower()
        value = match.group(3).strip()

        if key not in keys or not value.startswith('"'):
            out_lines.append(line)
            i += 1
            continue

        if value.endswith('"') and _count_unescaped_double_quotes(value) % 2 == 0:
            out_lines.append(line)
            i += 1
            continue

        collected = [value[1:] if value.startswith('"') else value]
        j = i + 1
        closed = value.endswith('"') and _count_unescaped_double_quotes(value) % 2 != 0
        while j < len(lines):
            segment = lines[j]
            if segment.endswith('"') and _count_unescaped_double_quotes(segment) % 2 != 0:
                collected.append(segment[:-1].strip())
                closed = True
                break
            collected.append(segment.strip())
            j += 1

        if not closed:
            inner = value[1:] if value.startswith('"') else value
            inner = inner[:-1] if inner.endswith('"') else inner
        else:
            inner = ' '.join(part for part in collected if part).strip()

        fixed_inner = _escape_unescaped_double_quotes(inner)
        if fixed_inner.endswith('\\'):
            # Prevent escaping the final quote of the rendered audit value.
            fixed_inner += '\\'
        fixed_value = f'"{fixed_inner}"'

        if fixed_value != value:
            changed = True
            out_lines.append(f'{prefix}{fixed_value}')
        else:
            out_lines.append(line)
        i = j + 1 if closed else i + 1

    repaired = '\n'.join(out_lines).rstrip() + '\n'
    return repaired, changed


def _repair_malformed_value_data_lines(text):
    changed = False
    out_lines = []
    for line in str(text or '').splitlines():
        match = re.match(r'^(\s*value_data\s*:\s*)(.+?)\s*$', line)
        if not match:
            out_lines.append(line)
            continue

        prefix = match.group(1)
        value = match.group(2)
        fixed = _repair_value_data_expression(value)
        if str(fixed) != str(value):
            changed = True
            out_lines.append(f'{prefix}{fixed}')
        else:
            out_lines.append(line)

    repaired = '\n'.join(out_lines).rstrip() + '\n'
    return repaired, changed


def _line_level_parse_errors(text):
    errors = []
    in_custom_item = False

    for line_no, line in enumerate(str(text or '').splitlines(), start=1):
        stripped = line.strip()

        if not stripped:
            continue

        if stripped.startswith('#'):
            continue

        if stripped.lower() == '<custom_item>':
            in_custom_item = True
            continue
        if stripped.lower() == '</custom_item>':
            in_custom_item = False
            continue

        if stripped.startswith('<') and stripped.endswith('>'):
            continue

        if not in_custom_item:
            continue

        kv = re.match(r'^\s*([A-Za-z_][A-Za-z0-9_]*)\s*:\s*(.+?)\s*$', line)
        if not kv:
            errors.append(f'line {line_no}: invalid custom_item field syntax')
            continue

        key = kv.group(1).lower()
        value = kv.group(2)
        key_upper = key.upper()

        value_stripped = value.strip()
        if key in REAL_KEYS and value_stripped.startswith('"'):
            if (not value_stripped.endswith('"')) or (_count_unescaped_double_quotes(value) % 2 != 0):
                errors.append(f'line {line_no}: unbalanced double quotes in {key}')

        if key in {'powershell_args', 'powershell_script'} and value_stripped.startswith('"') and value_stripped.endswith('"'):
            inner = value_stripped[1:-1]
            if inner and re.search(r'(?<!\\)"', inner):
                errors.append(f'line {line_no}: unescaped embedded double quote in {key}')

        if key == 'value_data' and key in REAL_KEYS:
            raw = value_stripped
            normalized = raw.replace(r'\&\&', '&&').replace(r'\|\|', '||')
            if ('&&' in normalized or '||' in normalized) and normalized.startswith('"') and normalized.endswith('"'):
                inner = normalized[1:-1]
                if re.search(r'(?<!\\)"', inner):
                    errors.append(f'line {line_no}: malformed value_data boolean expression')

    return errors


def _windows_expect_field_errors(text, platform_hint='UNKNOWN'):
    if not _is_windows_audit_platform(platform_hint):
        return []

    errors = []
    try:
        document, _, _ = _parse_document_for_platform(str(text or '').splitlines(), platform_hint)
    except Exception as exc:
        return [f'Could not parse document for Windows expect validation: {exc}']

    for idx, node in enumerate(document, start=1):
        if node.get('type') != 'custom_item':
            continue

        fields = node.get('fields', {})
        if _norm_upper_token(fields.get('type', '')) != 'AUDIT_POWERSHELL':
            continue
        if 'expect' not in fields:
            continue

        desc = _strip_outer_quotes_once(fields.get('description', ''))
        label = desc or f'custom_item #{idx}'
        errors.append(f'{label}: Windows AUDIT_POWERSHELL blocks must not emit expect')

    return errors


def _value_data_semantic_errors(text, platform_hint='UNKNOWN'):
    errors = []
    try:
        document, _, _ = _parse_document_for_platform(str(text or '').splitlines(), platform_hint)
    except Exception as exc:
        return [f'Could not parse document for semantic validation: {exc}']

    for idx, node in enumerate(document, start=1):
        if node.get('type') != 'custom_item':
            continue

        fields = node.get('fields', {})
        raw_value = str(fields.get('value_data', '') or '').strip()
        if not raw_value:
            continue

        check_type = _norm_upper_token(fields.get('check_type', ''))
        semantic_value = _normalize_boolean_value_data_expression(raw_value)
        candidate = semantic_value or _strip_outer_quotes_once(raw_value)

        # Many source audits use CHECK_REGEX with literal value_data paths.
        # Validate those as escaped literals to avoid false positives while
        # still compiling intentional regex patterns as-is.
        is_intentional_regex = bool(
            re.search(
                r'\^|\$|\[|\]|\(\?:|\(\?=|\(\?!|\|\||&&|\||\.\*|\.\+|\{\d+(,\d*)?\}',
                candidate,
            )
        )
        should_compile = bool(semantic_value or is_intentional_regex or check_type == 'CHECK_REGEX')
        if should_compile:
            pattern = candidate if (semantic_value or is_intentional_regex) else re.escape(candidate)
            try:
                re.compile(pattern)
            except re.error as exc:
                desc = _strip_outer_quotes_once(fields.get('description', ''))
                label = desc or f'custom_item #{idx}'
                errors.append(f'{label}: invalid value_data regex ({exc})')

    return errors


def _document_parse_errors(text, platform_hint='UNKNOWN'):
    try:
        _parse_document_for_platform(str(text or '').splitlines(), platform_hint)
    except Exception as exc:
        return [f'Could not fully parse document: {exc}']
    return []


def _repair_and_scan_audit_text(text, *, check_type_name=None, check_type_version=None, group_policy_name=None, platform_hint='UNKNOWN'):
    current = str(text or '')
    changed = False

    _, profile = _platform_parse_profile(platform_hint)
    repaired_text = _apply_parse_repair_pipeline(current, profile.get('repair_pipeline', ()))
    if repaired_text != current:
        current = repaired_text
        changed = True

    normalized = _normalize_tenable_audit_wrapper(
        current,
        check_type_name=check_type_name,
        check_type_version=check_type_version,
        group_policy_name=group_policy_name,
    )
    if normalized != current:
        current = normalized
        changed = True

    parse_errors = _line_level_parse_errors(current)
    if parse_errors:
        repaired_text = _apply_parse_repair_pipeline(current, profile.get('repair_pipeline', ()))
        if repaired_text != current:
            current = repaired_text
            changed = True
            parse_errors = _line_level_parse_errors(current)
    if parse_errors:
        return current, changed, 'line-level parse validation failed', parse_errors

    document_errors = _document_parse_errors(current, platform_hint=platform_hint)
    if document_errors:
        return current, changed, 'document parse validation failed', document_errors

    windows_expect_errors = _windows_expect_field_errors(current, platform_hint=platform_hint)
    if windows_expect_errors:
        return current, changed, 'windows audit structural validation failed', windows_expect_errors

    semantic_errors = _value_data_semantic_errors(current, platform_hint=platform_hint)
    if semantic_errors:
        return current, changed, 'value_data semantic validation failed', semantic_errors

    return current, changed, '', []


def validate_and_repair_audit_file(audit_path, *, check_type_name=None, check_type_version=None, group_policy_name=None, platform_hint=None):
    platform_for_parse = platform_hint or determine_platform_from_filename(audit_path)
    try:
        with open(audit_path, encoding='utf-8', errors='replace') as fh:
            current = _strip_bom_prefix(fh.read())
    except Exception:
        _record_validation_result(audit_path, 'failed', 'Could not read file for validation.')
        return False

    current, changed, error_label, scan_errors = _repair_and_scan_audit_text(
        current,
        check_type_name=check_type_name,
        check_type_version=check_type_version,
        group_policy_name=group_policy_name,
        platform_hint=platform_for_parse,
    )

    if changed:
        current = _strip_bom_prefix(current)
        _write_audit_text(audit_path, current)

    if scan_errors:
        detail = '\n'.join(scan_errors[:25])
        print(f'{error_label} for {audit_path}')
        print(detail)
        _record_validation_result(audit_path, 'failed', detail)
        return False

    code, output = _run_check_audit_in_docker(audit_path)
    if code == 127:
        print(f'Skipping check_audit for {audit_path}: {output}')
        _record_validation_result(audit_path, 'skipped', output)
        return False
    if code == 0:
        _record_validation_result(audit_path, 'passed', '')
        return True

    if output:
        print(f'check_audit failed for {audit_path}')
        print(output)

    repaired, changed, error_label, scan_errors = _repair_and_scan_audit_text(
        current,
        check_type_name=check_type_name,
        check_type_version=check_type_version,
        group_policy_name=group_policy_name,
        platform_hint=platform_for_parse,
    )

    if changed:
        _write_audit_text(audit_path, repaired)
        if scan_errors:
            detail = '\n'.join(scan_errors[:25])
            print(f'{error_label} for {audit_path}')
            print(detail)
            _record_validation_result(audit_path, 'failed', detail)
            return False
        code, output = _run_check_audit_in_docker(audit_path)
        if code == 0:
            print(f'check_audit passed after repair: {audit_path}')
            _record_validation_result(audit_path, 'passed', '')
            return True
        if output:
            print(f'check_audit still failed after repair for {audit_path}')
            print(output)

    _record_validation_result(audit_path, 'failed', output)
    return False


def _parse_document_for_platform(lines, platform_hint, source_path=''):
    platform_key, profile = _platform_parse_profile(platform_hint)
    strict_mode_label = 'windows-strict' if platform_key in {'MSSRV', 'MSWRK'} else f'{platform_key.lower()}-strict'
    lines = _sanitize_audit_lines(lines)
    source_text = ''.join(lines)

    try:
        document = parse_document(lines)
        return document, strict_mode_label, ''
    except Exception as strict_exc:
        if not profile.get('allow_repair_fallback', False):
            raise

        repaired_text = _apply_parse_repair_pipeline(source_text, profile.get('repair_pipeline', ()))
        repaired_lines = repaired_text.splitlines()

        document = parse_document(repaired_lines)
        note = f'strict parse failed; {platform_key} repaired fallback used ({strict_exc})'
        if source_path:
            print(f'WARNING: {source_path}: {note}')
        else:
            print(f'WARNING: {note}')
        repaired_mode_label = 'windows-repaired' if platform_key in {'MSSRV', 'MSWRK'} else f'{platform_key.lower()}-repaired'
        return document, repaired_mode_label, note


def _record_validation_result(audit_path, status, details=''):
    if not audit_path:
        return

    key = os.path.abspath(audit_path)
    existing = VALIDATION_RESULTS.get(key)
    rank = {'passed': 0, 'skipped': 1, 'failed': 2}
    new_rank = rank.get(status, 0)
    old_rank = rank.get((existing or {}).get('status', 'passed'), 0)

    if existing is None or new_rank >= old_rank:
        VALIDATION_RESULTS[key] = {
            'status': status,
            'details': (details or '').strip(),
        }


def _reset_validation_summary():
    VALIDATION_RESULTS.clear()


def _print_validation_summary():
    if not VALIDATION_RESULTS:
        print('\nValidation summary: no check_audit validations were executed in this run.')
        return

    totals = {'passed': 0, 'failed': 0, 'skipped': 0}
    for rec in VALIDATION_RESULTS.values():
        status = rec.get('status', 'passed')
        if status not in totals:
            continue
        totals[status] += 1

    print('\nValidation summary')
    print(f"  Passed : {totals['passed']}")
    print(f"  Failed : {totals['failed']}")
    print(f"  Skipped: {totals['skipped']}")

    if totals['failed']:
        print('\nFailed .audit files:')
        for path, rec in VALIDATION_RESULTS.items():
            if rec.get('status') != 'failed':
                continue
            print(f'  - {path}')
            details = rec.get('details', '')
            if details:
                first_line = details.splitlines()[0].strip()
                if first_line:
                    print(f'      reason: {first_line}')


def _extract_group_policy_name(audit_path):
    if not audit_path or not os.path.isfile(audit_path):
        return ''
    try:
        with open(audit_path, encoding='utf-8', errors='ignore') as fh:
            text = _strip_bom_prefix(fh.read(20000))
    except Exception:
        return ''

    m = re.search(r'<group_policy\s*:\s*"([^"]+)"\s*>', text, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip()
    m = re.search(r'<group_policy\s*:\s*"([^"]+)"\s*version\s*:\s*"([^"]+)"\s*>', text, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return ''


def _extract_check_type_meta(audit_path):
    if not audit_path or not os.path.isfile(audit_path):
        return '', '', ''
    try:
        with open(audit_path, encoding='utf-8', errors='ignore') as fh:
            text = _strip_bom_prefix(fh.read(20000))
    except Exception:
        return '', '', ''

    check_type_name = ''
    check_type_version = ''
    group_policy_name = ''

    m_ct = re.search(r'<check_type\s*:\s*"([^"]+)"(?:\s+version\s*:\s*"([^"]+)")?\s*>', text, flags=re.IGNORECASE)
    if m_ct:
        check_type_name = (m_ct.group(1) or '').strip()
        check_type_version = (m_ct.group(2) or '').strip()

    m_gp = re.search(r'<group_policy\s*:\s*"([^"]+)"\s*>', text, flags=re.IGNORECASE)
    if m_gp:
        group_policy_name = (m_gp.group(1) or '').strip()

    return check_type_name, check_type_version, group_policy_name


def _run_vscode_command(args):
    """Best-effort wrapper around VS Code CLI; never raises to callers."""
    code_exe = shutil.which('code')
    if not code_exe:
        return False, 'VS Code CLI not found in PATH.'

    try:
        proc = subprocess.run([code_exe] + args, capture_output=True, text=True)
    except Exception as exc:
        return False, str(exc)

    output = f"{proc.stdout or ''}\n{proc.stderr or ''}".strip()
    if proc.returncode != 0:
        return False, output or f'code exited with status {proc.returncode}'
    return True, output


def ensure_vscode_auditlang_extension(install_if_missing=False):
    ok, out = _run_vscode_command(['--list-extensions'])
    if not ok:
        print(f'VS Code extension check skipped: {out}')
        return False

    installed = any(
        line.strip().lower() == 'tenable.vscode-auditlang'
        for line in out.splitlines()
    )
    if installed:
        return True

    if not install_if_missing:
        print('Tenable AuditLang extension is not installed. Install with: code --install-extension Tenable.vscode-auditlang')
        return False

    ok, err = _run_vscode_command(['--install-extension', 'Tenable.vscode-auditlang'])
    if ok:
        print('Installed VS Code extension: Tenable.vscode-auditlang')
        return True

    print(f'Could not install Tenable AuditLang extension: {err}')
    return False


def open_audit_in_vscode_for_lint(audit_path):
    ok, err = _run_vscode_command(['--reuse-window', audit_path])
    if not ok:
        print(f'Could not open in VS Code: {audit_path} ({err})')
        return False

    print(f'Opened in VS Code for AuditLang diagnostics: {audit_path}')
    print('Use View -> Problems to review parse errors with line numbers.')
    return True


def _derive_group_policy_name_from_master(master_unique_path, platform):
    rows = _iter_workbook_rows(master_unique_path)
    source_counts = {}
    for rec in rows:
        plat = _normalize_platform_sheet_name(rec.get('sheet', ''))
        if plat != platform:
            continue
        src = str(rec.get('source_file', '')).strip()
        if not src:
            continue
        source_counts[src] = source_counts.get(src, 0) + 1

    if not source_counts:
        return ''

    best_source = max(source_counts.items(), key=lambda x: x[1])[0]
    source_path = _resolve_source_path_from_catalog(master_unique_path, best_source)
    return _extract_group_policy_name(source_path)


def _derive_warning_description_from_master(master_unique_path, platform):
    rows = _iter_workbook_rows(master_unique_path)
    source_counts = {}
    for rec in rows:
        plat = _normalize_platform_sheet_name(rec.get('sheet', ''))
        if plat != platform:
            continue
        src = str(rec.get('source_file', '')).strip()
        if not src:
            continue
        source_counts[src] = source_counts.get(src, 0) + 1

    if not source_counts:
        return f'"TARGET OS DOES NOT MATCH BASELINE - {platform}"'

    best_source = max(source_counts.items(), key=lambda x: x[1])[0]
    source_path = _resolve_source_path_from_catalog(master_unique_path, best_source)
    source_name = os.path.basename(source_path) if source_path else os.path.basename(best_source)
    group_policy = _extract_group_policy_name(source_path)

    if group_policy:
        return f'"{source_name} from {group_policy}"'
    return f'"{source_name} from {platform}"'


def _apply_mssrv_pass_target_overrides(text):
    def _rewrite_block(match):
        block = match.group(0)

        desc_match = re.search(r'^\s*description\s*:\s*"([^"]+)"', block, flags=re.MULTILINE)
        desc_inner = desc_match.group(1) if desc_match else ''
        desc_upper = _norm_upper_token(desc_inner)
        id_match = re.match(r'^(\d+\.\d+)\s*-', desc_inner)
        control_id = id_match.group(1) if id_match else ''

        ps_match = re.search(r'^\s*powershell_args\s*:\s*"([\s\S]*?)"\s*$', block, flags=re.MULTILINE)
        ps_inner = ps_match.group(1) if ps_match else ''
        ps_upper = _norm_upper_token(ps_inner)

        value_match = re.search(r'^(\s*value_data\s*:\s*)(.+?)\s*$', block, flags=re.MULTILINE)
        if not value_match:
            return block

        value_prefix = value_match.group(1)
        value_text = value_match.group(2).strip()
        value_inner = _strip_outer_quotes_once(value_text)

        replacement = value_text

        if 'USER_RIGHTS' in ps_upper and 'NO ONE' not in desc_upper and value_inner and 'NO_MEMBERS' not in value_inner and 'DENY LOG ON LOCALLY TO INCLUDE GUESTS' not in desc_upper:
            replacement = f'"{_allow_no_members_alternative(value_inner)}"'
        elif 'USER_RIGHTS' in ps_upper and 'NO ONE' not in desc_upper and value_inner and 'NO_MEMBERS' in value_inner and 'DENY LOG ON LOCALLY TO INCLUDE GUESTS' not in desc_upper:
            replacement = f'"{_normalize_no_members_token_regex(value_inner)}"'

        if replacement != value_text:
            block = re.sub(
                r'^\s*value_data\s*:\s*.+?\s*$',
                value_prefix + replacement,
                block,
                count=1,
                flags=re.MULTILINE,
            )

        return block

    return re.sub(r'<custom_item>[\s\S]*?</custom_item>', _rewrite_block, str(text or ''), flags=re.IGNORECASE)


def build_merged_master_audits(master_unique_path, other_unique_path, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    master_index = _load_unique_catalog_index(master_unique_path)
    other_index = _load_unique_catalog_index(other_unique_path)

    baseline_mssrv_index, baseline_mssrv_dupes = _load_scan_description_index(BASELINE_MSSRV_CSV_PATH, plugin_id=BASELINE_PLUGIN_ID_FILTER)
    baseline_mssrv_keys = list(baseline_mssrv_index.keys())
    if baseline_mssrv_keys:
        print(
            f'Loaded MSSRV baseline description keys: unique={len(baseline_mssrv_keys)} | '
            f'duplicate_rows_skipped={baseline_mssrv_dupes}'
        )
    else:
        print(f'WARNING: MSSRV baseline CSV not found or empty: {BASELINE_MSSRV_CSV_PATH}')

    written = []
    for platform in ('MSWRK', 'MSSRV'):
        warning_description = _derive_warning_description_from_master(master_unique_path, platform)
        group_policy_name = _derive_group_policy_name_from_master(master_unique_path, platform)
        master_items = master_index.get(platform, OrderedDict())
        other_items = other_index.get(platform, OrderedDict())

        all_keys = set(master_items.keys()) | set(other_items.keys())
        merged_rows = []
        for key in all_keys:
            master_source = master_items.get(key)
            other_source = other_items.get(key)
            source = master_source or other_source
            if source is None:
                continue
            master_fields = dict(master_source.get('fields', {})) if master_source else {}
            other_fields = dict(other_source.get('fields', {})) if other_source else {}
            preferred_fields = master_fields if master_source else other_fields
            fallback_fields = other_fields if master_source else master_fields
            fields = _merge_control_fields(preferred_fields, fallback_fields)
            if _is_target_os_applicability_control(fields):
                continue
            if _is_domain_controller_probe_control(fields):
                continue
            if _is_rapid7_insight_agent_control(fields):
                continue
            desc = _strip_outer_quotes_once(fields.get('description', ''))
            merged_rows.append({
                'key': key,
                'active': (key in master_items),
                'fields': fields,
                'description_sort': _norm_upper_token(desc),
            })

        merged_rows.sort(key=lambda r: (r['description_sort'], r['key']))
        merged_rows, dropped_dupes = _dedupe_merged_rows(merged_rows)
        prefilter_merged_rows = list(merged_rows)

        if platform == 'MSSRV' and baseline_mssrv_keys:
            merged_rows, missing_baseline_keys = _filter_mssrv_rows_to_baseline_descriptions(
                merged_rows,
                baseline_mssrv_keys,
            )
            selected_keys = {row.get('key') for row in merged_rows}
            commented_rows = [
                row for row in prefilter_merged_rows
                if (not row.get('active')) and row.get('key') not in selected_keys
            ]
            if commented_rows:
                merged_rows.extend(commented_rows)
                merged_rows.sort(key=lambda r: (r['description_sort'], r['key']))
            if missing_baseline_keys:
                known_meta_missing, actionable_missing = _partition_missing_baseline_keys(missing_baseline_keys)
                if actionable_missing:
                    print(
                        f'WARNING: MSSRV baseline descriptions missing from merged source set: '
                        f'{len(actionable_missing)}'
                        + (
                            f' (suppressed known meta patterns: {len(known_meta_missing)})'
                            if known_meta_missing else ''
                        )
                    )
                else:
                    print(
                        f'INFO: MSSRV baseline missing descriptions are known meta/non-control patterns: '
                        f'{len(known_meta_missing)} (warning suppressed)'
                    )
            print(
                f'MSSRV baseline filter applied: selected={len(merged_rows)} of '
                f'baseline_unique={len(baseline_mssrv_keys)}'
            )

        common_rows = []
        ms_only_rows = []
        dc_only_rows = []
        for row in merged_rows:
            fields = row['fields']
            # Keep this validation in both role-scoped sections.
            if _is_adcs_authorized_validation_control(fields):
                ms_only_rows.append(row)
                dc_only_rows.append(row)
            elif _is_dc_only_control(fields):
                dc_only_rows.append(row)
            elif _is_ms_only_control(fields):
                ms_only_rows.append(row)
            else:
                common_rows.append(row)

        common_active_rows = [r for r in common_rows if r['active']]
        common_inactive_rows = [r for r in common_rows if not r['active']]
        ms_only_active_rows = [r for r in ms_only_rows if r['active']]
        ms_only_inactive_rows = [r for r in ms_only_rows if not r['active']]
        dc_active_rows = [r for r in dc_only_rows if r['active']]
        dc_inactive_rows = [r for r in dc_only_rows if not r['active']]

        out_lines = []
        out_lines.append('# generated_by: merge-master-audits-v2')
        out_lines.append('<check_type:"Windows" version:"2">')
        effective_group_policy_name = group_policy_name or platform
        out_lines.append(f'<group_policy:"{effective_group_policy_name}">')
        out_lines.append('')
        ordinal = 0
        if platform == 'MSSRV':
            preconditional_rows = [
                row for row in merged_rows
                if _is_preconditional_inventory_control(row.get('fields', {}))
            ]
            if preconditional_rows:
                preconditional_keys = {row['key'] for row in preconditional_rows}
                common_active_rows = [r for r in common_active_rows if r['key'] not in preconditional_keys]
                common_inactive_rows = [r for r in common_inactive_rows if r['key'] not in preconditional_keys]
                ms_only_active_rows = [r for r in ms_only_active_rows if r['key'] not in preconditional_keys]
                ms_only_inactive_rows = [r for r in ms_only_inactive_rows if r['key'] not in preconditional_keys]
                dc_active_rows = [r for r in dc_active_rows if r['key'] not in preconditional_keys]
                dc_inactive_rows = [r for r in dc_inactive_rows if r['key'] not in preconditional_keys]

                for row in preconditional_rows:
                    pre_fields = dict(row.get('fields', {}))
                    pre_desc = _strip_outer_quotes_once(pre_fields.get('description', ''))
                    pre_desc = re.sub(r'^\d+\.\d+\s*-\s*(MSWRK|MSSRV)\s*-\s*', '', pre_desc, flags=re.IGNORECASE)
                    pre_desc = re.sub(r'^\d+\.\d+\s*-\s*', '', pre_desc)
                    pre_fields['description'] = f'"{pre_desc}"'

                    # Keep these informational controls ahead of any conditionals,
                    # but exclude them from the numbered control sequence.
                    block_lines = _render_custom_item_block(pre_fields, platform, ordinal, apply_prefix=False)
                    if row.get('active'):
                        out_lines.extend(block_lines)
                    else:
                        out_lines.extend(_comment_block_lines(block_lines))
                    out_lines.append('')

            ordinal = _append_condition_block(
                out_lines,
                platform=platform,
                active_rows=common_active_rows,
                inactive_rows=common_inactive_rows,
                warning_description=warning_description,
                ordinal_start=ordinal,
                dc_only=False,
                include_target_gate=True,
            )

            if ms_only_active_rows or ms_only_inactive_rows:
                ordinal = _append_domain_role_block(
                    out_lines,
                    title='MEMBER SERVER',
                    domain_roles=(2, 3),
                    description='Check if server is Member Server',
                    active_rows=ms_only_active_rows,
                    inactive_rows=ms_only_inactive_rows,
                    warning_description='"WARNING - TARGET OS DOES NOT MATCH BASELINE - MSSRV.MS"',
                    ordinal_start=ordinal,
                    passed_suffix='MSSRV.MS',
                )

            if dc_active_rows or dc_inactive_rows:
                ordinal = _append_domain_role_block(
                    out_lines,
                    title='DOMAIN CONTROLLER',
                    domain_roles=(4, 5),
                    description='Check if server is Domain Controller',
                    active_rows=dc_active_rows,
                    inactive_rows=dc_inactive_rows,
                    warning_description='"WARNING - TARGET OS DOES NOT MATCH BASELINE - MSSRV.DC"',
                    ordinal_start=ordinal,
                    passed_suffix='MSSRV.DC',
                )
        else:
            ordinal = _append_condition_block(
                out_lines,
                platform=platform,
                active_rows=[r for r in merged_rows if r['active']],
                inactive_rows=[r for r in merged_rows if not r['active']],
                warning_description=warning_description,
                ordinal_start=ordinal,
                dc_only=False,
                include_target_gate=True,
            )

        out_lines.append('')
        out_lines.append('</group_policy>')
        out_lines.append('</check_type>')

        out_text = '\n'.join(out_lines).rstrip() + '\n'
        out_text = re.sub(
            r'"0\.9999\s*-\s*(MSWRK|MSSRV)\s*-\s*(Windows(?: Workstation| Server) is installed)"',
            r'"\2"',
            out_text,
        )
        out_text = re.sub(
            r'("1\.\d{4}\s*-\s*(MSWRK|MSSRV)\s*-\s*)-\s*\2\s*-\s*',
            r'\1',
            out_text,
        )

        if platform == 'MSSRV':
            out_text = _apply_mssrv_pass_target_overrides(out_text)

        out_text = _strip_bom_prefix(out_text)
        out_path = _timestamped_output_path(os.path.join(output_dir, f'Merged_{platform}.audit'))
        _assert_no_encoding_markers(out_text, out_path)

        with open(out_path, 'w', encoding='utf-8') as fh:
            fh.write(out_text)

        validate_and_repair_audit_file(
            out_path,
            check_type_name='Windows',
            check_type_version='2',
            group_policy_name=effective_group_policy_name,
            platform_hint=platform,
        )

        active_count = len([r for r in merged_rows if r['active']])
        inactive_count = len([r for r in merged_rows if not r['active']])
        print(
            f'Wrote merged audit: {out_path} | active={active_count} | '
            f'inactive_commented={inactive_count} | deduped={dropped_dupes}'
        )
        written.append(out_path)

    return written
# =============================================================================
# PASS 1 â€” VARIABLE EXTRACTION
# =============================================================================

# WARNING: extract_variables and parse_document are imported from pysc_block_parser.
# They are intentionally left out of this module to prevent duplicated parsing logic.


# =============================================================================
# PASS 2 â€” PARSE STRUCTURE
# =============================================================================

# WARNING: parse_document is imported from pysc_block_parser.

# =============================================================================
# PASS 3 â€” TRANSFORM & EMIT (FIXED & STABLE)
# =============================================================================

def emit(document, variables, platform_hint=''):
    output = []
    rendered_blocks = []
    all_keys = []

    def _emit_description_platform_label(raw_desc):
        plat = _norm_upper_token(platform_hint)
        if not plat or plat == 'UNKNOWN':
            plat = _norm_upper_token(detect_platform(raw_desc))
        if plat == 'MSAZ':
            return 'AZURE'
        if not plat or plat == 'UNKNOWN':
            return 'CONTROL'
        return plat

    desc_counter = 0
    unknown_keys = set()

    # Build rendered blocks
    for node in document:
        if node["type"] not in ("custom_item", "report-warning", "report-passed"):
            continue

        if node["type"] == "report-passed":
            pairs = []
            # if "solution" in node["fields"]:
            #     print("\nSOLUTION VALUE:")
            #     print(repr(node["fields"]["solution"]))

            for k, v in node["fields"].items():
                if k in IGNORED_KEYS:
                    continue
                if k not in REAL_KEYS:
                    unknown_keys.add(k)
                    continue
                if k == "see_also":
                    pairs.append((k, f"\"{SEE_ALSO_REPLACEMENT}\""))
                elif k == "info":
                    info = normalize_info(v)
                    if info:
                        pairs.append((k, info))
                elif k == "reference":
                    raw_ref = resolve_variables(v, variables)
                    pairs.append((k, normalize_reference_or_passthrough(raw_ref)))
                elif k == "solution":
                    sol = normalize_solution(resolve_variables(v, variables))
                    if sol:
                        pairs.append((k, sol))
                elif k in AUDIT_QUOTED_STRING_FIELDS:
                    pairs.append((k, _render_audit_string_value(resolve_variables(v, variables))))
                else:
                    pairs.append((k, resolve_variables(v, variables)))
            pairs = enforce_solution_reference_adjacency(pairs)
            rendered_blocks.append(pairs)
            all_keys.extend(k for k, _ in pairs)
            continue

        pairs = []

        custom_fields = _convert_policy_item_to_audit_powershell(node["fields"])
        if _should_force_convert_non_powershell(platform_hint, node["fields"]):
            custom_fields = _convert_remaining_item_to_audit_powershell(custom_fields)
        custom_fields = _ensure_audit_powershell_metadata_fields(custom_fields)
        custom_fields = _finalize_audit_powershell_fields_for_emit(custom_fields, platform_hint)

        for k, v in custom_fields.items():
            if k in IGNORED_KEYS:
                continue
            if k not in REAL_KEYS:
                unknown_keys.add(k)
                continue

            if k == "see_also":
                pairs.append((k, f"\"{SEE_ALSO_REPLACEMENT}\""))

            elif k == "info":
                info = normalize_info(v)
                if info:
                    pairs.append((k, info))

            elif k == "reference":
                raw_ref = resolve_variables(v, variables)
                pairs.append((k, normalize_reference_or_passthrough(raw_ref)))

            elif k == "solution":
                sol = normalize_solution(resolve_variables(v, variables))
                if sol:
                    pairs.append((k, sol))

            elif k in AUDIT_QUOTED_STRING_FIELDS:
                pairs.append((k, _render_audit_string_value(resolve_variables(v, variables))))

            elif k == "description":
                desc_value = resolve_variables(v, variables)
                platform_label = _emit_description_platform_label(desc_value)
                desc = _prefixed_description(desc_value, platform_label, desc_counter)
                desc_counter += 1
                pairs.append((k, desc))

            else:
                pairs.append((k, resolve_variables(v, variables)))

        pairs = order_custom_item_pairs(pairs)
        rendered_blocks.append(pairs)
        all_keys.extend(k for k, _ in pairs)

    width = max(len(k) for k in all_keys) if all_keys else 0
    block_idx = 0

    # Emit final output
    for node in document:
        if node["type"] == "text":
            if not node["text"].lstrip().startswith("#"):
                output.append(resolve_variables(node["text"], variables))

        elif node["type"] == "report-passed":
            output.append('<report type:"PASSED">')
            for k, v in rendered_blocks[block_idx]:
                output.append(f"  {k.ljust(width)} : {v}")
            output.append("</report>")
            block_idx += 1
            continue

        elif node["type"] == "report-warning":
            output.append('<report type:"WARNING">')
            for k, v in rendered_blocks[block_idx]:
                output.append(f"  {k.ljust(width)} : {v}")
            output.append("</report>")
            block_idx += 1

        elif node["type"] == "custom_item":
            output.append("<custom_item>")
            for k, v in rendered_blocks[block_idx]:
                output.append(f"  {k.ljust(width)} : {v}")
            output.append("</custom_item>")
            block_idx += 1

    return output, unknown_keys

# =============================================================================
# MAIN
# =============================================================================

def _persist_key(key, set_name):
    """Add key to the named set in this script file for future runs."""
    script = os.path.abspath(__file__)
    with open(script, encoding="utf-8") as f:
        content = f.read()
    # Find the set block and insert the new key before the closing }
    pattern = rf'({set_name}\s*=\s*\{{)(.*?)(\}})'
    m = re.search(pattern, content, flags=re.DOTALL)
    if not m:
        print(f"  Could not find {set_name} â€” add '{key}' manually.")
        return
    prefix, body, closing = m.group(1), m.group(2), m.group(3)
    # Ensure trailing comma on last entry
    stripped = body.rstrip()
    if stripped and not stripped.endswith(','):
        body = stripped + ',\n'
    else:
        body = body.rstrip('\n') + '\n'
    new_block = f'{prefix}{body}    "{key}",\n{closing}'
    new_content = content[:m.start()] + new_block + content[m.end():]
    with open(script, "w", encoding="utf-8") as f:
        f.write(new_content)
    print(f"  '{key}' added to {set_name} in {os.path.basename(script)}.")


# =============================================================================
# MAIN
# =============================================================================

def _persist_key(key, set_name):
    """Add key to the named set in this script file for future runs."""
    script = os.path.abspath(__file__)
    with open(script, encoding="utf-8") as f:
        content = f.read()
    # Find the set block and insert the new key before the closing }
    pattern = rf'({set_name}\s*=\s*\{{)(.*?)(\}})'
    m = re.search(pattern, content, flags=re.DOTALL)
    if not m:
        print(f"  Could not find {set_name} â€” add '{key}' manually.")
        return
    prefix, body, closing = m.group(1), m.group(2), m.group(3)
    # Ensure trailing comma on last entry
    stripped = body.rstrip()
    if stripped and not stripped.endswith(','):
        body = stripped + ',\n'
    else:
        body = body.rstrip('\n') + '\n'
    new_block = f'{prefix}{body}    "{key}",\n{closing}'
    new_content = content[:m.start()] + new_block + content[m.end():]
    with open(script, "w", encoding="utf-8") as f:
        f.write(new_content)
    print(f"  '{key}' added to {set_name} in {os.path.basename(script)}.")

def process_file(infile, open_in_vscode=False, strict_mode=False):
    if not os.path.isfile(infile):
        print(f"ERROR: Input file does not exist: {infile}")
        return False

    base = os.path.splitext(os.path.basename(infile))[0]

    input_folder = os.path.dirname(infile)
    normalized_folder = os.path.join(input_folder, "Normalized")
    os.makedirs(normalized_folder, exist_ok=True)

    outfile = _timestamped_output_path(os.path.join(
        normalized_folder,
        f"{base}.audit"
    ))

    with open(infile, encoding="utf-8") as f:
        lines = _sanitize_audit_lines(f.readlines())

    source_check_type_name, source_check_type_version, source_group_policy_name = _extract_check_type_meta(infile)
    platform_hint = determine_platform_from_filename(infile)

    variables = extract_variables(lines)
    try:
        document, parse_mode, parse_note = _parse_document_for_platform(lines, platform_hint, infile)
    except Exception as exc:
        detail = f'parse failed for platform {platform_hint}: {exc}'
        print(f'ERROR: {detail}')
        _record_validation_result(infile, 'failed', detail)
        return False

    output, unknown_keys = emit(document, variables, platform_hint)
    output = _cleanup_normalized_output_lines(output)

    preflight_text, preflight_changed, preflight_error_label, preflight_errors = _repair_and_scan_audit_text(
        "\n".join(output) + "\n",
        check_type_name=source_check_type_name or None,
        check_type_version=source_check_type_version or None,
        group_policy_name=source_group_policy_name or None,
        platform_hint=platform_hint,
    )
    if preflight_changed:
        output = preflight_text.rstrip("\n").splitlines()
    if preflight_errors:
        print(f"ERROR: {preflight_error_label} for generated output: {infile}")
        for err in preflight_errors[:25]:
            print(err)
        _record_validation_result(infile, 'failed', '\n'.join(preflight_errors[:25]))
        return False

    node_type_counts = {}
    node_rows = []
    for idx, node in enumerate(document, start=1):
        node_type = node.get('type', 'unknown')
        node_type_counts[node_type] = node_type_counts.get(node_type, 0) + 1
        details = {}
        if node_type in ('custom_item', 'report-warning', 'report-passed'):
            details['description'] = node.get('fields', {}).get('description', '')
            details['field_count'] = len(node.get('fields', {}))
        elif node_type == 'text':
            details['text'] = node.get('text', '')[:200]
        node_rows.append({
            'node_index': idx,
            'node_type': node_type,
            'details': details,
        })

    safe_text = _strip_bom_prefix("\n".join(output) + "\n")
    _assert_no_encoding_markers(safe_text, outfile)
    _write_audit_text(outfile, safe_text)

    validate_and_repair_audit_file(
        outfile,
        check_type_name=source_check_type_name or None,
        check_type_version=source_check_type_version or None,
        group_policy_name=source_group_policy_name or None,
        platform_hint=platform_hint,
    )

    print("\nNormalized audit written to:")
    print(f"  {outfile}")

    if open_in_vscode:
        open_audit_in_vscode_for_lint(outfile)

    folder_key = os.path.abspath(os.path.dirname(infile))
    PARSING_RESULTS_BY_FOLDER.setdefault(folder_key, []).append({
        'audit_file': os.path.basename(infile),
        'sheet_name': os.path.splitext(os.path.basename(infile))[0],
        'source_folder': os.path.dirname(infile),
        'status': 'parsed',
        'node_count': len(document),
        'custom_item_count': node_type_counts.get('custom_item', 0),
        'report_count': node_type_counts.get('report-passed', 0) + node_type_counts.get('report-warning', 0),
        'text_count': node_type_counts.get('text', 0),
        'unknown_key_count': len(unknown_keys),
        'unknown_keys': '; '.join(sorted(unknown_keys)),
        'normalized_output': outfile,
        'notes': (
            f'Normalized with Tenable Audit Lang parser ({parse_mode})'
            + (f'; {parse_note}' if parse_note else '')
        ),
        'nodes': node_rows,
    })

    if unknown_keys:
        print("\nUnrecognized keys found â€” classify each:")
        reclassified = False

        for k in sorted(unknown_keys):
            while True:
                ans = input(f"  '{k}': (R)eal, (I)gnored, (S)kip? ").strip().upper()
                if ans in ('R', 'I', 'S'):
                    break

            if ans == 'R':
                REAL_KEYS.add(k)
                _persist_key(k, 'REAL_KEYS')
                reclassified = True

            elif ans == 'I':
                IGNORED_KEYS.add(k)
                _persist_key(k, 'IGNORED_KEYS')
                reclassified = True

        if reclassified:
            output, _ = emit(document, variables, platform_hint)
            output = _cleanup_normalized_output_lines(output)

            preflight_text, preflight_changed, preflight_error_label, preflight_errors = _repair_and_scan_audit_text(
                "\n".join(output) + "\n",
                check_type_name=source_check_type_name or None,
                check_type_version=source_check_type_version or None,
                group_policy_name=source_group_policy_name or None,
                platform_hint=platform_hint,
            )
            if preflight_changed:
                output = preflight_text.rstrip("\n").splitlines()
            if preflight_errors:
                print(f"ERROR: {preflight_error_label} for generated output: {infile}")
                for err in preflight_errors[:25]:
                    print(err)
                _record_validation_result(infile, 'failed', '\n'.join(preflight_errors[:25]))
                return False

            safe_text = _strip_bom_prefix("\n".join(output) + "\n")
            _assert_no_encoding_markers(safe_text, outfile)
            _write_audit_text(outfile, safe_text)

            validate_and_repair_audit_file(
                outfile,
                check_type_name=source_check_type_name or None,
                check_type_version=source_check_type_version or None,
                group_policy_name=source_group_policy_name or None,
                platform_hint=platform_hint,
            )

            print("\nRe-processed with updated key classifications.")

    return True

def process_folder(folder, open_in_vscode=False, strict_mode=False):
    audit_files = sorted(
        f for f in os.listdir(folder)
        if f.lower().endswith(".audit")
    )

    if not audit_files:
        print("No .audit files found.")
        return True

    print(f"\nFound {len(audit_files)} audit files.\n")

    failed_files = []

    for fname in audit_files:
        infile = os.path.join(folder, fname)

        print("-" * 60)
        print(f"Processing: {fname}")

        ok = process_file(infile, open_in_vscode=open_in_vscode, strict_mode=strict_mode)
        if not ok:
            failed_files.append(infile)

    if failed_files:
        print("\nPreflight/normalization failures:")
        for path in failed_files:
            print(f"  {path}")
        if strict_mode:
            print("\nERROR: strict mode enabled and one or more files failed preflight/normalization.")
            return False

    return True

def parse_cli_args():
    input_arg = None
    out_arg = None
    catalog_flag = False
    export_duplicates = False
    open_vscode = False
    install_auditlang = False
    strict_mode = False
    refresh_threat_intel = False
    write_docker_files_flag = False
    skip_next = False

    for i, arg in enumerate(sys.argv[1:]):
        if skip_next:
            skip_next = False
            continue
        if arg == '--catalog':
            catalog_flag = True
        elif arg == '--export-duplicates':
            export_duplicates = True
        elif arg == '--open-vscode':
            open_vscode = True
        elif arg == '--install-auditlang-ext':
            install_auditlang = True
        elif arg == '--strict':
            strict_mode = True
        elif arg == '--refresh-threat-intel':
            refresh_threat_intel = True
        elif arg == '--write-docker-files':
            write_docker_files_flag = True
        elif arg.startswith('-'):
            continue
        elif input_arg is None:
            input_arg = arg
        elif out_arg is None:
            out_arg = arg

    return input_arg, out_arg, catalog_flag, export_duplicates, open_vscode, install_auditlang, strict_mode, refresh_threat_intel, write_docker_files_flag


def write_docker_files(output_dir=None):
    """Write reproducible Docker build files next to the self-contained script."""
    target_dir = os.path.abspath(output_dir or SCRIPT_DIR)
    os.makedirs(target_dir, exist_ok=True)
    dockerfile_path = os.path.join(target_dir, 'Dockerfile')
    requirements_path = os.path.join(target_dir, 'requirements.txt')
    dockerfile = '''FROM python:3.13-slim

WORKDIR /app
COPY requirements.txt ./
RUN pip install --no-cache-dir -r requirements.txt
COPY ALL_AUDITS.py ./
ENTRYPOINT ["python", "/app/ALL_AUDITS.py"]
'''
    with open(requirements_path, 'w', encoding='utf-8', newline='\n') as handle:
        handle.write('openpyxl==3.1.5\n')
    with open(dockerfile_path, 'w', encoding='utf-8', newline='\n') as handle:
        handle.write(dockerfile)
    print(f'Wrote {dockerfile_path}')
    print(f'Wrote {requirements_path}')
    print('Build: docker build -t pysc-audit-normalizer:1.0 .')


def main():
    input_arg, out_arg, catalog_flag, export_duplicates, open_vscode, install_auditlang, strict_mode, refresh_threat_intel, write_docker_files_flag = parse_cli_args()

    if write_docker_files_flag:
        write_docker_files()
        return

    intel_data = _load_threat_intel_cache(force_refresh=refresh_threat_intel)
    taxonomy_count = len(intel_data.get('control_taxonomy', []) or [])
    external_count = len((intel_data.get('threat_by_control_id', {}) or {}).keys())
    print(
        'Threat intel loaded: '
        f'taxonomy={taxonomy_count} | '
        f'external_overrides={external_count} | '
        f'source={intel_data.get("source", "builtin")}'
    )

    vscode_ready = True
    if open_vscode:
        vscode_ready = ensure_vscode_auditlang_extension(install_if_missing=install_auditlang)
        if not vscode_ready:
            print('Continuing without VS Code AuditLang diagnostics.')

    if input_arg:
        input_arg = input_arg.strip().strip('"').strip("'")
        if os.path.isdir(input_arg):
            folder_ok = process_folder(input_arg, open_in_vscode=(open_vscode and vscode_ready), strict_mode=strict_mode)
            # auto-generate catalog for folder runs
            print('\nGenerating controls catalog...')
            outp = generate_catalog(input_arg, out_arg, os.path.join(input_arg, 'Normalized'))
            if export_duplicates:
                csvs = export_duplicates_csvs(outp)
                for p in csvs:
                    print(f'Wrote {p}')
            _run_merged_audit_generation()
            _write_parsing_results_for_folder(input_arg)
            if strict_mode and not folder_ok:
                raise RuntimeError('Strict mode: one or more files failed preflight/normalization.')
            return
        elif os.path.isfile(input_arg):
            file_ok = process_file(input_arg, open_in_vscode=(open_vscode and vscode_ready), strict_mode=strict_mode)
            if catalog_flag:
                folder = os.path.dirname(input_arg)
                print('\nGenerating controls catalog...')
                outp = generate_catalog(folder, out_arg, os.path.join(folder, 'Normalized'))
                if export_duplicates:
                    csvs = export_duplicates_csvs(outp)
                    for p in csvs:
                        print(f'Wrote {p}')
            _run_merged_audit_generation()
            _write_parsing_results_for_folder(os.path.dirname(input_arg))
            if strict_mode and not file_ok:
                raise RuntimeError('Strict mode: file failed preflight/normalization.')
            return
        else:
            print('ERROR: Path does not exist.')
            return

    # Default no-arg automation:
    # 1) Process actual_audit_inputs
    # 2) Process audit_inputs
    # 3) Generate crosswalk from both Unique_Controls_Catalog.xlsx outputs
    left_root = os.path.join(SCRIPT_DIR, 'actual_audit_inputs')
    right_root = os.path.join(SCRIPT_DIR, 'audit_inputs')

    for root in (left_root, right_root):
        if not os.path.isdir(root):
            print(f'ERROR: Default folder does not exist: {root}')
            return

    print(f'\nDefault run: processing {left_root}')
    left_ok = process_folder(left_root, open_in_vscode=(open_vscode and vscode_ready), strict_mode=strict_mode)
    print('\nGenerating controls catalog...')
    left_all = generate_catalog(left_root, None, os.path.join(left_root, 'Normalized'))
    if export_duplicates:
        csvs = export_duplicates_csvs(left_all)
        for p in csvs:
            print(f'Wrote {p}')
    _write_parsing_results_for_folder(left_root)

    print(f'\nDefault run: processing {right_root}')
    right_ok = process_folder(right_root, open_in_vscode=(open_vscode and vscode_ready), strict_mode=strict_mode)
    print('\nGenerating controls catalog...')
    right_all = generate_catalog(right_root, None, os.path.join(right_root, 'Normalized'))
    if export_duplicates:
        csvs = export_duplicates_csvs(right_all)
        for p in csvs:
            print(f'Wrote {p}')
    _write_parsing_results_for_folder(right_root)

    left_unique = _timestamped_output_path(os.path.join(left_root, 'Normalized', 'Unique_Controls_Catalog.xlsx'))
    right_unique = _timestamped_output_path(os.path.join(right_root, 'Normalized', 'Unique_Controls_Catalog.xlsx'))
    crosswalk_out = _timestamped_output_path(os.path.join(left_root, 'Normalized', 'Matched_Controls_Crosswalk.xlsx'))

    print('\nMatching controls across unique catalogs...')
    match_unique_catalogs(left_unique, right_unique, crosswalk_out)
    _run_merged_audit_generation()
    if strict_mode and (not left_ok or not right_ok):
        raise RuntimeError('Strict mode: one or more files failed preflight/normalization.')


def _run_merged_audit_generation():
    master_unique = _resolve_existing_or_latest_timestamped_path(os.path.join(SCRIPT_DIR, 'actual_audit_inputs', 'Normalized', 'Unique_Controls_Catalog.xlsx'))
    other_unique = _resolve_existing_or_latest_timestamped_path(os.path.join(SCRIPT_DIR, 'audit_inputs', 'Normalized', 'Unique_Controls_Catalog.xlsx'))
    out_dir = os.path.join(SCRIPT_DIR, 'actual_audit_inputs', 'Merged')

    if not (os.path.isfile(master_unique) and os.path.isfile(other_unique)):
        print('\nSkipping merged audit generation: unique catalogs are not ready yet.')
        return []

    print('\nGenerating merged audits...')
    written = build_merged_master_audits(master_unique, other_unique, out_dir)

    for merged_path in written:
        all_path = _write_uncommented_all_copy(merged_path)
        if all_path:
            platform_hint = determine_platform_from_filename(all_path)
            validate_and_repair_audit_file(
                all_path,
                check_type_name='Windows',
                check_type_version='2',
                group_policy_name=_extract_group_policy_name(all_path) or 'Windows Security Hardening Through Group Policy and Registry Settings',
                platform_hint=platform_hint,
            )
            print(f'Wrote all-controls audit: {all_path}')

    _stage_gap_analysis_files()

    if os.path.isfile(BASELINE_MSSRV_CSV_PATH) and os.path.isfile(MERGED_MSSRV_CSV_PATH):
        try:
            write_description_match_workbook(
                BASELINE_MSSRV_CSV_PATH,
                MERGED_MSSRV_CSV_PATH,
                DESCRIPTION_MATCH_XLSX_PATH,
            )
        except Exception as exc:
            print(f'WARNING: Could not write description match workbook: {exc}')
    else:
        print(
            'Skipping description match workbook: missing baseline or merged CSV '
            f'({BASELINE_MSSRV_CSV_PATH}, {MERGED_MSSRV_CSV_PATH})'
        )

    _run_production_reference_gap_analysis()

    return written

if __name__ == "__main__":
    _reset_validation_summary()
    try:
        if '--production-gap-analysis' in sys.argv:
            run_production_gap_analysis()
            sys.exit(0)

        if '--match-catalogs' in sys.argv:
            idx = sys.argv.index('--match-catalogs')
            left = None
            right = None
            out_xlsx = None
            if len(sys.argv) > idx + 1 and not sys.argv[idx + 1].startswith('-'):
                left = sys.argv[idx + 1]
            if len(sys.argv) > idx + 2 and not sys.argv[idx + 2].startswith('-'):
                right = sys.argv[idx + 2]
            if len(sys.argv) > idx + 3 and not sys.argv[idx + 3].startswith('-'):
                out_xlsx = sys.argv[idx + 3]

            if not left:
                left = _resolve_existing_or_latest_timestamped_path(os.path.join(SCRIPT_DIR, 'actual_audit_inputs', 'Normalized', 'Unique_Controls_Catalog.xlsx'))
            if not right:
                right = _resolve_existing_or_latest_timestamped_path(os.path.join(SCRIPT_DIR, 'audit_inputs', 'Normalized', 'Unique_Controls_Catalog.xlsx'))

            match_unique_catalogs(left, right, out_xlsx)
            _run_merged_audit_generation()
            sys.exit(0)

        if '--catalog' in sys.argv:
            # usage: python Normalize_Windows_AUDITS.py --catalog [input_folder] [output_file]
            idx = sys.argv.index('--catalog')
            input_arg = None
            out_arg = None
            if len(sys.argv) > idx + 1 and not sys.argv[idx + 1].startswith('-'):
                input_arg = sys.argv[idx + 1]
            if len(sys.argv) > idx + 2 and not sys.argv[idx + 2].startswith('-'):
                out_arg = sys.argv[idx + 2]

            if not input_arg:
                default_input = AUDIT_INPUTS_ROOT
                input_arg = default_input

            if not out_arg:
                # leave as None so generate_catalog will write to input_folder/Normalized
                out_arg = None

            outp = generate_catalog(input_arg, out_arg, os.path.join(input_arg, 'Normalized'))
            if '--export-duplicates' in sys.argv:
                # export duplicates CSVs next to the workbook
                csvs = export_duplicates_csvs(outp)
                for p in csvs:
                    print(f'Wrote {p}')
            _run_merged_audit_generation()
        else:
            main()
    except Exception:
        import traceback
        traceback.print_exc()
    finally:
        _print_validation_summary()


