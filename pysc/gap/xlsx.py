"""Excel export for the platform gap analysis.

Sheet structure follows the legacy interactive engine's workbook (NIST Gap
Analysis matrix, Summary, Reference Coverage, Inactive Controls, Inactive
Coverage Opportunities, Controls Not In Baseline) but is written once per run,
after the full derivation — not once per input file.
"""

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FILL_COVERED = PatternFill("solid", start_color="E2EFDA")
FILL_BASELINE = PatternFill("solid", start_color="A9D08E")
FILL_GAP = PatternFill("solid", start_color="FCE4D6")
FILL_RECOVERABLE = PatternFill("solid", start_color="DDEBF7")
FILL_HEADER = PatternFill("solid", start_color="1F4E78")
FONT_HEADER = Font(bold=True, color="FFFFFF")


def _clean(value):
    if value is None:
        return ""
    return ILLEGAL_CHARACTERS_RE.sub("", str(value))


def _write_sheet(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(vertical="top")
    for row in rows:
        ws.append([_clean(v) for v in row])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, _ in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        width = max(
            (len(str(c.value)) for c in ws[letter] if c.value is not None),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 80)


def export_workbook(analysis, output_file):
    """Write the six-sheet gap workbook for a PlatformGapAnalysis."""
    wb = Workbook()

    # --- Sheet 1: NIST Gap Analysis matrix -----------------------------------
    ws = wb.active
    ws.title = "NIST Gap Analysis"
    files = analysis.files
    headers = ["Control ID", "Control Title", "Family ID", "Family Name", "Overall"]
    for fa in files:
        headers += [f"{fa.short_name} Status", f"{fa.short_name} Controls"]
    matrix_rows = []
    fills = []  # parallel: fill for the Overall cell
    for control_id in sorted(analysis.target_baseline.keys()):
        family_id, family_name = analysis.catalog.family_of(control_id)
        covered_by_baseline = control_id in analysis.baseline_covered_set
        recoverable = control_id in analysis.inactive_coverage_opportunities
        covered_any = control_id in analysis.all_possible_controls
        if covered_by_baseline:
            overall, fill = "COVERED", FILL_BASELINE
        elif recoverable:
            overall, fill = "RECOVERABLE", FILL_RECOVERABLE
        elif covered_any:
            overall, fill = "GAP (candidate available)", FILL_COVERED
        else:
            overall, fill = "MISSING", FILL_GAP
        row = [control_id, analysis.catalog.title(control_id), family_id, family_name, overall]
        for fa in files:
            rules = fa.covered.get(control_id)
            row += ["COVERED" if rules else "-", " ".join(rules) if rules else ""]
        matrix_rows.append(row)
        fills.append(fill)
    _write_sheet(ws, headers, matrix_rows)
    for row_idx, fill in enumerate(fills, start=2):
        ws.cell(row=row_idx, column=5).fill = fill

    # --- Sheet 2: Summary -----------------------------------------------------
    ws = wb.create_sheet("Summary")
    rows = analysis.summary_rows()
    _write_sheet(ws, list(rows[0].keys()), [list(r.values()) for r in rows])

    # --- Sheet 3: Reference Coverage ------------------------------------------
    ws = wb.create_sheet("Reference Coverage")
    ref_rows = []
    for control_id in sorted(analysis.reference_counts):
        family_id, family_name = analysis.catalog.family_of(control_id)
        ref_rows.append(
            [
                control_id,
                analysis.catalog.title(control_id),
                family_id,
                family_name,
                analysis.reference_counts[control_id],
            ]
        )
    _write_sheet(
        ws,
        ["Control ID", "Control Title", "Family ID", "Family Name", "Referencing Checks"],
        ref_rows,
    )

    # --- Sheet 4: Inactive Controls -------------------------------------------
    ws = wb.create_sheet("Inactive Controls")
    _write_sheet(
        ws,
        ["Rule ID", "Description", "NIST Controls"],
        [
            [i["control_number"], i["description"], " ".join(i["controls"])]
            for i in analysis.baseline.inactive_checks
        ],
    )

    # --- Sheet 5: Inactive Coverage Opportunities ------------------------------
    ws = wb.create_sheet("Inactive Coverage Opportunities")
    _write_sheet(
        ws,
        ["Rule ID", "Description", "Recoverable Controls"],
        [
            [r["rule_id"], r["description"], " ".join(r["controls"])]
            for r in analysis.inactive_opportunity_rows
        ],
    )

    # --- Sheet 6: Controls Not In Baseline -------------------------------------
    ws = wb.create_sheet("Controls Not In Baseline")
    add_rows = []
    for control_id in sorted(analysis.additional_controls_not_present):
        family_id, family_name = analysis.catalog.family_of(control_id)
        sources = sorted(
            fa.short_name for fa in analysis.candidates if control_id in fa.covered_set
        )
        add_rows.append(
            [
                control_id,
                analysis.catalog.title(control_id),
                family_id,
                family_name,
                "; ".join(sources),
            ]
        )
    _write_sheet(
        ws,
        ["Control ID", "Control Title", "Family ID", "Family Name", "Available In"],
        add_rows,
    )

    wb.save(output_file)
    return output_file
