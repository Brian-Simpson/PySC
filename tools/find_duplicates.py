import openpyxl, json, re, sys
wb_path = sys.argv[1] if len(sys.argv)>1 else r"c:\PySC\audit_inputs\Normalized\controls_catalog.xlsx"
wb=openpyxl.load_workbook(wb_path, read_only=True)
out={}

def norm(s):
    if s is None:
        return ''
    s=str(s)
    s=s.strip().strip('"').strip("'")
    s=re.sub(r'\s+',' ',s)
    return s.lower()

for name in wb.sheetnames:
    ws=wb[name]
    rows=ws.iter_rows(values_only=True)
    try:
        headers=next(rows)
    except StopIteration:
        out[name]=[]
        continue
    hidx={ (h.lower() if h else ''):i for i,h in enumerate(headers) }
    desc_i=hidx.get('description')
    type_i=hidx.get('control_type')
    src_i=hidx.get('source_file')
    counts={}
    for r in rows:
        desc = norm(r[desc_i]) if desc_i is not None and desc_i<len(r) else ''
        ctype = norm(r[type_i]) if type_i is not None and type_i<len(r) else ''
        src = r[src_i] if src_i is not None and src_i<len(r) else ''
        key=(ctype, desc)
        counts.setdefault(key, []).append(src)
    dups=[{'control_type':k[0],'description':k[1],'count':len(v),'examples':v[:5]} for k,v in counts.items() if len(v)>1]
    dups.sort(key=lambda x:-x['count'])
    out[name]=dups

print(json.dumps(out))
