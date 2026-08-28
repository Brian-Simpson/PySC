import openpyxl, csv, re, sys, os, json

wb_path = sys.argv[1] if len(sys.argv)>1 else r"c:\PySC\audit_inputs\Normalized\controls_catalog.xlsx"
out_dir = os.path.dirname(wb_path)

wb = openpyxl.load_workbook(wb_path, read_only=True)

def norm(s):
    if s is None:
        return ''
    s = str(s)
    s = s.strip().strip('"').strip("'")
    s = re.sub(r'\s+', ' ', s)
    return s

for name in wb.sheetnames:
    ws = wb[name]
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [h.lower() if h else '' for h in next(rows)]
    except StopIteration:
        continue
    hidx = {h: i for i, h in enumerate(headers)}
    desc_i = hidx.get('description')
    type_i = hidx.get('control_type')
    src_i = hidx.get('source_file')
    info_i = hidx.get('info')
    ref_i = hidx.get('reference')
    see_i = hidx.get('see_also')
    show_i = hidx.get('show_output')
    cond_i = hidx.get('condition_type') or hidx.get('condition')
    rpt_type_i = hidx.get('report_type')
    rpt_desc_i = hidx.get('report_description')
    raw_i = hidx.get('raw_fields')

    counts = {}
    for r in rows:
        desc = norm(r[desc_i]) if desc_i is not None and desc_i < len(r) else ''
        ctype = norm(r[type_i]) if type_i is not None and type_i < len(r) else ''
        key = (ctype.lower(), desc.lower())
        counts.setdefault(key, []).append(r)

    duplicates = {k: v for k, v in counts.items() if len(v) > 1}
    if not duplicates:
        continue

    out_csv = os.path.join(out_dir, f'duplicates_{name}.csv')
    with open(out_csv, 'w', newline='', encoding='utf-8') as fh:
        writer = csv.writer(fh)
        hdr = ['sheet','control_type','description','info','reference','see_also','show_output','condition_type','report_type','report_description','raw_fields','count','examples']
        writer.writerow(hdr)
        for (ctype, desc), rows_list in sorted(duplicates.items(), key=lambda x: -len(x[1])):
            first = rows_list[0]
            info = first[info_i] if info_i is not None and info_i < len(first) else ''
            ref = first[ref_i] if ref_i is not None and ref_i < len(first) else ''
            see = first[see_i] if see_i is not None and see_i < len(first) else ''
            show = first[show_i] if show_i is not None and show_i < len(first) else ''
            cond = first[cond_i] if cond_i is not None and cond_i < len(first) else ''
            rpt_type = first[rpt_type_i] if rpt_type_i is not None and rpt_type_i < len(first) else ''
            rpt_desc = first[rpt_desc_i] if rpt_desc_i is not None and rpt_desc_i < len(first) else ''
            raw = first[raw_i] if raw_i is not None and raw_i < len(first) else ''
            examples = []
            for r in rows_list[:10]:
                src = r[src_i] if src_i is not None and src_i < len(r) else ''
                examples.append(src)
            writer.writerow([name, ctype, desc, info, ref, see, show, cond, rpt_type, rpt_desc, raw, len(rows_list), ';'.join(examples)])

    print(f'Wrote {out_csv}')
