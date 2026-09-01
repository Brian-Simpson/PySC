"""Phase 4 tests: enterprise driver, history store, matrix + dashboard builds."""

from pathlib import Path

import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
CATALOG = REPO_ROOT / "NIST_SP-800-53_rev5_catalog.json"

from pysc.gap.engine import analyze_files  # noqa: E402
from pysc.gap.enterprise import DETECTION_TO_PLATFORM, EnterpriseGapResult  # noqa: E402
from pysc.history import HistoryStore  # noqa: E402
from pysc.report.excel_util import sanitize_for_excel  # noqa: E402
from pysc.report.html import _PLATFORM_COLORS, build_dashboard  # noqa: E402
from pysc.report.matrix import build_matrix  # noqa: E402

BASELINE_AUDIT = '''<check_type:"Windows" version:"2">
<custom_item>
  type        : AUDIT_POWERSHELL
  description : "1.0001 - TEST - Active check"
  reference   : "NIST 800-53r5|AC-2"
</custom_item>
#<custom_item>
#  type        : AUDIT_POWERSHELL
#  description : "1.0002 - TEST - Commented check"
#  reference   : "NIST 800-53r5|IA-5"
#</custom_item>
</check_type>
'''

CANDIDATE_AUDIT = '''<check_type:"Windows" version:"2">
<custom_item>
  type        : AUDIT_POWERSHELL
  description : "2.1 Candidate covers IA-5"
  reference   : "NIST 800-53r5|IA-5"
</custom_item>
<custom_item>
  type        : AUDIT_POWERSHELL
  description : "2.2 Candidate covers SC-7"
  reference   : "NIST 800-53r5|SC-7"
</custom_item>
</check_type>
'''


@pytest.fixture(scope="module")
def result(tmp_path_factory):
    root = tmp_path_factory.mktemp("enterprise")
    baseline = root / "HTH_TEST_BASELINE.audit"
    candidate = root / "CIS_Candidate.audit"
    baseline.write_text(BASELINE_AUDIT, encoding="utf-8")
    candidate.write_text(CANDIDATE_AUDIT, encoding="utf-8")
    analysis = analyze_files(baseline, [candidate], CATALOG)
    return EnterpriseGapResult(
        {"MSSRV": analysis},
        {"NetACI": ["CIS_ACI_candidate.audit"]},
        ["weird_file.audit"],
    )


def test_detection_mapping_covers_all_configured_platforms():
    from pysc.config import load_config

    cfg = load_config(REPO_ROOT / "pysc.toml")
    mapped = set(DETECTION_TO_PLATFORM.values())
    configured = set(cfg.platforms())
    # Every mapped code must exist in config; NetACI has no filename detection.
    assert mapped <= configured
    assert configured - mapped == {"NetACI"}


def test_history_roundtrip(result, tmp_path):
    store = HistoryStore(tmp_path / "hist.sqlite")
    run1 = store.record_enterprise_run(result, notes="t1")
    run2 = store.record_enterprise_run(result, notes="t2")
    assert run2 == run1 + 1

    rows = store.platform_trend("MSSRV")
    assert len(rows) == 2
    _run, _ts, platform, covered, recoverable, total = rows[0]
    assert platform == "MSSRV"
    assert covered >= 1          # AC-2
    assert recoverable >= 1      # IA-5 via commented check
    assert total > 1000          # full base catalog

    out = store.export_csv(tmp_path / "hist.csv")
    text = Path(out).read_text(encoding="utf-8")
    assert text.splitlines()[0].startswith("run_id,ts,platform,family")
    store.close()


def test_matrix_build(result, tmp_path):
    store = HistoryStore(tmp_path / "hist.sqlite")
    store.record_enterprise_run(result)
    out = build_matrix(result, tmp_path / "matrix.xlsx", history=store)
    store.close()

    from openpyxl import load_workbook

    wb = load_workbook(out)
    assert wb.sheetnames == [
        "Executive_Summary", "Platform_Family_Coverage", "NIST_Matrix",
        "Priority_Gaps", "CIS_Variances", "Attack_Vectors", "Trend",
    ]
    ws = wb["Executive_Summary"]
    platforms = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert "MSSRV" in platforms and "NetACI" in platforms
    trend = wb["Trend"]
    assert trend.max_row >= 2  # snapshot present


def test_dashboard_build(result, tmp_path):
    out = build_dashboard(result, tmp_path / "dash.html")
    text = Path(out).read_text(encoding="utf-8")
    assert "<!DOCTYPE html>" in text
    assert "MSSRV" in text
    # Executive scope: no missing-baseline reporting on the dashboard.
    assert "NO BASELINE" not in text
    assert "without a baseline" not in text
    # Ranked remediation priorities are front and center.
    assert "Top remediation priorities" in text
    assert "Import candidate check" in text or "Un-comment existing check" in text
    assert "prefers-color-scheme: dark" in text
    assert "Coverage % by platform and NIST family" in text
    assert "background: var(--platform-color)" in text
    assert f"--platform-color:{_PLATFORM_COLORS['MSSRV']}" in text


def test_dashboard_platform_colors_are_distinct():
    assert len(_PLATFORM_COLORS) == len(set(_PLATFORM_COLORS.values()))


def test_priority_gap_rows(result):
    from pysc.report.priority import priority_gap_rows

    rows = priority_gap_rows(result)
    # Synthetic MSSRV analysis: IA-5 recoverable (priority family, weight 3),
    # SC-7 missing (priority family, weight 3 x2 = 6, ranked first).
    assert rows[0]["control_id"] == "SC-7" and rows[0]["score"] == 6
    assert rows[0]["action"] == "Import candidate check"
    ia5 = next(r for r in rows if r["control_id"] == "IA-5")
    assert ia5["score"] == 3 and ia5["action"] == "Un-comment existing check"
    assert priority_gap_rows(result, limit=1) == rows[:1]


def test_attack_vectors_from_gaps(result, tmp_path):
    from pysc.nist.attack import attack_vectors_for_gaps

    # Synthetic mapping: gaps are IA-5 (recoverable) and SC-7 (missing).
    mappings = {
        "IA-5": [("T1110", "Brute Force"), ("T1110.001", "Password Guessing")],
        "SC-7": [("T1110", "Brute Force"), ("T1046", "Network Service Discovery")],
        "AC-2": [("T1136", "Create Account")],  # covered control -> not exposed
    }
    vectors = attack_vectors_for_gaps(result, mappings)
    by_id = {v["technique_id"]: v for v in vectors}
    # Brute Force weakened via BOTH gap controls, sub-technique rolled up.
    assert by_id["T1110"]["controls"] == ["IA-5", "SC-7"]
    assert by_id["T1110"]["sub_technique_count"] == 1
    assert by_id["T1110"]["technique_name"] == "Brute Force"
    assert by_id["T1046"]["controls"] == ["SC-7"]
    # AC-2 is covered by the baseline -> Create Account is NOT exposed.
    assert "T1136" not in by_id
    # Ranked first: most platforms, then most weakened controls.
    assert vectors[0]["technique_id"] == "T1110"

    # Dashboard renders the section when mappings are supplied.
    out = build_dashboard(result, tmp_path / "dash2.html", attack_mappings=mappings)
    text = Path(out).read_text(encoding="utf-8")
    assert "Common attack vectors exposed by current gaps" in text
    assert "Brute Force" in text
    assert "Enterprise coverage trend" not in text


def test_sanitize_for_excel_formula_guard():
    assert sanitize_for_excel("=1+2") == "'=1+2"
    assert sanitize_for_excel("@cmd") == "'@cmd"
    assert sanitize_for_excel("plain") == "plain"
    assert sanitize_for_excel(None) == ""
