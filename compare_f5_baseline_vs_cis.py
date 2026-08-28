#!/usr/bin/env python3
"""Compare active F5 audit controls against a Baseline audit file.

The baseline is the .audit file whose name contains "Baseline". Each active
<custom_item> or <item> block is identified by its f5_command and
json_transform values. Results are written to an Excel workbook.
"""

from __future__ import annotations

import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BLOCK_RE = re.compile(
    r"(?ms)^(?!\s*#)\s*<(?P<tag>custom_item|item)\b[^>]*>"
    r"(?P<body>.*?)</(?P=tag)\s*>",
    re.IGNORECASE,
)
FIELD_RE = re.compile(r"^\s*([A-Za-z0-9_-]+)\s*:\s*(.*?)\s*$")
REFERENCE_RE = re.compile(r"(?im)^\s*reference\s*:\s*(.*?)\s*$")
SIGNATURE_FIELDS = ("f5_command", "json_transform")


def canonical_value(value: str) -> str:
    """Normalize equivalent audit values for signature comparison."""
    value = value.strip()
    if len(value) >= 2 and value[0] == value[-1] == '"':
        value = value[1:-1]
    value = value.replace(r"\"", '"')
    return re.sub(r"\s+", " ", value).strip()


def parse_blocks(path: Path) -> List[Dict[str, str]]:
    text = path.read_text(encoding="utf-8", errors="replace")
    blocks: List[Dict[str, str]] = []
    for match in BLOCK_RE.finditer(text):
        fields: Dict[str, str] = {
            "_tag": match.group("tag").lower(),
            "_raw": match.group(0).strip(),
        }
        preceding_references = list(REFERENCE_RE.finditer(text, 0, match.start()))
        inherited_reference = preceding_references[-1].group(1) if preceding_references else ""
        for line in match.group("body").splitlines():
            field = FIELD_RE.match(line)
            if field:
                fields[field.group(1).lower()] = field.group(2)
        if "reference" not in fields and inherited_reference:
            fields["reference"] = inherited_reference
        if all(field in fields for field in SIGNATURE_FIELDS):
            blocks.append(fields)
    return blocks


def orphaned_blocks(
    baseline: List[Dict[str, str]],
    comparisons: List[Tuple[Path, List[Dict[str, str]]]],
) -> List[Dict[str, str]]:
    """Return comparison controls that have no matching baseline control."""
    baseline_counts = Counter(signature(block) for block in baseline)
    orphaned: List[Dict[str, str]] = []
    for _, blocks in comparisons:
        remaining = Counter(baseline_counts)
        for block in blocks:
            key = signature(block)
            if key is not None and remaining[key] > 0:
                remaining[key] -= 1
            else:
                orphaned.append(block)
    return orphaned


def build_audit_with_orphans(
    output_path: Path,
    baseline_path: Path,
    baseline: List[Dict[str, str]],
    comparisons: List[Tuple[Path, List[Dict[str, str]]]],
) -> int:
    """Copy the baseline and append orphan controls before its closing tag."""
    baseline_text = baseline_path.read_text(encoding="utf-8", errors="replace")
    closing_tag = re.compile(r"(?im)^[ \t]*</check_type>[ \t]*$")
    match = closing_tag.search(baseline_text)
    if match is None:
        raise ValueError(f"Baseline file has no </check_type> closing tag: {baseline_path}")

    orphaned = orphaned_blocks(baseline, comparisons)
    if orphaned:
        appended = "\n\n" + "\n\n".join(block["_raw"] for block in orphaned) + "\n\n"
        output_text = baseline_text[:match.start()] + appended + baseline_text[match.start():]
    else:
        output_text = baseline_text
    with output_path.open("w", encoding="utf-8", newline="") as output_file:
        output_file.write(output_text)
    return len(orphaned)


def signature(block: Dict[str, str]) -> Tuple[str, str]:
    return tuple(canonical_value(block[field]) for field in SIGNATURE_FIELDS)  # type: ignore[return-value]


def control_label(block: Dict[str, str]) -> str:
    return canonical_value(block.get("description", "")) or "(no description)"

def control_reference(block: Dict[str, str]) -> str:
    reference = canonical_value(block.get("reference", ""))
    prefix = "NIST 800-53r5|"
    if reference.startswith(prefix):
        return reference[len(prefix):]
    controls = re.findall(r"800-53r5\|([^,\s\"]+)", reference, re.IGNORECASE)
    return " ".join(controls)


def autosize(sheet) -> None:
    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(width + 2, 70)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def style_header(row) -> None:
    for cell in row:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")


def build_workbook(
    output_path: Path,
    baseline_path: Path,
    baseline: List[Dict[str, str]],
    comparisons: List[Tuple[Path, List[Dict[str, str]]]],
) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Audit file", "Role", "Active controls", "Matching controls", "Orphaned controls"])
    style_header(summary[1])
    summary.append([baseline_path.name, "Baseline", len(baseline), "N/A", "N/A"])

    baseline_counts = Counter(signature(block) for block in baseline)
    details = workbook.create_sheet("Control Details")
    details.append(["Audit file", "Status", "Description", "Reference", "f5_command", "json_transform"])
    style_header(details[1])

    for block in baseline:
        details.append([
            baseline_path.name, "Baseline", control_label(block),
            control_reference(block),
            canonical_value(block["f5_command"]),
            canonical_value(block["json_transform"]),
        ])

    for path, blocks in comparisons:
        remaining = Counter(baseline_counts)
        matching = 0
        for block in blocks:
            key = signature(block)
            status = "Orphaned"
            if remaining[key] > 0:
                remaining[key] -= 1
                matching += 1
                status = "Matching"
            details.append([
                path.name, status, control_label(block),
                control_reference(block),
                canonical_value(block["f5_command"]),
                canonical_value(block["json_transform"]),
            ])
        summary.append([path.name, "Comparison", len(blocks), matching, len(blocks) - matching])

    notes = workbook.create_sheet("Read Me")
    notes.append(["Comparison rules"])
    style_header(notes[1])
    notes.append(["Baseline", f"{baseline_path.name} (filename contains 'Baseline')"])
    notes.append(["Active control", "A non-commented <custom_item> or <item> block."])
    notes.append(["Control identity", "Equivalent f5_command and json_transform values."])
    notes.append(["Orphaned control", "A comparison-file control with no matching baseline control."])

    for sheet in workbook.worksheets:
        autosize(sheet)
    workbook.save(output_path)


DEFAULT_DIRECTORY = Path(r"C:\PySC")


def choose_directory() -> Path:
    entered = input(
        f"Enter the directory containing the .audit files [{DEFAULT_DIRECTORY}]: "
    ).strip().strip('"')
    directory = Path(entered) if entered else DEFAULT_DIRECTORY
    if not directory.is_dir():
        raise ValueError(f"Directory does not exist: {directory}")
    return directory


def main() -> int:
    try:
        directory = choose_directory()
        audit_files = sorted(directory.glob("*.audit"))
        baseline_files = [path for path in audit_files if "baseline" in path.name.lower()]
        if len(baseline_files) != 1:
            raise ValueError(
                f"Expected exactly one .audit filename containing 'Baseline'; found {len(baseline_files)}."
            )

        baseline_path = baseline_files[0]
        comparison_paths = [
            path for path in audit_files
            if path != baseline_path and not path.name.startswith("F5_Audit_with_orphaned_controls_")
        ]
        baseline = parse_blocks(baseline_path)
        comparisons = [(path, parse_blocks(path)) for path in comparison_paths]
        timestamp = datetime.now().strftime("%y%m%d%H%M")
        output_path = directory / f"Baseline_vs_CIS_Comparison_{timestamp}.xlsx"
        build_workbook(output_path, baseline_path, baseline, comparisons)
        audit_output_path = directory / f"F5_Audit_with_orphaned_controls_{timestamp}.audit"
        orphan_count = build_audit_with_orphans(
            audit_output_path, baseline_path, baseline, comparisons
        )
        print(f"Baseline controls: {len(baseline)}")
        print(f"Compared files: {len(comparisons)}")
        print(f"Excel report: {output_path}")
        print(f"Audit with orphaned controls ({orphan_count}): {audit_output_path}")
        return 0
    except (OSError, ValueError) as error:
        print(f"Error: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
