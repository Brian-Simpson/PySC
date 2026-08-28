import csv
import json
import os
import re
from collections import defaultdict
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import subprocess
import sys
from pathlib import Path

LOCAL_CATALOG_NAME = "NIST_SP-800-53_rev5_catalog.json"

NIST_FAMILIES = {
    "AC": "Access Control",
    "AT": "Awareness and Training",
    "AU": "Audit and Accountability",
    "CA": "Assessment, Authorization, and Monitoring",
    "CM": "Configuration Management",
    "CP": "Contingency Planning",
    "IA": "Identification and Authentication",
    "IR": "Incident Response",
    "MA": "Maintenance",
    "MP": "Media Protection",
    "PE": "Physical and Environmental Protection",
    "PL": "Planning",
    "PM": "Program Management",
    "PS": "Personnel Security",
    "RA": "Risk Assessment",
    "SA": "System and Services Acquisition",
    "SC": "System and Communications Protection",
    "SI": "System and Information Integrity",
    "SR": "Supply Chain Risk Management",
}


def load_local_oscal_baseline(script_dir: str) -> tuple[dict, dict]:
    """Loads the manually downloaded NIST OSCAL catalog file from the script directory."""
    local_path = os.path.join(script_dir, LOCAL_CATALOG_NAME)

    if not os.path.exists(local_path):
        print(f"\n[-] ERROR: Catalog file not found at: {local_path}")
        print(f"[-] Please ensure '{LOCAL_CATALOG_NAME}' is placed in your folder: {script_dir}")
        exit(1)

    print(f"\n[+] Successfully located local OSCAL baseline file: {LOCAL_CATALOG_NAME}")
    try:
        with open(local_path, "r", encoding="utf-8") as f:
            catalog_data = json.load(f)
    except Exception as e:
        print(f"[-] Error parsing JSON structure: {e}")
        exit(1)

    all_controls = {}
    control_parents = {}

    def recurse_controls(control_list, parent_id=None):
        for ctrl in control_list:
            ctrl_id = ctrl.get("id", "").upper()
            title = ctrl.get("title", "No Title Available")
            if ctrl_id:
                all_controls[ctrl_id] = title
                if parent_id:
                    control_parents[ctrl_id] = parent_id
            if "controls" in ctrl:
                recurse_controls(
                    ctrl["controls"],
                    parent_id=ctrl_id if not parent_id else parent_id,
                )

    groups = catalog_data.get("catalog", {}).get("groups", [])
    for group in groups:
        if "controls" in group:
            recurse_controls(group["controls"])

    print(f"[+] Loaded {len(all_controls)} global controls and enhancements from OSCAL database.")
    return all_controls, control_parents


def extract_control_number(description: str) -> str:
    """Extracts only the description identifier number like 1.0022 or 10203 from audit tags."""
    match = re.match(r"^([\d\.]+)", description.strip())
    if match:
        return match.group(1).rstrip(".")
    return description


def extract_audit_items(file_path: str):
    """Parses Nessus .audit file blocks and extracts descriptions and compliance tags."""
    if not os.path.exists(file_path):
        print(f"[-] Error: Target audit file not found at '{file_path}'")
        return None

    if not os.path.isfile(file_path):
        print(f"[-] Error: '{file_path}' is a directory, not an audit file")
        return None

    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        content = f.read()

    # Remove commented-out custom_item blocks
    content = re.sub(
        r"(?ms)^\s*#\s*<custom_item>.*?^\s*#\s*</custom_item>\s*$",
        "",
        content,
    )

    block_pattern = re.compile(
        r"<(?:custom_item|item|report)>([\s\S]*?)</(?:custom_item|item|report)>",
        re.MULTILINE,
        )

    items = []
    for match in block_pattern.finditer(content):
        block_text = match.group(1)

        desc_match = re.search( r'description\s*:\s*"([^"]*)"', block_text, re.DOTALL )
        description = ( desc_match.group(1).strip() if desc_match else "Unnamed Check" )

        ref_match = re.search(r'reference\s*:\s*["\'](.*?)["\']', block_text)
        references = ref_match.group(1) if ref_match else ""

        nist_controls = set()
        nist_pattern = re.compile(
            r"(?:800-53r5|800-53|NIST[\s\-_]SP[\s\-_]800-53(?:r5| Rev\.? 5)?)\|?([A-Z]{2}-\d+(?:\(\d+\))?)",
            re.IGNORECASE,
        )
        for found in nist_pattern.finditer(references + " " + block_text):
            control_id = found.group(1).upper()
            nist_controls.add(control_id)

        control_num = extract_control_number(description)

        items.append(
            {"control_number": control_num,
              "description": description,  
              "controls": list(nist_controls)}
        )
    return items

def extract_inactive_audit_items(file_path):

    with open(
        file_path,
        "r",
        encoding="utf-8",
        errors="ignore"
    ) as f:
        content = f.read()

    inactive_items = []

    inactive_pattern = re.compile(
        r"(?ms)^\s*#\s*<custom_item>(.*?)^\s*#\s*</custom_item>\s*$"
    )

    nist_pattern = re.compile(
        r"(?:800-53r5|800-53|NIST[\s\-_]SP[\s\-_]800-53(?:r5| Rev\.? 5)?)\|?([A-Z]{2}-\d+(?:\(\d+\))?)",
        re.IGNORECASE,
    )

    for match in inactive_pattern.finditer(content):

        block_text = re.sub(
            r"(?m)^\s*#\s?",
            "",
            match.group(1)
        )

        desc_match = re.search(
            r'description\s*:\s*"(.*?)"',
            block_text
        )

        description = (
            desc_match.group(1)
            if desc_match
            else ""
        )

        control_number = extract_control_number(
            description
        )

        refs = set()

        for ref_match in nist_pattern.finditer(
            block_text
        ):
            refs.add(
                ref_match.group(1).upper()
            )

        inactive_items.append({
            "control_number": control_number,
            "description": description,
            "controls": sorted(refs)
        })

    return inactive_items

def analyze_single_file(audit_items: list, target_baseline: dict, control_parents: dict):
    """Maps a single file's parsed components to the target baseline set."""
    covered_controls = defaultdict(list)
    for item in audit_items:
        for ctrl in item["controls"]:
            normalized_ctrl = re.sub(r"-0(\d)", r"-\1", ctrl)
            if item["control_number"] not in covered_controls[normalized_ctrl]:
                covered_controls[normalized_ctrl].append(item["control_number"])
            if normalized_ctrl in control_parents:
                parent = control_parents[normalized_ctrl]
                if item["control_number"] not in covered_controls[parent]:
                    covered_controls[parent].append(item["control_number"])
            elif "(" in normalized_ctrl:
                base_ctrl = normalized_ctrl.split("(")[0]
                if item["control_number"] not in covered_controls[base_ctrl]:
                    covered_controls[base_ctrl].append(item["control_number"])
    return covered_controls


def try_float(val):
    """Sort CIS rule IDs naturally."""

    try:
        return (
            0,
            tuple(
                int(x)
                for x in str(val).split(".")
            )
        )

    except ValueError:

        return (
            1,
            str(val)
        )

def clean_excel_text(value):

    if value is None:
        return ""

    return ILLEGAL_CHARACTERS_RE.sub(
        "",
        str(value)
    )

def export_excel_matrix(results_by_file: dict, file_metadata: list, target_baseline: dict, output_file: str,
                         summary_rows: list, reference_counts: dict, inactive_controls: list = None,
                           analysis_metrics: dict = None, inactive_opportunity_rows: list = None, additional_control_rows: list = None):
    """Generates an Excel comparison matrix supporting one or multiple side-by-side audit files."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NIST Gap Analysis"
    ws.sheet_view.showGridLines = True

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    covered_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Green
    gap_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")     # Soft Orange/Red
    partial_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Soft Light Blue
    
    # NEW: Custom 20% Darker Green Fill used specifically for files containing "baseline" in their name
    baseline_covered_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")

    # Base structural headers
    headers = [
        "Control ID",
        "Control Title",
        "Family ID",
        "Family Name",
        "Overall"
    ]

    for file_info in file_metadata:
        short_name = file_info["short_name"]
        headers.extend([f"{short_name} Status", f"{short_name} Controls"])
        
    ws.append(headers)

    # Format header row
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for ctrl_id in sorted(target_baseline.keys()):
        family_id = ctrl_id.split("-")[0] if "-" in ctrl_id else "PM"
        family_name = NIST_FAMILIES.get(family_id, "Program Management")
        title = target_baseline[ctrl_id]

        row_data = [ctrl_id, title, family_id, family_name]

        # Holds all rule IDs from all files for this control
        overall_controls = set()
        file_statuses = []

        # Gather status and matched numbers per individual file context
        for file_info in file_metadata:
            f_id = file_info["id"]
            covered_map = results_by_file[f_id]
            
            if ctrl_id in covered_map:
                rule_ids = sorted(covered_map[ctrl_id], key=try_float)
                status = "COVERED"
                controls_str = " | ".join(rule_ids)

                overall_controls.update(rule_ids)
            else:
                status = "GAP / MISSING"
                controls_str = "None"
                
            row_data.extend([status, controls_str])
            file_statuses.append(status)

        overall_value = (
            " | ".join(
                sorted(
                    overall_controls,
                    key=lambda x: str(x)
                )
            )
            if overall_controls
            else "None"
        )

        row_data.insert(4, overall_value)

        ws.append(row_data)
        current_row = ws.max_row

        # Determine if there is partial coverage (divergence between files)
        has_covered = "COVERED" in file_statuses
        has_gap = "GAP / MISSING" in file_statuses
        is_partial_row = has_covered and has_gap and len(file_metadata) > 1

        # Base alignments for standard metadata cells
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.cell(row=current_row, column=2).alignment = left_align
        ws.cell(row=current_row, column=3).alignment = center_align
        ws.cell(row=current_row, column=4).alignment = left_align
        ws.cell(row=current_row, column=5).alignment = left_align

        # Process and color individual cell sets dynamically
        col_pointer = 6
        for idx, file_info in enumerate(file_metadata):
            status = file_statuses[idx]
            status_cell = ws.cell(row=current_row, column=col_pointer)
            status_cell.alignment = center_align
            
            # Check filename context dynamically (Case-Insensitive string matching)
            is_baseline_file = "baseline" in file_info["short_name"].lower()

            if is_partial_row:
                status_cell.fill = partial_fill
            else:
                if status == "COVERED":
                    # NEW: Apply 20% darker green if it's a baseline file, otherwise standard green
                    status_cell.fill = baseline_covered_fill if is_baseline_file else covered_fill
                else:
                    status_cell.fill = gap_fill
            
            ws.cell(row=current_row, column=col_pointer + 1).alignment = left_align
            col_pointer += 2

        # Apply whole row highlights for divergence or row-wide baseline matches
        if is_partial_row:
            for col_idx in range(1, len(row_data) + 1):
                ws.cell(row=current_row, column=col_idx).fill = partial_fill
        else:
            # NEW: If all compared files are COVERED and any file contains "baseline", 
            # highlight the metadata blocks to the 20% darker green
            all_covered = all(s == "COVERED" for s in file_statuses)
            any_baseline_file = any("baseline" in f["short_name"].lower() for f in file_metadata)
            
            if all_covered and any_baseline_file:
                for col_idx in range(1, len(row_data) + 1):
                    # Check individual file status column blocks to retain precise green variants
                    if col_idx >= 6:
                        file_idx = (col_idx - 6) // 2
                        if "baseline" in file_metadata[file_idx]["short_name"].lower():
                            ws.cell(row=current_row, column=col_idx).fill = baseline_covered_fill
                        else:
                            ws.cell(row=current_row, column=col_idx).fill = covered_fill
                    else:
                        ws.cell(row=current_row, column=col_idx).fill = baseline_covered_fill

    # Fit column widths gracefully
    for col in ws.columns:

        max_len = max(
            len(str(cell.value or ""))
            for cell in col
        )

        col_letter = get_column_letter(
            col[0].column
        )

        ws.column_dimensions[
            col_letter
        ].width = max(max_len + 3, 12)

    # ==========================================================
    # SUMMARY SHEET
    # ==========================================================

    if summary_rows:

        summary_ws = wb.create_sheet("Summary")

        headers = [
            "File",
            "Checks Parsed",
            "Current Coverage",
            "Current Coverage %",
            "Individual Coverage",
            "Individual Coverage %",
            "Highest Potential",
            "Highest Potential %"
        ]

        for col_num, header in enumerate(headers, start=1):

            cell = summary_ws.cell(
                row=1,
                column=col_num
            )

            cell.value = header

            cell.font = Font(bold=True)

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="D9EAD3"
            )

        row_num = 2

        for row in summary_rows:

            summary_ws.cell(row_num, 1).value = row["File"]
            summary_ws.cell(row_num, 2).value = row["Checks Parsed"]
            summary_ws.cell(row_num, 3).value = row["Current Coverage"]
            summary_ws.cell(row_num, 4).value = row["Current Coverage %"]
            summary_ws.cell(row_num, 5).value = row["Individual Coverage"]
            summary_ws.cell(row_num, 6).value = row["Individual Coverage %"]
            summary_ws.cell(row_num, 7).value = row["Highest Potential"]
            summary_ws.cell(row_num, 8).value = row["Highest Potential %"]

            row_num += 1

        # ------------------------------------------------------
        # Coverage Opportunity Analysis
        # ------------------------------------------------------

        row_num += 2

        summary_ws.cell(
            row=row_num,
            column=1
        ).value = "Coverage Opportunity Analysis"

        summary_ws.cell(
            row=row_num,
            column=1
        ).font = Font(
            bold=True,
            size=12
        )

        summary_ws.cell(
            row=row_num,
            column=1
        ).fill = PatternFill(
            fill_type="solid",
            fgColor="BDD7EE"
        )

        row_num += 1

        analysis_rows = [
            (
                "Baseline Controls",
                analysis_metrics["baseline_controls"]
            ),
            (
                "All Possible Controls",
                analysis_metrics["all_possible_controls"]
            ),
            (
                "Coverage Opportunities",
                analysis_metrics["coverage_opportunities"]
            ),
            (
                "Generated Gap Rule IDs",
                analysis_metrics["gap_rule_ids"]
            ),
            (
                "Current Coverage %",
                analysis_metrics["baseline_pct"]
            ),
            (
                "Highest Potential %",
                analysis_metrics["highest_potential_pct"]
            ),
            (
                "Potential Improvement %",
                round(
                    analysis_metrics["highest_potential_pct"]
                    - analysis_metrics["baseline_pct"],
                    2
                )
            )
        ]

        for label, value in analysis_rows:

            summary_ws.cell(
                row=row_num,
                column=1
            ).value = label

            summary_ws.cell(
                row=row_num,
                column=1
            ).font = Font(bold=True)

            summary_ws.cell(
                row=row_num,
                column=2
            ).value = value

            row_num += 1

        # ------------------------------------------------------
        # Auto-size columns
        # ------------------------------------------------------

        for column in summary_ws.columns:

            max_length = 0

            column_letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                if cell.value:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            summary_ws.column_dimensions[
                column_letter
            ].width = max_length + 4
            
        # ==========================================================
        # REFERENCE COVERAGE SHEET
        # ==========================================================

        reference_ws = wb.create_sheet("Reference Coverage")

        headers = [
            "Control ID",
            "Control Title",
            "Family ID",
            "Family Name",
            "Count"
        ]

        for col_num, header in enumerate(headers, start=1):

            cell = reference_ws.cell(
                row=1,
                column=col_num
            )

            cell.value = header
            cell.font = Font(bold=True)

        row_num = 2

        for control_id, count in sorted(
            reference_counts.items(),
            key=lambda x: x[1],
            reverse=True
        ):

            family_id = (
                control_id.split("-")[0]
                if "-" in control_id
                else "PM"
            )

            family_name = NIST_FAMILIES.get(
                family_id,
                "Program Management"
            )

            reference_ws.cell(
                row_num,
                1
            ).value = control_id

            reference_ws.cell(
                row_num,
                2
            ).value = target_baseline.get(
                control_id,
                ""
            )

            reference_ws.cell(
                row_num,
                3
            ).value = family_id

            reference_ws.cell(
                row_num,
                4
            ).value = family_name

            reference_ws.cell(
                row_num,
                5
            ).value = count

            row_num += 1

        for column in reference_ws.columns:

            max_length = 0

            column_letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                if cell.value:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            reference_ws.column_dimensions[
                column_letter
            ].width = max_length + 4

        # ==========================================================
        # INACTIVE CONTROLS SHEET
        # ==========================================================

        if inactive_controls:

            inactive_ws = wb.create_sheet(
                "Inactive Controls"
            )

            headers = [
                "Control Number",
                "Description",
                "NIST Controls",
                "Count of NIST Controls"
            ]

            for col_num, header in enumerate(
                headers,
                start=1
            ):

                cell = inactive_ws.cell(
                    row=1,
                    column=col_num
                )

                cell.value = header
                cell.font = Font(
                    bold=True
                )

            row_num = 2

            for item in inactive_controls:

                inactive_ws.cell(
                    row=row_num,
                    column=1
                ).value = item.get(
                    "control_number",
                    ""
                )

                inactive_ws.cell(
                    row=row_num,
                    column=2
                ).value = clean_excel_text(
                    item.get(
                        "description",
                        ""
                    )
                )

                inactive_ws.cell(
                    row=row_num,
                    column=3
                ).value = ", ".join(
                    item.get(
                        "controls",
                        []
                    )
                )

                inactive_ws.cell(
                    row=row_num,
                    column=4
                ).value = len(
                    item.get(
                        "controls",
                        []
                    )
                )

                row_num += 1

            for column in inactive_ws.columns:

                max_length = 0

                col_letter = get_column_letter(
                    column[0].column
                )

                for cell in column:

                    if cell.value:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                inactive_ws.column_dimensions[
                    col_letter
                ].width = min(
                    max_length + 4,
                    100
                )

        # ==========================================================
        # INACTIVE COVERAGE OPPORTUNITIES
        # ==========================================================

        if inactive_opportunity_rows:

            inactive_ws = wb.create_sheet(
                "Inactive Coverage Opportunities"
            )

            headers = [
                "NIST Control",
                "NIST Title",
                "Baseline Rule ID",
                "Baseline Rule Description",
                "Coverage Category"
            ]

            for col_num, header in enumerate(
                headers,
                start=1
            ):

                cell = inactive_ws.cell(
                    row=1,
                    column=col_num
                )

                cell.value = header

                cell.font = Font(
                    bold=True
                )

                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="F4CCCC"
                )

            row_num = 2

            for item in inactive_opportunity_rows:

                for ctrl in item["controls"]:

                    inactive_ws.cell(
                        row=row_num,
                        column=1
                    ).value = ctrl

                    inactive_ws.cell(
                        row=row_num,
                        column=2
                    ).value = target_baseline.get(
                        ctrl,
                        ""
                    )

                    inactive_ws.cell(
                        row=row_num,
                        column=3
                    ).value = item["rule_id"]

                    inactive_ws.cell(
                        row=row_num,
                        column=4
                    ).value = item["description"]

                    inactive_ws.cell(
                        row=row_num,
                        column=5
                    ).value = "Recoverable"

                    row_num += 1

            for column in inactive_ws.columns:

                max_length = 0

                column_letter = get_column_letter(
                    column[0].column
                )

                for cell in column:

                    if cell.value:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                inactive_ws.column_dimensions[
                    column_letter
                ].width = min(
                    max_length + 4,
                    80
                )

        # ==========================================================
        # Controls Not In Baseline
        # ==========================================================

        baseline_type = "UNKNOWN"

        for meta in file_metadata:

            short_name = meta.get(
                "short_name",
                ""
            ).upper()

            if short_name.endswith(
                "_BASELINE"
            ):

                baseline_type = short_name.replace(
                    "_BASELINE",
                    ""
                )

                break

        if additional_control_rows:

            add_ws = wb.create_sheet(
                "Controls Not In Baseline"
            )

            headers = [
                "NIST Control",
                "NIST Title",
                "Setting",
                "Type",
                "Source Audit File",
                "Source Rule ID",
                "Source Description"
            ]

            for col_num, header in enumerate(
                headers,
                start=1
            ):

                cell = add_ws.cell(
                    row=1,
                    column=col_num
                )

                cell.value = header

                cell.font = Font(
                    bold=True
                )

                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="FFF2CC"
                )

            row_num = 2

            seen_descriptions = set()

            for item in additional_control_rows:

                description = item["description"]

                if baseline_type == "MSSRV":

                    match = re.search(
                        r"Ensure '(.*?)'\s+(?:is set to(?: include)*|to include)",
                        description*
                        re.IGNORECASE
                    )

                    if match:
                        normalized_description = match.group(1)
                    else:
                     normalized_description = description

                else:

                    normalized_description = description

                description_upper = description.upper()

                if baseline_type == "MSSRV":

                    if "(DC ONLY)" in description_upper:

                        control_type = "DC Only"

                    elif "(MS ONLY)" in description_upper:

                        control_type = "MS Only"

                    else:

                        control_type = "ALL"

                else:

                    control_type = baseline_type

                dedup_key = (
                    item["control"],
                    normalized_description
                )

                if dedup_key in seen_descriptions:
                    continue

                seen_descriptions.add(
                    dedup_key
                )
   
                add_ws.cell(
                    row=row_num,
                    column=1
                ).value = item["control"]

                add_ws.cell(
                    row=row_num,
                    column=2
                ).value = item["title"]

                add_ws.cell(
                    row=row_num,
                    column=3
                ).value = normalized_description

                add_ws.cell(
                    row=row_num,
                    column=4
                ).value = control_type

                add_ws.cell(
                    row=row_num,
                    column=5
                ).value = item["audit_file"]

                add_ws.cell(
                    row=row_num,
                    column=6
                ).value = item["rule_id"]

                add_ws.cell(
                    row=row_num,
                    column=7
                ).value = item["description"]

                row_num += 1

            for column in add_ws.columns:

                max_length = 0

                column_letter = get_column_letter(
                    column[0].column
                )

                for cell in column:

                    if cell.value:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                add_ws.column_dimensions[
                    column_letter
                ].width = min(
                    max_length + 4,
                    120
                )
                
        wb.save(output_file)

def export_gap_controls(
    controls,
    output_folder
):
    """
    Export NIST coverage opportunities
    directly to controls.txt.
    """

    import os
    import subprocess
    import sys

    output_file = os.path.join(
        output_folder,
        "controls.txt"
    )

    controls = sorted(controls)

    with open(
        output_file,
        "w",
        encoding="utf-8"
    ) as f:

        for control in controls:

            f.write(
                f"{control}\n"
            )

    print()
    print(
        f"[+] Exported {len(controls)} controls"
    )

    print(
        "[+] Controls file written to:"
    )

    print(
        f"    {output_file}"
    )

    gap_script = os.path.join(
        os.path.dirname(
            os.path.abspath(__file__)
        ),
        "Gap Controls.py"
    )

    if os.path.exists(
        gap_script
    ):

        print()
        print(
            "[+] Launching Gap Controls.py..."
        )

        subprocess.run(
            [
                sys.executable,
                gap_script,
                output_folder,
                output_file
            ]
        )



def main():

    print("=" * 60)
    print(" NIST SP 800-53 Rev 5 OSCAL Offline Gap Engine ")
    print("=" * 60)

    script_dir = os.path.dirname(
        os.path.abspath(__file__)
    )

    print("\nChoose Operational Mode:")
    print("  1. Single File Evaluation")
    print("  2. Batch Directory Evaluation (All .audit files in a folder)")

    mode = input(
    "[?] Select Mode [1-2] (Default 2): "
    ).strip() or "2"

    file_metadata = []
    audit_dir = None

    # ----------------------------------------------------------
    # Batch Mode
    # ----------------------------------------------------------

    if mode == "2":

        audit_dir = input(
            "[?] Enter the path to the folder containing your .audit files:\n--> "
        ).strip(r'"\' ')

        if not audit_dir:
            audit_dir = r"C:\PySC\Gap\MSSRV"

        if not os.path.isdir(audit_dir):

            print(
                f"[-] Error: '{audit_dir}' is not a valid directory."
            )
            return

        idx = 0

        for entry in os.listdir(audit_dir):

            if entry.lower().endswith(".audit"):

                full_path = os.path.join(
                    audit_dir,
                    entry
                )

                base_name, _ = os.path.splitext(entry)

                file_metadata.append(
                    {
                        "id": idx,
                        "path": full_path,
                        "short_name": base_name
                    }
                )

                idx += 1

        print(
            f"[+] Found {len(file_metadata)} source .audit file(s) inside the targeted directory."
        )

    # ----------------------------------------------------------
    # Single File Mode
    # ----------------------------------------------------------

    else:

        audit_path = input(
            "[?] Enter the path to your .audit file: "
        ).strip(r'"\' ')

        audit_dir = os.path.dirname(audit_path)

        base_name, _ = os.path.splitext(
            os.path.basename(audit_path)
        )

        file_metadata.append(
            {
                "id": 0,
                "path": audit_path,
                "short_name": base_name
            }
        )



    # ----------------------------------------------------------
    # Parse Audit Files
    # ----------------------------------------------------------

    results_by_file = {}
    valid_files = []
    inactive_controls = []
    baseline_inactive_controls = []

    for file_info in file_metadata:

        print(
            f"[+] Ingesting and parsing: "
            f"{file_info['short_name']}..."
        )

        print(file_info["path"])

        items = extract_audit_items(
            file_info["path"]
        )

        if "baseline" in file_info["short_name"].lower():

            print(
                f"[+] Baseline file detected: "
                f"{file_info['short_name']}"
            )

            baseline_inactive_controls = (
                extract_inactive_audit_items(
                    file_info["path"]
                )
            )

            inactive_controls = baseline_inactive_controls

            print(
                f"[+] Found {len(inactive_controls)} inactive controls"
            )

            active_refs = set()

            for item in items:

                for ctrl in item["controls"]:
                    active_refs.add(ctrl)

        valuable_inactive_controls = []
        for item in inactive_controls:

            new_refs = set(
                item["controls"]
            ) - active_refs

            if new_refs:

                item["new_controls"] = sorted(
                    new_refs
                )

                valuable_inactive_controls.append(
                    item
                )



        if items:

            file_info["items_count"] = len(items)

            valid_files.append(file_info)

            results_by_file[
                file_info["id"]
            ] = items
            
        else:

            print(
                f"[-] Skipping file due to ingestion error: "
                f"{file_info['path']}"
            )

    if not valid_files:

        print(
            "[-] Error: No valid audit files were successfully parsed."
        )

        return

    # ----------------------------------------------------------
    # Load OSCAL Catalog
    # ----------------------------------------------------------

    all_controls, control_parents = (
        load_local_oscal_baseline(
            script_dir
        )
    )

    print("\nSelect your Target Operational Baseline Profile:")
    print("  1. High-Impact Core Controls Only")
    print("  2. Moderate-Impact Core Controls Only")
    print("  3. Low-Impact Core Controls Only")
    print("  4. Full Catalog Complete Evaluation")

    baseline_choice = input(
        "[?] Choose baseline level [1-4] (Default 4): "
    ).strip() or "4"



    target_baseline = {
        k: v
        for k, v in all_controls.items()
        if "(" not in k
    }

    # ----------------------------------------------------------
    # Coverage Summary
    # ----------------------------------------------------------

    final_covered_maps = {}

    file_summaries = []

    reference_counts = {}

    # ----------------------------------------------------------
    # Determine highest potential coverage
    # ----------------------------------------------------------

    all_possible_controls = set()

    # Coverage provided by the baseline audit
    baseline_coverage_count = 0

    print("\n" + "=" * 60)
    print(" AUDIT ASSESSMENT REPORT SUMMARY")
    print("=" * 60)

    for file_info in valid_files:

        f_id = file_info["id"]

        raw_items = results_by_file[f_id]

        for item in raw_items:

            for ctrl in item["controls"]:

                normalized_ctrl = re.sub(
                    r"-0(\d)",
                    r"-\1",
                    ctrl
                )

                reference_counts[normalized_ctrl] = (
                    reference_counts.get(
                        normalized_ctrl,
                        0
                    ) + 1
                )

        covered_map = analyze_single_file(
            raw_items,
            target_baseline,
            control_parents
        )

        if file_info["short_name"] == valid_files[0]["short_name"]:

            print()

        all_possible_controls.update(
            covered_map.keys()
        )

        final_covered_maps[f_id] = covered_map

        baseline_set = set(
            target_baseline.keys()
        )

        covered_set = set(
            covered_map.keys()
        )

        covered_count = len(
            baseline_set.intersection(
                covered_set
            )
        )
        
        if "baseline" in file_info["short_name"].lower():

            baseline_coverage_count = covered_count

            baseline_covered_set = set(
                covered_map.keys()
            )

        coverage_pct = round(
            (covered_count / len(baseline_set)) * 100,
            2
        ) if baseline_set else 0

        potential_count = len(covered_map)

        potential_pct = round(
            (potential_count / len(baseline_set)) * 100,
            2
        ) if baseline_set else 0

        highest_potential_count = len(
            baseline_set.intersection(
                all_possible_controls
            )
        )

        highest_potential_pct = round(
            (
                highest_potential_count
                / len(baseline_set)
            ) * 100,
            2
        ) if baseline_set else 0

        baseline_pct = round(
            (
                baseline_coverage_count
                / len(baseline_set)
            ) * 100,
            2
        ) if baseline_set else 0

        file_summaries.append({
            "File": file_info["short_name"],
            "Checks Parsed": file_info["items_count"],
            "Individual Coverage": covered_count,
            "Individual Coverage %": coverage_pct
        })

        print(f"File: {file_info['short_name']}")
        print(f"  -> Total Checks Parsed : {file_info['items_count']}")

        print(
            f"  -> Current Coverage    : "
            f"{baseline_coverage_count} / {len(baseline_set)} "
            f"({baseline_pct}%)"
        )

        print(
            f"  -> Individual Coverage : "
            f"{covered_count} / {len(baseline_set)} "
            f"({coverage_pct}%)"
        )

        print(
            f"  -> Highest Potential   : "
            f"{highest_potential_count} / {len(baseline_set)} "
            f"({highest_potential_pct}%)"
        )
        print()

    print("=" * 60)

    highest_potential_count = len(
        baseline_set.intersection(
            all_possible_controls
        )
    )

    coverage_opportunities = (
        baseline_set.intersection(
            all_possible_controls
        )
        - baseline_covered_set
    )

    inactive_nist_controls = set()

    for item in baseline_inactive_controls:

        for ctrl in item["controls"]:

            normalized_ctrl = re.sub(
                r"-0(\d)",
                r"-\1",
                ctrl
            )

            inactive_nist_controls.add(
                normalized_ctrl
            )

    inactive_coverage_opportunities = ( 
        inactive_nist_controls & coverage_opportunities
    )

    additional_controls_not_present = (
        coverage_opportunities
        - inactive_coverage_opportunities
    )



    inactive_opportunity_rows = []

    for item in baseline_inactive_controls:

        matching_controls = []

        for ctrl in item["controls"]:

            normalized_ctrl = re.sub(
                r"-0(\d)",
                r"-\1",
                ctrl
            )

            if normalized_ctrl in inactive_coverage_opportunities:

                matching_controls.append(
                    normalized_ctrl
                )

        if matching_controls:

            inactive_opportunity_rows.append({
                "rule_id":
                    item["control_number"],

                "description":
                    item.get(
                        "description",
                        ""
                    ),

                "controls":
                    sorted(
                        set(
                            matching_controls
                        )
                    )
            })

    print(
        "\n[+] Inactive Coverage Opportunity Controls:"
    )

    for ctrl in sorted(
        inactive_coverage_opportunities
    ):

        title = target_baseline.get(
            ctrl,
            "Unknown Title"
        )

        print(
            f"    {ctrl} : {title}"
        )

    inactive_coverage_rows = []

    for ctrl in sorted(
        inactive_coverage_opportunities
    ):

        inactive_coverage_rows.append(
            {
                "Control ID": ctrl,
                "Title": target_baseline.get(
                    ctrl,
                    "Unknown Title"
                )
            }
        )

    print(
        f"[+] Inactive NIST Controls : "
        f"{len(inactive_nist_controls)}"
    )

    print(
        f"[+] Recoverable Controls : "
        f"{len(inactive_coverage_opportunities)}"
    )

    print(
        f"[+] Recoverable Rule IDs : "
        f"{len(inactive_opportunity_rows)}"
    )

    # ----------------------------------------------------------
    # Build Inactive Coverage Opportunity Details
    # ----------------------------------------------------------

    inactive_opportunity_rows = []

    for item in baseline_inactive_controls:

        matching_controls = []

        for ctrl in item["controls"]:

            normalized_ctrl = re.sub(
                r"-0(\d)",
                r"-\1",
                ctrl
            )

            if normalized_ctrl in inactive_coverage_opportunities:

                matching_controls.append(
                    normalized_ctrl
                )

        if matching_controls:

            inactive_opportunity_rows.append({
                "rule_id":
                    item.get(
                        "control_number",
                        ""
                    ),

                "description":
                    item.get(
                        "description",
                        ""
                    ),

                "controls":
                    sorted(
                        set(
                            matching_controls
                        )
                    )
            })
    
    gap_rule_ids = set()

    for ctrl in coverage_opportunities:

        for file_id, covered_map in final_covered_maps.items():

            if ctrl in covered_map:

                gap_rule_ids.update(
                    covered_map[ctrl]
                )

    print(
        f"[+] Coverage Opportunities : "
        f"{len(coverage_opportunities)} NIST controls"
    )

    print(
        f"[+] Gap Rule IDs : "
        f"{len(gap_rule_ids)} audit checks"
    )

    highest_potential_pct = round(
        (
            highest_potential_count
            / len(baseline_set)
        ) * 100,
        2
    ) if baseline_set else 0

    baseline_pct = round(
        (
            baseline_coverage_count
            / len(baseline_set)
        ) * 100,
        2
    ) if baseline_set else 0

    summary_rows = []

    for item in file_summaries:

        summary_rows.append({
            "File": item["File"],
            "Checks Parsed": item["Checks Parsed"],

            "Current Coverage":
                baseline_coverage_count,

            "Current Coverage %":
                baseline_pct,

            "Individual Coverage":
                item["Individual Coverage"],

            "Individual Coverage %":
                item["Individual Coverage %"],

            "Highest Potential":
                highest_potential_count,

            "Highest Potential %":
                highest_potential_pct
        })

    
    # ----------------------------------------------------------
    # Generate Excel
    # ----------------------------------------------------------

    timestamp = datetime.now().strftime(
        "%y%m%d%H%M"
    )

    if len(valid_files) == 1:

        output_filename = (
            f"{valid_files[0]['short_name']}"
            f"_processed_{timestamp}.xlsx"
        )

    else:

        output_filename = (
            f"NIST_Audit_Batch_Comparison_"
            f"processed_{timestamp}.xlsx"
        )

    xlsx_out = os.path.join(
        audit_dir,
        output_filename
    )

    highest_potential_count = len(
        baseline_set.intersection(
            all_possible_controls
        )
    )

    highest_potential_pct = round(
        (
            highest_potential_count
            / len(baseline_set)
        ) * 100,
        2
    ) if baseline_set else 0

    baseline_pct = round(
        (
            baseline_coverage_count
            / len(baseline_set)
        ) * 100,
        2
    ) if baseline_set else 0

    summary_rows = []

    for item in file_summaries:

        summary_rows.append({
            "File": item["File"],
            "Checks Parsed": item["Checks Parsed"],

            "Current Coverage":
                baseline_coverage_count,

            "Current Coverage %":
                baseline_pct,

            "Individual Coverage":
                item["Individual Coverage"],

            "Individual Coverage %":
                item["Individual Coverage %"],

            "Highest Potential":
                highest_potential_count,

            "Highest Potential %":
                highest_potential_pct
        })

    for control_id, count in sorted(
        reference_counts.items(),
        key=lambda x: x[1],
        reverse=True
    )[:10]:

        analysis_metrics = {
            "baseline_controls": len(baseline_covered_set),
            "all_possible_controls": len(all_possible_controls),
            "coverage_opportunities": len(coverage_opportunities),
            "gap_rule_ids": len(gap_rule_ids),
            "baseline_pct": baseline_pct,
            "highest_potential_pct": highest_potential_pct,
        }



    additional_control_rows = []

    for file_info in valid_files:

        if "baseline" in file_info["short_name"].lower():
            continue

        f_id = file_info["id"]

        raw_items = results_by_file[f_id]

        for item in raw_items:

            for ctrl in item["controls"]:

                normalized_ctrl = re.sub(
                    r"-0(\d)",
                    r"-\1",
                    ctrl
                )

                if normalized_ctrl in additional_controls_not_present:

                    additional_control_rows.append({

                        "control":
                            normalized_ctrl,

                        "title":
                            target_baseline.get(
                                normalized_ctrl,
                                ""
                            ),

                        "audit_file":
                            file_info["short_name"] + ".audit",

                        "rule_id":
                            item.get(
                                "control_number",
                                ""
                            ),

                        "description":
                            item.get(
                                "description",
                                ""
                            )
                    })

        export_excel_matrix(
            final_covered_maps,
            valid_files,
            target_baseline,
            xlsx_out,
            summary_rows,
            reference_counts,
            inactive_controls,
            analysis_metrics,
            inactive_opportunity_rows,
            additional_control_rows
        )

    print(
        f"\n[+] Comprehensive batch matrix written out to:\n"
        f" {xlsx_out}\n"
    )


def export_blue_overall_controls(xlsx_file, output_folder):
    """
    Export controls from blue-highlighted rows in the Overall column,
    write controls.txt to the selected audit folder,
    and launch Gap Controls.py.
    """

    import os
    import re
    import subprocess
    import sys
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb["NIST Gap Analysis"]

    overall_col = None

    # Find Overall column
    for col in range(1, ws.max_column + 1):

        header = str(
            ws.cell(1, col).value or ""
        ).strip()

        if header == "Overall":
            overall_col = col
            break

    if overall_col is None:
        print("[-] Could not find Overall column")
        return

    controls = set()

    # Process data rows
    for row in range(2, ws.max_row + 1):

        row_is_blue = False

        for col in range(1, ws.max_column + 1):

            cell = ws.cell(row, col)

            try:
                color = str(
                    cell.fill.fgColor.rgb or ""
                ).upper()
            except Exception:
                color = ""

        overall_value = str(
            ws.cell(row, overall_col).value or ""
        ).strip()

        if (
            not overall_value
            or overall_value.upper() == "NONE"
        ):
            continue

        overall_value = str(
            ws.cell(row, overall_col).value or ""
        ).strip()


        overall_value = str(
            ws.cell(row, overall_col).value or ""
        ).strip()

        if not overall_value:
            continue

        if overall_value.upper() == "NONE":
            continue

        for control in overall_value.split("|"):

            control = control.strip()

            # Remove embedded whitespace
            control = re.sub(r"\s+", "", control)

            if control:
                controls.add(control)

    def sort_key(value):
        try:
            return [int(x) for x in value.split(".")]
        except Exception:
            return [value]

    controls = sorted(
        controls,
        key=sort_key
    )

    output_file = os.path.join(
        output_folder,
        "controls.txt"
    )

    with open(
        output_file,
        "w",
        encoding="utf-8"
    ) as f:

        for control in controls:
            f.write(control + "\n")

    print()
    print(f"[+] Exported {len(controls)} controls")
    print("[+] Controls file written to:")
    print(f"    {output_file}")

    # ----------------------------------------------------------
    # Launch Gap Controls.py
    # ----------------------------------------------------------

    gap_script = r"C:\PySC\Gap Controls.py"

    print()
    print("[+] Launching Gap Controls.py...")

    if not os.path.exists(gap_script):
        print(f"[-] Could not find: {gap_script}")
        return

    result = subprocess.run(
        [
            sys.executable,
            gap_script,
            output_folder
        ],
        cwd=output_folder
    )

    print()
    
    # print(
    #     f"[+] Gap Controls.py exited with code "
    #     f"{result.returncode}"
    # )

if __name__ == "__main__":
    main()